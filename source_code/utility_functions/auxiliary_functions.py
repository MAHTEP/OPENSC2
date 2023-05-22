import bisect
import numpy as np
from openpyxl import load_workbook
import pandas as pd
from collections import namedtuple
from scipy import interpolate
from typing import Union
import warnings


def check_repeated_headings(input_file, sheet):
    """[summary]

    Args:
        input_file ([type]): [description]
        sheet ([type]): [description]

    Raises:
        ValueError: [description]
    """

    # Get the columns names that user can define (except the first four ones that are fixed). The variable columns is a tuple.
    columns = list(
        sheet.iter_rows(
            min_row=3,
            max_row=3,
            min_col=sheet.min_column,
            max_col=sheet.max_column,
            values_only=True,
        )
    )[0][4:]
    # Buil dictionay exploiting dict comprehension: each key as the numer of repetitions of the column as the corresponding value.
    dict_colum = {column: columns.count(column) for column in columns}
    # Raise error message
    if max(list(dict_colum.values())) > 1:
        raise ValueError(
            f"ERROR! Different objects of the same kind ({sheet['A1'].value}) can not have the same identifier.\nUser defines the following:\n{dict_colum.items()}.\nPlease check the headers in sheet {sheet.title} of file {input_file}"
        )


# End fuction check_repeated_headings.


def check_headers(cond, path_input, path_operation, sheet_input, sheet_operation):
    """[summary]

    Args:
        cond ([type]): [description]
        path_input ([type]): [description]
        path_operation ([type]): [description]
        sheet_input ([type]): [description]
        sheet_operation ([type]): [description]

    Raises:
        SyntaxError: [description]
    """
    header_input = list(
        pd.read_excel(
            path_input, sheet_name=sheet_input.title, skiprows=2, header=0, index_col=0
        ).columns
    )
    header_operation = list(
        pd.read_excel(
            path_operation,
            sheet_name=sheet_operation.title,
            skiprows=2,
            header=0,
            index_col=0,
        ).columns
    )

    for ii in range(len(header_input)):
        if header_input[ii] != header_operation[ii]:
            raise SyntaxError(
                f"ERROR in method {cond.__init__.__name__} of class {cond.__class__.__name__}: object headers in {sheet_input} of file {cond.file_input['STRUCTURE_ELEMENTS']} and in {sheet_operation} of file {cond.file_input['OPERATION']} should be the same!\nCompare the third row, column {ii+1} of those sheets.\n"
            )


# End function check_headers


def check_object_number(object, path_1, path_2, sheet_1, sheet_2):
    """[summary]

    Raises:
        ValueError: [description]
    """
    dict_method = dict(
        Simulation="conductor_instance", Conductor="conductor_components_instance"
    )
    if int(sheet_1.cell(row=1, column=2).value) != int(
        sheet_2.cell(row=1, column=2).value
    ):
        raise ValueError(
            f"ERROR in class {object.__class__.__name__} method {dict_method[object.__class__.__name__]}: number of objects defined in file {path_1} sheet {sheet_1.title} and in file {path_2} sheet {sheet_2.title} must be the same.\nPlease compare cell B1 of those sheets.\n"
        )


# End function check_object_number.

def check_costheta(object,path:str,sheet):
    """Function that checks COSTHETA value in sheet conductor_input.xlsx for instances of class StrandComponent.

    Args:
        object (StrandComponent): instance of class StrandComponent.
        path (str): path to the input file with information on the value of COSTHETA.
        sheet (_type_): workbook sheet.

    Raises:
        ValueError: if object.inputs["COSTETA"] = 0.0 or if abs(object.inputs["COSTETA"]) > 1.0
    """
    
    if np.isclose(object.inputs["COSTETA"], 0.0) or abs(object.inputs["COSTETA"]) > 1.0:
        raise ValueError(
            f"ERROR in class {object.__class__.__name__} method __init__: cos(theta) must be in the range (0,1]. Current value is {object.inputs['COSTETA']}.\nPlease, check sheet {sheet.title} in file {path}.\n"
        )

def set_diagnostic(vv, **kwargs):
    """[summary]

    Args:
        vv ([type]): [description]

    Returns:
        [type]: [description]
    """

    # Keep only the positive values, negative values means that user do not want to save spatial distribution (or time evolutions) of variables at that times (or spatial coordinate). Positive values are sorted in ascendent order.
    vv = np.sort(vv[vv >= 0])
    # add to array vv the lower boundary (lb) and the upper boundary (ub) to save data by default also at these times (or spatial coordinates)
    vv = np.concatenate((vv, np.array([kwargs["lb"], kwargs["ub"]], dtype=float)))
    return np.unique(vv)


# End function set_diagnostic.


def load_auxiliary_files(file_path, sheetname):
    """Function that load the auxiliary input file as a data frame

    Args:
        file_path (_type_): _description_
        sheetname (_type_): _description_

    Returns:
        _type_: _description_
    """
    wb = load_workbook(file_path, data_only=True)
    sheet = wb[sheetname]
    return (
        pd.read_excel(file_path, sheet_name=sheetname, header=None),
        sheet.cell(1, 1).value,
    )


# End function load_auxiliary_files


def build_interpolator(df, interpolation_kind="linear"):
    """Function that builds the interpolator for the data loaded from auxiliary input files. If data are costant in time or in space 1D interpolate.interp1d is used, if values describe a dependence in both time and space interpolate.interp1d is used.

    Args:
        df (_type_): _description_
        interpolation_kind (str, optional): _description_. Defaults to "linear".

    Returns:
        _type_: _description_
    """
    # The first column of the dataframe stores the space points used in the 
    # interpolation function starting from the second row.
    strand_space_points = df.iloc[1:,0].to_numpy(dtype=float)

    # The first row of the dataframe stores the time points used in the 
    # interpolation function starting from the second column.
    strand_time_points = df.iloc[0,1:].to_numpy(dtype=float)

    # Get the known values to be used in the interpolation, stored from the 
    # second row second column of the data frame.
    known_val = df.iloc[1:,1:].to_numpy(dtype=float)
    
    # Remark: value in position df.iloc[0,0] is not used; it is a kind of flag not well understood up to now.

    if strand_time_points.size == 1 and strand_space_points.size > 1:
        # Values are constant in time but not in space.
        known_val = known_val.reshape(known_val.shape[0])
        return (
            interpolate.interp1d(
                strand_space_points,
                known_val,
                bounds_error=False,
                fill_value=known_val[-1],
                kind=interpolation_kind,
            ),
            "space_only",
        )

    elif strand_time_points.size > 1 and strand_space_points.size == 1:
        # Values are constant in space but not in time.
        known_val = known_val.reshape(known_val.shape[1])
        return (
            interpolate.interp1d(
                strand_time_points,
                known_val,
                bounds_error=False,
                fill_value=known_val[-1],
                kind=interpolation_kind,
            ),
            "time_only",
        )

    elif strand_time_points.size > 1 and strand_space_points.size > 1:
        return (
            interpolate.interp2d(
                strand_time_points,
                strand_space_points,
                known_val,
                kind=interpolation_kind,
            ),
            "space_and_time",
        )


# End function build_interpolator


def do_interpolation(interpolator, zcoord, time_step, kind):
    """Functin that makes the interpolation of the data loaded from auxiliary input files accorting to the kind flag.

    Args:
        interpolator (_type_): _description_
        zcoord (_type_): _description_
        time_step (_type_): _description_
        kind (_type_): _description_

    Returns:
        _type_: _description_
    """

    if kind == "space_only":
        # Values are constant in time but not in space.
        return interpolator(zcoord)

    elif kind == "time_only":
        # Values are constant in space but not in time.
        return interpolator(time_step)

    elif kind == "space_and_time":
        return interpolator(time_step, zcoord).reshape(zcoord.shape)


# End function do_interpolation


def with_read_csv(fname, sheet, delimiter=";"):
    """[summary]

    Args:
        fname ([type]): [description]
        sheet ([type]): [description]
        delimiter (str, optional): [description]. Defaults to ";".

    Returns:
        [type]: [description]
    """
    # Works for a file with the spatial discretization for each conductor and
    # for a single file with a sheet for each conductor. The sheets in the file
    # must have 3 columns with the same number of rows with the following
    # content:
    # column 1: coordinate along x direction
    # column 2: coordinate along y direction
    # column 3: coordinate along z direction

    # Load the column of the file (exension dat, csv, or tsv) with the spatial
    # discretization of the conductor named conductor.name as a series; then
    # convert to numpy dropping NaN values.
    vector = (
        pd.read_csv(fname, sheet_name=sheet, sep=delimiter, squeeze=True)
        .dropna()
        .to_numpy()
    )
    return vector, vector.shape[0]

    # End function with_read_csv.


def with_read_excel(fname, sheet, delimiter=";"):
    """[summary]

    Args:
        fname ([type]): [description]
        sheet ([type]): [description]
        delimiter ([type]): [description]

    Returns:
        [type]: [description]
    """
    # Works for a file with the spatial discretization for each conductor and
    # for a single file with a sheet for each conductor. The sheets in the file
    # must have 3 columns with the same number of rows with the following
    # content:
    # column 1: coordinate along x direction
    # column 2: coordinate along y direction
    # column 3: coordinate along z direction

    # Load the column of the file .xlsv with the spatial discretization of the
    # conductor named conductor.name as a series; then convert to numpy
    # dropping NaN values.
    vector = pd.read_excel(fname, sheet_name=sheet, squeeze=True).dropna().to_numpy()
    return vector, vector.shape[0]

    # End function with_read_csv.


def get_from_xlsx(conductor, f_path, comp, flag_name, *INTIAL):

    """
    ##############################################################################
    #              Get_from_xlsx(file_path, comp)
    ##############################################################################
    #
    # Function that perform interpolation on data read from .xlsx files to
    # initialize SolidComponent python objects parameters like magnetic field,
    # magnetic field gradient, strain and external heat.
    #
    ##############################################################################
    # VARIABLE    I/O    TYPE              DESCRIPTION                      UNIT
    # --------------------------------------------------------------------------
    # f_path      I      string            path of the file to be read      -
    # comp        I      object            python object of
    #                                      class SolidComponent            -
    # value       O      np array float    vector of interpolation
    #                                      results                          many
    # flag        O      scalar integer    flag that specifies data units   -
    ##############################################################################
    # Invoked functions/methods: Read_interp_file, interpolation
    #
    ##############################################################################
    # N.B. no self parameters in input.
    # N.B. In case of bfield.xlsx and alphaB.xlsx if flag == 1 value unit are T
    # and T/m respectively, if flag == 2 value is per current value (i.e T/A
    # and T/m/A).
    # N.B. value has the same shape of zcoord and is an np array
    ##############################################################################
    #
    # Author D. Placido Polito 06/2020
    #
    ##############################################################################
    """

    # Function Read_interp_file starts here. (cdp)
    # Read data from .xlsx file
    # INTIAL[0]
    if len(INTIAL) == 0:  # interpolation in time and space
        [flag, tt, matrix, xx, sheet] = read_interp_file(
            f_path, comp, action="time_&_space"
        )
        # Perform loop on mesh nodes and perform interpolation in both time and \
        # space.
        value = interpolation(
            conductor,
            comp,
            matrix,
            tt,
            f_path,
            sheet,
            xx,
            action="time_&_space",
            Flag_name=flag_name,
        )
    else:  # interpolation in time
        [flag, tt, matrix, sheet] = read_interp_file(
            f_path, comp, INTIAL[0], action="time"
        )
        # Perform loop on mesh nodes and perform interpolation in time.
        value = interpolation(conductor, comp, matrix, tt, f_path, sheet, action="time")
    return [value, flag]


def interpolation(conductor, comp, MM, tvec, f_path, sheet, *xvec, **options):

    """
    ##############################################################################
    #              Interpolation(conductor, MM, tvec, xvec, zcoord)
    ##############################################################################
    #
    # Function that interpolate on MM @ t = time and x = zcoord (cdp)
    #
    ##############################################################################
    # VARIABLE    I/O    TYPE                  DESCRIPTION                  UNIT
    # -------------------------------------------------------------------------
    # MM          I      np 2 dimensional      data to be interpolated
    #                    array float           read from file .xlsx         T\W
    # tvec        I      np array float        time vector @ which data
    #                                          are available                s
    # xvec        I      np array float        space vector @ which data
    #                                          are available                m
    # time        I      scalar float          time @ which interpolation
    #                                          is carried out               s
    # zcoord*     I      np array float        spatial coordinate @ which
    #                                          interpolation is carried out m
    # yy          O      np array float        interpolation result
    #                                          vector                       many
    ##############################################################################
    #
    # Invoked functions/methods: none
    #
    ##############################################################################
    # N.B yy has the same shape of zcoord and is an np array
    # * zcoord is given by conductor.zcoord
    ##############################################################################
    # Author D. Placido Polito 06/2020
    #
    ##############################################################################
    """

    # Function interpolation starts here. (cdp)

    if options.get("action") == "time_&_space":
        # time-space interpolation (cdp, 07/2020)
        # Check on xvec and tvec values: first value must be >= 0 and they should \
        # be sorted from min to max value (cdp, 06/2020)
        xvec = np.array(xvec[0])
        if (xvec[0] >= 0 and all(np.diff(xvec) >= 0)) and (
            tvec[0] >= 0 and all(np.diff(tvec) >= 0)
        ):
            # Check that time @ which I perform interpolation is within tvec values, \
            # if not yy is given by spatial interpolation @ time = tvec[0] or \
            # time = tvec[-1] (cdp, 07/2020)
            yy = np.zeros(conductor.grid_features["N_nod"], dtype=float)
            lower_bound = np.min(
                np.nonzero(conductor.grid_features["zcoord"] >= xvec[0])
            )
            upper_bound = np.max(
                np.nonzero(conductor.grid_features["zcoord"] <= xvec[-1])
            )
            if options["Flag_name"] == "IQFUN":
                # Interpolation to get external flux (cdp, 11/2020)
                if comp.operations["IQFUN"] == -1:
                    # square wave in time and space (cdp, 11/2020)
                    if (
                        conductor.cond_time[-1] >= tvec[0]
                        and conductor.cond_time[-1] < tvec[-1]
                        and conductor.cond_time[-1] >= 0
                    ):
                        # search in time (cdp, 11/2020)
                        # array smart (cdp, 11/2020)
                        ii = np.max(np.nonzero(tvec <= conductor.cond_time[-1]))
                        # Square wave in time and space: at any times zcoord < lower_bound \
                        # and zcoord > upper_bound are equal to 0.0 (cdp, 10/2020)
                        if conductor.cond_time[-1] <= tvec[-2]:
                            yy[lower_bound : upper_bound + 1] = MM[0, ii]
                        else:
                            yy[lower_bound : upper_bound + 1] = MM[0, ii + 1]
                        # end if (cdp, 10/2020)
                    # end if conductor.cond_time[-1] (cdp, 11/2020)
                # end if comp.operations["IQFUN"]
            else:
                if (
                    conductor.cond_time[-1] >= tvec[0]
                    and conductor.cond_time[-1] < tvec[-1]
                    and conductor.cond_time[-1] >= 0
                ):
                    yy[0:lower_bound] = MM[0, ii] + (
                        conductor.cond_time[-1] - tvec[ii]
                    ) / (tvec[ii + 1] - tvec[ii]) * (MM[0, ii + 1] - MM[0, ii])
                    for kk in range(len(xvec) - 1):
                        lb = np.min(
                            np.nonzero(conductor.grid_features["zcoord"] >= xvec[kk])
                        )
                        ub = np.max(
                            np.nonzero(
                                conductor.grid_features["zcoord"] <= xvec[kk + 1]
                            )
                        )
                        if (ii < tvec.shape[0] - 1) and (ub <= upper_bound):
                            fx = (
                                conductor.grid_features["zcoord"][lb : ub + 1]
                                - xvec[kk]
                            ) / (xvec[kk + 1] - xvec[kk])
                            # interpolation in time (cdp)
                            yy_t_j = MM[kk, ii] + (
                                conductor.cond_time[-1] - tvec[ii]
                            ) / (tvec[ii + 1] - tvec[ii]) * (
                                MM[kk, ii + 1] - MM[kk, ii]
                            )
                            yy_t_jplus1 = MM[kk + 1, ii] + (
                                conductor.cond_time[-1] - tvec[ii]
                            ) / (tvec[ii + 1] - tvec[ii]) * (
                                MM[kk + 1, ii + 1] - MM[kk + 1, ii]
                            )
                            # interpolation in space (cdp)
                            yy[lb : ub + 1] = yy_t_j + (yy_t_jplus1 - yy_t_j) * fx
                            if ub == upper_bound:
                                yy[ub] = MM[kk + 1, ii] + (
                                    conductor.cond_time[-1] - tvec[ii]
                                ) / (tvec[ii + 1] - tvec[ii]) * (
                                    MM[kk + 1, ii + 1] - MM[kk + 1, ii]
                                )
                        elif (ii == tvec.shape[0] - 1) and (ub <= upper_bound):
                            fx = (
                                conductor.grid_features["zcoord"][lb : ub + 1]
                                - xvec[kk]
                            ) / (xvec[kk + 1] - xvec[kk])
                            yy[lb : ub + 1] = (
                                MM[kk, ii] + (MM[kk + 1, ii] - MM[kk, ii]) * fx
                            )
                            if ub == upper_bound:
                                yy[ub] = MM[kk + 1, ii]
                    yy[upper_bound + 1 : len(yy)] = MM[kk + 1, ii] + (
                        conductor.cond_time[-1] - tvec[ii]
                    ) / (tvec[ii + 1] - tvec[ii]) * (
                        MM[kk + 1, ii + 1] - MM[kk + 1, ii]
                    )
                elif conductor.cond_time[-1] < tvec[0] and conductor.cond_time[-1] >= 0:
                    # time > 0 always (cdp, 07/2020)
                    yy[0:lower_bound] = MM[0, 0]
                    for kk in range(len(xvec) - 1):
                        lb = np.min(
                            np.nonzero(conductor.grid_features["zcoord"] >= xvec[kk])
                        )
                        ub = np.max(
                            np.nonzero(
                                conductor.grid_features["zcoord"] <= xvec[kk + 1]
                            )
                        )
                        fx = (
                            conductor.grid_features["zcoord"][lb : ub + 1] - xvec[kk]
                        ) / (xvec[kk + 1] - xvec[kk])
                        # spatial interpolation @ time = tvec[0] (cdp, 07/2020)
                        yy[lb : ub + 1] = MM[kk, 0] + (MM[kk + 1, 0] - MM[kk, 0]) * fx
                    yy[upper_bound + 1 : len(yy)] = MM[kk + 1, 0]
                elif conductor.cond_time[-1] >= tvec[-1]:
                    # time > 0 always (cdp, 07/2020)
                    yy[0:lower_bound] = MM[0, -1]
                    for kk in range(len(xvec) - 1):
                        lb = np.min(
                            np.nonzero(conductor.grid_features["zcoord"] >= xvec[kk])
                        )
                        ub = np.max(
                            np.nonzero(
                                conductor.grid_features["zcoord"] <= xvec[kk + 1]
                            )
                        )
                        fx = (
                            conductor.grid_features["zcoord"][lb : ub + 1] - xvec[kk]
                        ) / (xvec[kk + 1] - xvec[kk])
                        # spatial interpolation @ time = tvec[-1] (cdp, 07/2020)
                        yy[lb : ub + 1] = (
                            MM[kk, -1] + (MM[kk + 1, -1] - MM[kk, -1]) * fx
                        )
                    yy[upper_bound + 1 : len(yy)] = MM[kk + 1, -1]
                else:
                    raise ValueError(
                        f"ERROR in {interpolation.__name__}: time must \
                        be positive scalar!\ntime = {conductor.cond_time[-1]}"
                    )
        else:
            raise ValueError(
                f"ERROR in {interpolation.__name__}: xvec and/or \
                         tvec are not sorted or have values < 0.0!\nCheck \
                         sheet {sheet} in {f_path}\n"
            )
    elif options.get("action") == "time":
        # time-only interpolation (cdp, 07/2020)
        if tvec[0] >= 0 and all(np.diff(tvec) >= 0):
            yy = np.zeros(len(MM[:, 0]))  # one value for each row on MM
            if (
                conductor.cond_time[-1] >= tvec[0]
                and conductor.cond_time[-1] < tvec[-1]
                and conductor.cond_time[-1] >= 0
            ):
                # search in time (cdp)
                # array smart (cdp)
                ii = np.max(np.nonzero(tvec <= conductor.cond_time[-1]))
                for kk in range(len(yy)):
                    yy[kk] = MM[kk, ii] + (conductor.cond_time[-1] - tvec[ii]) / (
                        tvec[ii + 1] - tvec[ii]
                    ) * (MM[kk, ii + 1] - MM[kk, ii])
            elif (
                conductor.cond_time[-1] < tvec[0] and conductor.cond_time[-1] >= 0
            ):  # time > 0 always (cdp, 07/2020)
                for kk in range(len(yy)):
                    yy[kk] = MM[kk, 0]
            elif conductor.cond_time[-1] >= tvec[-1]:
                # time > 0 always (cdp, 07/2020)
                for kk in range(len(yy)):
                    yy[kk] = MM[kk, -1]
            else:
                raise ValueError(
                    f"ERROR in {interpolation.__name__}: time must \
                        be positive scalar!\ntime = {conductor.cond_time[-1]}"
                )
        else:
            raise ValueError(
                f"ERROR in {interpolation.__name__}: tvec is not \
                         sorted or have values < 0.0 !\nCheck sheet {sheet} \
                         in {f_path}\n"
            )
    return yy


def read_interp_file(file_path, comp, *INTIAL, **options):

    """
    ##############################################################################
    #              Read_interp_file(file_path, comp)
    ##############################################################################
    #
    # Function that reads data to be interpolated from .xlsx files.
    #
    ##############################################################################
    # VARIABLE    I/O    TYPE                  DESCRIPTION                  UNIT
    # --------------------------------------------------------------------------
    # file_path   I      string                path of the file to be read  -
    # comp        I      object                python object of
    #                                          class SolidComponent    -
    # tt          O      np array float        time values vector to be
    #                                          used in interpolation        s
    # xx          O      np array float        space values vector to be
    #                                          used in interpolation        m
    # Data_matrix O      np 2 dimensional      matrix of data to be
    #                    array float           used in interpolation        many
    # flag        O      scalar integer        flag that specifies
    #                                          data units                   -
    ##############################################################################
    #
    # Invoked functions/methods: none
    #
    ##############################################################################
    # N.B. no self parameters in input.
    # N.B. In case of bfield.xlsx and alphaB.xlsx if flag == 1 data unit are T
    # and T/m respectively, if flag == 2 data are per current value (i.e T/A
    # and T/m/A)
    ##############################################################################
    #
    # Author D. Placido Polito 06/2020
    #
    ##############################################################################
    """

    # Function Read_interp_file starts here. (cdp)

    ff = load_workbook(file_path, data_only=True)
    sheet = ff[comp.identifier]
    flag = sheet.cell(1, 1).value  # read the element in cell (1,1) of .xlsx file
    # get sheet boundary to read data properly (cdp, 07/2020)
    r_first = sheet.min_row + 1  # first row to be read (cdp, 07/2020)
    r_last = sheet.max_row  # last row to be read (cdp, 07/2020)
    c_first = sheet.min_column  # first column to be read (cdp, 07/2020)
    c_last = sheet.max_column  # last column to be read (cdp, 07/2020)
    r_start = r_first + 1

    if options.get("action") == "time_&_space":
        # interpolation in both time and space (cdp, 07/2020)
        c_start = c_first + 1
        # Data matrix initialization
        Data_matrix = np.zeros((r_last - r_first, c_last - c_first))
        # for loop to read tt values form file
        for cc in sheet.iter_rows(
            min_row=r_first,
            max_row=r_first,
            min_col=c_start,
            max_col=c_last,
            values_only=True,
        ):
            # get time @ which data are given in a single shot (cdp, 07/2020)
            tt = np.array(cc)
        # for loop to read xx values form file
        for rr in sheet.iter_cols(
            min_row=r_start,
            max_row=r_last,
            min_col=c_first,
            max_col=c_first,
            values_only=True,
        ):
            # get spatial coordinates @ which data are given in a single shot \
            # (cdp, 07/2020)
            xx = np.array(rr)
        # Data_matrix construction
        for rr in range(r_start, r_last + 1):
            for cc in sheet.iter_rows(
                min_row=rr,
                max_row=rr,
                min_col=c_start,
                max_col=c_last,
                values_only=True,
            ):
                Data_matrix[rr - r_start, :] = np.array(cc)
        return [flag, tt, Data_matrix, xx, sheet]
    elif options.get("action") == "time":
        # interpolation in time, only for flow variables at channel inlet \
        # (cdp, 07/2020)
        c_start = c_first + 2
        # for loop to read tt values form file
        for cc in sheet.iter_rows(
            min_row=r_first,
            max_row=r_first,
            min_col=c_start,
            max_col=c_last,
            values_only=True,
        ):
            # get time @ which data are given in a single shot (cdp, 07/2020)
            tt = np.array(cc)
        # read input data names
        for rr in sheet.iter_cols(
            min_row=r_start,
            max_row=r_last,
            min_col=c_first,
            max_col=c_first,
            values_only=True,
        ):
            val_name = np.array(rr, dtype=str)
        if INTIAL[0] == -1:
            # get TEMINL, TEMOUT, PREINL and PREOUT rows (cdp, 07/2020)
            req_variable = np.array(["TEMINL", "TEMOUT", "PREINL", "PREOUT"], dtype=str)
            row_ind = np.zeros(req_variable.shape, dtype=int)
            for ii in range(len(req_variable)):
                row_ind[ii] = int(np.nonzero(val_name == req_variable[ii])[0] + r_start)
        elif INTIAL[0] == -2:
            # get TEMINL, TEMOUT, PREINL and MDTIN rows (cdp, 07/2020)
            req_variable = np.array(["TEMINL", "TEMOUT", "PREINL", "MDTIN"], dtype=str)
            row_ind = np.zeros(req_variable.shape, dtype=int)
            for ii in range(len(req_variable)):
                row_ind[ii] = int(np.nonzero(val_name == req_variable[ii])[0] + r_start)
        elif INTIAL[0] == -5:
            # get TEMINL, TEMOUT, PREOUT ans MDTIN rows (cdp, 07/2020)
            req_variable = np.array(["TEMINL", "TEMOUT", "PREOUT", "MDTIN"], dtype=str)
            row_ind = np.zeros(req_variable.shape, dtype=int)
            for ii in range(len(req_variable)):
                row_ind[ii] = int(np.nonzero(val_name == req_variable[ii])[0] + r_start)

        # create Data_matrix by row: each row correspond to a different input \
        # values according to the above if else series. (cdp, 07/2020)

        # Data_matrix initialization (cdp, 07/2020)
        Data_matrix = np.zeros((row_ind.shape[0], tt.shape[0]))
        for ii in range(len(row_ind)):
            for values in sheet.iter_rows(
                min_row=row_ind[ii],
                max_row=row_ind[ii],
                min_col=c_start,
                max_col=c_last,
                values_only=True,
            ):
                Data_matrix[ii, :] = np.array(values)
        return [flag, tt, Data_matrix, sheet]

def filter_component(iterable:Union[list,tuple],item:object)->tuple:
    """Function that returns a new collection of tuple (index,object) with all the objects except the one in item; index are the locations of the objects in iterable. The returned collection (tuple) preserves the order of the items in iterable.

    Args:
        iterable (Union[list,tuple]): collection of object to be filtered.
        item (object): element to be removed from iterable.

    Raises:
        TypeError: if iterable is not a list or a tuple.

    Returns:
        tuple: collection of tuple (index,object). All objects from iterable exctep the element passed as item are collected. Index stores the original location in iterable; items order is preserved.
    """

    # Namedtuple constructor.
    Obj_info = namedtuple("Obj_info",["idx","obj"])
    # Sanity check.
    if isinstance(iterable, (list,tuple)):
        # Return a new collection of (index,obj) without item preserving the 
        # order.
        return tuple(Obj_info(index,obj) for index,obj in enumerate(iterable) if obj != item)
    else:
        raise TypeError(f"Iterable should be of type list or tuple; current type is {type(iterable)}")