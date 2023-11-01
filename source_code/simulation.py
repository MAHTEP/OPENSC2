from decimal import Decimal
from openpyxl import load_workbook
import numpy as np
import pandas as pd
import os
from stat import S_IREAD, S_IRGRP, S_IROTH, S_IWUSR
from typing import Union
import warnings

import cProfile, pstats
from pstats import SortKey
from line_profiler import LineProfiler

from conductor import Conductor
from conductor_flags import IOP_NOT_DEFINED
from environment import Environment
from utility_functions.auxiliary_functions import (
    check_repeated_headings,
    check_object_number,
    with_read_csv,
    with_read_excel,
)
from utility_functions.transient_solution_functions import get_time_step, step
from utility_functions.output import (
    save_simulation_space,
    reorganize_spatial_distribution,
    save_simulation_time,
    save_properties,
)
from utility_functions.plots import (
    plot_properties,
    make_plots,
    create_real_time_plots,
    update_real_time_plots,
)
from simulation_global_info import MLT_DEFAULT_VALUE

class Simulation:

    # Current working directory
    CWD = os.getcwd()

    def __init__(self, base_path):

        # Current working directory: SCMagnetCode (cdp, 10/2020)
        # self.cwd = os.getcwd()
        # Ask User the name of the cable. (cdp, 10/2020)
        self.dict_path = dict(
            Current_work_dir=self.CWD,
            Results_dir=os.path.join(self.CWD, "..", "Simulations_results"),
        )
        # Create directory Simulations_results if it does not exist yet
        os.makedirs(self.dict_path["Results_dir"], exist_ok=True)
        self.basePath = base_path
        # loop inside self.basePath (cdp, 10/2020)
        input_files = os.listdir(self.basePath)
        for f_name in input_files:
            if "transitory_input" in f_name:
                self.starter_file = f_name
        
        # Path of the starting file (master input file)
        self.starter_file_path = os.path.join(self.basePath, self.starter_file)
    
        # Load input file transitory_input.xlsx and convert to a dictionary.
        self.transient_input = pd.read_excel(
            self.starter_file_path,
            sheet_name="TRANSIENT",
            skiprows=1,
            header=0,
            index_col=0,
            usecols=["Variable name", "Value"],
        )["Value"].to_dict()
        self.flag_start = False
        # get the order of maginitude of the minimum time step to make proper 
        # rounds to when saving data and figures of solution spatial 
        # distribution at default or User defined times.
        self.__count_sigfigs(self.transient_input["STPMIN"])
        self.list_of_Conductors = list()
        # Define the environment object as an attribute of the simulation
        self.environment = Environment(
            os.path.join(self.basePath, self.transient_input["ENVIRONMENT"])
        )
        # Define dictionary with default values for functions with_read_csv and with_read_excel.
        self.default_vals = dict(dat="\t", csv=";", tsv="\t", xlsx=0)
        # Define dictionaty with the pointers to the function to be called according to the file extension if some quantities should be loaded from an auxiliary input file.
        self.func_open_aux = dict(
            dat=with_read_csv,
            csv=with_read_csv,
            tsv=with_read_csv,
            xlsx=with_read_excel,
        )

        self.fluid_prop_aliases = dict(
            isobaric_expansion_coefficient="isobaric_expansion_coefficient",
            isothermal_compressibility="isothermal_compressibility",  # 1/Pa
            Prandtl="Prandtl",
            total_density="Dmass",
            total_dynamic_viscosity="viscosity",
            total_enthalpy="Hmass",
            total_isobaric_specific_heat="Cpmass",
            total_isochoric_specific_heat="Cvmass",
            total_speed_of_sound="speed_of_sound",
            total_thermal_conductivity="conductivity",
        )

        # Apply default values to keys MLT_UPPER and MLT_LOWER if user 
        # specifies value none in input file transitory_input.xlsx (i.e. use 
        # the default value).
        self.transient_input.update(
            {
                key: val for key,val in MLT_DEFAULT_VALUE.values() 
                if (
                    isinstance(self.transient_input[key],str) 
                    and self.transient_input[key].lower() == "none"
                )
            }
        )

    # end method __init__ (cdp, 06/2020)

    def __count_sigfigs(self,num:Union[int,float]):
        """Private method that counts the number of significant digits.
        N.B: quite general but does not work for all cases; see 
        https://stackoverflow.com/questions/8101353/counting-significant-figures-in-python

        Args:
            num Union[int,float]: number of which find the significant digits
        """

        numstr = str(num)
        decimal_norm = Decimal(numstr).normalize()
        if str(decimal_norm) == "0":
            # Deal with cases "0", "00", "00...0"
            self.n_digit_time = 0
        else:
            if 0.0 < num < 1.0:
                # Not the correct number of significant digits but rather the 
                # number of digits to be displayed.
                # E.g. 0.01 gives n_digit_time = 2 rather than n_digit_time = 1 
                # (which is the correct one).
                self.n_digit_time = abs(decimal_norm.as_tuple().exponent)
            else:
                # Deal with cases specified in https://stackoverflow.com/questions/8101353/counting-significant-figures-in-python
                self.n_digit_time = len(Decimal(numstr).as_tuple().digits)

    def conductor_instance(self):
        # Load spread-sheet conductor_definition.xlsx
        conductor_defn = os.path.join(self.basePath, self.transient_input["MAGNET"])
        dict_file = dict()
        # Loop to get the correct name of file conductor_grid.xlsx
        for f_name in os.listdir(self.basePath):
            if "grid" in f_name:
                dict_file["GRID"] = os.path.join(self.basePath, f_name)
            elif "diagnostic" in f_name:
                dict_file["Spatial_distribution"] = os.path.join(self.basePath, f_name)
                dict_file["Time_evolution"] = os.path.join(self.basePath, f_name)
            # End if "grid".
        # End for f_name.
        # Load workbook conductor_definition.xlsx.
        conductorsSpec = load_workbook(conductor_defn, data_only=True)
        list_conductor_sheet = [
            conductorsSpec["CONDUCTOR_files"],
            conductorsSpec["CONDUCTOR_input"],
            conductorsSpec["CONDUCTOR_operation"],
        ]

        # Load the workbook in file conductor_grid.xlsx.
        gridCond = load_workbook(dict_file["GRID"], data_only=True)
        # Load the workbook in file conductor_diagnostic.xlsx.
        wb_diagno = load_workbook(dict_file["Spatial_distribution"], data_only=True)
        # Loop to check if user define conductors with the same identifier.
        for sheet in list_conductor_sheet:
            check_repeated_headings(conductor_defn, sheet)
        # End for sheet.
        # Check if user define conductors with the same identifier in file conductor_grid.xlsx.
        check_repeated_headings(dict_file["GRID"], gridCond["GRID"])
        for sheet in wb_diagno:
            # Check if user define conductors with the same identifier in file conductor_diagnostic.xlsx.
            check_repeated_headings(dict_file["Spatial_distribution"], sheet)
        # End for sheet.
        # Check if the number of defined conductors is consitent in file conductor_definition.xlsx and in file conductor_gri.xlsx.

        for cond_sheet in list_conductor_sheet:
            for sheet in [
                gridCond["GRID"],
                wb_diagno["Spatial_distribution"],
                wb_diagno["Time_evolution"],
            ]:
                check_object_number(
                    self,
                    self.transient_input["MAGNET"],
                    dict_file[sheet.title],
                    cond_sheet,
                    sheet,
                )
            # End for sheet.
        # End for cond_sheet.

        self.numObj = int(list_conductor_sheet[0].cell(row=1, column=2).value)
        # LOAD MAIN CONDUCTORS PARAMETERS
        for ii in range(1, 1 + self.numObj):
            conductor = Conductor(self, list_conductor_sheet, ii)
            self.list_of_Conductors.append(conductor)
        # end for ii (cdp, 12/2020)
        self.contactBetweenConductors = pd.read_excel(
            conductor_defn, sheet_name="CONDUCTOR_coupling", header=0, index_col=0
        )

        # Loop to create the attributes required to make the real time plots (shortly rtp).
        for conductor in self.list_of_Conductors:
            create_real_time_plots(self, conductor)
        # End for conductor.

    # end method Conductor_instance

    def conductor_initialization(self, gui):
        for cond in self.list_of_Conductors:
            # ** INITIALIZATION **
            # s time @ which simulation is started (cdp, 07/2020)
            self.simulation_time = [0.0]
            self.num_step = 0
            cond.initialization(self, gui)
            # Use electric method only if needed, i.e., user specifies a 
            # current.
            if cond.inputs["I0_OP_MODE"] != IOP_NOT_DEFINED:
                # Call to electric_electric method allows to define, 
                # initialize, solve and reorganize the electric problem.
                cond.eval_total_operating_current()
                cond.electric_method()
            else:
                # Quick fix to the following silent bug: solid components 
                # thermophysical, electromagnetic and critical properties 
                # do not update at each time step if flag 
                # conductor.inputs ["I0_OP_MODE"] == IOP_NOT_DEFINED. 
                # This will not cause the simulation to stop, but will give 
                # reasonable but wrong output. Since all solid component 
                # properties are updated in method operating_conditions_em of 
                # class Conductor, a quick fix is to explicitly call this 
                # method when flag 
                # conductor.inputs["I0_OP_MODE"] == IOP_NOT_DEFINED and let 
                # method electric_method of class Conductor call it in all the 
                # other cases.
                # A better fix involves a complete refactoring of at least 
                # methods operating_conditions_th and operating_conditions_em 
                # of class Conductor and should be done later.
                cond.operating_conditions_em()

            # plot conductor initialization spatial distribution (cdp, 12/2020)
            plot_properties(self, cond)
            save_simulation_time(self, cond)
            # ** END INITIALIZATION **
        # end for cond (cdp, 12/202)
        # dictionary declaration (cdp,07/2020)
        self.dict_qsource = dict()
        if self.numObj == 1:
            # There is only 1 Conductor object, exploit cond instantiated above.
            # The number of columns is equal to the number of SolidComponent 
            # equations to exploit the new function build_svec in 
            # utility_functions/step_matrix_construction.py. Read the docstring 
            # for further details.
            self.dict_qsource[cond.identifier] = np.zeros(
                (
                    cond.grid_features["N_nod"],
                    cond.dict_N_equation["SolidComponent"],
                )
            )
        else:
            # more than one Conductor object (cdp,07/2020)
            for rr in range(self.numObj):
                cond_r = self.list_of_Conductors[rr]
                if all(self.contactBetweenConductors.iloc[rr, :]) == 0:
                    # There is not contact between cond_r and all the others.
                    # Consider all the columns in order to not miss the info on 
                    # last raw (otherwise the last conductor will not be added 
                    # as key of the dictionary).
                    # The number of columns is equal to the number of 
                    # SolidComponent equations to exploit the new function 
                    # build_svec in 
                    # utility_functions/step_matrix_construction.py. Read the 
                    # docstring for further details.
                    self.dict_qsource[cond_r.identifier] = np.zeros(
                        (
                            cond_r.grid_features["N_nod"],
                            cond_r.dict_N_equation["SolidComponent"],
                        )
                    )
                else:
                    for cc in range(rr + 1, self.numObj):
                        cond_c = self.list_of_Conductors[cc]
                        if self.contactBetweenConductors.iat[rr, cc] == 1:
                            # there is contact between cond_r and cond_c (cdp,07/2020)
                            # la colonna non nulla è solo quella del jacket esterno \
                            # eventualmente in contatto con un altro jacket esterno di un \
                            # altro conduttore. Nel caso il conduttore sia costituito da un \
                            # solo jacket la matrice degenera in un vettore colonna. \
                            # (cdp, 08/2020)
                            # N.B. controllare anche nella chiamata a step come passare self.dict_qsource.
                            raise ValueError(
                                "Multi conductors with thermal contact: not yet managed the situation, need to build dict_qsource properly!!"
                            )
                        else:
                            # There is not contact between cond_r and cond_c (cdp,07/2020)
                            # Proposta di soluzione ma va studiata decisamente meglio!
                            # N.B. controllare anche nella chiamata a step come passare self.dict_qsource.
                            # The number of columns is equal to the number of 
                            # SolidComponent equations to exploit the new 
                            # function build_svec in 
                            # utility_functions/step_matrix_construction.py. 
                            # Read the docstring for further details.
                            self.dict_qsource[
                                f"{cond_r.identifier}_{cond_c.identifier}"
                            ] = np.zeros(
                                (
                                    cond_r.grid_features["N_nod"],
                                    cond_r.dict_N_equation["SolidComponent"],
                                )
                            )
                            raise ValueError(
                                "Multi conductors with thermal contact: not yet managed the situation, need to build dict_qsource properly!!"
                            )
                        # End if self.contactBetweenConductors[rr, cc].
                    # End for cc.
                # End if all(self.contactBetweenConductors[rr, rr + 1:])
                # End for rr.
        # end if numObj

    # end method Conductor_initialization

    def conductor_solution(self, gui):
        # ** TRANSIENT SOLUTION **
        num_step_store = 100
        count_store = 1
        stoptime = 0  # flag to stop simulation if some problems (like quench) \
        # arise (cdp, 07/2020)
        # Time step initialization (cdp, 08/2020)
        # time_step = transient_input["STPMIN"]
        # Search if User asks to save the solution spatial distribution at TEND \
        # (cdp, 10/2020)
        # loop on conductors (cdp, 10/2020)
        for conductor in self.list_of_Conductors:
            # Compute radiative heat exchanged between jackets.
            conductor.compute_radiative_heat_exhange_jk()
            # Compute radiative heat exchanged outer jacket and environment.
            conductor.compute_heat_exchange_jk_env(self.environment)
            # get the times at which users saves the solution spatial distribution \
            # (cdp, 10/2020)
            # list_values = list(conductor.dict_Space_save.values())
            # Save of the solution spatial distribution at 0.0 s (cdp, 12/2020)
            save_simulation_space(
                conductor,
                self.dict_path[
                    f"Output_Spatial_distribution_{conductor.identifier}_dir"
                ],
                abs(self.n_digit_time),
            )
        # end for ii (cdp, 10/2020)
        # while loop to solve transient at each timestep (cdp, 07/2020)
        while (
            self.simulation_time[-1]
            < self.transient_input["TEND"] - 1e-5 * self.transient_input["STPMIN"]
            and stoptime == 0
        ):
            self.num_step = self.num_step + 1
            time_step = np.zeros(self.numObj)
            for ii, conductor in enumerate(self.list_of_Conductors):
                
                time_step[ii] = conductor.time_step
                # Increase time (cdp, 08/2020)
                conductor.cond_time.append(
                    conductor.cond_time[-1] + conductor.time_step
                )
                # update time step (cdp, 08/2020)
                conductor.cond_num_step = conductor.cond_num_step + 1
                # before calling Conductor method initialization adapt mesh if \
                # necessary as foreseen by ITYMSH. To do later (cdp, 07/2020)
                # se ho nuova griglia calcolare coefficenti, temperature, pressioni, \
                # parametri adimensionati sfruttando np.interp (in fortrand è adaptm)
            # End for conductor (cdp, 08/2020)
            # Evaluate simulation time and simulation time step (cdp, 08/2020)
            self.simulation_time_step = np.max(time_step)
            self.simulation_time.append(
                self.simulation_time[-1] + self.simulation_time_step
            )
            print(
                f"Simulation time: {self.simulation_time[-1]:.{self.n_digit_time}f} s; {self.simulation_time[-1]/self.transient_input['TEND']*100:5.2f} %"
            )
            for conductor in self.list_of_Conductors:
                
                # Use electric method only if needed, i.e., user specifies a 
                # current.
                if conductor.inputs["I0_OP_MODE"] != IOP_NOT_DEFINED:
                    # Call to electric_electric method allows to define, 
                    # initialize, solve and reorganize the electric problem.
                    conductor.electric_method()
                else:
                    # Quick fix to the following silent bug: solid components 
                    # thermophysical, electromagnetic and critical properties 
                    # do not update at each time step if flag 
                    # conductor.inputs ["I0_OP_MODE"] == IOP_NOT_DEFINED. 
                    # This will not cause the simulation to stop, but will give 
                    # reasonable but wrong output. Since all solid component 
                    # properties are updated in method operating_conditions_em 
                    # of class Conductor, a quick fix is to explicitly call 
                    # this method when flag 
                    # conductor.inputs["I0_OP_MODE"] == IOP_NOT_DEFINED and let 
                    # method electric_method of class Conductor call it in all 
                    # the other cases.
                    # A better fix involves a complete refactoring of at least 
                    # methods operating_conditions_th and 
                    # operating_conditions_em of class Conductor and should be 
                    # done later.
                    conductor.operating_conditions_em()
                    
                # Evaluate thermal hydraulic properties and quantities in Gauss 
                # points, method __eval_Gauss_point_th is invoked inside method 
                # operating_conditions_th. Method operating_conditions_th is 
                # called at each time step before function step because the 
                # method for the integration in time is implicit.
                conductor.operating_conditions_th(self)
                
                conductor.build_heat_source(self)
                # call step to solve the problem @ new timestep (cdp, 07/2020)
                step(
                    conductor,
                    self.environment,
                    self.dict_qsource[conductor.identifier],
                    self.num_step,
                )

                # Call function get_time_step to compute the new time step used 
                # to compute the next value of the time at which the next 
                # thermal-hydraulic solution will be evaluated. Moving the call 
                # to function get_time_step after the call to funciton step 
                # allows to avoid checking if num_step > 1 for each time step 
                # and for each conductor. The initial value of the time step 
                # for each conductor is defined at conductor instance and it is 
                # equal to the minimum time step.
                conductor.time_step = get_time_step(
                    conductor,
                    self.transient_input,
                )

                # Loop on FluidComponent (cdp, 10/2020)
                for fluid_comp in conductor.inventory["FluidComponent"].collection:
                    # compute density and mass flow rate in nodal points with the
                    # updated FluidComponent temperature and velocity (nodal = True by default)
                    fluid_comp.coolant._compute_density_and_mass_flow_rates_nodal_gauss(
                        conductor
                    )
                    # Enthalpy balance: sum((mdot*w)_out - (mdot*w)_inl), used to check \
                    # the imposition of SolidComponent temperature initial spatial \
                    # distribution (cdp, 12/2020)
                    conductor.enthalpy_balance = (
                        conductor.enthalpy_balance
                        + conductor.time_step
                        * (
                            fluid_comp.coolant.dict_node_pt["mass_flow_rate"][-1]
                            * fluid_comp.coolant.dict_node_pt["total_enthalpy"][-1]
                            - fluid_comp.coolant.dict_node_pt["mass_flow_rate"][0]
                            * fluid_comp.coolant.dict_node_pt["total_enthalpy"][0]
                        )
                    )
                    conductor.enthalpy_out = (
                        conductor.enthalpy_out
                        + conductor.time_step
                        * fluid_comp.coolant.dict_node_pt["mass_flow_rate"][-1]
                        * fluid_comp.coolant.dict_node_pt["total_enthalpy"][-1]
                    )
                    conductor.enthalpy_inl = (
                        conductor.enthalpy_inl
                        + conductor.time_step
                        * fluid_comp.coolant.dict_node_pt["mass_flow_rate"][0]
                        * fluid_comp.coolant.dict_node_pt["total_enthalpy"][0]
                    )
                # end for fluid_comp (cdp, 10/2020)

                # Compute radiative heat exchanged between jackets.
                conductor.compute_radiative_heat_exhange_jk()
                # Compute radiative heat exchanged outer jacket and environment.
                conductor.compute_heat_exchange_jk_env(self.environment)

                update_real_time_plots(conductor)

                if self.num_step == num_step_store * count_store:
                    # Update counter to store the state of the simulation, still to come \
                    # (cdp, 08/2020)
                    count_store = count_store + 1
                if (
                    conductor.i_save < len(conductor.Space_save) - 1
                    and abs(
                        conductor.cond_time[-1] - conductor.Space_save[conductor.i_save]
                    )
                    < conductor.time_step / 2.0
                ):
                    # save simulation spatial distribution at user defined time steps \
                    # (cdp, 08/2020)
                    save_simulation_space(
                        conductor,
                        self.dict_path[
                            f"Output_Spatial_distribution_{conductor.identifier}_dir"
                        ],
                        abs(self.n_digit_time),
                    )
                # end if isave
                # Save variables time evolution at given spatial coordinates \
                # (cdp, 08/2020)
                save_simulation_time(self, conductor)
                # call sensor to plot results at any time the user asks (cdp, 07/2020)
            # End for conductor (cdp, 07/2020)
        # end while (cdp, 07/2020)
        print("End simulation called " + self.transient_input["SIMULATION"] + "\n")

    # end method Conductor_solution (cdp, 09/2020)

    def conductor_post_processing(self):
        # loop to save the norm of the solution at the end of the transient for 
        # each conductor, usefull to make space convergence
        # t_end = np.array([self.transient_input["TEND"]])
        for cond in self.list_of_Conductors:
            cond.post_processing(self)

            # save simulation spatial distribution at TEND.
            save_simulation_space(
                cond,
                self.dict_path[f"Output_Spatial_distribution_{cond.identifier}_dir"],
                abs(self.n_digit_time),
            )
            # Call function Save_properties to save the conductor final 
            # solution.
            save_properties(
                cond, self.dict_path[f"Output_Solution_{cond.identifier}_dir"]
            )
            print("Saved final solution\n")

            reorganize_spatial_distribution(
                cond,
                self.dict_path[f"Output_Spatial_distribution_{cond.identifier}_dir"],
                self.n_digit_time,
            )
            # Plot conductor solution spatial distribution (cdp, 12/2020)
            plot_properties(self, cond, what="solution")
        # end for cond (cdp, 12/2020)
        # Call function Make_plots to make plot of spatial distribution and time \
        # evolution (cdp, 11/2020)
        make_plots(self, kind="Space_distr")
        make_plots(self, kind="Time_evol")

    # end method Conductor_post_processing (cdp, 09/2020)

    def _make_directories(self, list_key_val, exist_ok=False):
        """[summary]

        Args:
            list_key_val ([type]): [description]
            exist_ok (bool, optional): [description]. Defaults to False.
        """
        # Loop to create the folders.
        for ii in range(len(list_key_val)):
            # Create the folders in list_key_val[ii][0].
            os.makedirs(self.dict_path[list_key_val[ii][0]], exist_ok=exist_ok)
        # End for ii.

    # End method _make_directories.

    def _space_convergence_paths(self, list_folder):
        """Private method that builds the paths for the space convergence analysis invocking the private method _update_dict_path_and_make_dirs."""
        # list_folder = ["output", "figures"]
        str_dir = (
            "TEND_"
            + f"{self.transient_input['TEND']}_"
            + "STPMIN_"
            + f"{self.transient_input['STPMIN']}"
        )
        # Build list_key_val exploiting list comprehension. List of tuples: index [0] is the key of the dictionary, index [1] is the corresponding value that is the path to Output or Figures sub directories.
        list_key_val = [
            (
                f"Space_conv_{folder}_dir",
                os.path.join(
                    self.dict_path["Sub_dir"],
                    "Space_convergence",
                    str_dir,
                    folder.capitalize(),
                ),
            )
            for folder in list_folder
        ]
        # Update the dictionary self.dict_path with keys and values from the dictionary comprehension.
        self.dict_path.update(
            {
                list_key_val[ii][0]: list_key_val[ii][1]
                for ii in range(len(list_key_val))
            }
        )
        # Make the directories invoking method _make_directories.
        self._make_directories(list_key_val, exist_ok=True)

    # End method _space_convergence_paths.

    def _time_convergence_paths(self, list_folder):
        """Private method that builds the paths for the time convergence analysis invocking the private method _update_dict_path_and_make_dirs. The value of NELEMS is the one of the first defined conductor."""
        # list_folder = ["output", "figures"]
        str_dir = (
            "TEND_"
            + f"{self.transient_input['TEND']}_"
            + "NELEMS_"
            + f"{self.list_of_Conductors[0].grid_input['NELEMS']}"
        )
        # Build list_key_val exploiting list comprehension. List of tuples: index [0] is the key of the dictionary, index [1] is the corresponding value that is the path to Output or Figures sub directories.
        list_key_val = [
            (
                f"Time_conv_{folder}_dir",
                os.path.join(
                    self.dict_path["Sub_dir"],
                    "Time_convergence",
                    str_dir,
                    folder.capitalize(),
                ),
            )
            for folder in list_folder
        ]
        # Update the dictionary self.dict_path with keys and values from the dictionary comprehension.
        self.dict_path.update(
            {
                list_key_val[ii][0]: list_key_val[ii][1]
                for ii in range(len(list_key_val))
            }
        )
        # Make the directories invoking method _make_directories.
        self._make_directories(list_key_val, exist_ok=True)

        # Path of the directory to save the comparison of the time convergence \
        # with the same method, comparison is by NELEMS (cdp, 11/2020) [seems to not be used, I do not remember the aim of this!! 08/07/2020]
        # self.dict_path["Time_conv_comp_METHOD"] = os.path.join(\
        # 																		self.dict_path["Main_dir"], "METHOD")
        # Path of the directory to save the comparison of the time convergence \
        # with the same NELEMS, comparison is by METHODS (cdp, 11/2020) [seems to not be used, I do not remember the aim of this!! 08/07/2020]
        # self.dict_path["Time_conv_comp_NELEMS"] = os.path.join(\
        # 																		self.dict_path["Main_dir"], "NELEMS")

    # End method _time_convergence_paths.

    def _subfolders_paths(self, list_folder, list_f_names, dict_make, dict_benchmark):
        """[summary]

        Args:
            list_folder ([type]): [description]
            list_f_names ([type]): [description]
            dict_make ([type]): [description]
            dict_benchmark ([type]): [description]
        """
        # Loop to create sub folders initialization, Spatial_distribution, Time_evolution and Benchmark in Output and Figures directories; each folder in f_names_list, will contain folder conductor.identifier.
        for f_name in list_f_names:
            for conductor in self.list_of_Conductors:
                # Build list_key_val exploiting list comprehension. List of tuples: index [0] is the key of the dictionary, index [1] is the corresponding value that is the path to Output or Figures sub directories.
                list_key_val = [
                    (
                        f"{folder.capitalize()}_{f_name}_{conductor.identifier}_dir",
                        os.path.join(
                            self.dict_path["Sub_dir"],
                            self.transient_input["SIMULATION"],
                            folder.capitalize(),
                            f_name,
                            conductor.identifier,
                        ),
                    )
                    for folder in list_folder
                ]
                # Update the dictionary self.dict_path with keys and values from dictionary comprehension, which may be either a mapping or an iterable of key/value pairs. The values of dictionary comprehension take priority when self.dict_path0 and other share keys.
                self.dict_path.update(
                    {
                        list_key_val[ii][0]: list_key_val[ii][1]
                        for ii in range(len(list_key_val))
                    }
                )
                # Invocke method self._make_warnings if path exists, self._make_directories if path does not exist.
                dict_make[os.path.exists(self.dict_path[list_key_val[0][0]])](
                    list_key_val
                )
                # Create benchmark directory if f_name is "Spatial_distribution" or "Time_evolution"
                dict_benchmark[f_name](conductor, list_folder, f_name)
            # End for conductor.
        # End for f_name.

    # End method _subfolders_paths.

    def _make_warnings(self, list_key_val):
        """[summary]

        Args:
            list_key_val ([type]): [description]
        """
        # Da sistemare nella GUI!
        warnings.warn(
            f"Directories\n{self.dict_path[list_key_val[0][0]]}\n{self.dict_path[list_key_val[1][0]]} already exist.\nProbably you have already performed a simulation with the same input data.\nPlease check and confirm if you want to continue with the simulation or not."
        )

    # End method _make_warnings.

    def _benchmark_paths(self, conductor, list_folder, f_name):
        """[summary]

        Args:
            conductor ([type]): [description]
            list_folder ([type]): [description]
            f_name ([type]): [description]
        """
        # Build list_key_val exploiting list comprehension. List of tuples: index [0] is the key of the dictionary, index [1] is the corresponding value that is the path to Output or Figures sub directories.
        list_key_val = [
            (
                f"{folder.capitalize()}_{f_name}_benchmark_{conductor.identifier}_dir",
                os.path.join(
                    self.dict_path["Sub_dir"],
                    self.transient_input["SIMULATION"],
                    folder.capitalize(),
                    f_name,
                    "Benchmark",
                    conductor.identifier,
                ),
            )
            for folder in list_folder
        ]
        # Update the dictionary self.dict_path with keys and values from dictionary comprehension.
        self.dict_path.update(
            {
                list_key_val[ii][0]: list_key_val[ii][1]
                for ii in range(len(list_key_val))
            }
        )
        # Make the directories invoking method _make_directories.
        self._make_directories(list_key_val, exist_ok=True)
        # End method _benchmark_paths.

    def _do_nothing(self, conductor, list_folder, f_name):
        # Do nothing method, indroduced to use the dictionary dict_benchmark when creating folders.
        pass

    def simulation_folders_manager(self):
        """[summary]"""
        # Define the dummy dictionary with the integration methods (da sistemare, non utilizzare i numeri come keys)
        dict_int_method = dict(BE="BE", CN="CN", AM4="AM4")
        # Update dictionary self.dict_path
        self.dict_path["Sub_dir"] = os.path.join(
            self.dict_path["Main_dir"],
            dict_int_method[self.list_of_Conductors[0].inputs["METHOD"]],
        )
        list_folder = ["output", "figures"]
        # Create paths and folders with method _space_convergence_paths
        self._space_convergence_paths(list_folder)
        # Create paths and folders with method _time_convergence_paths
        self._time_convergence_paths(list_folder)
        # Print a warning if os.path.exists() returns True, build the directories if returns False.
        dict_make = {True: self._make_warnings, False: self._make_directories}
        list_f_names = [
            "Initialization",
            "Spatial_distribution",
            "Time_evolution",
            "Solution",
        ]
        dict_benchmark = dict(
            Initialization=self._do_nothing,
            Spatial_distribution=self._benchmark_paths,
            Time_evolution=self._benchmark_paths,
            Solution=self._do_nothing,
        )
        # Create subfolders path invocking method _subfolders_paths
        self._subfolders_paths(list_folder, list_f_names, dict_make, dict_benchmark)

        # Path to save the input files of the simulation in read olny mode as
        # metadata for the simulation itself.
        self.dict_path["Save_input"] = os.path.join(
            self.dict_path["Sub_dir"],
            self.transient_input["SIMULATION"],
            self.basePath.split("/")[-1],
        )
        os.makedirs(self.dict_path["Save_input"], exist_ok=True)

    # End method Simulation_folders_manager.

    def save_input_files(self):
        """Method that saves the input file of the simulation as .xlsx files in read only mode. These files are metadata for the simulation output."""
        load_paths = list()
        save_paths = list()
        filenames = list()

        # Load and save paths for transitory_input file.
        load_transient_input = os.path.join(self.basePath, self.starter_file)
        load_paths.append(os.path.join(self.basePath, self.starter_file))
        filenames.append(self.starter_file)

        # Load file transitory_input.
        transient_input = pd.read_excel(
            load_transient_input,
            sheet_name="TRANSIENT",
            skiprows=1,
            header=0,
            index_col=0,
        )

        # Load paths of environment_input and conductor_definition files.
        for index in ["ENVIRONMENT", "MAGNET"]:
            load_paths.append(
                os.path.join(self.basePath, transient_input.loc[index, "Value"])
            )
            filenames.append(transient_input.loc[index, "Value"])

        # Load file conductor_definition
        conductors = pd.read_excel(
            load_paths[-1],
            sheet_name=None,
            header=0,
            index_col=0,
            skiprows=2,
        )

        # Get all the other input file names.
        # -3 and not -4 since one column of the input file becomes the index of
        # the data frame and should not be considered.
        n_cond = conductors["CONDUCTOR_files"].shape[1] - 3
        # Empty set to store default (necessary or primary) input file names.
        default_files = set()
        # Empty set to store auxiliary input file names.
        aux_files = set()

        files_index_name = conductors["CONDUCTOR_files"].index.to_list()
        for ii in range(n_cond):
            for idx, fname in enumerate(conductors["CONDUCTOR_files"].iloc[:, ii + 3]):
                # Check if file is classified as default or auxiliary (keyword 
                # EXTERNAL in variable name).
                if "EXTERNAL" in files_index_name[idx]:
                    # Auxiliary input file.
                    aux_files.add(fname)
                else:
                    # Default input file.
                    default_files.add(fname)
        
        aux_files.discard("none")
        default_files.discard("none")

        del conductors, transient_input

        # Complete load_paths list
        for fname in default_files.union(aux_files):
            load_paths.append(os.path.join(self.basePath, fname))
            filenames.append(fname)

        for ii, fname in enumerate(filenames):
            # Build save_paths from load_paths.
            save_paths.append(os.path.join(self.dict_path["Save_input"], fname))
            if os.path.exists(save_paths[-1]):
                # Makes the file read/write for the owner if the file
                # already exists
                os.chmod(save_paths[-1], S_IWUSR | S_IREAD)

            if (
                "coupling" in fname
                or "environment_input" in fname
                or "transitory_input" in fname
            ):
                skip_rows = 1
            elif fname in aux_files:
                skip_rows = 0
            else:
                skip_rows = 2

            # Load input file
            dff = pd.read_excel(
                load_paths[ii],
                sheet_name=None,
                header=0,
                index_col=0,
                skiprows=skip_rows,
            )
            # Save input file
            with pd.ExcelWriter(save_paths[-1]) as writer:
                for key, df in dff.items():
                    df.to_excel(writer, sheet_name=key)

        # Convert saved files to read only mode.
        for path in save_paths:
            os.chmod(path, S_IREAD | S_IRGRP | S_IROTH)

    # End method save_input_files

    def __profiling(self, conductor):
        """Temporary private method to performe code profiling.

        Args:
            conductor (obj): object of type conductor.
        """

        if conductor.cond_num_step == 1:
            path_stat = os.path.join("D:/refactoring/function_step", "profiling/before")
            # Name of the binary file storing outcome from cProfile
            file_name_cp_bin = os.path.join(path_stat,"case_3_stats_cprofiler")
            # Name of the text file storing outcome from cProfile processed with pstats.
            file_name_cp_ana = os.path.join(path_stat,"case_3_stats_cprofiler_processed.txt")
            # Name of the binary file storing outcome from line_profiler.
            file_name_lp_bin = os.path.join(path_stat,"case_3_stats_line_profiler.py.lprof")
            # Name of the text file storing outcome from line_profiler.
            file_name_lp_ana = os.path.join(path_stat,"case_3_stats_line_profiler_processed.txt")

            os.makedirs(path_stat, exist_ok=True)
            # Context manager to make the profiling of function step.
            with cProfile.Profile() as pr:
                step(
                    conductor,
                    self.environment,
                    self.dict_qsource[conductor.identifier],
                    self.num_step,
                )
                # Save row profiling outcomes inf file_name (binary).
                pr.dump_stats(file_name_cp_bin)
                # Context manager to write a prcessed .txt file with profiling outcomes.
                with open(file_name_cp_ana,"w+") as ff:
                    # instance of pstats class with methods to manipulate and print the data saved into a profile results file.
                    pp = pstats.Stats(pr,stream=ff)
                    # Remove extraneous path fromm all the module names.
                    pp.strip_dirs()
                    # Sort all the entries according to the cumulative time and number of calls.
                    pp.sort_stats(SortKey.CUMULATIVE,SortKey.CALLS)
                    # Print out the statistics to stdout (mandatory to not get an empty file).
                    pp.print_stats()

            # Instance of the line profiler class.
            lp = LineProfiler()
            # Get a wrapper of function step sutable to perform line profiling.
            step_wrapped = lp(step)
            # Make line profilig on function step.
            lp.enable_by_count()
            
            step_wrapped(
                conductor,
                self.environment,
                self.dict_qsource[conductor.identifier],
                self.num_step,
            )
            
            lp.disable_by_count()
            # Save binary file with the outcomes of the line profiler.
            lp.dump_stats(file_name_lp_bin)
