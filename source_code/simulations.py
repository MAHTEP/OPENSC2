from openpyxl import load_workbook
import numpy as np
import pandas as pd
import os
import warnings

from conductors import Conductors
from environment import Environment
from UtilityFunctions.auxiliary_functions import (
    check_repeated_headings,
    check_object_number,
    with_read_csv,
    with_read_excel,
)
from UtilityFunctions.transient_solution_functions import get_time_step, step
from UtilityFunctions.output import (
    save_simulation_space,
    reorganize_spatial_distribution,
    save_simulation_time,
    save_properties,
)
from UtilityFunctions.plots import (
    plot_properties,
    make_plots,
    create_real_time_plots,
    update_real_time_plots,
)


class Simulations:

    # Current working directory
    CWD = os.getcwd()

    def __init__(self, base_path):

        # Current working directory: SCMagnetCode (cdp, 10/2020)
        # self.cwd = os.getcwd()
        # Ask User the name of the cable. (cdp, 10/2020)
        self.dict_path = dict(
            Current_work_dir=self.CWD,
            Results_dir=os.path.join(self.CWD,"..","Simulations_results"),
        )
        # Create directory Simulations_results if it does not exist yet
        os.makedirs(self.dict_path["Results_dir"], exist_ok=True)
        self.basePath = base_path
        # loop inside self.basePath (cdp, 10/2020)
        input_files = os.listdir(self.basePath)
        for f_name in input_files:
            if "transitory_input" in f_name:
                self.starter_file = f_name
        # Load input file transitory_input.xlsx and convert to a dictionary.
        self.transient_input = pd.read_excel(
            os.path.join(self.basePath, self.starter_file),
            sheet_name="TRANSIENT",
            skiprows=1,
            header=0,
            index_col=0,
            usecols=["Variable name", "Value"],
        )["Value"].to_dict()
        self.flag_start = False
        # get the order of maginitude of the minimum time step to make proper rounds to when saving data and figures of solution spatial distribution at default or User defined times.
        self.n_digit = abs(int(np.floor(np.log10(self.transient_input["STPMIN"]))))
        self.list_of_Conductors = list()
        # Define the environment object as an attribute of the simulation
        self.environment = Environment(
            os.path.join(self.basePath, self.transient_input["ENVIRONMENT"])
        )
        # Define dictionary with default values for functions with_read_csv and with_read_excel.
        self.default_vals = dict(dat="\t", csv=";", tsv="\t", xlsx=0)
        # Define dictionaty with the pointers to the function to be called according to the file extension if some quantities should be loaded from an auxiliary input file.
        self.func_open_aux = dict(
            dat=with_read_csv,
            csv=with_read_csv,
            tsv=with_read_csv,
            xlsx=with_read_excel,
        )

        self.fluid_prop_aliases = dict(
            isobaric_expansion_coefficient="isobaric_expansion_coefficient",
            isothermal_compressibility="isothermal_compressibility",  # 1/Pa
            Prandtl="Prandtl",
            total_density="Dmass",
            total_dynamic_viscosity="viscosity",
            total_enthalpy="Hmass",
            total_isobaric_specific_heat="Cpmass",
            total_isochoric_specific_heat="Cvmass",
            total_speed_of_sound="speed_of_sound",
            total_thermal_conductivity="conductivity",
        )

    # end method __init__ (cdp, 06/2020)

    def conductor_instance(self):
        # Load spread-sheet conductor_definition.xlsx
        conductor_defn = os.path.join(self.basePath, self.transient_input["MAGNET"])
        dict_file = dict()
        # Loop to get the correct name of file conductor_grid.xlsx
        for f_name in os.listdir(self.basePath):
            if "grid" in f_name:
                dict_file["GRID"] = os.path.join(self.basePath, f_name)
            elif "diagnostic" in f_name:
                dict_file["Space"] = os.path.join(self.basePath, f_name)
                dict_file["Time"] = os.path.join(self.basePath, f_name)
            # End if "grid".
        # End for f_name.
        # Load workbook conductor_definition.xlsx.
        conductorsSpec = load_workbook(conductor_defn, data_only=True)
        list_conductor_sheet = [
            conductorsSpec["CONDUCTOR_files"],
            conductorsSpec["CONDUCTOR_input"],
        ]

        # Load the workbook in file conductor_grid.xlsx.
        gridCond = load_workbook(dict_file["GRID"], data_only=True)
        # Load the workbook in file conductor_diagnostic.xlsx.
        wb_diagno = load_workbook(dict_file["Space"], data_only=True)
        # Loop to check if user define conductors with the same identifier.
        for sheet in list_conductor_sheet:
            check_repeated_headings(conductor_defn, sheet)
        # End for sheet.
        # Check if user define conductors with the same identifier in file conductor_grid.xlsx.
        check_repeated_headings(dict_file["GRID"], gridCond["GRID"])
        for sheet in wb_diagno:
            # Check if user define conductors with the same identifier in file conductor_diagnostic.xlsx.
            check_repeated_headings(dict_file["Space"], sheet)
        # End for sheet.
        # Check if the number of defined conductors is consited in file conductor_definition.xlsx and in file conductor_gri.xlsx.

        for cond_sheet in list_conductor_sheet:
            for sheet in [gridCond["GRID"], wb_diagno["Space"], wb_diagno["Time"]]:
                check_object_number(
                    self,
                    self.transient_input["MAGNET"],
                    dict_file[sheet.title],
                    cond_sheet,
                    sheet,
                )
            # End for sheet.
        # End for cond_sheet.

        self.numObj = int(list_conductor_sheet[0].cell(row=1, column=2).value)
        # LOAD MAIN CONDUCTORS PARAMETERS
        for ii in range(1, 1 + self.numObj):
            conductor = Conductors(self, list_conductor_sheet, ii)
            self.list_of_Conductors.append(conductor)
        # end for ii (cdp, 12/2020)
        self.contactBetweenConductors = pd.read_excel(
            conductor_defn, sheet_name="CONDUCTOR_COUPLING", header=0, index_col=0
        )

        # Loop to create the attributes required to make the real time plots (shortly rtp).
        for conductor in self.list_of_Conductors:
            create_real_time_plots(self, conductor)
        # End for conductor.

    # end method Conductor_instance

    def conductor_initialization(self, gui):
        for cond in self.list_of_Conductors:
            # ** INITIALIZATION **
            # s time @ which simulation is started (cdp, 07/2020)
            self.simulation_time = [0.0]
            self.num_step = 0
            cond.initialization(self, gui)
            # plot conductor initialization spatial distribution (cdp, 12/2020)
            plot_properties(self, cond)
            save_simulation_time(self, cond)
            # ** END INITIALIZATION **
        # end for cond (cdp, 12/202)
        # dictionary declaration (cdp,07/2020)
        self.dict_qsource = dict()
        if self.numObj == 1:
            # There is only 1 Conductor object, exploit cond instantiated above (cdp,07/2020)
            self.dict_qsource[cond.ID] = np.zeros(
                (cond.dict_discretization["N_nod"], cond.dict_N_equation["Jacket"])
            )
        else:
            # more than one Conductors object (cdp,07/2020)
            for rr in range(self.numObj):
                cond_r = self.list_of_Conductors[rr]
                if all(self.contactBetweenConductors.iloc[rr, :]) == 0:
                    # There is not contact between cond_r and all the others.
                    # Consider all the columns in order to not miss the info on last raw (otherwise the last conductor will not be added as key of the dictionary).
                    self.dict_qsource[cond_r.ID] = np.zeros(
                        (
                            cond_r.dict_discretization["N_nod"],
                            cond_r.dict_N_equation["Jacket"],
                        )
                    )
                else:
                    for cc in range(rr + 1, self.numObj):
                        cond_c = self.list_of_Conductors[cc]
                        if self.contactBetweenConductors.iat[rr, cc] == 1:
                            # there is contact between cond_r and cond_c (cdp,07/2020)
                            # la colonna non nulla è solo quella del jacket esterno \
                            # eventualmente in contatto con un altro jacket esterno di un \
                            # altro conduttore. Nel caso il conduttore sia costituito da un \
                            # solo jacket la matrice degenera in un vettore colonna. \
                            # (cdp, 08/2020)
                            # N.B. controllare anche nella chiamata a step come passare self.dict_qsource.
                            raise ValueError(
                                "Multi conductors with thermal contact: not yet managed the situation, need to build dict_qsource properly!!"
                            )
                        else:
                            # There is not contact between cond_r and cond_c (cdp,07/2020)
                            # Proposta di soluzione ma va studiata decisamente meglio!
                            # N.B. controllare anche nella chiamata a step come passare self.dict_qsource.
                            self.dict_qsource[f"{cond_r.ID}_{cond_c.ID}"] = np.zeros(
                                (
                                    cond_r.dict_discretization["N_nod"],
                                    cond_r.dict_N_equation["Jacket"],
                                )
                            )
                            raise ValueError(
                                "Multi conductors with thermal contact: not yet managed the situation, need to build dict_qsource properly!!"
                            )
                        # End if self.contactBetweenConductors[rr, cc].
                    # End for cc.
                # End if all(self.contactBetweenConductors[rr, rr + 1:])
                # End for rr.
        # end if numObj

    # end method Conductor_initialization

    def conductor_solution(self, gui):
        # ** TRANSIENT SOLUTION **
        num_step_store = 100
        count_store = 1
        stoptime = 0  # flag to stop simulation if some problems (like quench) \
        # arise (cdp, 07/2020)
        # Time step initialization (cdp, 08/2020)
        # time_step = transient_input["STPMIN"]
        # Search if User asks to save the solution spatial distribution at TEND \
        # (cdp, 10/2020)
        # loop on conductors (cdp, 10/2020)
        for ii in range(self.numObj):
            conductor = self.list_of_Conductors[ii]
            # Compute radiative heat exchanged between jackets.
            conductor.compute_radiative_heat_exhange_jk()
            # Compute radiative heat exchanged outer jacket and environment.
            conductor.compute_heat_exchange_jk_env(self.environment)
            # get the times at which users saves the solution spatial distribution \
            # (cdp, 10/2020)
            # list_values = list(conductor.dict_Space_save.values())
            # Save of the solution spatial distribution at 0.0 s (cdp, 12/2020)
            save_simulation_space(
                conductor,
                self.dict_path[f"Output_Space_{conductor.ID}_dir"],
                abs(self.n_digit),
            )
        # end for ii (cdp, 10/2020)
        # while loop to solve transient at each timestep (cdp, 07/2020)
        while (
            self.simulation_time[-1]
            < self.transient_input["TEND"] - 1e-5 * self.transient_input["STPMIN"]
            and stoptime == 0
        ):
            self.num_step = self.num_step + 1
            time_step = np.zeros(self.numObj)
            for ii in range(self.numObj):
                conductor = self.list_of_Conductors[ii]
                # Call function Get_time_step to select new time step (cdp, 08/2020)
                get_time_step(conductor, self.transient_input, self.num_step)
                time_step[ii] = conductor.time_step
                # Increase time (cdp, 08/2020)
                conductor.cond_time.append(
                    conductor.cond_time[-1] + conductor.time_step
                )
                # update time step (cdp, 08/2020)
                conductor.cond_num_step = conductor.cond_num_step + 1
                # before calling Conductors method initialization adapt mesh if \
                # necessary as foreseen by ITYMSH. To do later (cdp, 07/2020)
                # se ho nuova griglia calcolare coefficenti, temperature, pressioni, \
                # parametri adimensionati sfruttando np.interp (in fortrand è adaptm)
            # End for conductor (cdp, 08/2020)
            # Evaluate simulation time and simulation time step (cdp, 08/2020)
            self.simulation_time_step = np.max(time_step)
            self.simulation_time.append(
                self.simulation_time[-1] + self.simulation_time_step
            )
            print(
                f"Simulation time: {self.simulation_time[-1]:.{self.n_digit}f} s; {self.simulation_time[-1]/self.transient_input['TEND']*100:5.2f} %"
            )
            for conductor in self.list_of_Conductors:
                # evaluate properties and quantities in Gauss points, method \
                # Eval_Gauss_point is invoked inside method operating_conditions\
                # (cdp, 07/2020)
                # if self.num_step > 1:
                conductor.operating_conditions(self)
                # end if self.num_step
                # call step to solve the problem @ new timestep (cdp, 07/2020)
                step(
                    conductor,
                    self.environment,
                    self.dict_qsource[conductor.ID],
                    self.num_step,
                )
                # Loop on FluidComponents (cdp, 10/2020)
                for fluid_comp in conductor.dict_obj_inventory["FluidComponents"][
                    "Objects"
                ]:
                    # compute density and mass flow rate in nodal points with the
                    # updated FluidComponents temperature and velocity (nodal = True by default)
                    fluid_comp.coolant._compute_density_and_mass_flow_rates_nodal_gauss(
                        conductor
                    )
                    # Enthalpy balance: sum((mdot*w)_out - (mdot*w)_inl), used to check \
                    # the imposition of SolidComponents temperature initial spatial \
                    # distribution (cdp, 12/2020)
                    conductor.enthalpy_balance = (
                        conductor.enthalpy_balance
                        + conductor.time_step
                        * (
                            fluid_comp.coolant.dict_node_pt["mass_flow_rate"][-1]
                            * fluid_comp.coolant.dict_node_pt["total_enthalpy"][-1]
                            - fluid_comp.coolant.dict_node_pt["mass_flow_rate"][0]
                            * fluid_comp.coolant.dict_node_pt["total_enthalpy"][0]
                        )
                    )
                    conductor.enthalpy_out = (
                        conductor.enthalpy_out
                        + conductor.time_step
                        * fluid_comp.coolant.dict_node_pt["mass_flow_rate"][-1]
                        * fluid_comp.coolant.dict_node_pt["total_enthalpy"][-1]
                    )
                    conductor.enthalpy_inl = (
                        conductor.enthalpy_inl
                        + conductor.time_step
                        * fluid_comp.coolant.dict_node_pt["mass_flow_rate"][0]
                        * fluid_comp.coolant.dict_node_pt["total_enthalpy"][0]
                    )
                # end for fluid_comp (cdp, 10/2020)
                # Loop on SolidComponents (cdp, 01/2021)
                # for s_comp in conductor.dict_obj_inventory["SolidComponents"]\
                # 	["Objects"]:
                # 	if s_comp.NAME != "STR_STAB" and s_comp.NAME != "Z_JACKET":
                # 		# Call Get_superconductor_critical_prop to evaluate \
                # 		# MixSCStabilizer and/or SuperConductor properties in nodal points.
                # 		# In questo modo le calclo 2 volte, qui e alla successiva chiamata \
                # 		# dell methodo Operating conditions, da sistemare per ottimizzare.
                # 		s_comp.Get_superconductor_critical_prop(self, Where = "nodal")
                # 	# compute, average density, thermal conductivity, specifi heat at \
                # 	# constant pressure and electrical resistivity with the updated \
                # 	# Solidcomponents temperature in nodal points (cdp, 01/2021)
                # 	s_comp.Eval_sol_comp_properties(conductor.dict_obj_inventory, \
                # 		Where = "nodal")
                ## end for s_comp (cdp, 01/2021)

                # Compute radiative heat exchanged between jackets.
                conductor.compute_radiative_heat_exhange_jk()
                # Compute radiative heat exchanged outer jacket and environment.
                conductor.compute_heat_exchange_jk_env(self.environment)

                update_real_time_plots(conductor)

                if self.num_step == num_step_store * count_store:
                    # Update counter to store the state of the simulation, still to come \
                    # (cdp, 08/2020)
                    count_store = count_store + 1
                if (
                    conductor.i_save < len(conductor.Space_save) - 1
                    and abs(
                        conductor.cond_time[-1] - conductor.Space_save[conductor.i_save]
                    )
                    < conductor.time_step / 2.0
                ):
                    # save simulation spatial distribution at user defined time steps \
                    # (cdp, 08/2020)
                    save_simulation_space(
                        conductor,
                        self.dict_path[f"Output_Space_{conductor.ID}_dir"],
                        abs(self.n_digit),
                    )
                # end if isave
                ##
                # end if conductor.i_save_space[i_save]
                # for key in list(conductor.dict_Space_save.keys()):
                # 	if "done" not in key:
                # 		# do not consider the flags (cdp, 08/2020)
                # 		if conductor.dict_Space_save[key] > 0:
                # 			conductor.dict_Space_save[key] = round(\
                # 																conductor.dict_Space_save[key], 14) - \
                # 																conductor.time_step
                # 		# reached time values to save the solution and plot it \
                # 			# cdp, 08/2020)
                # 			flag_rename = Save_simulation_space(conductor, \
                # 									self.dict_path[f"Output_Space_{conductor.ID}_dir"], \
                # 									Who = "User")
                # 			# Saved the solution, this time is no longer considered \
                # 			# (cdp, 08/2020)
                # 			conductor.dict_Space_save[f"{key}_done"] = 1
                # 			if flag_rename == False:
                # 				# Update the list of times at which make plots of the solution \
                # 				# spatial distributions only if flag_rename is False; if it is \
                # 				# True the update is already done above in the Default part \
                # 				# (cdp, 10/2020)
                # 				conductor.plot_space_distribution.append(\
                # 					conductor.cond_time[-1])

                # Save variables time evolution at given spatial coordinates \
                # (cdp, 08/2020)
                save_simulation_time(self, conductor)
                # call sensor to plot results at any time the user asks (cdp, 07/2020)
            # End for conductor (cdp, 07/2020)
        # end while (cdp, 07/2020)
        # Loop on conductors (cdp, 12/2020)
        for cond in self.list_of_Conductors:
            # save simulation spatial distribution at TEND (cdp, 01/2021)
            save_simulation_space(
                cond, self.dict_path[f"Output_Space_{cond.ID}_dir"], abs(self.n_digit)
            )
            # Call function Save_properties to save the conductor final solution \
            # (cdp, 12/2020)
            save_properties(cond, self.dict_path[f"Output_Solution_{cond.ID}_dir"])
        # end for cond (cdp, 12/2020)
        print("Saved final solution\n")
        print("End simulation called " + self.transient_input["SIMULATION"] + "\n")

    # end method Conductor_solution (cdp, 09/2020)

    def conductor_post_processing(self):
        # loop to save the norm of the solution at the end of the transient for \
        # each conductor, usefull to make space convergence (cdp, 09/2020)
        # t_end = np.array([self.transient_input["TEND"]])
        for cond in self.list_of_Conductors:
            cond.post_processing(self)
            reorganize_spatial_distribution(
                cond, self.dict_path[f"Output_Space_{cond.ID}_dir"], self.n_digit
            )
            # Plot conductor solution spatial distribution (cdp, 12/2020)
            plot_properties(self, cond, what="solution")
        # end for cond (cdp, 12/2020)
        # Call function Make_plots to make plot of spatial distribution and time \
        # evolution (cdp, 11/2020)
        make_plots(self, kind="Space_distr")
        make_plots(self, kind="Time_evol")

    # end method Conductor_post_processing (cdp, 09/2020)

    def _make_directories(self, list_key_val, exist_ok=False):
        """[summary]

        Args:
            list_key_val ([type]): [description]
            exist_ok (bool, optional): [description]. Defaults to False.
        """
        # Loop to create the folders.
        for ii in range(len(list_key_val)):
            # Create the folders in list_key_val[ii][0].
            os.makedirs(self.dict_path[list_key_val[ii][0]], exist_ok=exist_ok)
        # End for ii.

    # End method _make_directories.

    def _space_convergence_paths(self, list_folder):
        """Private method that builds the paths for the space convergence analysis invocking the private method _update_dict_path_and_make_dirs."""
        # list_folder = ["output", "figures"]
        str_dir = (
            "TEND_"
            + f"{self.transient_input['TEND']}_"
            + "STPMIN_"
            + f"{self.transient_input['STPMIN']}"
        )
        # Build list_key_val exploiting list comprehension. List of tuples: index [0] is the key of the dictionary, index [1] is the corresponding value that is the path to Output or Figures sub directories.
        list_key_val = [
            (
                f"Space_conv_{folder}_dir",
                os.path.join(
                    self.dict_path["Sub_dir"],
                    "Space_convergence",
                    str_dir,
                    folder.capitalize(),
                ),
            )
            for folder in list_folder
        ]
        # Update the dictionary self.dict_path with keys and values from the dictionary comprehension.
        self.dict_path.update(
            {
                list_key_val[ii][0]: list_key_val[ii][1]
                for ii in range(len(list_key_val))
            }
        )
        # Make the directories invoking method _make_directories.
        self._make_directories(list_key_val, exist_ok=True)

    # End method _space_convergence_paths.

    def _time_convergence_paths(self, list_folder):
        """Private method that builds the paths for the time convergence analysis invocking the private method _update_dict_path_and_make_dirs. The value of NELEMS is the one of the first defined conductor."""
        # list_folder = ["output", "figures"]
        str_dir = (
            "TEND_"
            + f"{self.transient_input['TEND']}_"
            + "NELEMS_"
            + f"{self.list_of_Conductors[0].dict_discretization['Grid_input']['NELEMS']}"
        )
        # Build list_key_val exploiting list comprehension. List of tuples: index [0] is the key of the dictionary, index [1] is the corresponding value that is the path to Output or Figures sub directories.
        list_key_val = [
            (
                f"Time_conv_{folder}_dir",
                os.path.join(
                    self.dict_path["Sub_dir"],
                    "Time_convergence",
                    str_dir,
                    folder.capitalize(),
                ),
            )
            for folder in list_folder
        ]
        # Update the dictionary self.dict_path with keys and values from the dictionary comprehension.
        self.dict_path.update(
            {
                list_key_val[ii][0]: list_key_val[ii][1]
                for ii in range(len(list_key_val))
            }
        )
        # Make the directories invoking method _make_directories.
        self._make_directories(list_key_val, exist_ok=True)

        # Path of the directory to save the comparison of the time convergence \
        # with the same method, comparison is by NELEMS (cdp, 11/2020) [seems to not be used, I do not remember the aim of this!! 08/07/2020]
        # self.dict_path["Time_conv_comp_METHOD"] = os.path.join(\
        # 																		self.dict_path["Main_dir"], "METHOD")
        # Path of the directory to save the comparison of the time convergence \
        # with the same NELEMS, comparison is by METHODS (cdp, 11/2020) [seems to not be used, I do not remember the aim of this!! 08/07/2020]
        # self.dict_path["Time_conv_comp_NELEMS"] = os.path.join(\
        # 																		self.dict_path["Main_dir"], "NELEMS")

    # End method _time_convergence_paths.

    def _subfolders_paths(self, list_folder, list_f_names, dict_make, dict_benchmark):
        """[summary]

        Args:
            list_folder ([type]): [description]
            list_f_names ([type]): [description]
            dict_make ([type]): [description]
            dict_benchmark ([type]): [description]
        """
        # Loop to create sub folders initialization, Space, Time and Benchmark in Output and Figures directories; each folder in f_names_list, will contain folder conductor.ID.
        for f_name in list_f_names:
            for conductor in self.list_of_Conductors:
                # Build list_key_val exploiting list comprehension. List of tuples: index [0] is the key of the dictionary, index [1] is the corresponding value that is the path to Output or Figures sub directories.
                list_key_val = [
                    (
                        f"{folder.capitalize()}_{f_name}_{conductor.ID}_dir",
                        os.path.join(
                            self.dict_path["Sub_dir"],
                            self.transient_input["SIMULATION"],
                            folder.capitalize(),
                            f_name,
                            conductor.ID,
                        ),
                    )
                    for folder in list_folder
                ]
                # Update the dictionary self.dict_path with keys and values from dictionary comprehension, which may be either a mapping or an iterable of key/value pairs. The values of dictionary comprehension take priority when self.dict_path0 and other share keys.
                self.dict_path.update(
                    {
                        list_key_val[ii][0]: list_key_val[ii][1]
                        for ii in range(len(list_key_val))
                    }
                )
                # Invocke method self._make_warnings if path exists, self._make_directories if path does not exist.
                dict_make[os.path.exists(self.dict_path[list_key_val[0][0]])](
                    list_key_val
                )
                # Create benchmark directory if f_name is "Space" ot "Time"
                dict_benchmark[f_name](conductor, list_folder, f_name)
            # End for conductor.
        # End for f_name.

    # End method _subfolders_paths.

    def _make_warnings(self, list_key_val):
        """[summary]

        Args:
            list_key_val ([type]): [description]
        """
        # Da sistemare nella GUI!
        warnings.warn(
            f"Directories\n{self.dict_path[list_key_val[0][0]]}\n{self.dict_path[list_key_val[1][0]]} already exist.\nProbably you have already performed a simulation with the same input data.\nPlease check and confirm if you want to continue with the simulation or not."
        )

    # End method _make_warnings.

    def _benchmark_paths(self, conductor, list_folder, f_name):
        """[summary]

        Args:
            conductor ([type]): [description]
            list_folder ([type]): [description]
            f_name ([type]): [description]
        """
        # Build list_key_val exploiting list comprehension. List of tuples: index [0] is the key of the dictionary, index [1] is the corresponding value that is the path to Output or Figures sub directories.
        list_key_val = [
            (
                f"{folder.capitalize()}_{f_name}_benchmark_{conductor.ID}_dir",
                os.path.join(
                    self.dict_path["Sub_dir"],
                    self.transient_input["SIMULATION"],
                    folder.capitalize(),
                    f_name,
                    "Benchmark",
                    conductor.ID,
                ),
            )
            for folder in list_folder
        ]
        # Update the dictionary self.dict_path with keys and values from dictionary comprehension.
        self.dict_path.update(
            {
                list_key_val[ii][0]: list_key_val[ii][1]
                for ii in range(len(list_key_val))
            }
        )
        # Make the directories invoking method _make_directories.
        self._make_directories(list_key_val, exist_ok=True)
        # End method _benchmark_paths.

    def _do_nothing(self, conductor, list_folder, f_name):
        # Do nothing method, indroduced to use the dictionary dict_benchmark when creating folders.
        pass

    def simulation_folders_manager(self):
        """[summary]"""
        # Define the dummy dictionary with the integration methods (da sistemare, non utilizzare i numeri come keys)
        dict_int_method = dict(BE="BE", CN="CN", AM4="AM4")
        # Update dictionary self.dict_path
        self.dict_path["Sub_dir"] = os.path.join(
            self.dict_path["Main_dir"],
            dict_int_method[self.list_of_Conductors[0].dict_input["METHOD"]],
        )
        list_folder = ["output", "figures"]
        # Create paths and folders with method _space_convergence_paths
        self._space_convergence_paths(list_folder)
        # Create paths and folders with method _time_convergence_paths
        self._time_convergence_paths(list_folder)
        # Print a warning if os.path.exists() returns True, build the directories if returns False.
        dict_make = {True: self._make_warnings, False: self._make_directories}
        list_f_names = ["Initialization", "Space", "Time", "Solution"]
        dict_benchmark = dict(
            Initialization=self._do_nothing,
            Space=self._benchmark_paths,
            Time=self._benchmark_paths,
            Solution=self._do_nothing,
        )
        # Create subfolders path invocking method _subfolders_paths
        self._subfolders_paths(list_folder, list_f_names, dict_make, dict_benchmark)

        # Path to save the input files of the simulation in read olny mode as 
        # metadata for the simulation itself.
        self.dict_path["Save_input"] = os.path.join(self.dict_path["Sub_dir"], self.transient_input["SIMULATION"], self.basePath.split("/")[-1])
        os.makedirs(self.dict_path["Save_input"], exist_ok=True)

    # End method Simulation_folders_manager.

        ##############################################################################

    ## This code was part of the old version of method Simulation_folders_manager, may be useful in future so I do not delete it. ##

    # Gives an error due to a too long path: it thakes into account of \
    # other features of the simulations together with the method, that is \
    # the one considered below to shorten the path (cpd, 10/2020)
    # block = list()
    # block.append("XLENGTH_" + str(cond.dict_input["XLENGTH"]))
    # if cond.self.dict_input["METHOD"] == "BE":
    # 	block.append("BE")
    # elif cond.self.dict_input["METHOD"] == "CN]":
    # 	block.append("CN")
    # elif cond.self.dict_input["METHOD"] == "AM4":
    # 	block.append("AM4")
    # end if cond.dict_input["METHOD"] (cdp, 10/2020)
    # block.append("IOP0_TOT_" + str(cond.dict_input["IOP0_TOT"]))
    # INTIAL_val = str()
    # for ii in range(cond.dict_obj_inventory["FluidComponents"]["Number"]):
    # 	fluid_comp = cond.dict_obj_inventory["FluidComponents"]["Objects"][ii]
    # 	if cond.dict_obj_inventory["FluidComponents"]["Number"] == 1:
    # 		INTIAL_val = INTIAL_val + str(fluid_comp.coolant.dict_operation["INTIAL"])
    # 	elif cond.dict_obj_inventory["FluidComponents"]["Number"] > 1:
    # 		if ii < cond.dict_obj_inventory["FluidComponents"]["Number"] - 1:
    # 			INTIAL_val = INTIAL_val + str(fluid_comp.coolant.dict_operation["INTIAL"]) + ","
    # 		elif ii == cond.dict_obj_inventory["FluidComponents"]["Number"] - 1:
    # 			INTIAL_val = INTIAL_val + str(fluid_comp.coolant.dict_operation["INTIAL"])
    # 		# end if ii (cdp, 10/2020)
    # 	# end if cond.dict_obj_inventory["FluidComponents"]["Number"] \
    # 	# (cdp, 10/2020)
    ## end for ii (cdp, 10/2020)
    # block.append(f"INTIAL_({INTIAL_val})")
    # XQEND_val = str()
    # for ii in range(cond.dict_obj_inventory["SolidComponents"]["Number"]):
    # 	s_comp = cond.dict_obj_inventory["SolidComponents"]["Objects"][ii]
    # 	if cond.dict_obj_inventory["SolidComponents"]["Number"] == 1:
    # 		XQEND_val = XQEND_val + str(s_comp.dict_operation["XQEND"])
    # 	elif cond.dict_obj_inventory["SolidComponents"]["Number"] > 1:
    # 		if ii < cond.dict_obj_inventory["SolidComponents"]["Number"] - 1:
    # 			XQEND_val = XQEND_val + str(s_comp.dict_operation["XQEND"]) + ","
    # 		elif ii == cond.dict_obj_inventory["SolidComponents"]["Number"] - 1:
    # 			XQEND_val = XQEND_val + str(s_comp.dict_operation["XQEND"])
    # 		# end if ii (cdp, 10/2020)
    # 	# end if cond.dict_obj_inventory["SolidComponents"]["Number"] \
    # 	# (cdp, 10/2020)
    ## end for ii (cdp, 10/2020)
    # block.append(f"XQEND_({XQEND_val})")
    ## Join the block together using _ ad separator (cdp, 10/2020)
    # self.dict_path["Sub_dir"] = "_".join([f"{val}" for val in block])
