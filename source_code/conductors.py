# Import packages
import logging
import logging.config
from openpyxl import load_workbook
import numpy as np
from scipy import constants
import pandas as pd
import os
import sys
import warnings

# import classes
from fluid_components import FluidComponents
from jacket import Jacket
from mix_sc_stabilizer import MixSCStabilizer
from stabilizer import Stabilizer
from super_conductor import SuperConductor

# import functions
from UtilityFunctions.auxiliary_functions import (
    check_repeated_headings,
    check_headers,
    check_object_number,
    set_diagnostic,
)
from UtilityFunctions.initialization_functions import conductor_spatial_discretization
from UtilityFunctions.gen_flow import gen_flow
from UtilityFunctions.output import save_properties, save_convergence_data
from UtilityFunctions.plots import update_real_time_plots, create_legend_rtp
from UtilityFunctions.solid_components_initialization import (
    solid_components_temperature_initialization,
)

# Stainless Steel properties
from Properties_of_materials.stainless_steel import thermal_conductivity_ss

logging.config.fileConfig(fname='logging_electric_module.conf', disable_existing_loggers=True)

# Get the logger specified in the file
consolelogger = logging.getLogger("consoleLogger")
filelogger = logging.getLogger("fileLogger")

class Conductors:

    BASE_PATH = ""

    ## computed in initialization
    # Remove from here if actually used to avoid problems!
    # MDTINL = ""
    # IBDINL = ""
    # IBDOUT = ""
    # inHard = "" # old: list() (cdp 06/2020)
    # MAXDOF = "" not used at the moment

    # Minimum absolute pressure difference between two channels in hydraulic \
    # parallel (threshold)
    Delta_p_min = 1e-4  # Pa
    # Localized pressure drop coefficient beywen two channels in hydraulic \
    # parallel
    k_loc = 1.0  # ~
    # Lambda velocity parameter (--> 0 if porous wall, --> 1 if helicoidal wall)
    lambda_v = 1.0  # ~

    KIND = "Conductor"
    CHUNCK_SIZE = 100

    def __init__(self, simulation, sheetConductorsList, ICOND):

        self.BASE_PATH = simulation.basePath
        self.ICOND = ICOND
        self.NAME = sheetConductorsList[0].cell(row=1, column=1).value
        # get channels ID consistently with user definition (cdp, 09/2020)
        self.ID = sheetConductorsList[0].cell(row=3, column=4 + self.ICOND).value
        # Get the number of the conductor from the ID.
        self.number = int(self.ID.split("_")[1])
        # file_input dictionary initialization (cdp, 06/2020)
        self.file_input = dict()
        # inputs dictionary initialization (cdp, 06/2020)
        self.inputs = dict()
        # Load the sheet CONDUCTOR_files form file conducor_definition.xlsx as a disctionary.
        self.file_input = pd.read_excel(
            os.path.join(self.BASE_PATH, simulation.transient_input["MAGNET"]),
            sheet_name=sheetConductorsList[0].title,
            skiprows=2,
            header=0,
            index_col=0,
            usecols=["Variable name", self.ID],
        )[self.ID].to_dict()
        # Load the sheet CONDUCTOR_input form file conducor_definition.xlsx as a disctionary.
        self.inputs = pd.read_excel(
            os.path.join(self.BASE_PATH, simulation.transient_input["MAGNET"]),
            sheet_name=sheetConductorsList[1].title,
            skiprows=2,
            header=0,
            index_col=0,
            usecols=["Variable name", self.ID],
        )[self.ID].to_dict()
        # Get the user defined name of the conductor (default is ID)
        self.name = self.inputs["NAME"]
        # Delete key NAME from dictionary self.inputs
        del self.inputs["NAME"]

        _ = dict(BE=1.0, CE=0.5)
        self.electric_theta = _[self.inputs["ELECTRIC_METHOD"]]
        consolelogger.debug(f"Defined electric_theta\n")

        # Load the sheet CONDUCTOR_operation form file conducor_definition.xlsx as a disctionary.
        self.operations = pd.read_excel(
            os.path.join(self.BASE_PATH, simulation.transient_input["MAGNET"]),
            sheet_name=sheetConductorsList[2].title,
            skiprows=2,
            header=0,
            index_col=0,
            usecols=["Variable name", self.ID],
        )[self.ID].to_dict()
        consolelogger.debug(f"Loaded sheet CONDUCTOR_operation from file conductor_definition\n")

        # Load all the sheets in file conductor_coupling.xlsx as a dictionary of dataframes.
        self.dict_df_coupling = pd.read_excel(
            os.path.join(self.BASE_PATH, self.file_input["STRUCTURE_COUPLING"]),
            sheet_name=None,
            skiprows=1,
            header=0,
            index_col=0,
        )

        # Dictionary declaration (cdp, 09/2020)
        self.dict_obj_inventory = dict()
        # call method Conductor_components_instance to make instance of conductor components (cdp, 11/2020)
        self.conductor_components_instance(simulation)

        self.dict_topology = dict()  # dictionary declaration (cdp, 09/2020)
        self.dict_interf_peri = dict()  # dictionary declaration (cdp, 07/2020)
        # Call method Get_conductor_topology to evaluate conductor topology: \
        # interfaces between channels, channels and solid components and between \
        # solid components(cdp, 09/2020)
        self.get_conductor_topology(simulation.environment)

        self.dict_node_pt = dict()
        self.dict_Gauss_pt = dict()

        # CREATE grid for the i-th conductor
        self.dict_discretization = dict()
        self.dict_discretization["Grid_input"] = pd.read_excel(
            os.path.join(self.BASE_PATH, self.file_input["GRID_DEFINITION"]),
            sheet_name="GRID",
            skiprows=2,
            header=0,
            index_col=0,
            usecols=["Variable name", self.ID],
            dtype="object",
        )[self.ID].to_dict()

        self.heat_rad_jk = dict()
        self.heat_exchange_jk_env = dict()

        # **NUMERICS**
        # evaluate value of theta_method according to flag METHOD (cdo, 08/2020)
        if self.inputs["METHOD"] == "BE":
            # Backward Euler (cdp, 10/2020)
            self.theta_method = 1.0
        elif self.inputs["METHOD"] == "CN":
            # Crank-Nicolson (cdp, 10/2020)
            self.theta_method = 0.5
        elif self.inputs["METHOD"] == "AM4":
            # Adams-Moulton 4 (cdp, 10/2020)
            # questo valore è provvisorio e sicuramente poco corretto, da ragionare e approfondire
            self.theta_method = 1.0 / 24.0

        ## Evaluate parameters useful in function \
        # Transient_solution_functions.py\STEP (cdp, 07/2020)
        # dict_N_equation keys meaning:
        # ["FluidComponents"]: total number of equations for FluidComponents \
        # objects (cdp, 07/2020);
        # ["Strands"]: total number of equations for Strands objects (cdp, 07/2020);
        # ["Jacket"]: total number of equations for Jacket objects (cdp, 07/2020);
        # ["SolidComponents"]: total number of equations for SolidComponents \
        # objects (cdp, 07/2020);
        # ["NODOFS"]: total number of equations for for each node, i.e. Number Of \
        # Degrees Of Freedom, given by: \
        # 3*(number of channels) + (number of strands) + (number of jackets) \
        # (cdp, 07/2020);
        self.dict_N_equation = dict(
            FluidComponents=3 * self.dict_obj_inventory["FluidComponents"]["Number"],
            Strands=self.dict_obj_inventory["Strands"]["Number"],
            Jacket=self.dict_obj_inventory["Jacket"]["Number"],
            SolidComponents=self.dict_obj_inventory["SolidComponents"]["Number"],
        )
        # necessary since it is not allowed to use the value of a dictionary key \
        # before that the dictionary is fully defined (cdp, 09/2020)
        self.dict_N_equation.update(
            NODOFS=self.dict_N_equation["FluidComponents"]
            + self.dict_N_equation["SolidComponents"]
        )
        # dict_band keys meaning:
        # ["Half"]: half band width, including main diagonal (IEDOFS) (cdp, 09/2020)
        # ["Main_diag"]: main diagonal index within the band (IHBAND) (cdp, 09/2020)
        # ["Full"]: full band width, including main diagonal (IBWIDT) (cdp, 09/2020)
        self.dict_band = dict(
            Half=2 * self.dict_N_equation["NODOFS"],
            Main_diag=2 * self.dict_N_equation["NODOFS"] - 1,
            Full=4 * self.dict_N_equation["NODOFS"] - 1,
        )
        # self.MAXDOF = self.dict_N_equation["NODOFS"]*MAXNOD
        self.EQTEIG = np.zeros(self.dict_N_equation["NODOFS"])
        # dict_norm keys meaning:
        # ["Solution"]: norm of the solution (cdp, 09/2020)
        # ["Change"]: norm of the solution variation wrt the previous time step \
        # (cdp, 09/2020)
        self.dict_norm = dict(
            Solution=np.zeros(self.dict_N_equation["NODOFS"]),
            Change=np.zeros(self.dict_N_equation["NODOFS"]),
        )

        # evaluate attribute EIGTIM exploiting method Aprior (cdp, 08/2020)
        self.aprior()
        path_diagnostic = os.path.join(self.BASE_PATH, self.file_input["OUTPUT"])
        # Load the content of column self.ID of sheet Space in file conductors_disgnostic.xlsx as a series and convert to numpy array of float.
        self.Space_save = (
            pd.read_excel(
                path_diagnostic,
                sheet_name="Spatial_distribution",
                skiprows=2,
                header=0,
                usecols=[self.ID],
                squeeze=True,
            )
            .dropna()
            .to_numpy()
            .astype(float)
        )
        # Adjust the user defined diagnostic.
        self.Space_save = set_diagnostic(
            self.Space_save, lb=0.0, ub=simulation.transient_input["TEND"]
        )
        # Check on spatial distribution diagnostic.
        if self.Space_save.max() > simulation.transient_input["TEND"]:
            raise ValueError(
                f"File {self.file_input['OUTPUT']}, sheet Space, conductor {self.ID}: impossible to save spatial distributions at time {self.Space_save.max()} s since it is larger than the end time of the simulation {simulation.transient_input['TEND']} s.\n"
            )
        # End if self.Space_save.max() > simulation.transient_input["TEND"]
        # index pointer to save solution spatial distribution (cdp, 12/2020)
        self.i_save = 0
        # list of number of time steps at wich save the spatial discretization
        self.num_step_save = np.zeros(self.Space_save.shape, dtype=int)
        # Load the content of column self.ID of sheet Time in file conductors_disgnostic.xlsx as a series and convert to numpy array of float.
        self.Time_save = (
            pd.read_excel(
                path_diagnostic,
                sheet_name="Time_evolution",
                skiprows=2,
                header=0,
                usecols=[self.ID],
                squeeze=True,
            )
            .dropna()
            .to_numpy()
            .astype(float)
        )
        # Adjust the user defined diagnostic.
        self.Time_save = set_diagnostic(
            self.Time_save, lb=0.0, ub=self.inputs["XLENGTH"]
        )
        # Check on time evolution diagnostic.
        if self.Time_save.max() > self.inputs["XLENGTH"]:
            raise ValueError(
                f"File {self.file_input['OUTPUT']}, sheet Time, conductor {self.ID}: impossible to save time evolutions at axial coordinate {self.Time_save.max()} s since it is ouside the computational domain of the simulation [0, {self.inputs['XLENGTH']}] m.\n"
            )
        # End if self.Time_save.max() > self.inputs["XLENGTH"]

        # declare dictionaries to store Figure and axes objects to constructi real \
        # time figures (cdp, 10/2020)
        self.dict_Figure_animation = dict(T_max=dict(), mfr=dict())
        self.dict_axes_animation = dict(T_max=dict(), mfr=dict())
        self.dict_canvas = dict(T_max=dict(), mfr=dict())
        self.color = ["r*", "bo"]

    # end method __init__ (cdp, 11/2020)

    def __str__(self):
        pass

    def __repr__(self):
        return f"{self.__class__.__name__}(Type: {self.KIND}, ID: {self.ID})"

    def conductor_components_instance(self, simulation):
        """
        Method that makes instances of the conductor components defined with workbook conductorc_input.xlsx and conductcor_operation.xlsx (cdp, 11/2020)
        """
        dict_file_path = dict(
            input=os.path.join(self.BASE_PATH, self.file_input["STRUCTURE_ELEMENTS"]),
            operation=os.path.join(self.BASE_PATH, self.file_input["OPERATION"]),
        )
        # Load workbook conductor_input.xlsx.
        wb_input = load_workbook(dict_file_path["input"], data_only=True)
        # Load workbook conductor_operation.xlsx.
        wb_operations = load_workbook(dict_file_path["operation"], data_only=True)

        listOfComponents = wb_input.get_sheet_names()
        self.dict_obj_inventory["FluidComponents"] = dict()
        self.dict_obj_inventory["MixSCStabilizer"] = dict()
        self.dict_obj_inventory["Stabilizer"] = dict()
        self.dict_obj_inventory["SuperConductor"] = dict()
        self.dict_obj_inventory["Jacket"] = dict()
        self.dict_obj_inventory["Strands"] = dict()
        self.dict_obj_inventory["SolidComponents"] = dict()
        self.dict_obj_inventory["Conductor_components"] = dict()
        self.dict_obj_inventory["Strands"]["Objects"] = list()
        self.dict_obj_inventory["SolidComponents"]["Objects"] = list()
        self.dict_obj_inventory["Conductor_components"]["Objects"] = list()
        for sheetID in listOfComponents:
            sheet = wb_input[sheetID]
            sheetOpar = wb_operations[sheetID]
            # Call function check_object_number to check that the number of objects defined in sheets of file conductor_input.xlsx and conductor_operation.xlsx are the same.
            check_object_number(
                self,
                dict_file_path["input"],
                dict_file_path["operation"],
                sheet,
                sheetOpar,
            )
            # Call function check_repeated_headings to check if there are repeaded headings in sheet of file conductor_input.xlsx.
            check_repeated_headings(dict_file_path["input"], sheet)
            # Call function check_repeated_headings to check if there are repeaded headings in sheet of file conductor_operation.xlsx.
            check_repeated_headings(dict_file_path["operation"], sheetOpar)
            # Call function check_headers to check if sheets of file conductor_input.xlsx and conductor_operation.xlsx have exactly the same headers.
            check_headers(
                self,
                dict_file_path["input"],
                dict_file_path["operation"],
                sheet,
                sheetOpar,
            )
            kindObj = sheet.cell(row=1, column=1).value  # sheet["A1"].value
            numObj = int(sheet.cell(row=1, column=2).value)  # sheet["B1"].value
            if kindObj == "CHAN":
                # Define and assign values to FluidComponents dictionary keys: "Name", \
                # "Number" and "Objects" are respectively object name, total number of \
                # objects and a list of objects(cdp, 09/2020)
                self.dict_obj_inventory["FluidComponents"]["Name"] = kindObj
                self.dict_obj_inventory["FluidComponents"]["Number"] = numObj
                self.dict_obj_inventory["FluidComponents"]["Objects"] = list()
                for ii in range(1, 1 + numObj):
                    # ["FluidComponents"]["Objects"]: list of FluidComponents objects;
                    # ["Conductor_components"]["Objects"] list of all objects \
                    # (cdp, 09/2020)
                    self.dict_obj_inventory["FluidComponents"]["Objects"].append(
                        FluidComponents(sheet, sheetOpar, ii, dict_file_path)
                    )
                    self.dict_obj_inventory["Conductor_components"]["Objects"].append(
                        self.dict_obj_inventory["FluidComponents"]["Objects"][ii - 1]
                    )
                # end for ii (cdp, 09/2020)
            elif kindObj == "STR_MIX":
                # Define and assign values to MixSCStabilizer dictionary keys: "Name", \
                # "Number" and "Objects" are respectively object name, total number of \
                # objects and a list of objects(cdp, 09/2020)
                self.dict_obj_inventory["MixSCStabilizer"]["Name"] = kindObj
                self.dict_obj_inventory["MixSCStabilizer"]["Number"] = numObj
                self.dict_obj_inventory["MixSCStabilizer"]["Objects"] = list()
                for ii in range(1, 1 + numObj):
                    # ["MixSCStabilizer"]["Objects"]: list of MixSCStabilizer objects;
                    # ["Strands"]["Objects"]: list of Strands objects;
                    # ["SolidComponents"]["Objects"]: list of SolidComponents objects;
                    # ["Conductor_components"]["Objects"]: list of all objects
                    # (cdp, 09/2020)
                    self.dict_obj_inventory["MixSCStabilizer"]["Objects"].append(
                        MixSCStabilizer(simulation, sheet, ii, kindObj, dict_file_path)
                    )
                    self.dict_obj_inventory["Strands"]["Objects"].append(
                        self.dict_obj_inventory["MixSCStabilizer"]["Objects"][ii - 1]
                    )
                    self.dict_obj_inventory["SolidComponents"]["Objects"].append(
                        self.dict_obj_inventory["MixSCStabilizer"]["Objects"][ii - 1]
                    )
                    self.dict_obj_inventory["Conductor_components"]["Objects"].append(
                        self.dict_obj_inventory["MixSCStabilizer"]["Objects"][ii - 1]
                    )
                # end for ii (cdp, 09/2020)
            elif kindObj == "STR_SC":
                # Define and assign values to SuperConductor dictionary keys: "Name", \
                # "Number" and "Objects" are respectively object name, total number of \
                # objects and a list of objects(cdp, 09/2020)
                self.dict_obj_inventory["SuperConductor"]["Name"] = kindObj
                self.dict_obj_inventory["SuperConductor"]["Number"] = numObj
                self.dict_obj_inventory["SuperConductor"]["Objects"] = list()
                for ii in range(1, 1 + numObj):
                    # ["SuperConductor"]["Objects"]: list of SuperConductor objects;
                    # ["Strands"]["Objects"]: list of Strands objects;
                    # ["SolidComponents"]["Objects"]: list of SolidComponents objects;
                    # ["Conductor_components"]["Objects"]: list of all objects
                    # (cdp, 09/2020)
                    self.dict_obj_inventory["SuperConductor"]["Objects"].append(
                        SuperConductor(simulation, sheet, ii, kindObj, dict_file_path)
                    )
                    self.dict_obj_inventory["Strands"]["Objects"].append(
                        self.dict_obj_inventory["SuperConductor"]["Objects"][ii - 1]
                    )
                    self.dict_obj_inventory["SolidComponents"]["Objects"].append(
                        self.dict_obj_inventory["SuperConductor"]["Objects"][ii - 1]
                    )
                    self.dict_obj_inventory["Conductor_components"]["Objects"].append(
                        self.dict_obj_inventory["SuperConductor"]["Objects"][ii - 1]
                    )
                # end for ii (cdp, 09/2020)
            elif kindObj == "STR_STAB":
                # Define and assign values to Stabilizer dictionary keys: "Name", \
                # "Number" and "Objects" are respectively object name, total number of \
                # objects and a list of objects(cdp, 09/2020)
                self.dict_obj_inventory["Stabilizer"]["Name"] = kindObj
                self.dict_obj_inventory["Stabilizer"]["Number"] = numObj
                self.dict_obj_inventory["Stabilizer"]["Objects"] = list()
                for ii in range(1, 1 + numObj):
                    # ["Stabilizer"]["Objects"]: list of Stabilizer objects;
                    # ["Strands"]["Objects"]: list of Strands objects;
                    # ["SolidComponents"]["Objects"]: list of SolidComponents objects;
                    # ["Conductor_components"]["Objects"]: list of all objects
                    # (cdp, 09/2020)
                    self.dict_obj_inventory["Stabilizer"]["Objects"].append(
                        Stabilizer(simulation, sheet, ii, kindObj, dict_file_path)
                    )
                    self.dict_obj_inventory["Strands"]["Objects"].append(
                        self.dict_obj_inventory["Stabilizer"]["Objects"][ii - 1]
                    )
                    self.dict_obj_inventory["SolidComponents"]["Objects"].append(
                        self.dict_obj_inventory["Stabilizer"]["Objects"][ii - 1]
                    )
                    self.dict_obj_inventory["Conductor_components"]["Objects"].append(
                        self.dict_obj_inventory["Stabilizer"]["Objects"][ii - 1]
                    )
                # end for ii (cdp, 09/2020)
            elif kindObj == "Z_JACKET":
                # Define and assign values to Jacket dictionary keys: "Name", \
                # "Number" and "Objects" are respectively object name, total number of \
                # objects and a list of objects(cdp, 09/2020)
                self.dict_obj_inventory["Jacket"]["Name"] = kindObj
                self.dict_obj_inventory["Jacket"]["Number"] = numObj
                self.dict_obj_inventory["Jacket"]["Objects"] = list()
                # Z_JACKET since it must be the last object in the list (cdp, 06/2020)
                for ii in range(1, 1 + numObj):
                    # ["Jacket"]["Objects"]: list of Jacket objects;
                    # ["SolidComponents"]["Objects"]: list of SolidComponents objects;
                    # ["Conductor_components"]["Objects"]: list of all objects
                    # (cdp, 09/2020)
                    self.dict_obj_inventory["Jacket"]["Objects"].append(
                        Jacket(simulation, sheet, ii, kindObj, dict_file_path)
                    )
                    self.dict_obj_inventory["SolidComponents"]["Objects"].append(
                        self.dict_obj_inventory["Jacket"]["Objects"][ii - 1]
                    )
                    self.dict_obj_inventory["Conductor_components"]["Objects"].append(
                        self.dict_obj_inventory["Jacket"]["Objects"][ii - 1]
                    )
                # end for ii (cdp, 09/2020)
            else:
                raise NameError(
                    f"ERROR in method {self.__init__.__name__} of class {self.__class__.__name__}: kind object {kindObj} does not exist.\nPlease check cell A1 in sheet {sheet} of file {self.file_input['STRUCTURE_ELEMENTS']}.\n"
                )
            # end if kindObj (cdp, 11/2020)
        # end for sheetID (cdp, 11/2020)

        # Total number of Strands objects (cdp, 09/2020)
        self.dict_obj_inventory["Strands"]["Number"] = (
            self.dict_obj_inventory["MixSCStabilizer"]["Number"]
            + self.dict_obj_inventory["Stabilizer"]["Number"]
            + self.dict_obj_inventory["SuperConductor"]["Number"]
        )
        # Total number of SolidComponents objects (cdp, 09/2020)
        self.dict_obj_inventory["SolidComponents"]["Number"] = (
            self.dict_obj_inventory["Strands"]["Number"]
            + self.dict_obj_inventory["Jacket"]["Number"]
        )
        # Total number of Conductor component objects (cdp, 09/2020)
        self.dict_obj_inventory["Conductor_components"]["Number"] = (
            self.dict_obj_inventory["FluidComponents"]["Number"]
            + self.dict_obj_inventory["SolidComponents"]["Number"]
        )

    # end method Conductor_components_instance (cdp, 11/2020)

    def conductors_coupling(self):
        pass

    def get_conductor_topology(self, environment):

        """
        Method that evaluate the detailed conductor topology (dict_topology) together with the contatc perimeter (dict_interf_peri). Possible configurations are:
        1) channels in hydraulic parallel;
        2) channels not in hydraulic parallel but in thermal contact;
        3) stand alone channels (no mass or energy exchange);
        4) channels in thermal contact with solid components;
        5) channels not in thermal contact with solid components;
        6) contact between solid components.

        dict_topology describes the full conductor topology, i.e. interface between channels, channels and solid components, solid components as well as the isolated channels; it is organized into three sub dictionaries accessed by keys "ch_ch"; "ch_sol"; "sol_sol".
        Each sub dictionary is characterized by a series of strings that uniquely determines which components are in contact, and a list of object constituted by all the components in contact. Keys of this dictionaries are the ID of the first object in alphabetical order constituting the interface.

        dict_interf_peri holds the values of the contact perimenter. It is also subdivided into three sub dictionaries with the same name as above. It is important to notice that in case of contact between channels, the "Open" and "Close" keys are introduced.
        (cdp, 09/2020)
        """

        # nested dictionaries declarations (cdp, 09/2020)
        self.dict_topology["ch_ch"] = dict()
        self.dict_topology["ch_ch"]["Hydraulic_parallel"] = dict()
        self.dict_topology["ch_ch"]["Thermal_contact"] = dict()
        self.dict_topology["Standalone_channels"] = list()
        self.dict_interf_peri["ch_ch"] = dict()
        self.dict_interf_peri["ch_ch"]["Open"] = dict()
        self.dict_interf_peri["ch_ch"]["Close"] = dict()
        self.dict_topology["ch_sol"] = dict()
        self.dict_interf_peri["ch_sol"] = dict()
        self.dict_topology["sol_sol"] = dict()
        self.dict_interf_peri["sol_sol"] = dict()
        self.dict_interf_peri["env_sol"] = dict()

        # Call method Get_hydraulic_parallel to obtain the channels subdivision \
        # into groups of channels that are in hydraulic parallel.
        self.get_hydraulic_parallel()

        # Nested loop channel-channel (cdp, 09/2020)
        for rr in range(self.dict_obj_inventory["FluidComponents"]["Number"]):
            fluid_comp_r = self.dict_obj_inventory["FluidComponents"]["Objects"][rr]
            for cc in range(
                rr + 1, self.dict_obj_inventory["FluidComponents"]["Number"]
            ):
                fluid_comp_c = self.dict_obj_inventory["FluidComponents"]["Objects"][cc]
                if (
                    self.dict_df_coupling["contact_perimeter_flag"].at[
                        fluid_comp_r.ID, fluid_comp_c.ID
                    ]
                    == 1
                ):
                    # There is at least thermal contact between fluid_comp_r and fluid_comp_c (cdp, 09/2020)
                    # Assign the contact perimeter value (cdp, 09/2020)
                    self.dict_interf_peri["ch_ch"]["Open"][
                        f"{fluid_comp_r.ID}_{fluid_comp_c.ID}"
                    ] = (
                        self.dict_df_coupling["contact_perimeter"].at[
                            fluid_comp_r.ID, fluid_comp_c.ID
                        ]
                        * self.dict_df_coupling["open_perimeter_fract"].at[
                            fluid_comp_r.ID, fluid_comp_c.ID
                        ]
                    )
                    self.dict_interf_peri["ch_ch"]["Close"][
                        f"{fluid_comp_r.ID}_{fluid_comp_c.ID}"
                    ] = self.dict_df_coupling["contact_perimeter"].at[
                        fluid_comp_r.ID, fluid_comp_c.ID
                    ] * (
                        1.0
                        - self.dict_df_coupling["open_perimeter_fract"].at[
                            fluid_comp_r.ID, fluid_comp_c.ID
                        ]
                    )
                    if cc == rr + 1:
                        # declare dictionary flag_found (cdp, 09/2020)
                        flag_found = dict()
                    # Invoke method Get_thermal_contact_channels to search for channels \
                    # that are only in thermal contact (cdp, 09/2020)
                    flag_found = self.get_thermal_contact_channels(
                        rr, cc, fluid_comp_r, fluid_comp_c, flag_found
                    )
                # end self.dict_df_coupling["contact_perimeter_flag"].at[fluid_comp_r.ID, fluid_comp_c.ID] == 1 (cdp, 09/2020)
            # end for cc (cdp, 09/2020)
            if (
                self.dict_topology["ch_ch"]["Thermal_contact"].get(fluid_comp_r.ID)
                != None
            ):
                # key fluid_comp_r.ID exists (cdp, 09/2020)
                if (
                    len(
                        list(
                            self.dict_topology["ch_ch"]["Thermal_contact"][
                                fluid_comp_r.ID
                            ].keys()
                        )
                    )
                    - 3
                    > 0
                ):
                    # There are channels that are in thermal contact (cdp, 09/2020)
                    # Update the number of channels in thermal contact with fluid_comp_r: it \
                    # is the length of list group. The actual number of thermal contacts \
                    # can be larger, since some channels that are in thermal contact may \
                    # belong to groups of channels in hydraulic parallel; it is keep \
                    # into account by the key Actual_number. (cdp, 09/2020)
                    self.dict_topology["ch_ch"]["Thermal_contact"][
                        fluid_comp_r.ID
                    ].update(
                        Number=len(
                            self.dict_topology["ch_ch"]["Thermal_contact"][
                                fluid_comp_r.ID
                            ]["Group"]
                        )
                    )
                    # Assign values to key Actual_number (cdp, 09/2020)
                    self.dict_topology["ch_ch"]["Thermal_contact"][
                        fluid_comp_r.ID
                    ].update(
                        Actual_number=len(
                            list(
                                self.dict_topology["ch_ch"]["Thermal_contact"][
                                    fluid_comp_r.ID
                                ].keys()
                            )
                        )
                        - 3
                        + 1
                    )
                else:
                    # There are not channels that are in thermal contact remove \
                    # key fluid_comp_r.ID from dictionary \
                    # self.dict_topology["ch_ch"]["Thermal_contact"] (cdp, 09/2020)
                    self.dict_topology["ch_ch"]["Thermal_contact"].pop(fluid_comp_r.ID)
            # end if self.dict_topology["ch_ch"]\
            # ["Thermal_contact"].get(fluid_comp_r.ID) != None (cdp, 09/2020)
        # end for rr (cdp, 09/2020)
        # Call method Find_Standalone_channels to search for eventually stand \
        # alone channels (not in hydraulic parallel) (cdp, 09/2020)
        self.find_standalone_channels()
        # dummy dictionary to store channel-solid topology (cdp, 09/2020)
        dict_topology_dummy_ch_sol = dict()
        # dummy to optimize nested loop (cdp, 09/2020)
        dict_chan_s_comp_contact = dict()
        # Nested loop channel-solid (cdp, 09/2020)
        for rr in range(self.dict_obj_inventory["FluidComponents"]["Number"]):
            fluid_comp_r = self.dict_obj_inventory["FluidComponents"]["Objects"][rr]
            # List linked channels-solid initialization (cdp, 09/2020)
            list_linked_chan_sol = list()
            # Nested dictionary in dict_topology_dummy_ch_sol declaration \
            # dict_topology_dummy_ch_sol
            dict_topology_dummy_ch_sol[fluid_comp_r.ID] = dict()
            for cc in range(self.dict_obj_inventory["SolidComponents"]["Number"]):
                s_comp_c = self.dict_obj_inventory["SolidComponents"]["Objects"][cc]
                if (
                    self.dict_df_coupling["contact_perimeter_flag"].at[
                        fluid_comp_r.ID, s_comp_c.ID
                    ]
                    == 1
                ):
                    # There is contact between fluid_comp_r and s_comp_c (cdp, 09/2020)
                    self.dict_interf_peri["ch_sol"][
                        f"{fluid_comp_r.ID}_{s_comp_c.ID}"
                    ] = self.dict_df_coupling["contact_perimeter"].at[
                        fluid_comp_r.ID, s_comp_c.ID
                    ]
                    # Interface identification (cdp, 09/2020)
                    dict_topology_dummy_ch_sol[fluid_comp_r.ID][
                        s_comp_c.ID
                    ] = f"{fluid_comp_r.ID}_{s_comp_c.ID}"
                    # Call method Chan_sol_interfaces (cdp, 09/2020)
                    [
                        dict_chan_s_comp_contact,
                        list_linked_chan_sol,
                    ] = self.chan_sol_interfaces(
                        fluid_comp_r,
                        s_comp_c,
                        dict_chan_s_comp_contact,
                        list_linked_chan_sol,
                    )
                # end if self.dict_df_coupling["contact_perimeter_flag"].at[fluid_comp_r.ID, s_comp_c.ID] == 1: (cdp, 09/2020)
            # end for cc (cdp, 09/2020)
            # Call method Update_interface_dictionary to update dictionaries \
            # (cdp, 09/2020)
            [
                dict_topology_dummy_ch_sol,
                dict_chan_s_comp_contact,
            ] = self.update_interface_dictionary(
                fluid_comp_r,
                dict_topology_dummy_ch_sol,
                dict_chan_s_comp_contact,
                list_linked_chan_sol,
            )
        # end for rr (cdp, 09/2020)
        self.dict_topology.update(ch_sol=dict_topology_dummy_ch_sol)
        # dummy dictionary to store solid-solid topology (cdp, 09/2020)
        dict_topology_dummy_sol = dict()
        # dummy to optimize nested loop (cdp, 09/2020)
        dict_s_comps_contact = dict()
        # Nested loop solid-solid (cdp, 09/2020)
        for rr in range(self.dict_obj_inventory["SolidComponents"]["Number"]):
            s_comp_r = self.dict_obj_inventory["SolidComponents"]["Objects"][rr]
            # List linked solids initialization (cdp, 09/2020)
            list_linked_solids = list()
            # Nested dictionary in dict_topology_dummy_sol declaration \
            # dict_topology_dummy_sol
            dict_topology_dummy_sol[s_comp_r.ID] = dict()
            for cc in range(
                rr + 1, self.dict_obj_inventory["SolidComponents"]["Number"]
            ):
                s_comp_c = self.dict_obj_inventory["SolidComponents"]["Objects"][cc]
                if (
                    self.dict_df_coupling["contact_perimeter_flag"].at[
                        s_comp_r.ID, s_comp_c.ID
                    ]
                    == 1
                ):
                    # There is contact between s_comp_r and s_comp_c (cdp, 09/2020)
                    self.dict_interf_peri["sol_sol"][
                        f"{s_comp_r.ID}_{s_comp_c.ID}"
                    ] = self.dict_df_coupling["contact_perimeter"].at[
                        s_comp_r.ID, s_comp_c.ID
                    ]
                    # Interface identification (cdp, 09/2020)
                    dict_topology_dummy_sol[s_comp_r.ID][
                        s_comp_c.ID
                    ] = f"{s_comp_r.ID}_{s_comp_c.ID}"
                    # Call method Chan_sol_interfaces (cdp, 09/2020)
                    [
                        dict_s_comps_contact,
                        list_linked_solids,
                    ] = self.chan_sol_interfaces(
                        s_comp_r, s_comp_c, dict_s_comps_contact, list_linked_solids
                    )
                # end if self.dict_df_coupling["contact_perimeter_flag"].iat[rr, cc] == 1: (cdp, 09/2020)
            # end for cc (cdp, 09/2020)
            # Call method Update_interface_dictionary to update dictionaries \
            # (cdp, 09/2020)
            [
                dict_topology_dummy_sol,
                dict_s_comps_contact,
            ] = self.update_interface_dictionary(
                s_comp_r,
                dict_topology_dummy_sol,
                dict_s_comps_contact,
                list_linked_solids,
            )
            if (
                self.dict_df_coupling["contact_perimeter_flag"].at[
                    environment.KIND, s_comp_r.ID
                ]
                == 1
            ):
                if (
                    s_comp_r.inputs["Jacket_kind"] == "outer_insulation"
                    or s_comp_r.inputs["Jacket_kind"] == "whole_enclosure"
                ):
                    # There is an interface between environment and s_comp_r.
                    self.dict_interf_peri["env_sol"][
                        f"{environment.KIND}_{s_comp_r.ID}"
                    ] = self.dict_df_coupling["contact_perimeter"].at[
                        environment.KIND, s_comp_r.ID
                    ]
                else:
                    # Raise error
                    raise os.error(
                        f"Jacket of kind {s_comp_r.inputs['Jacket_kind']} can not have and interface with the environment.\n"
                    )
                # End if s_comp_r.inputs["Jacket_kind"]
        # end for rr (cdp, 09/2020)
        self.dict_topology.update(sol_sol=dict_topology_dummy_sol)

    def chan_sol_interfaces(
        self, comp_r, comp_c, dict_comp_interface, list_linked_comp
    ):

        """
        Method that evaluates interfaces between channels and solid components or between solids, and list them in a list of objects to be assigned to dict_topology. (cdp, 09/2020)
        """

        if dict_comp_interface.get(comp_r.ID) == None:
            # No key called comp_r.ID in dictionary dict_comp_interface \
            # (cdp, 09/2020)
            dict_comp_interface[comp_r.ID] = list()
            # In this case necessarily we store both comp_r and comp_c \
            # (cdp, 09/2020)
            list_linked_comp.append(comp_r)
            list_linked_comp.append(comp_c)
        else:  # key comp_r.ID already exist in dict_comp_interface
            # In this case store necessarily only comp_c (cdp, 09/2020)
            list_linked_comp.append(comp_c)
        # end if dict_comp_interface.get(comp_r.ID) (cdp, 09/2020)
        return [dict_comp_interface, list_linked_comp]

    # end method Chan_sol_interfaces (cdp, 09/2020)

    def find_standalone_channels(self):

        """
        Method that searchs for possible isolated (not in hydraulic parallel) channels: search is on each channel in order to not miss anything (cdp, 09/2020)
        """

        # crate dictionary used to understand if channel is or not a stand alone one
        check_found = dict()

        ii = -1
        while ii < self.dict_obj_inventory["FluidComponents"]["Number"] - 1:
            ii = ii + 1
            fluid_comp = self.dict_obj_inventory["FluidComponents"]["Objects"][ii]
            # loop on reference channels (cdp, 09/2020)
            check_found[fluid_comp.ID] = dict(
                Hydraulic_parallel=False, Thermal_contact=False
            )
            for fluid_comp_ref in list(
                self.dict_topology["ch_ch"]["Hydraulic_parallel"].keys()
            ):
                # Search in Hydraulic parallel groups (cdp, 09/2020)
                if check_found[fluid_comp.ID]["Hydraulic_parallel"] == False:
                    if (
                        fluid_comp
                        in self.dict_topology["ch_ch"]["Hydraulic_parallel"][
                            fluid_comp_ref
                        ]["Group"]
                    ):
                        # channel fluid_comp constitutes a group of channels in hydraulic \
                        # parallel thus it can not be a stand alone channel (cdp, 09/2020)
                        # Update dictionart check_found (cdp, 09/2020)
                        check_found[fluid_comp.ID].update(Hydraulic_parallel=True)
            if check_found[fluid_comp.ID]["Hydraulic_parallel"] == False:
                # Channel fluid_comp is not inside Hydraulic parallel groups (cdp, 09/2020)
                for fluid_comp_ref in list(
                    self.dict_topology["ch_ch"]["Thermal_contact"].keys()
                ):
                    # Search in Hydraulic parallel groups (cdp, 09/2020)
                    if check_found[fluid_comp.ID]["Thermal_contact"] == False:
                        if (
                            fluid_comp
                            in self.dict_topology["ch_ch"]["Thermal_contact"][
                                fluid_comp_ref
                            ]["Group"]
                        ):
                            # channel fluid_comp constitutes a thermal contact thus it can not be \
                            # a stand alone channel (cdp, 09/2020)
                            # Update dictionart check_found (cdp, 09/2020)
                            check_found[fluid_comp.ID].update(Thermal_contact=True)
            if (
                check_found[fluid_comp.ID]["Hydraulic_parallel"] == False
                and check_found[fluid_comp.ID]["Thermal_contact"] == False
            ):
                # fluid_comp is a stand alone channel since it does not belong to a group of \
                # channels in hydraulic parallel and it does not constitute a thermal \
                # contact (cdp, 09/2020)
                self.dict_topology["Standalone_channels"].append(fluid_comp)

    # 	N_channel_no_par = len(self.dict_topology["Standalone_channels"])
    # 	if N_channel_no_par == 0:
    # 		print("There are no isolated channels\n")
    # 	elif N_channel_no_par > 0 and N_channel_no_par < \
    # 			 self.dict_obj_inventory["FluidComponents"]["Number"]:
    # 		if N_channel_no_par == 1:
    # 			print(f"""There is {N_channel_no_par} channel that is not in hydraulic parallel: {self.dict_topology["Standalone_channels"][0].ID}\n""")
    # 		else:
    # 			print(f"""There are {N_channel_no_par} channels that are not in hydraulic parallel: {self.dict_topology["Standalone_channels"][:].ID}\n""")
    # 	elif N_channel_no_par == \
    # 			 self.dict_obj_inventory["FluidComponents"]["Number"]:
    # 		print("All channels are isolated\n")
    # 	else:
    # 		print(f"Something does not work\n")
    # end method Find_Standalone_channels (cdp, 09/2020)

    def update_interface_dictionary(
        self, comp, dict_topology_dummy, dict_contacts, list_contacts
    ):

        dict_contacts[comp.ID] = list_contacts
        dict_topology_dummy[comp.ID].update(Group=list_contacts)
        dict_topology_dummy[comp.ID].update(Number=len(list_contacts))
        if dict_topology_dummy[comp.ID]["Number"] == 0:
            # Removed empty keys from dictionaries (cdp, 09/2020)
            dict_topology_dummy.pop(comp.ID)
            dict_contacts.pop(comp.ID)
        return [dict_topology_dummy, dict_contacts]

    # end method Update_interface_dictionary (cdp, 09/2020)

    def get_hydraulic_parallel(self):

        """
        Method that interprets the information in table self.dict_df_coupling["open_perimeter_fract"] understanding if there are channels in hydraulic parallel and how the are organized into groups. The method returns a dictionary with:
        1) the ID of the reference channel of each group
        2) a list of all the channels that belongs to a group
        3) for each group the IDs of all the linked channels organized into lists
        (cdp, 09/2020)
        """

        full_ind = dict()
        # get row and column index of non zero matrix elements (cdp, 09/2020)
        [full_ind["row"], full_ind["col"]] = np.nonzero(
            self.dict_df_coupling["open_perimeter_fract"].iloc[1:, 1:].to_numpy()
        )
        # USEFUL QUANTITIES AND VARIABLES (cdp, 09/2020)
        # array of the not considered array (cdp, 09/2020)
        already = dict(no=np.unique(np.union1d(full_ind["row"], full_ind["col"])))
        # list of the already considered array (cdp, 09/2020)
        already["yes"] = -1 * np.ones(already["no"].shape, dtype=int)
        # index to used to update key "yes" (cdp, 09/2020)
        already["ii_yes"] = 0
        # Define dictionary dict_topology. This is different from \
        # self.dict_topology which is a class Conductors attribute. A the end of \
        # this method self.dict_topology will be updated with the values in \
        # dict_topology (cdp, 09/2020)
        dict_topology = dict()
        # Define dictionary check to be sure that for each channel both searchs \
        # are performed (cdp, 09/2020)
        check = dict()

        Total_connections = len(full_ind["row"])
        total_connections_counter = 0
        group_counter = 0
        # loop until all channel connections are realized (cdp, 09/2020)
        while total_connections_counter < Total_connections:
            # get the reference channel: it is the one characterized by the minimum \
            # index value in array already["no"] (cdp, 09/2020)
            fluid_comp_ref_row_ind = min(already["no"])
            fluid_comp_ref = self.dict_obj_inventory["FluidComponents"]["Objects"][
                fluid_comp_ref_row_ind
            ]
            # update dictionary already (cdp, 09/2020)
            already["no"] = np.delete(already["no"], 0, 0)
            if fluid_comp_ref_row_ind not in already["yes"]:
                already["yes"][already["ii_yes"]] = fluid_comp_ref_row_ind
                already.update(ii_yes=already["ii_yes"] + 1)
            # Construct check dictionary (cdp, 09/2020)
            check[fluid_comp_ref.ID] = dict()
            # Get minimum and maximum index of array full_ind["row"] that correspond \
            # to the reference channel (cdp, 09/2020)
            boundary = dict(
                fluid_comp_ref_lower=min(
                    np.nonzero(full_ind["row"] == fluid_comp_ref_row_ind)[0]
                ),
                fluid_comp_ref_upper=max(
                    np.nonzero(full_ind["row"] == fluid_comp_ref_row_ind)[0]
                ),
            )
            # get all the channels that are directly in contact with reference \
            # channel (cdp, 09/2020)
            ind_direct = full_ind["col"][
                boundary["fluid_comp_ref_lower"] : boundary["fluid_comp_ref_upper"] + 1
            ]
            # Update dictionary dict_topology
            dict_topology[fluid_comp_ref.ID] = dict(
                Ref_channel=fluid_comp_ref.ID, Group=list(), Number=0
            )
            dict_topology[fluid_comp_ref.ID]["Group"].append(fluid_comp_ref)
            for ch_index in ind_direct:
                # get channel (cdp, 09/2020)
                fluid_comp = self.dict_obj_inventory["FluidComponents"]["Objects"][
                    ch_index
                ]
                # Construct check dictionary (cdp, 09/2020)
                check[fluid_comp_ref.ID][fluid_comp.ID] = dict(row=False, col=False)
                # find the index in array already["no"] of the element that must be \
                # deleted (cdp, 09/2020)
                i_del = np.nonzero(already["no"] == ch_index)[0]
                # update total_connections_counter (cdp, 09/2020)
                total_connections_counter = total_connections_counter + 1
                # update dictionary already (cdp, 09/2020)
                already["no"] = np.delete(already["no"], i_del, 0)
                if ch_index not in already["yes"]:
                    already["yes"][already["ii_yes"]] = ch_index
                    already.update(ii_yes=already["ii_yes"] + 1)
                dict_topology[fluid_comp_ref.ID][fluid_comp.ID] = [
                    f"{fluid_comp_ref.ID}_{fluid_comp.ID}"
                ]
                dict_topology[fluid_comp_ref.ID]["Group"].append(fluid_comp)
            # end for ii (cdp, 09/2020)
            # for each channel that is in direct contact with the reference one, \
            # search if it is in contact with other channels, constituting an \
            # indirect contact with the reference channel. This is done in a \
            # different loop because total_connections_counter must be fully updated \
            # (cdp, 09/2020)
            for ch_index in ind_direct:
                # get channel (cdp, 09/2020)
                fluid_comp = self.dict_obj_inventory["FluidComponents"]["Objects"][
                    ch_index
                ]
                # Initialize key value "variable_lower" of dictionary boundary. This \
                # parameter is used to look only in the region of not directly \
                # connected channels and is updated to consider only the data below \
                # this index value. Initialization must be done at each iteration in \
                # order to not miss some index during the search. (cdp, 09/2020)
                boundary.update(variable_lower=boundary["fluid_comp_ref_upper"] + 1)
                if check[fluid_comp_ref.ID][fluid_comp.ID]["col"] == False:
                    # The search in array full_ind["col"] is not performed yet \
                    # (cdp, 09/2020)
                    total_connections_counter = self.search_on_ind_col(
                        ch_index,
                        full_ind,
                        dict_topology,
                        fluid_comp_ref,
                        fluid_comp,
                        check,
                        already,
                        total_connections_counter,
                        boundary,
                    )
                if check[fluid_comp_ref.ID][fluid_comp.ID]["row"] == False:
                    # The search in array full_ind["row"] is not performed yet \
                    # (cdp, 09/2020)
                    total_connections_counter = self.search_on_ind_row(
                        ch_index,
                        full_ind,
                        dict_topology,
                        fluid_comp_ref,
                        fluid_comp,
                        check,
                        already,
                        total_connections_counter,
                        boundary,
                    )
            # end for (cdp, 09/2020)
            # Sort list Group by channel ID (cdp, 09/2020)
            dict_topology[fluid_comp_ref.ID].update(
                Group=sorted(
                    dict_topology[fluid_comp_ref.ID]["Group"], key=lambda ch: ch.ID
                )
            )
            # Get the number of channels that are in hydraulic parallel for each \
            # reference channel (cdp, 10/2020)
            dict_topology[fluid_comp_ref.ID].update(
                Number=len(dict_topology[fluid_comp_ref.ID]["Group"])
            )
            if total_connections_counter == Total_connections:
                if group_counter == 0:
                    # Only one group of channels in hydraulic parallel (cdp, 09/2020)
                    group_counter = group_counter + 1
                    print(
                        f"There is only {group_counter} group of channels in hydraulic parallel.\n"
                    )
                elif group_counter > 0:
                    # There are a least two groups of channels in hydraulic parallel \
                    # (cdp, 09/2020)
                    group_counter = group_counter + 1
                    print(
                        f"There are {group_counter} groups of channels in hydraulic parallel\n"
                    )
            elif total_connections_counter < Total_connections:
                # The number of groups of channels in hydraulic parallel is > 1.
                # Repeat the above procedure. Keep in mind that the reference channel \
                # is evaluated as the one with the minimum ID between the ones that \
                # are not connected yet (cdp, 09/2020)
                group_counter = group_counter + 1
            elif total_connections_counter > Total_connections:
                # Something odd occurs
                raise ValueError(
                    "ERROR! The counter of the total connections can not be larger than the number of total connections! Something odd occurs!\n"
                )
            # end if total_connections_counter (cdp, 09/2020)
        # end while (cdp, 09/2020)
        # Update key Hydraulic_parallel of dictionary dict_topology (cdp, 09/2020)
        self.dict_topology["ch_ch"].update(Hydraulic_parallel=dict_topology)

    # end method Get_hydraulic_parallel (cdp, 09/2020)

    def search_on_ind_col(
        self,
        ch_index,
        full_ind,
        dict_topology,
        fluid_comp_ref,
        fluid_comp_c,
        check,
        already,
        total_connections_counter,
        boundary,
    ):

        """
        method that search in array full_ind["col"] if there are other recall to channel fluid_comp_c (cdp, 09/2020)
        """

        # N.B ch_index is the channel that is in contact with the fluid_comp_ref or \
        # another channel (cdp, 09/2020)

        # The search in array full_ind["col"] will be performed so flag check \
        # [fluid_comp_ref.ID][fluid_comp.ID]["col"] can be set to True. (cdp, 09/2020)
        check[fluid_comp_ref.ID][fluid_comp_c.ID].update(col=True)
        # search for all the values that are equal to ch_index in array \
        # full_ind["col"], excluding the index of the direct contact region, and
        # store the corresponding indices (cdp, 09/2020)
        ind_found = (
            np.nonzero(full_ind["col"][boundary["variable_lower"] :] == ch_index)[0]
            + boundary["variable_lower"]
        )
        if len(ind_found) == 0:
            if check[fluid_comp_ref.ID][fluid_comp_c.ID]["row"] == False:
                # Search if there is some value equal to ch_index in the array \
                # full_ind["row"] calling method Search_on_ind_row (cdp, 09/2020)
                total_connections_counter = self.search_on_ind_row(
                    ch_index,
                    full_ind,
                    dict_topology,
                    fluid_comp_ref,
                    fluid_comp_c,
                    check,
                    already,
                    total_connections_counter,
                    boundary,
                )
            elif check[fluid_comp_ref.ID][fluid_comp_c.ID]["row"] == True:
                # no channel with smaller channel index is connected with ch_index \
                # (cdp, 09/2020)
                print(f"No other channels are connected to {fluid_comp_c.ID}\n")
        elif len(ind_found) > 0:
            # there is at least one channel with smaller channel index that is \
            # connected with ch_index (cdp, 09/2020)
            # update key "variable_lower": the next search is done from this index \
            # value up to the end of arrays full_ind["row"] and full_ind["col"] \
            # (cdp, 09/2020)
            boundary.update(variable_lower=ind_found[0] + 1)
            # get the index of the linked channel(s) (col -> row) (cdp, 09/2020)
            ind_link = full_ind["row"][ind_found[0 : len(ind_found)]]
            # loop linked channels (cdp, 09/2020)
            jj = -1
            while jj < len(ind_link) - 1:
                jj = jj + 1
                ch_index = ind_link[jj]
                # get channel
                fluid_comp_r = self.dict_obj_inventory["FluidComponents"]["Objects"][
                    ch_index
                ]
                if (
                    dict_topology[fluid_comp_ref.ID].get(fluid_comp_r.ID) != None
                    and f"{fluid_comp_r.ID}_{fluid_comp_c.ID}"
                    in dict_topology[fluid_comp_ref.ID][fluid_comp_r.ID]
                ):
                    # Skip operations since the connection is already preformed \
                    # (cdp, 09/2020)
                    jj = jj + 1
                else:
                    if check[fluid_comp_ref.ID].get(fluid_comp_r.ID) == None:
                        # Update dictionary check: add key fluid_comp_r.ID (cdp, 09/2020)
                        check[fluid_comp_ref.ID][fluid_comp_r.ID] = dict(
                            row=False, col=False
                        )
                    # construct channels link (cdp, 09/2020)
                    if dict_topology[fluid_comp_ref.ID].get(fluid_comp_r.ID) == None:
                        # key fluid_comp_r.ID does not exist, so it is added to the dictionary \
                        # and a list is created (cdp, 09/2020)
                        dict_topology[fluid_comp_ref.ID][fluid_comp_r.ID] = [
                            f"{fluid_comp_r.ID}_{fluid_comp_c.ID}"
                        ]
                        print(
                            dict_topology[fluid_comp_ref.ID][fluid_comp_r.ID][-1] + "\n"
                        )
                        # update total_connections_counter (cdp, 09/2020)
                        total_connections_counter = total_connections_counter + 1
                    else:
                        # key fluid_comp_r.ID exists (cdp, 09/2020)
                        if (
                            f"{fluid_comp_r.ID}_{fluid_comp_c.ID}"
                            not in dict_topology[fluid_comp_ref.ID][fluid_comp_r.ID]
                        ):
                            # List is updated only if the contact ID is not already in the \
                            # list (cdp, 09/2020)
                            dict_topology[fluid_comp_ref.ID][fluid_comp_r.ID].append(
                                f"{fluid_comp_r.ID}_{fluid_comp_c.ID}"
                            )
                            print(
                                dict_topology[fluid_comp_ref.ID][fluid_comp_r.ID][-1]
                                + "\n"
                            )
                            # update total_connections_counter (cdp, 09/2020)
                            total_connections_counter = total_connections_counter + 1
                    # end if dict_topology[fluid_comp_ref.ID].get(fluid_comp_r.ID) == None \
                    # (cdp, 09/2020)
                    # find the index in array already["no"] of the element that must be \
                    # deleted (cdp, 09/2020)
                    i_del = np.nonzero(already["no"] == ch_index)[0]
                    # update dictionary already (cdp, 09/2020)
                    already["no"] = np.delete(already["no"], i_del, 0)
                    if ch_index not in already["yes"]:
                        already["yes"][already["ii_yes"]] = ch_index
                        already.update(ii_yes=already["ii_yes"] + 1)
                    if fluid_comp_r not in dict_topology[fluid_comp_ref.ID]["Group"]:
                        # Add channel fluid_comp_r to list Group (cdp, 09/2020)
                        dict_topology[fluid_comp_ref.ID]["Group"].append(fluid_comp_r)
                    # call method Search_on_ind_row with to search if channel fluid_comp_r \
                    # is linked to other channels (cdp, 09/2020)
                    total_connections_counter = self.search_on_ind_row(
                        ch_index,
                        full_ind,
                        dict_topology,
                        fluid_comp_ref,
                        fluid_comp_r,
                        check,
                        already,
                        total_connections_counter,
                        boundary,
                    )
                # end if dict_topology[fluid_comp_ref.ID].get(fluid_comp_r.ID) != None and \
                # f"{fluid_comp_r.ID}_{fluid_comp_c.ID}" in dict_topology[fluid_comp_ref.ID][fluid_comp_r.ID] \
                # (cdp, 09/2020)
            # end while (cdp, 09/2020)
        # end if len(ind_found) (cdp, 09/2020)
        return total_connections_counter

    # end method Search_on_ind_col (cdp, 09/2020)

    def search_on_ind_row(
        self,
        ch_index,
        full_ind,
        dict_topology,
        fluid_comp_ref,
        fluid_comp_r,
        check,
        already,
        total_connections_counter,
        boundary,
    ):

        """
        Method that search in array full_ind["row"] if there are other recall to channel fluid_comp_c (cdp, 09/2020)
        """

        # The search in array full_ind["row"] will be performed so flag check \
        # [fluid_comp_ref.ID][fluid_comp.ID]["row"] can be set to True. (cdp, 09/2020)
        check[fluid_comp_ref.ID][fluid_comp_r.ID].update(row=True)

        # search for all the values that are equal to ch_index in array \
        # full_ind["row"], excluding the index of the direct contact region, and \
        # store the corresponding indices (cdp, 09/2020)
        ind_found = (
            np.nonzero(full_ind["row"][boundary["variable_lower"] :] == ch_index)[0]
            + boundary["variable_lower"]
        )
        if len(ind_found) == 0:
            if check[fluid_comp_ref.ID][fluid_comp_r.ID]["col"] == False:
                # Search if there is some value equal to ch_index in the array \
                # full_ind["col"] calling method Search_on_ind_col (cdp, 09/2020)
                total_connections_counter = self.search_on_ind_col(
                    ch_index,
                    full_ind,
                    dict_topology,
                    fluid_comp_ref,
                    fluid_comp_r,
                    check,
                    already,
                    total_connections_counter,
                    boundary,
                )
            elif check[fluid_comp_ref.ID][fluid_comp_r.ID]["col"] == True:
                # no channel with larger channel index is connected with ch_index \
                # (cdp, 09/2020)
                print(f"No other channels are connected to {fluid_comp_r.ID}\n")
        elif len(ind_found) > 0:
            # there is at least one channel with larger channel index that is \
            # connected with ch_index (cdp, 09/2020)
            # update key "variable_lower": the next search is done from this index \
            # value up to the end of arrays full_ind["row"] and full_ind["col"] \
            # (cdp, 09/2020)
            boundary.update(variable_lower=ind_found[0] + 1)
            # get the index of the linked channel(s) (row -> col) (cdp, 09/2020)
            ind_link = full_ind["col"][ind_found[0 : len(ind_found)]]
            # loop linked channels (cdp, 09/2020)
            jj = -1
            while jj < len(ind_link) - 1:
                jj = jj + 1
                ch_index = ind_link[jj]
                # get channel (cdp, 09/2020)
                fluid_comp_c = self.dict_obj_inventory["FluidComponents"]["Objects"][
                    ch_index
                ]
                if (
                    dict_topology[fluid_comp_ref.ID].get(fluid_comp_c.ID) != None
                    and f"{fluid_comp_c.ID}_{fluid_comp_r.ID}"
                    in dict_topology[fluid_comp_ref.ID][fluid_comp_c.ID]
                ):
                    # Skip operations since the connection is already preformed \
                    # (cdp, 09/2020)
                    jj = jj + 1
                else:
                    if check[fluid_comp_ref.ID].get(fluid_comp_c.ID) == None:
                        # Update dictionary check: add key fluid_comp_c.ID (cdp, 09/2020)
                        check[fluid_comp_ref.ID][fluid_comp_c.ID] = dict(
                            row=False, col=False
                        )
                    # construct channels link (cdp, 09/2020)
                    if dict_topology[fluid_comp_ref.ID].get(fluid_comp_r.ID) == None:
                        # key fluid_comp_c.ID does not exist, so it is added to the dictionary \
                        # and a list is created (cdp, 09/2020)
                        dict_topology[fluid_comp_ref.ID][fluid_comp_r.ID] = [
                            f"{fluid_comp_r.ID}_{fluid_comp_c.ID}"
                        ]
                        print(
                            dict_topology[fluid_comp_ref.ID][fluid_comp_r.ID][-1] + "\n"
                        )
                        # update total_connections_counter (cdp, 09/2020)
                        total_connections_counter = total_connections_counter + 1
                    else:
                        # key fluid_comp_c.ID exists (cdp, 09/2020)
                        if (
                            f"{fluid_comp_r.ID}_{fluid_comp_c.ID}"
                            not in dict_topology[fluid_comp_ref.ID][fluid_comp_r.ID]
                        ):
                            # List is updated only if the contact ID is not already in the \
                            # list (cdp, 09/2020)
                            dict_topology[fluid_comp_ref.ID][fluid_comp_r.ID].append(
                                f"{fluid_comp_r.ID}_{fluid_comp_c.ID}"
                            )
                            print(
                                dict_topology[fluid_comp_ref.ID][fluid_comp_r.ID][-1]
                                + "\n"
                            )
                            # update total_connections_counter (cdp, 09/2020)
                            total_connections_counter = total_connections_counter + 1
                    # end if dict_topology[fluid_comp_ref.ID].get(fluid_comp_c.ID) == None \
                    # (cdp, 09/2020)
                    # find the index in array already["no"] of the element that must be \
                    # deleted (cdp, 09/2020)
                    i_del = np.nonzero(already["no"] == ch_index)[0]
                    # update dictionary already (cdp, 09/2020)
                    already["no"] = np.delete(already["no"], i_del, 0)
                    if ch_index not in already["yes"]:
                        already["yes"][already["ii_yes"]] = ch_index
                        already.update(ii_yes=already["ii_yes"] + 1)
                    if fluid_comp_c not in dict_topology[fluid_comp_ref.ID]["Group"]:
                        # Add channel fluid_comp_c to list Group (cdp, 09/2020)
                        dict_topology[fluid_comp_ref.ID]["Group"].append(fluid_comp_c)
                    # call method Search_on_ind_col to search if channel fluid_comp_c is \
                    # linked to other channels (cdp, 09/2020)
                    total_connections_counter = self.search_on_ind_col(
                        ch_index,
                        full_ind,
                        dict_topology,
                        fluid_comp_ref,
                        fluid_comp_c,
                        check,
                        already,
                        total_connections_counter,
                        boundary,
                    )
                # end if dict_topology[fluid_comp_ref.ID].get(fluid_comp_c.ID) != None and \
                # f"{fluid_comp_c.ID}_{fluid_comp_r.ID}" in dict_topology[fluid_comp_ref.ID][fluid_comp_c.ID] \
                # (cdp, 09/2020)
            # end while (cdp, 09/2020)
        # end if len(ind_found) (cdp, 09/2020)
        return total_connections_counter

    # end method Search_on_ind_row (cdp, 09/2020)

    def get_thermal_contact_channels(
        self, rr, cc, fluid_comp_r, fluid_comp_c, flag_found
    ):

        """
        Method that recognize if there are some channels that are only in thermal contact with other channels. If one or both of the two considered channels belongs also to two different groups of channels in hydraulic parallel, they are not included in the list called Group, however the thermal contact is indicated in a suitable key-value pair. This is because channels that have both the properties of being in thermal contact and in hydraulic parallel should be threated considering the latter, while flow initialization is performed. (cdp, 09/2020)
        """

        if self.dict_topology["ch_ch"]["Thermal_contact"].get(fluid_comp_r.ID) == None:
            # Update dictionary self.dict_topology["ch_ch"]["Thermal_contact"] \
            # (cdp, 09/2020)
            # key fluid_comp_r.ID does not already exist (cdp, 09/2020)
            self.dict_topology["ch_ch"]["Thermal_contact"][fluid_comp_r.ID] = dict(
                Group=list(), Number=0, Actual_number=0
            )
        if (
            self.dict_df_coupling["open_perimeter_fract"].at[
                fluid_comp_r.ID, fluid_comp_c.ID
            ]
            == 0.0
        ):
            # There is only thermal contact between fluid_comp_r and fluid_comp_c \
            # (cdp, 09/2020)
            # Update dictionary self.dict_topology["ch_ch"]["Thermal_contact"] \
            # (cdp, 09/2020)
            self.dict_topology["ch_ch"]["Thermal_contact"][fluid_comp_r.ID][
                fluid_comp_c.ID
            ] = f"{fluid_comp_r.ID}_{fluid_comp_c.ID}"
            if len(list(self.dict_topology["ch_ch"]["Hydraulic_parallel"].keys())) > 0:
                # There is at least one group of channels in hydraulic parallel \
                # (cdp, 09/2020)
                if flag_found.get(fluid_comp_r.ID) == None:
                    # initialize key fluid_comp_r.ID of dictionary flag_found to False \
                    # only once (cdp, 09/2020)
                    flag_found[fluid_comp_r.ID] = False
                # initialize key fluid_comp_c.ID of dictionary flag_found to False for \
                # each value of cc (cdp, 09/2020)
                flag_found[fluid_comp_c.ID] = False
                # Following lines check if one of this two channels belongs to a
                # group of channels in hydraulic parallel. In this case only the
                # one that does not belong to the group is added to the list of
                # channels with only thermal contact. (cdp, 09/2020)
                for fluid_comp_ref in list(
                    self.dict_topology["ch_ch"]["Hydraulic_parallel"].keys()
                ):
                    # Search if fluid_comp_r and fluid_comp_c are already inserted into two
                    # different groups of channels in parallel. (cdp, 09/2020)
                    if (
                        flag_found[fluid_comp_r.ID] == False
                        and fluid_comp_r
                        in self.dict_topology["ch_ch"]["Hydraulic_parallel"][
                            fluid_comp_ref
                        ]["Group"]
                    ):
                        # fluid_comp_r is in found in a group of channels in hydraulic \
                        # parallel (cdp, 09/2020)
                        flag_found[fluid_comp_r.ID] = True
                    # end if cc == rr + 1 (cdp, 09/2020)
                    if (
                        flag_found[fluid_comp_c.ID] == False
                        and fluid_comp_c
                        in self.dict_topology["ch_ch"]["Hydraulic_parallel"][
                            fluid_comp_ref
                        ]["Group"]
                    ):
                        # fluid_comp_c is in found in a group of channels in hydraulic \
                        # parallel (cdp, 09/2020)
                        flag_found[fluid_comp_c.ID] = True
                # end for fluid_comp_ref (cdp, 09/2020)
                if flag_found[fluid_comp_r.ID] == False:
                    # fluid_comp_r is not in hydraulic parallel with other channels \
                    # (cdp, 09/2020)
                    self.dict_topology["ch_ch"]["Thermal_contact"][fluid_comp_r.ID][
                        "Group"
                    ].append(fluid_comp_r)
                if flag_found[fluid_comp_c.ID] == False:
                    # fluid_comp_c is not in hydraulic parallel with other channels
                    # (cdp, 09/2020)
                    self.dict_topology["ch_ch"]["Thermal_contact"][fluid_comp_r.ID][
                        "Group"
                    ].append(fluid_comp_c)
            else:
                # there are no groups of channels in hydraulic parallel \
                # (cdp, 09/2020)
                if cc == rr + 1:
                    # Add fluid_comp_r to the list Group only once (cdp, 09/2020)
                    self.dict_topology["ch_ch"]["Thermal_contact"][fluid_comp_r.ID][
                        "Group"
                    ].append(fluid_comp_r)
                # Add channel fluid_comp_c to the list Group (cdp, 09/2020)
                self.dict_topology["ch_ch"]["Thermal_contact"][fluid_comp_r.ID][
                    "Group"
                ].append(fluid_comp_c)
            # end if len(list(self.dict_topology["ch_ch"]\
            # ["Hydraulic_parallel"].keys())) > 0 (cdp, 09/2020)
        # end if self.dict_df_coupling["open_perimeter_fract"].iat[rr, cc] == 0.0 (cdp, 09/2020)
        return flag_found

    # end method Get_thermal_contact_channels (cdp, 09/2020)

    ############################################################################### this method initialize the Conductor on the base of the input parameters
    ##############################################################################

    def initialization(self, simulation, gui):

        time_simulation = simulation.simulation_time[-1]
        sim_name = simulation.transient_input["SIMULATION"]
        # Construct spatial discretization
        conductor_spatial_discretization(simulation, self)
        # Total number of equations for each conductor (cdp, 09/2020)
        self.dict_N_equation["Total"] = (
            self.dict_N_equation["NODOFS"] * self.dict_discretization["N_nod"]
        )
        # initialize conductor time values, it can be different for different \
        # conductors since the conductor time step can be different (cdp, 10/202)
        self.cond_time = [time_simulation]
        # Initialize conductor time step counter
        self.cond_num_step = 0

        """
    if (imsourcefun.eq.0) then
      premed(icond)=dble((preinl(icond)+preout(icond))/2.)
      temmed(icond)=dble((teminl(icond)+temout(icond))/2.)
    endif
      CALL GETsource  (NNODES,ncond) ### si trova in 
    ENTALINLET = hhe(premed(icond),temmed(icond))
    Questo per il momento non lo considero!!!!! (cdp)
    """

        gen_flow(self)
        # C*SET THE INITIAL VALUE OF THE FLOW VARIABLES (LINEAR P AND T)

        temp_ave = np.zeros(self.dict_discretization["N_nod"])
        self.enthalpy_balance = 0.0
        self.enthalpy_out = 0.0
        self.enthalpy_inl = 0.0
        # For each channel evaluate following fluid properties (array):velocity, \
        # pressure, temperature, density, Reynolds, Prandtl (cdp, 06/2020)
        # N.B. questo loop si potrebbe fare usando map.
        for fluid_comp in self.dict_obj_inventory["FluidComponents"]["Objects"]:
            # Compute pressure, temperature and velocity in nodal points according to the initial conditions
            fluid_comp.coolant._eval_nodal_pressure_temperature_velocity_initialization(
                self
            )
            # Evaluate the other coolant properties in nodal points (nodal = True by default)
            fluid_comp.coolant._eval_properties_nodal_gauss(
                self, simulation.fluid_prop_aliases
            )
            # Evaluate the density (if necessary) and the mass flow rate in points (nodal = True by default)
            fluid_comp.coolant._compute_density_and_mass_flow_rates_nodal_gauss(self)
            temp_ave = (
                temp_ave
                + fluid_comp.coolant.dict_node_pt["temperature"]
                / self.dict_obj_inventory["FluidComponents"]["Number"]
            )
            # Enthaly balance: dt*sum((mdot*w)_out - (mdot*w)_inl), used to check \
            # the imposition of SolidComponents temperature initial spatial \
            # distribution (cdp, 12/2020)
            # N.B. queste istruzioni posso inserirle in un metodo della classe.
            self.enthalpy_balance = self.enthalpy_balance + simulation.transient_input[
                "STPMIN"
            ] * (
                fluid_comp.coolant.dict_node_pt["mass_flow_rate"][-1]
                * fluid_comp.coolant.dict_node_pt["total_enthalpy"][-1]
                - fluid_comp.coolant.dict_node_pt["mass_flow_rate"][0]
                * fluid_comp.coolant.dict_node_pt["total_enthalpy"][0]
            )
            self.enthalpy_out = (
                self.enthalpy_out
                + simulation.transient_input["STPMIN"]
                * fluid_comp.coolant.dict_node_pt["mass_flow_rate"][-1]
                * fluid_comp.coolant.dict_node_pt["total_enthalpy"][-1]
            )
            self.enthalpy_inl = (
                self.enthalpy_inl
                + simulation.transient_input["STPMIN"]
                * fluid_comp.coolant.dict_node_pt["mass_flow_rate"][0]
                * fluid_comp.coolant.dict_node_pt["total_enthalpy"][0]
            )

        ## For each solid component evaluate temperature (cdp, 07/2020)
        ## If needed read only the sub matrix describing channel - solid objects \
        ## contact (cdp, 07/2020)
        ## nested loop on channel - solid objects (cpd 07/2020)
        # for cc in range(self.dict_obj_inventory["SolidComponents"]["Number"]):
        # 	s_comp = self.dict_obj_inventory["SolidComponents"]["Objects"][cc]
        # 	weight = self.dict_df_coupling["contact_perimeter"][0:\
        # 					 self.dict_obj_inventory["FluidComponents"]["Number"], \
        # 					 cc + self.dict_obj_inventory["FluidComponents"]["Number"]]
        # 	s_comp.dict_node_pt["temperature"] = \
        # 																np.zeros(self.dict_discretization["N_nod"])
        # 	if np.sum(weight) > 0:
        # 		# evaluate SolidComponents temperature as the weighted average on \
        # 		# conctat_perimeter with channels (cpd 07/2020)
        # 		for rr in range(self.dict_obj_inventory["FluidComponents"]["Number"]):
        # 			fluid_comp = self.dict_obj_inventory["FluidComponents"]["Objects"][rr]
        # 			s_comp.dict_node_pt["temperature"] = \
        # 																	s_comp.dict_node_pt["temperature"] + \
        # 																	fluid_comp.coolant.dict_node_pt["temperature"]*\
        # 																	weight[rr]/np.sum(weight)
        # 	else:
        # 		# evaluate SolidComponents temperature as the algebraic average of \
        # 		# channels temperature (cdp, 07/2020)
        # 		s_comp.dict_node_pt["temperature"] = temp_ave
        # 		# scalar (cdp, 07/2020)
        # 		s_comp.TEMOUT = s_comp.dict_node_pt["temperature"][-1]
        # 	# call function f_plot to make property plots (cdp, 07/2020)
        # 	#f_plot(s_comp, self.dict_discretization["xcoord"])

        # Call function SolidComponents_T_initialization to initialize \
        # SolidComponents temperature spatial distribution from FluidComponents \
        # temperature or from input values according to flag INTIAL (cdp, 12/2020)
        solid_components_temperature_initialization(self)

        # Nested loop jacket - jacket.
        for rr in range(self.dict_obj_inventory["Jacket"]["Number"]):
            jacket_r = self.dict_obj_inventory["Jacket"]["Objects"][rr]
            # np array of shape (Node, 1) to avoid broadcasting error.
            jacket_r.radiative_heat_env = np.zeros(
                (jacket_r.dict_node_pt["temperature"].size, 1)
            )
            for cc in range(rr + 1, self.dict_obj_inventory["Jacket"]["Number"]):
                jacket_c = self.dict_obj_inventory["Jacket"]["Objects"][cc]
                jacket_r.radiative_heat_inn[f"{jacket_r.ID}_{jacket_c.ID}"] = np.zeros(
                    (jacket_r.dict_node_pt["temperature"].size, 1)
                )
                jacket_c.radiative_heat_inn[f"{jacket_r.ID}_{jacket_c.ID}"] = np.zeros(
                    (jacket_c.dict_node_pt["temperature"].size, 1)
                )
            # End for cc.
        # End for rr.

        #### INITIAL OPERATING CURRENT
        self.IOP_TOT = self.inputs["I0_OP_TOT"]
        # IOP=IOP0(icond)

        # call functions get_current, get_magnetic_field, get_magnetic_field_gradient, \
        # get_superconductor_critical_prop, get_heat, jhtflx_new_0, \
        # set_energy_counters, Get_transp_coeff @ time = 0 to perform \
        # initialization (cdp, 07/2020)
        self.operating_conditions(simulation)
        # Loop on SolidComponents (cdp, 01/2021)
        # N.B. questo loop si potrebbe fare usando map.
        for s_comp in self.dict_obj_inventory["SolidComponents"]["Objects"]:
            # compute, average density, thermal conductivity, specifi heat at \
            # constant pressure and electrical resistivity at initial \
            # SolidComponents temperature in nodal points (cdp, 01/2021)
            s_comp.eval_sol_comp_properties(self.dict_obj_inventory)
        # end for s_comp.

        # Initialize the Energy of the SolidComponents (cdp, 12/2020)
        self.E_sol_ini = 0.0
        self.E_sol_fin = 0.0
        self.E_str_ini = 0.0
        self.E_str_fin = 0.0
        self.E_jk_ini = 0.0
        self.E_jk_fin = 0.0
        # Loop on SolidComponents to evaluate the total initial energy of \
        # SolidComponents, used to check the imposition of SolidComponents \
        # temperature initial spatial distribution (cdp, 12/2020)
        # N.B. questo loop si potrebbe fare usando map.
        for s_comp in self.dict_obj_inventory["SolidComponents"]["Objects"]:
            # N.B. queste istruzioni posso inserirle in un metodo della classe.
            self.E_sol_ini = self.E_sol_ini + s_comp.inputs["CROSSECTION"] * np.sum(
                (
                    self.dict_discretization["xcoord"][
                        1 : self.dict_discretization["N_nod"]
                    ]
                    - self.dict_discretization["xcoord"][0:-1]
                )
                * s_comp.dict_Gauss_pt["total_density"]
                * s_comp.dict_Gauss_pt["total_isobaric_specific_heat"]
                * s_comp.dict_Gauss_pt["temperature"]
            )
            if s_comp.NAME != "Z_JACKET":
                self.E_str_ini = self.E_str_ini + s_comp.inputs[
                    "CROSSECTION"
                ] * np.sum(
                    (
                        self.dict_discretization["xcoord"][
                            1 : self.dict_discretization["N_nod"]
                        ]
                        - self.dict_discretization["xcoord"][0:-1]
                    )
                    * s_comp.dict_Gauss_pt["total_density"]
                    * s_comp.dict_Gauss_pt["total_isobaric_specific_heat"]
                    * s_comp.dict_Gauss_pt["temperature"]
                )
            else:
                self.E_jk_ini = self.E_jk_ini + s_comp.inputs[
                    "CROSSECTION"
                ] * np.sum(
                    (
                        self.dict_discretization["xcoord"][
                            1 : self.dict_discretization["N_nod"]
                        ]
                        - self.dict_discretization["xcoord"][0:-1]
                    )
                    * s_comp.dict_Gauss_pt["total_density"]
                    * s_comp.dict_Gauss_pt["total_isobaric_specific_heat"]
                    * s_comp.dict_Gauss_pt["temperature"]
                )
        # end for s_comp (cdp, 12/2020)

        # Construct and initialize dictionary dict_Step to correctly apply the \
        # method that solves the transient (cdp, 10/2020)
        if self.inputs["METHOD"] == "BE" or self.inputs["METHOD"] == "CN":
            # Backward Euler or Crank-Nicolson (cdp, 10/2020)
            self.dict_Step = dict(
                SYSLOD=np.zeros((self.dict_N_equation["Total"], 2)),
                SYSVAR=np.zeros((self.dict_N_equation["Total"], 1)),
            )
        elif self.inputs["METHOD"] == "AM4":
            # Adams-Moulton order 4 (cdp, 10/2020)
            self.dict_Step = dict(
                SYSLOD=np.zeros((self.dict_N_equation["Total"], 4)),
                SYSVAR=np.zeros((self.dict_N_equation["Total"], 3)),
                AM4_AA=np.zeros(
                    (self.dict_band["Full"], self.dict_N_equation["Total"], 4)
                ),
            )
        # end if self.inputs

        # Assign initial values to key SYSVAR (cdp, 10/2020)
        for jj in range(self.dict_obj_inventory["FluidComponents"]["Number"]):
            fluid_comp = self.dict_obj_inventory["FluidComponents"]["Objects"][jj]
            # velocity (cdp, 10/2020)
            self.dict_Step["SYSVAR"][
                jj : self.dict_N_equation["Total"] : self.dict_N_equation["NODOFS"], 0
            ] = fluid_comp.coolant.dict_node_pt["velocity"]
            # pressure (cdp, 10/2020)
            self.dict_Step["SYSVAR"][
                jj
                + self.dict_obj_inventory["FluidComponents"][
                    "Number"
                ] : self.dict_N_equation["Total"] : self.dict_N_equation["NODOFS"],
                0,
            ] = fluid_comp.coolant.dict_node_pt["pressure"]
            # temperature (cdp, 10/2020)
            self.dict_Step["SYSVAR"][
                jj
                + 2
                * self.dict_obj_inventory["FluidComponents"][
                    "Number"
                ] : self.dict_N_equation["Total"] : self.dict_N_equation["NODOFS"],
                0,
            ] = fluid_comp.coolant.dict_node_pt["temperature"]
        # end for jj (cdp, 10/2020)
        for ll in range(self.dict_obj_inventory["SolidComponents"]["Number"]):
            comp = self.dict_obj_inventory["SolidComponents"]["Objects"][ll]
            # solid components temperature (cdp, 10/2020)
            self.dict_Step["SYSVAR"][
                ll
                + self.dict_N_equation["FluidComponents"] : self.dict_N_equation[
                    "Total"
                ] : self.dict_N_equation["NODOFS"],
                0,
            ] = comp.dict_node_pt["temperature"]
        # end for ll (cdp, 10/2020)
        if self.dict_Step["SYSVAR"].shape[-1] > 1:
            # if this is true, it means that an higher order method than \
            # Crank-Nicolson is applied to solve the transient (cdp, 10/2020)
            for cc in range(1, self.dict_Step["SYSVAR"].shape[-1]):
                # Copy the values of the first colum in all the other columns, like \
                # they are the results of a dummy initial steady state (cdp, 10/2020)
                self.dict_Step["SYSVAR"][:, cc] = self.dict_Step["SYSVAR"][:, 0].copy()
            # end for cc (cdp, 10/2020)
        # end if self.dict_Step["SYSVAR"].shape[-1] (cdp, 10/2020)

        # Call function Save_properties to save conductor inizialization
        save_properties(
            self, simulation.dict_path[f"Output_Initialization_{self.ID}_dir"]
        )
        print("Saved initialization")

        # Call function update_real_time_plot
        update_real_time_plots(self)
        create_legend_rtp(self)

    # end method initialization

    ################################################################################

    def operating_conditions(self, simulation):

        """
        Method tha calls functions get_current, get_magnetic_field, get_magnetic_field_gradient, get_superconductor_critical_prop, get_heat, jhtflx_new_0, set_energy_counters, Get_transp_coeff at each time step to evaluate properies and quantities in each node of spatial discretization.
        """

        # Evaluate transport coefficients in nodal points.
        self.get_transp_coeff(simulation)

        # Loop on Strands objects (csp, 07/2020)
        for strand in self.dict_obj_inventory["Strands"]["Objects"]:
            strand.get_current(self)
            # MAGNETIC FIELD AS A FUNCTION OF POSITION
            # call method get_magnetic_field
            strand.get_magnetic_field(self)
            # call method get_magnetic_field_gradient for each Strands object (cdp, 06/2020)
            strand.get_magnetic_field_gradient(self)
            if strand.NAME != self.dict_obj_inventory["Stabilizer"]["Name"]:
                if strand.inputs["ISUPERCONDUCTOR"] == "Nb3Sn":
                    # mix or superconducor strands objects made of Nb3Sn (cdp, 08/2020)
                    # call method get_eps to evaluate strain
                    strand.get_eps(self)
                # end if strand.inputs["ISUPERCONDUCTOR"] (cdp, 08/2020)
                # Call get_superconductor_critical_prop to evaluate MixSCStabilizer \
                # and/or SuperConductor properties in nodal points. Added to allow \
                # storage of current sharing temperature time evolution values in \
                # user defined nodal points (cdp, 08/2020)
                strand.get_superconductor_critical_prop(self)
                if (
                    strand.operations["TCS_EVALUATION"] == False
                    and self.cond_num_step == 0
                ):
                    # Evaluate current sharing temperature only at the first time step.
                    strand.get_tcs()
                elif strand.operations["TCS_EVALUATION"] == True:
                    # Evaluate current sharing temperature at each time step.
                    strand.get_tcs()
            # end if strand.NAME != self.dict_obj_inventory["Stabilizer"]["Name"] \
            # (cdp, 08/2020)
            if self.cond_num_step == 0 and strand.operations["IQFUN"] == 0:
                # call method get_heat only once to initialize key EXTFLX of dictionary \
                # dict_node_pt to zeros (cdp, 11/2020)
                strand.get_heat(self)
            elif strand.operations["IQFUN"] != 0:
                # call method get_heat to evaluate external heating only if heating is on \
                # (cdp, 10/2020)
                strand.get_heat(self)
            # end if strand.operations["IQFUN"] (cdp, 10/2020)
            # call method jhtflx_new_0 to initialize JHTFLX to zeros for each \
            # conductor solid components (cdp, 06/2020)
            strand.jhtflx_new_0(self)
            # call set_energy_counters to initialize EEXT and EJHT to zeros for each \
            # conductor solid components (cdp, 06/2020)
            strand.set_energy_counters(self)
        # end for strand (cdp,07/2020)
        # Loop on Jackets objects
        for rr in range(self.dict_obj_inventory["Jacket"]["Number"]):
            jacket = self.dict_obj_inventory["Jacket"]["Objects"][rr]
            jacket.get_current(self)
            # MAGNETIC FIELD AS A FUNCTION OF POSITION
            # call method get_magnetic_field
            jacket.get_magnetic_field(self)
            if self.cond_num_step == 0 and jacket.operations["IQFUN"] == 0:
                # call method get_heat only once to initialize key EXTFLX of dictionary \
                # dict_node_pt to zeros (cdp, 11/2020)
                jacket.get_heat(self)
            elif jacket.operations["IQFUN"] != 0:
                # call method get_heat to evaluate external heating only if heating is on\
                # (cdp, 10/2020)
                jacket.get_heat(self)
            # end if jacket.operations["IQFUN"] (cdp, 10/2020)
            # call method jhtflx_new_0 to initialize JHTFLX to zeros for each \
            # conductor solid components (cdp, 06/2020)
            jacket.jhtflx_new_0(self)
            # call set_energy_counters to initialize EEXT and EJHT to zeros for each \
            # conductor solid components (cdp, 06/2020)
            jacket.set_energy_counters(self)
            if (
                self.dict_df_coupling["contact_perimeter_flag"].at[
                    simulation.environment.KIND, jacket.ID
                ]
                == 1
            ):
                # Evaluate the external heat by radiation in nodal points.
                jacket._radiative_source_therm_env(self, simulation.environment)
            # End if self.dict_df_coupling["contact_perimeter_flag"]
            for cc in range(rr + 1, self.dict_obj_inventory["Jacket"]["Number"]):
                jacket_c = self.dict_obj_inventory["Jacket"]["Objects"][cc]
                if (
                    abs(self.dict_df_coupling["HTC_choice"].at[jacket.ID, jacket_c.ID])
                    == 3
                ):
                    # Evaluate the inner heat exchange by radiation in nodal points.
                    jacket._radiative_heat_exc_inner(self, jacket_c)
                    jacket_c._radiative_heat_exc_inner(self, jacket)
                # End if abs.
            # End for cc.
        # End for rr.

        self.eval_gauss_point(simulation)

    # end Operating_conditions

    def eval_gauss_point(self, simulation):

        """
        Method that evaluates material properties and coefficients at the Gauss point, i.e at the centre of the element.
        """

        # PROPERTIES AT THE GAUSS POINT (CENTRE OF THE ELEMENT)

        # Fluid components (cdp, 07/2020)
        for fluid_comp in self.dict_obj_inventory["FluidComponents"]["Objects"]:
            # Evaluate coolant properties in Gauss points
            fluid_comp.coolant._eval_properties_nodal_gauss(
                self, simulation.fluid_prop_aliases, nodal=False
            )

        # Jacket (cdp, 07/2020)
        for jacket in self.dict_obj_inventory["Jacket"]["Objects"]:
            jacket.dict_Gauss_pt["temperature"] = (
                np.abs(
                    jacket.dict_node_pt["temperature"][:-1]
                    + jacket.dict_node_pt["temperature"][1:]
                )
                / 2.0
            )
            jacket.get_magnetic_field(self, nodal=False)

            jacket.dict_Gauss_pt["Q1"] = (
                jacket.dict_node_pt["JHTFLX"][:-1] + jacket.dict_node_pt["EXTFLX"][:-1]
            )
            jacket.dict_Gauss_pt["Q2"] = (
                jacket.dict_node_pt["JHTFLX"][1:] + jacket.dict_node_pt["EXTFLX"][1:]
            )
            # Add the radiative heat contribution with the environment.
            jacket.dict_Gauss_pt["Q1"] = (
                jacket.dict_Gauss_pt["Q1"] + jacket.radiative_heat_env[:-1]
            )
            jacket.dict_Gauss_pt["Q2"] = (
                jacket.dict_Gauss_pt["Q2"] + jacket.radiative_heat_env[1:]
            )
            # End if self.dict_df_coupling["contact_perimeter_flag"].
        # End for jacket.

        # Nested loop jacket - jacket.
        for rr in range(self.dict_obj_inventory["Jacket"]["Number"]):
            jacket = self.dict_obj_inventory["Jacket"]["Objects"][rr]
            for cc in range(rr + 1, self.dict_obj_inventory["Jacket"]["Number"]):
                jacket_c = self.dict_obj_inventory["Jacket"]["Objects"][cc]
                key = f"{jacket.ID}_{jacket_c.ID}"
                # Add the radiative heat contribution between inner surface of the enclosure and inner jackets.
                jacket.dict_Gauss_pt["Q1"] = (
                    jacket.dict_Gauss_pt["Q1"] + jacket.radiative_heat_inn[key][:-1]
                )
                jacket.dict_Gauss_pt["Q2"] = (
                    jacket.dict_Gauss_pt["Q2"] + jacket.radiative_heat_inn[key][1:]
                )

                jacket_c.dict_Gauss_pt["Q1"] = (
                    jacket_c.dict_Gauss_pt["Q1"] + jacket_c.radiative_heat_inn[key][:-1]
                )
                jacket_c.dict_Gauss_pt["Q2"] = (
                    jacket_c.dict_Gauss_pt["Q2"] + jacket_c.radiative_heat_inn[key][1:]
                )
                # End if abs().
            # End for cc.
            jacket.eval_sol_comp_properties(self.dict_obj_inventory, nodal=False)
        # end for rr.

        # Strands (cdp, 07/2020)
        for strand in self.dict_obj_inventory["Strands"]["Objects"]:
            strand.dict_Gauss_pt["temperature"] = (
                np.abs(
                    strand.dict_node_pt["temperature"][:-1]
                    + strand.dict_node_pt["temperature"][1:]
                )
                / 2.0
            )
            strand.get_magnetic_field(self, nodal=False)
            strand.dict_Gauss_pt["Q1"] = (
                strand.dict_node_pt["JHTFLX"][:-1] + strand.dict_node_pt["EXTFLX"][:-1]
            )
            strand.dict_Gauss_pt["Q2"] = (
                strand.dict_node_pt["JHTFLX"][1:] + strand.dict_node_pt["EXTFLX"][1:]
            )
            # call method get_magnetic_field_gradient for each Strands object (cdp, 06/2020)
            strand.get_magnetic_field_gradient(self, nodal=False)
            # only for MixSCStabilizer and SuperConductor objects (cdp, 07/2020)
            if strand.NAME != self.dict_obj_inventory["Stabilizer"]["Name"]:
                if strand.inputs["ISUPERCONDUCTOR"] == "Nb3Sn":
                    # mix or superconducor strands objects made of Nb3Sn (cdp, 08/2020)
                    # call method get_eps to evaluate strain
                    strand.get_eps(self, nodal=False)
                # end if strand.inputs["ISUPERCONDUCTOR"] (cdp, 08/2020)
                # Call get_superconductor_critical_prop to evaluate MixSCStabilizer \
                # and/or SuperConductor properties in the Gauss point (cdp, 07/2020)
                strand.get_superconductor_critical_prop(self, nodal=False)
                if (
                    strand.operations["TCS_EVALUATION"] == False
                    and self.cond_num_step == 0
                ):
                    # Evaluate current sharing temperature only at the first time step.
                    strand.get_tcs(nodal=False)
                elif strand.operations["TCS_EVALUATION"] == True:
                    # Evaluate current sharing temperature at each time step.
                    strand.get_tcs(nodal=False)
            # end if strand.NAME != self.dict_obj_inventory["Stabilizer"]["Name"] \
            # (cdp, 08/2020)
            # Evaluate SolidComponents properties
            strand.eval_sol_comp_properties(self.dict_obj_inventory, nodal=False)

        # call method Get_transp_coeff to evaluate transport properties (heat \
        # transfer coefficient and friction factor) in each Gauss point \
        # (cdp, 07/2020)
        self.get_transp_coeff(simulation, flag_nodal=False)

    # end Eval_Gauss_point

    def post_processing(self, simulation):

        # bozza della funzione Post_processing

        # Loop on FluidComponents (cdp, 07/2020)
        for fluid_comp in self.dict_obj_inventory["FluidComponents"]["Objects"]:
            # Evaluate the other coolant properties in nodal points (nodal = True by default)
            fluid_comp.coolant._eval_properties_nodal_gauss(
                self, simulation.fluid_prop_aliases
            )
        # Loop on Jackets (cdp, 07/2020)
        for jacket in self.dict_obj_inventory["Jacket"]["Objects"]:
            jacket.get_current(self)
            # MAGNETIC FIELD AS A FUNCTION OF POSITION
            # call method get_magnetic_field
            jacket.get_magnetic_field(self)
        # end for jacket (cdp, 07/2020)
        # Loop on Strands (cdp, 07/2020)
        for strand in self.dict_obj_inventory["Strands"]["Objects"]:
            # call method get_magnetic_field_gradient for each Strands object (cdp, 06/2020)
            strand.get_magnetic_field_gradient(self)
            # questa è la parte che credo sia rilevante
            # only for MixSCStabilizer and SuperConductor objects (cdp, 07/2020)
            if strand.NAME != self.dict_obj_inventory["Stabilizer"]["Name"]:
                if strand.inputs["ISUPERCONDUCTOR"] == "Nb3Sn":
                    # mix or superconducor strands objects made of Nb3Sn (cdp, 08/2020)
                    # call method get_eps to evaluate strain
                    strand.get_eps(self)
                # end if strand.inputs["ISUPERCONDUCTOR"] (cdp, 08/2020)
                # Call get_superconductor_critical_prop to evaluate MixSCStabilizer \
                # and/or SuperConductor properties in nodal points (cdp, 07/2020)
                strand.get_superconductor_critical_prop(self)
                # Evaluate current sharing temperature
                strand.get_tcs()

            # end if strand.NAME != self.dict_obj_inventory["Stabilizer"]["Name"] \
            # (cdp, 08/2020)

        # Loop on SolidComponents to evaluate the total final energy of \
        # SolidComponents, used to check the imposition of SolidComponents \
        # temperature initial spatial distribution (cdp, 12/2020)
        for s_comp in self.dict_obj_inventory["SolidComponents"]["Objects"]:
            self.E_sol_fin = self.E_sol_fin + s_comp.inputs["CROSSECTION"] * np.sum(
                (
                    self.dict_discretization["xcoord"][
                        1 : self.dict_discretization["N_nod"]
                    ]
                    - self.dict_discretization["xcoord"][0:-1]
                )
                * s_comp.dict_Gauss_pt["total_density"]
                * s_comp.dict_Gauss_pt["total_isobaric_specific_heat"]
                * s_comp.dict_Gauss_pt["temperature"]
            )
            if s_comp.NAME != "Z_JACKET":
                self.E_str_fin = self.E_str_fin + s_comp.inputs[
                    "CROSSECTION"
                ] * np.sum(
                    (
                        self.dict_discretization["xcoord"][
                            1 : self.dict_discretization["N_nod"]
                        ]
                        - self.dict_discretization["xcoord"][0:-1]
                    )
                    * s_comp.dict_Gauss_pt["total_density"]
                    * s_comp.dict_Gauss_pt["total_isobaric_specific_heat"]
                    * s_comp.dict_Gauss_pt["temperature"]
                )
            else:
                self.E_jk_fin = self.E_jk_fin + s_comp.inputs[
                    "CROSSECTION"
                ] * np.sum(
                    (
                        self.dict_discretization["xcoord"][
                            1 : self.dict_discretization["N_nod"]
                        ]
                        - self.dict_discretization["xcoord"][0:-1]
                    )
                    * s_comp.dict_Gauss_pt["total_density"]
                    * s_comp.dict_Gauss_pt["total_isobaric_specific_heat"]
                    * s_comp.dict_Gauss_pt["temperature"]
                )
        # end for s_comp (cdp, 12/2020)
        # call method Get_transp_coeff to evaluate transport properties (heat \
        # transfer coefficient and friction factor) in each nodal point \
        # (cdp, 06/2020)
        self.get_transp_coeff(simulation)

        # call method Mass_Energy_balance to get data for the space convergence \
        # (cdp, 09/2020)
        self.mass_energy_balance(simulation)
        # call function Save_convergence_data to save solution spatial \
        # distribution at TEND, together with the mass and energy balance results, \
        # to make the space convergence analisys (cdp, 12/2020)
        save_convergence_data(self, simulation.dict_path["Space_conv_output_dir"])
        # call function Save_convergence_data to save solution spatial \
        # distribution at TEND, together with the mass and energy balance results \
        # to make the time convergence analisys (cdp, 12/2020)
        save_convergence_data(
            self,
            simulation.dict_path["Time_conv_output_dir"],
            abs(simulation.n_digit),
            space_conv=False,
        )

    # end Post_processing

    def get_transp_coeff(self, simulation, flag_nodal=True):  # (cpd 06/2020)

        """
    Method that compute channels friction factor and heat transfer coefficient \
    between channels, channel and solid component, solid components, both in \
    nodal and Gauss points. (cpd 06/2020)
    """
        # Properties evaluation in each nodal point (cdp, 07/2020)
        if flag_nodal:
            dict_dummy = self.dict_node_pt
            self.dict_node_pt = self.eval_transp_coeff(simulation, dict_dummy)
        # Properties evaluation in each Gauss point (cdp, 07/2020)
        elif flag_nodal == False:
            dict_dummy = self.dict_Gauss_pt
            self.dict_Gauss_pt = self.eval_transp_coeff(
                simulation, dict_dummy, flag_nodal=False
            )

    # end method Get_transp_coeff (cdp, 09/2020)

    def eval_transp_coeff(self, simulation, dict_dummy, flag_nodal=True):

        """
    Method that actually computes channels friction factor and heat transfer \
    coefficient between channels, channel and solid component, solid \
    components, both in nodal and Gauss points. (cpd 06/2020)
    """

        # loop to evaluate htc_steady for each channel according to its geometry (cpd 06/2020)
        for rr in range(self.dict_obj_inventory["FluidComponents"]["Number"]):
            fluid_comp = self.dict_obj_inventory["FluidComponents"]["Objects"][rr]
            # Define dictionary to select nodal or gauss properties according to the value of flag_nodal.0
            dict_dummy_chan = {
                True: fluid_comp.coolant.dict_node_pt,
                False: fluid_comp.coolant.dict_Gauss_pt,
            }
            # Evaluate steady state heat transfer coefficient for each channel.
            fluid_comp.channel.eval_steady_state_htc(
                dict_dummy_chan[flag_nodal], nodal=flag_nodal
            )
            # Evaluate total friction factor for each channel.
            fluid_comp.channel.eval_friction_factor(
                dict_dummy_chan[flag_nodal]["Reynolds"], nodal=flag_nodal
            )
        # end for loop fluid_comp.

        # htc dummy dictionary and its sub-dictionary declaration (cpd 09/2020)
        dict_dummy["HTC"] = dict()
        dict_dummy["HTC"]["ch_ch"] = dict()
        dict_dummy["HTC"]["ch_ch"]["Open"] = dict()
        dict_dummy["HTC"]["ch_ch"]["Close"] = dict()
        dict_dummy["HTC"]["ch_sol"] = dict()
        dict_dummy["HTC"]["sol_sol"] = dict()
        dict_dummy["HTC"]["sol_sol"]["cond"] = dict()
        dict_dummy["HTC"]["sol_sol"]["rad"] = dict()
        dict_dummy["HTC"]["env_sol"] = dict()
        # Counters to check the number of the different possible kinds of interfaces (cdp, 09/2020)
        htc_len = 0
        for rr in range(self.dict_obj_inventory["FluidComponents"]["Number"]):
            fluid_comp_r = self.dict_obj_inventory["FluidComponents"]["Objects"][rr]
            dict_dummy_chan_r = {
                True: fluid_comp_r.coolant.dict_node_pt,
                False: fluid_comp_r.coolant.dict_Gauss_pt,
            }
            # Read the submatrix containing information about channel - solid objects iterfaces (cdp, 06/2020)
            # nested loop on channel - solid objects (cpd 06/2020)
            for s_comp in self.dict_obj_inventory["SolidComponents"]["Objects"]:
                dict_dummy_comp = {
                    True: s_comp.dict_node_pt,
                    False: s_comp.dict_Gauss_pt,
                }
                # Rationale: compute dictionary vaules only if there is an interface \
                # (cdp, 09/2020)
                if (
                    self.dict_df_coupling["contact_perimeter_flag"].at[
                        fluid_comp_r.ID, s_comp.ID
                    ]
                    == 1
                ):
                    htc_len = htc_len + 1
                    # new channel-solid interface (cdp, 09/2020)
                    htc_Kapitza = np.zeros(
                        dict_dummy_chan_r[flag_nodal]["temperature"].shape
                    )
                    htc_transient = np.zeros(
                        dict_dummy_chan_r[flag_nodal]["temperature"].shape
                    )
                    htc_full_transient = np.zeros(
                        dict_dummy_chan_r[flag_nodal]["temperature"].shape
                    )
                    if (
                        self.dict_df_coupling["HTC_choice"].at[
                            fluid_comp_r.ID, s_comp.ID
                        ]
                        == 2
                    ):
                        htc_Kapitza = (
                            200.0
                            * (
                                dict_dummy_comp[flag_nodal]["temperature"]
                                + dict_dummy_chan_r[flag_nodal]["temperature"]
                            )
                            * (
                                dict_dummy_comp[flag_nodal]["temperature"] ** 2
                                + dict_dummy_chan_r[flag_nodal]["temperature"] ** 2
                            )
                        )
                        if self.cond_time[-1] > s_comp.operations["TQBEG"]:
                            # implementation fully correct only for fully implicit method \
                            # (cdp, 06/2020)
                            htc_transient = np.sqrt(
                                (
                                    dict_dummy_chan_r[flag_nodal][
                                        "total_thermal_conductivity"
                                    ]
                                    * dict_dummy_chan_r[flag_nodal]["total_density"]
                                    * dict_dummy_chan_r[flag_nodal][
                                        "total_isobaric_specific_heat"
                                    ]
                                )
                                / (
                                    np.pi
                                    * (
                                        self.cond_time[-1]
                                        - s_comp.operations["TQBEG"]
                                    )
                                )
                            )
                            htc_full_transient = (htc_Kapitza * htc_transient) / (
                                htc_Kapitza + htc_transient
                            )
                        mlt = self.dict_df_coupling["HTC_multiplier"].at[
                            fluid_comp_r.ID, s_comp.ID
                        ]
                        # Assign to the HTC key of dictionary dict_dummy the dictionary whit the information about heat trasfer coefficient betweent channel fluid_comp_r and solid s_comp. Interface identification is given by the key name itself: f"{fluid_comp_r.ID}_{s_comp.ID}". This inner dictionary consists of a single key-value pair. (cdp, 07/2020)
                        dict_dummy["HTC"]["ch_sol"][
                            self.dict_topology["ch_sol"][fluid_comp_r.ID][s_comp.ID]
                        ] = np.maximum(
                            fluid_comp_r.channel.dict_htc_steady[flag_nodal] * mlt,
                            htc_full_transient,
                        )
                    elif (
                        self.dict_df_coupling["HTC_choice"].at[
                            fluid_comp_r.ID, s_comp.ID
                        ]
                        == -2
                    ):
                        dict_dummy["HTC"]["ch_sol"][
                            self.dict_topology["ch_sol"][fluid_comp_r.ID][s_comp.ID]
                        ] = self.dict_df_coupling["contact_HTC"].at[
                            fluid_comp_r.ID, s_comp.ID
                        ] * np.ones(
                            dict_dummy_chan_r[flag_nodal]["temperature"].shape
                        )
            # end loop on SolidComponents
            # nested loop on channel - channel objects (cdp, 06/2020)
            for cc in range(
                rr + 1, self.dict_obj_inventory["FluidComponents"]["Number"]
            ):
                fluid_comp_c = self.dict_obj_inventory["FluidComponents"]["Objects"][cc]
                dict_dummy_chan_c = {
                    True: fluid_comp_c.coolant.dict_node_pt,
                    False: fluid_comp_c.coolant.dict_Gauss_pt,
                }
                if (
                    self.dict_df_coupling["contact_perimeter_flag"].at[
                        fluid_comp_r.ID, fluid_comp_c.ID
                    ]
                    == 1
                ):
                    # new channel-channel interface (cdp, 09/2020)
                    htc_len = htc_len + 1
                    # Construct interface name: it can be found also in dict_topology["ch_ch"] but a search in dictionaties "Hydraulic_parallel" and "Thermal_contact" should be performed, which makes thinks not easy to do; it is simpler to construct interface names combining channels ID (cdp, 09/2020)
                    interface_name = f"{fluid_comp_r.ID}_{fluid_comp_c.ID}"
                    if (
                        self.dict_df_coupling["HTC_choice"].at[
                            fluid_comp_r.ID, fluid_comp_c.ID
                        ]
                        == 2
                    ):
                        # dummy
                        htc1 = fluid_comp_r.channel.dict_htc_steady[flag_nodal]
                        # dummy
                        htc2 = fluid_comp_c.channel.dict_htc_steady[flag_nodal]
                        mlt = self.dict_df_coupling["HTC_multiplier"].at[
                            fluid_comp_r.ID, fluid_comp_c.ID
                        ]
                        dict_dummy["HTC"]["ch_ch"]["Open"][interface_name] = (
                            mlt * htc1 * htc2 / (htc1 + htc2)
                        )
                        cond_interface = thermal_conductivity_ss(
                            (
                                dict_dummy_chan_r[flag_nodal]["temperature"]
                                + dict_dummy_chan_c[flag_nodal]["temperature"]
                            )
                            / 2
                        )
                        R_wall = (
                            self.dict_df_coupling["interf_thickness"].at[
                                fluid_comp_r.ID, fluid_comp_c.ID
                            ]
                            / cond_interface
                        )
                        dict_dummy["HTC"]["ch_ch"]["Close"][interface_name] = mlt / (
                            1 / htc1 + 1 / htc2 + R_wall
                        )
                    elif (
                        self.dict_df_coupling["HTC_choice"].at[
                            fluid_comp_r.ID, fluid_comp_c.ID
                        ]
                        == -2
                    ):
                        # in this case it is assumed that both open and close hct have the same value (cdp, 07/2020)
                        dict_dummy["HTC"]["ch_ch"]["Open"][
                            interface_name
                        ] = self.dict_df_coupling["contact_HTC"].at[
                            fluid_comp_r.ID, fluid_comp_c.ID
                        ] * np.ones(
                            dict_dummy_chan_r[flag_nodal]["temperature"].shape
                        )
                        dict_dummy["HTC"]["ch_ch"]["Close"][
                            interface_name
                        ] = self.dict_df_coupling["contact_HTC"].at[
                            fluid_comp_r.ID, fluid_comp_c.ID
                        ] * np.ones(
                            dict_dummy_chan_r[flag_nodal]["temperature"].shape
                        )
            # end for loop cc
        # end for loop rr
        # nested loop on solid - solid objects (cdp, 06/2020)
        for rr in range(self.dict_obj_inventory["SolidComponents"]["Number"]):
            s_comp_r = self.dict_obj_inventory["SolidComponents"]["Objects"][rr]
            dict_dummy_comp_r = {
                True: s_comp_r.dict_node_pt,
                False: s_comp_r.dict_Gauss_pt,
            }
            for cc in range(
                rr + 1, self.dict_obj_inventory["SolidComponents"]["Number"]
            ):
                s_comp_c = self.dict_obj_inventory["SolidComponents"]["Objects"][cc]
                dict_dummy_comp_c = {
                    True: s_comp_c.dict_node_pt,
                    False: s_comp_c.dict_Gauss_pt,
                }
                if (
                    self.dict_df_coupling["contact_perimeter_flag"].at[
                        s_comp_r.ID, s_comp_c.ID
                    ]
                    == 1
                ):
                    dict_dummy["HTC"]["sol_sol"]["cond"][
                        self.dict_topology["sol_sol"][s_comp_r.ID][s_comp_c.ID]
                    ] = np.zeros(dict_dummy_comp_r[flag_nodal]["temperature"].shape)
                    dict_dummy["HTC"]["sol_sol"]["rad"][
                        self.dict_topology["sol_sol"][s_comp_r.ID][s_comp_c.ID]
                    ] = np.zeros(dict_dummy_comp_r[flag_nodal]["temperature"].shape)

                    # New solid-solid interface (cdp, 09/2020)
                    htc_len = htc_len + 1
                    if (
                        self.dict_df_coupling["HTC_choice"].at[s_comp_r.ID, s_comp_c.ID]
                        == 1
                    ):
                        # Thermal contact.
                        htc_solid = 500.0
                        mlt = self.dict_df_coupling["HTC_multiplier"].at[
                            s_comp_r.ID, s_comp_c.ID
                        ]
                        dict_dummy["HTC"]["sol_sol"]["cond"][
                            self.dict_topology["sol_sol"][s_comp_r.ID][s_comp_c.ID]
                        ] = (
                            mlt
                            * htc_solid
                            * np.ones(
                                dict_dummy_comp_r[flag_nodal]["temperature"].shape
                            )
                        )
                    elif (
                        self.dict_df_coupling["HTC_choice"].at[s_comp_r.ID, s_comp_c.ID]
                        == -1
                    ):
                        # Thermal contact.
                        dict_dummy["HTC"]["sol_sol"]["cond"][
                            self.dict_topology["sol_sol"][s_comp_r.ID][s_comp_c.ID]
                        ] = self.dict_df_coupling["contact_HTC"].at[
                            s_comp_r.ID, s_comp_c.ID
                        ] * np.ones(
                            dict_dummy_comp_r[flag_nodal]["temperature"].shape
                        )
                    elif (
                        self.dict_df_coupling["HTC_choice"].at[s_comp_r.ID, s_comp_c.ID]
                        == 3
                    ):
                        # Radiative heat transfer.
                        if (
                            s_comp_r.inputs["Emissivity"] > 0.0
                            and s_comp_c.inputs["Emissivity"] > 0.0
                            and self.dict_df_coupling["view_factors"].at[
                                s_comp_r.ID, s_comp_c.ID
                            ]
                            > 0.0
                        ):
                            # Evaluate the radiative heat transfer coefficient (assume that sr is the inner convex surface and sc the outer not convex surface, only in the comment below):
                            # A_sr*sigma*(T_sr^2 + T_sc^2)*(T_sr + T_sc)/((1 - emissivity_sr)/emissivity_sr + 1/F_sr_cs + (1 - emissivity_sc)/emissivity_sc*(A_sr/A_sc))
                            # Reciprocal of the view factor.
                            view_factor_rec = np.reciprocal(
                                self.dict_df_coupling["view_factors"].at[
                                    s_comp_r.ID, s_comp_c.ID
                                ]
                            )
                            if (
                                s_comp_r.inputs["Outer_perimeter"]
                                < s_comp_c.inputs["Inner_perimeter"]
                            ):
                                # Set the contact perimeter to the correct value (overwrite the value assigned in input file conductor_coupling.xlsx)
                                self.dict_df_coupling["contact_perimeter"].at[
                                    s_comp_r.ID, s_comp_c.ID
                                ] = s_comp_r.inputs["Outer_perimeter"]
                                dict_dummy["HTC"]["sol_sol"]["rad"][
                                    self.dict_topology["sol_sol"][s_comp_r.ID][
                                        s_comp_c.ID
                                    ]
                                ] = self._inner_radiative_htc(
                                    s_comp_r,
                                    s_comp_c,
                                    dict_dummy_comp_r[flag_nodal]["temperature"],
                                    dict_dummy_comp_c[flag_nodal]["temperature"],
                                    view_factor_rec,
                                )
                            elif (
                                s_comp_c.inputs["Outer_perimeter"]
                                < s_comp_r.inputs["Inner_perimeter"]
                            ):
                                # Set the contact perimeter to the correct value (overwrite the value assigned in input file conductor_coupling.xlsx)
                                self.dict_df_coupling["contact_perimeter"].at[
                                    s_comp_r.ID, s_comp_c.ID
                                ] = s_comp_c.inputs["Outer_perimeter"]
                                dict_dummy["HTC"]["sol_sol"]["rad"][
                                    self.dict_topology["sol_sol"][s_comp_r.ID][
                                        s_comp_c.ID
                                    ]
                                ] = self._inner_radiative_htc(
                                    s_comp_c,
                                    s_comp_r,
                                    dict_dummy_comp_c[flag_nodal]["temperature"],
                                    dict_dummy_comp_r[flag_nodal]["temperature"],
                                    view_factor_rec,
                                )
                            # End if s_comp_r.inputs["Outer_perimeter"].
                        # End if emissivity.
                    elif (
                        self.dict_df_coupling["HTC_choice"].at[s_comp_r.ID, s_comp_c.ID]
                        == -3
                    ):
                        # Radiative heat transfer from sheet contact_HTC of file conductor_coupling.xlsx.
                        dict_dummy["HTC"]["sol_sol"]["rad"][
                            self.dict_topology["sol_sol"][s_comp_r.ID][s_comp_c.ID]
                        ] = self.dict_df_coupling["contact_HTC"].at[
                            s_comp_r.ID, s_comp_c.ID
                        ] * np.ones(
                            dict_dummy_comp_r[flag_nodal]["temperature"].shape
                        )
                    # End if self.dict_df_coupling["HTC_choice"].at[s_comp_r.ID, s_comp_c.ID]
            # end for loop cc

            key = f"{simulation.environment.KIND}_{s_comp_r.ID}"
            if (
                self.dict_df_coupling["contact_perimeter_flag"].at[
                    simulation.environment.KIND, s_comp_r.ID
                ]
                == 1
            ):
                # New environment-solid interface
                htc_len = htc_len + 1
                if (
                    s_comp_r.inputs["Jacket_kind"] == "outer_insulation"
                    or s_comp_r.inputs["Jacket_kind"] == "whole_enclosure"
                ):
                    # Heat transfer with the environment by radiation and/or by convection.
                    # Initialize dictionary.
                    dict_dummy["HTC"]["env_sol"][key] = dict(
                        conv=np.zeros(
                            dict_dummy_comp_r[flag_nodal]["temperature"].size
                        ),
                        rad=np.zeros(dict_dummy_comp_r[flag_nodal]["temperature"].size),
                    )
                    if (
                        self.dict_df_coupling["HTC_choice"].at[
                            simulation.environment.KIND, s_comp_r.ID
                        ]
                        == 2
                    ):
                        # Heat transfer by convection: heat transfer coefficient evaluated from air properties.
                        if self.inputs["Is_rectangular"]:
                            # Rectangular conductor
                            dict_dummy["HTC"]["env_sol"][key]["conv"] = dict(
                                side=np.zeros(
                                    dict_dummy_comp_r[flag_nodal]["temperature"].size
                                ),
                                bottom=np.zeros(
                                    dict_dummy_comp_r[flag_nodal]["temperature"].size
                                ),
                                top=np.zeros(
                                    dict_dummy_comp_r[flag_nodal]["temperature"].size
                                ),
                            )
                            # Evaluate side bottom and top surfaces htc.
                            (
                                dict_dummy["HTC"]["env_sol"][key]["conv"]["side"],
                                dict_dummy["HTC"]["env_sol"][key]["conv"]["bottom"],
                                dict_dummy["HTC"]["env_sol"][key]["conv"]["top"],
                            ) = simulation.environment.eval_heat_transfer_coefficient(
                                self, dict_dummy_comp_r[flag_nodal]["temperature"]
                            )
                        else:
                            # Circular conductor
                            if flag_nodal == False:
                                # Compute only in gauss node to avoid error
                                dict_dummy["HTC"]["env_sol"][key]["conv"] = (
                                    simulation.environment.eval_heat_transfer_coefficient(
                                        self,
                                        dict_dummy_comp_r[flag_nodal]["temperature"],
                                    )
                                    * self.inputs["Phi_conv"]
                                )
                    elif (
                        self.dict_df_coupling["HTC_choice"].at[
                            simulation.environment.KIND, s_comp_r.ID
                        ]
                        == -2
                    ):
                        # Heat transfer by convection: from sheet contact_HTC of file conductor_coupling.xlsx.
                        dict_dummy["HTC"]["env_sol"][key]["conv"] = (
                            self.dict_df_coupling["contact_HTC"].at[
                                simulation.environment.KIND, s_comp_r.ID
                            ]
                            * self.inputs["Phi_conv"]
                            * np.ones(
                                dict_dummy_comp_r[flag_nodal]["temperature"].shape
                            )
                        )
                    elif (
                        self.dict_df_coupling["HTC_choice"].at[
                            simulation.environment.KIND, s_comp_r.ID
                        ]
                        == 3
                    ):
                        # Heat transfer by radiation.
                        # Evaluate radiative heat transfer coefficient invoking method eval_weighted_radiative_htc.
                        dict_dummy["HTC"]["env_sol"][key][
                            "rad"
                        ] = self.eval_weighted_radiative_htc(
                            simulation,
                            s_comp_r,
                            dict_dummy_comp_r[flag_nodal]["temperature"],
                        )
                    elif (
                        self.dict_df_coupling["HTC_choice"].at[
                            simulation.environment.KIND, s_comp_r.ID
                        ]
                        == -3
                    ):
                        # Heat transfer by radiation: from sheet contact_HTC of file conductor_coupling.xlsx.
                        dict_dummy["HTC"]["env_sol"][key]["rad"] = (
                            self.dict_df_coupling["contact_HTC"].at[
                                simulation.environment.KIND, s_comp_r.ID
                            ]
                            * self.inputs["Phi_rad"]
                            * np.ones(
                                dict_dummy_comp_r[flag_nodal]["temperature"].shape
                            )
                        )
                    elif (
                        self.dict_df_coupling["HTC_choice"].at[
                            simulation.environment.KIND, s_comp_r.ID
                        ]
                        == 4
                    ):
                        # Heat transfer by radiation and free convection: code evaluation
                        dict_dummy["HTC"]["env_sol"][key]["conv"] = (
                            simulation.environment.eval_heat_transfer_coefficient(
                                self, dict_dummy_comp_r[flag_nodal]["temperature"]
                            )
                            * self.inputs["Phi_conv"]
                        )

                        dict_dummy["HTC"]["env_sol"][key][
                            "rad"
                        ] = self.eval_weighted_radiative_htc(
                            simulation,
                            s_comp_r,
                            dict_dummy_comp_r[flag_nodal]["temperature"],
                        )
                    elif (
                        self.dict_df_coupling["HTC_choice"].at[
                            simulation.environment.KIND, s_comp_r.ID
                        ]
                        == -4
                    ):
                        # Heat transfer by radiation and free convection: from sheet contact_HTC of file conductor_coupling.xlsx.
                        # Questo va ragionato meglio: secondo me devo trovare il modo di distinguere i due contributi anche se in input sono dati come valore complessivo.
                        dict_dummy["HTC"]["env_sol"][key][
                            "conv"
                        ] = self.dict_df_coupling["contact_HTC"].at[
                            simulation.environment.KIND, s_comp_r.ID
                        ] * np.ones(
                            dict_dummy_comp_r[flag_nodal]["temperature"].shape
                        )
                    # End if self.dict_df_coupling["HTC_choice"].at[simulation.environment.KIND, s_comp_r.ID]
                else:
                    # Raise error
                    raise os.error(
                        f"Jacket of kind {s_comp_r.inputs['Jacket_kind']} can not exchange heat by radiation and/or convection with the environment.\n"
                    )
                # End if s_comp_r.inputs["Jacket_kind"]
            # End if self.dict_df_coupling["contact_perimeter_flag"].at[simulation.environment.KIND, s_comp_r.ID]
        # end for loop rr

        # Check number of evaluated interface htc (cdp, 06/2020)

        if htc_len != self.dict_df_coupling["contact_perimeter_flag"].to_numpy().sum():
            raise ValueError(
                f"ERROR!!! Number of interface and number of \
        evaluated interface htc mismatch: {self.dict_df_coupling['contact_perimeter_flag'].to_numpy().sum()} != {htc_len}"
            )

        return dict_dummy

    # end method Eval_transp_coeff (cdp, 09/2020)

    def eval_weighted_radiative_htc(self, simulation, s_comp, temperature_jk):
        # Evaluate the weighted external radiative heat transfer coefficient: htc_rad*Phi_rad
        if s_comp.inputs["Emissivity"] <= 0.0:
            # Set the radiative heat transfer coefficient to zero: no radiative heat transfer.
            return np.zeros(temperature_jk.shape)  # W/m^2/K
        else:
            return (
                constants.Stefan_Boltzmann
                * s_comp.inputs["Emissivity"]
                * (
                    temperature_jk ** 2
                    + simulation.environment.inputs["Temperature"] ** 2
                )
                * (temperature_jk + simulation.environment.inputs["Temperature"])
                * self.inputs["Phi_rad"]
            )  # W/m^2/K
        # End if s_comp_r.inputs["Emissivity"] <= 0.

    # End method eval_weighted_radiative_htc.

    def _inner_radiative_htc(self, jk_i, jk_j, temp_i, temp_j, view_factor_rec):

        return (
            constants.Stefan_Boltzmann
            * (temp_i ** 2 + temp_j ** 2)
            * (temp_i + temp_j)
            / (
                (1.0 - jk_i.inputs["Emissivity"]) / jk_i.inputs["Emissivity"]
                + view_factor_rec
                + (1.0 - jk_j.inputs["Emissivity"])
                / jk_j.inputs["Emissivity"]
                * (
                    jk_i.inputs["Outer_perimeter"]
                    / jk_j.inputs["Inner_perimeter"]
                )
            )
        )  # W/m^2/K

    # End method _inner_radiative_htc

    def aprior(self):

        # C * COMPUTE BY BISECTION THE PRODUCT LAMBDA * DELTA T TO ACHIEVE THE
        # C * DESIRED RELATIVE ACCURACY ON THE TIME STEP
        TIMACC = 1e-3
        X1 = 1.0
        X2 = 0.0
        RES = 1.0
        while abs(RES) > 1.0e-2 * TIMACC:
            self.EIGTIM = 0.5 * (X1 + X2)
            Y1 = (1.0 - (1.0 - self.theta_method) * self.EIGTIM) / (
                1.0 + self.theta_method * self.EIGTIM
            )
            Y2 = np.exp(-self.EIGTIM)
            RES = abs((Y1 / Y2) - 1.0) - TIMACC
            if RES >= 0.0:
                X1 = self.EIGTIM
            elif RES < 0.0:
                X2 = self.EIGTIM

    # end Aprior

    def mass_energy_balance(self, simulation):

        """
        Function that operform mass and energy balance on conductor (cdp, 09/2020)
        """
        self.mass_balance = 0.0  # mass balance initialization (cdp, 09/2020)
        self.energy_balance = 0.0  # energy balance initialization (cdp, 09/2020)
        self.inner_pow = 0.0
        self.outer_pow = 0.0
        for fluid_comp in self.dict_obj_inventory["FluidComponents"]["Objects"]:
            # Mass balance (cdp, 09/2020)
            self.mass_balance = self.mass_balance + self.time_step * (
                fluid_comp.coolant.dict_node_pt["mass_flow_rate"][0]
                - fluid_comp.coolant.dict_node_pt["mass_flow_rate"][-1]
            )
            # Energy balance: sum(mdot_inl*(w_inl + v_inl^2/2) - \
            # mdot_out*(w_out + v_out^2/2)) (cdp, 09/2020)
            self.energy_balance = self.energy_balance + self.time_step * (
                fluid_comp.coolant.dict_node_pt["mass_flow_rate"][0]
                * (
                    fluid_comp.coolant.dict_node_pt["total_enthalpy"][0]
                    + fluid_comp.coolant.dict_node_pt["velocity"][0] ** 2 / 2.0
                )
                - fluid_comp.coolant.dict_node_pt["mass_flow_rate"][-1]
                * (
                    fluid_comp.coolant.dict_node_pt["total_enthalpy"][-1]
                    + fluid_comp.coolant.dict_node_pt["velocity"][-1] ** 2 / 2.0
                )
            )
            self.inner_pow = self.inner_pow + fluid_comp.coolant.dict_node_pt[
                "mass_flow_rate"
            ][0] * (
                fluid_comp.coolant.dict_node_pt["total_enthalpy"][0]
                + fluid_comp.coolant.dict_node_pt["velocity"][0] ** 2 / 2.0
            )
            self.outer_pow = self.outer_pow + fluid_comp.coolant.dict_node_pt[
                "mass_flow_rate"
            ][-1] * (
                fluid_comp.coolant.dict_node_pt["total_enthalpy"][-1]
                + fluid_comp.coolant.dict_node_pt["velocity"][-1] ** 2 / 2.0
            )
        # End for fluid_comp.
        for jacket in self.dict_obj_inventory["Jacket"]["Objects"]:
            if (
                self.dict_df_coupling["contact_perimeter_flag"].loc[
                    simulation.environment.KIND, jacket.ID
                ]
                == 1
            ):
                key = f"{simulation.environment.KIND}_{jacket.ID}"
                self.energy_balance = (
                    self.energy_balance
                    + self.time_step
                    * jacket.inputs["Outer_perimeter"]
                    * np.sum(
                        (
                            self.dict_Gauss_pt["HTC"]["env_sol"][key]["conv"]
                            + self.dict_Gauss_pt["HTC"]["env_sol"][key]["rad"]
                        )
                        * self.dict_discretization["Delta_x"]
                        * (
                            simulation.environment.inputs["Temperature"]
                            - jacket.dict_Gauss_pt["temperature"]
                        )
                    )
                )
                self.inner_pow = self.inner_pow + jacket.inputs[
                    "Outer_perimeter"
                ] * np.sum(
                    (
                        self.dict_Gauss_pt["HTC"]["env_sol"][key]["conv"]
                        + self.dict_Gauss_pt["HTC"]["env_sol"][key]["rad"]
                    )
                    * self.dict_discretization["Delta_x"]
                    * (
                        simulation.environment.inputs["Temperature"]
                        - jacket.dict_Gauss_pt["temperature"]
                    )
                )
            # End if.
        # End jacket.
        # Energy balance to check the correct management of SolidComponents forced \
        # initial temperature distribution (cdp, 12/2020)
        # E_residual = (self.E_sol_ini - self.E_sol_fin) - \
        #              (self.enthalpy_inl - self.enthalpy_out)
        # print(f"E_sol_ini = {self.E_sol_ini} J\n")
        # print(f"E_sol_fin = {self.E_sol_fin} J\n")
        # print(f"E_str_ini = {self.E_str_ini} J\n")
        # print(f"E_str_fin = {self.E_str_fin} J\n")
        # print(f"E_jk_ini = {self.E_jk_ini} J\n")
        # print(f"E_jk_fin = {self.E_jk_fin} J\n")
        # print(f"enthalpy_inl = {self.enthalpy_inl} J\n")
        # print(f"enthalpy_out = {self.enthalpy_out} J\n")
        # print(f"enthalpy_balance = {self.enthalpy_balance} J\n")
        # print(f"E_residual = {E_residual} J\n")

        print(f"Energy balance = {self.energy_balance} J\n")
        print(f"Outer inner power ratio % = {1e2*self.outer_pow/self.inner_pow} ~")

    # end method Mass_Energy_balance (cdp, 09/2020)

    def compute_radiative_heat_exhange_jk(self):
        """Method that evaluates the radiative heat exchanged by radiation between jackets."""
        # Nested loop on jackets.
        for rr in range(self.dict_obj_inventory["Jacket"]["Number"]):
            jk_r = self.dict_obj_inventory["Jacket"]["Objects"][rr]
            for cc in range(rr + 1, self.dict_obj_inventory["Jacket"]["Number"]):
                jk_c = self.dict_obj_inventory["Jacket"]["Objects"][cc]
                if abs(self.dict_df_coupling["HTC_choice"].at[jk_r.ID, jk_c.ID]) == 3:
                    self.heat_rad_jk[f"{jk_r.ID}_{jk_c.ID}"] = (
                        self.dict_df_coupling["contact_perimeter"].at[jk_r.ID, jk_c.ID]
                        * self.dict_discretization["Delta_x"]
                        * self.dict_Gauss_pt["HTC"]["sol_sol"]["rad"][
                            self.dict_topology["sol_sol"][jk_r.ID][jk_c.ID]
                        ]
                        * (
                            jk_r.dict_Gauss_pt["temperature"]
                            - jk_c.dict_Gauss_pt["temperature"]
                        )
                    )  # W
                # End if abs.
            # End for cc.
        # End for rr.

    # End method _compute_radiative_heat_exhange_jk.

    def compute_heat_exchange_jk_env(self, environment):
        """Method that computes the heat exchange between the outer surface of the conductor and the environment by convection and or radiation.

        Args:
            environment ([type]): [description]
        """
        for jk in self.dict_obj_inventory["Jacket"]["Objects"]:
            if (
                self.dict_df_coupling["contact_perimeter_flag"].at[
                    environment.KIND, jk.ID
                ]
                == 1
            ):
                key = f"{environment.KIND}_{jk.ID}"
                self.heat_exchange_jk_env[key] = (
                    self.dict_df_coupling["contact_perimeter"].at[
                        environment.KIND, jk.ID
                    ]
                    * self.dict_discretization["Delta_x"]
                    * (
                        self.dict_Gauss_pt["HTC"]["env_sol"][key]["conv"]
                        + self.dict_Gauss_pt["HTC"]["env_sol"][key]["rad"]
                    )
                    * (
                        environment.inputs["Temperature"]
                        - jk.dict_Gauss_pt["temperature"]
                    )
                )  # W
            # End if self.dict_df_coupling["contact_perimeter_flag"].
        # End for jk.

    # End method _compute_heat_exchange_environment.

    def load_user_defined_quantity(self, simulation, key, col_name):
        """[summary]

        Args:
            conductor ([type]): [description]

        Returns:
            [type]: [description]
        """
        file_extension = self.file_input[key].split(".")[1]
        # Build the path to the file
        fname = os.path.join(self.BASE_PATH, self.file_input[key])
        # Call function used to open the file
        return simulation.func_open_aux[file_extension](
            fname, col_name, simulation.default_vals[file_extension]
        )

        # End method load_user_defined_quantity.
