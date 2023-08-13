# Import packages
from decimal import Decimal
import logging
from typing_extensions import Self
from openpyxl import load_workbook
import numpy as np
from scipy.sparse import coo_matrix, csr_matrix, lil_matrix, diags
from scipy import constants, integrate, interpolate
import pandas as pd
import os
from collections import namedtuple
import warnings

# import classes
from component_collection import ComponentCollection
from conductor_flags import (
    ANALYTICAL_INDUCTANCE,
    APPROXIMATE_INDUCTANCE,
    ELECTRIC_CONDUCTANCE_UNIT_LENGTH,
    ELECTRIC_CONDUCTANCE_NOT_UNIT_LENGTH,
    IOP_NOT_DEFINED,
    IOP_CONSTANT,
    IOP_FROM_EXT_FUNCTION,
    IOP_FROM_FILE,
    SELF_INDUCTANCE_MODE_1,
    SELF_INDUCTANCE_MODE_2,
    STATIC_ELECTRIC_SOLVER,
    ELECTRIC_TIME_STEP_NUMBER,
    VARIABLE_CONTACT_PERIMETER,
    CONSTANT_CONTACT_PERIMETER,
)
from fluid_component import FluidComponent
from jacket_component import JacketComponent
from stack_component import StackComponent
from strand_component import StrandComponent
from strand_mixed_component import StrandMixedComponent
from strand_stabilizer_component import StrandStabilizerComponent

# import functions
from utility_functions.auxiliary_functions import (
    check_repeated_headings,
    check_headers,
    check_object_number,
    set_diagnostic,
)
from utility_functions.electric_auxiliary_functions import (
    custom_current_function,
    electric_steady_state_solution,
    electric_transient_solution,
)
from utility_functions.initialization_functions import (
    build_coordinates_of_barycenter,
    check_max_node_number,
    conductor_spatial_discretization,
    user_defined_grid,
)
from utility_functions.gen_flow import gen_flow
from utility_functions.output import (
    save_properties,
    save_convergence_data,
    save_geometry_discretization,
)
from utility_functions.plots import update_real_time_plots, create_legend_rtp
from utility_functions.solid_components_initialization import (
    solid_components_temperature_initialization,
)

# Stainless Steel properties
from properties_of_materials.stainless_steel import thermal_conductivity_ss

# Get the logger specified in the file
conductorlogger = logging.getLogger("opensc2Logger.conductor")


class Conductor:

    BASE_PATH = ""

    ## computed in initialization
    # Remove from here if actually used to avoid problems!
    # MDTINL = ""
    # IBDINL = ""
    # IBDOUT = ""
    # inHard = "" # old: list() (cdp 06/2020)
    # MAXDOF = "" not used at the moment

    # Minimum absolute pressure difference between two channels in hydraulic \
    # parallel (threshold)
    Delta_p_min = 1e-4  # Pa
    # Localized pressure drop coefficient beywen two channels in hydraulic \
    # parallel
    k_loc = 1.0  # ~
    # Lambda velocity parameter (--> 0 if porous wall, --> 1 if helicoidal wall)
    lambda_v = 1.0  # ~

    KIND = "Conductor"
    CHUNCK_SIZE = 100

    def __init__(self: Self, simulation: object, sheetConductorsList: list, ICOND: int):
        """Makes an instance of class conductor.

        Args:
            self (Self): conductor object.
            simulation (object): simulation object.
            sheetConductorsList (list): list of sheets available in input file conductor_definition.
            ICOND (int): conductor counter.
        """

        self.BASE_PATH = simulation.basePath
        self.ICOND = ICOND
        self.name = sheetConductorsList[0].cell(row=1, column=1).value
        # get channels ID consistently with user definition (cdp, 09/2020)
        self.identifier = (
            sheetConductorsList[0].cell(row=3, column=4 + self.ICOND).value
        )
        # Get the number of the conductor from the ID.
        self.number = int(self.identifier.split("_")[1])
        # file_input dictionary initialization (cdp, 06/2020)
        self.file_input = dict()
        # inputs dictionary initialization (cdp, 06/2020)
        self.inputs = dict()
        self.workbook_name = os.path.join(
            self.BASE_PATH, simulation.transient_input["MAGNET"]
        )
        self.workbook_sheet_name = [sheet.title for sheet in sheetConductorsList]
        # Load the sheet CONDUCTOR_files form file conducor_definition.xlsx as a disctionary.
        self.file_input = pd.read_excel(
            self.workbook_name,
            sheet_name=self.workbook_sheet_name[0],
            skiprows=2,
            header=0,
            index_col=0,
            usecols=["Variable name", self.identifier],
        )[self.identifier].to_dict()
        # Load the sheet CONDUCTOR_input form file conducor_definition.xlsx as a disctionary.
        self.inputs = pd.read_excel(
            self.workbook_name,
            sheet_name=self.workbook_sheet_name[1],
            skiprows=2,
            header=0,
            index_col=0,
            usecols=["Variable name", self.identifier],
        )[self.identifier].to_dict()
        # Get the user defined name of the conductor (default is identifier)
        self.name = self.inputs["NAME"]
        # Delete key NAME from dictionary self.inputs
        del self.inputs["NAME"]

        # Temporary solution to mangage input file loading, strange behavior: 1
        # are converted to True but 0 not converted to False. Works with no
        # conversion only if the first value (XLENGHT) is set to 1.
        if self.inputs["UPWIND"] == True:
            self.inputs["UPWIND"] = 1

        for key in ["I0_OP_MODE","ELECTRIC_TIME_STEP"]:
            if isinstance(self.inputs[key],str) and self.inputs[key].lower() == "none":
                self.inputs[key] = None
                
        # Set total current to 0.0 A if user specifies no current with flag 
        # I0_OP_MODE.
        if self.inputs["I0_OP_MODE"] is IOP_NOT_DEFINED:
            self.inputs["I0_OP_TOT"] = 0.0

        # Load the sheet CONDUCTOR_operation form file conducor_definition.xlsx as a disctionary.
        self.operations = pd.read_excel(
            self.workbook_name,
            sheet_name=self.workbook_sheet_name[2],
            skiprows=2,
            header=0,
            index_col=0,
            usecols=["Variable name", self.identifier],
        )[self.identifier].to_dict()
        conductorlogger.debug(
            f"Loaded sheet CONDUCTOR_operation from file conductor_definition\n"
        )

        # Temporary solution to mangage input file loading, strange behavior: 1
        # are converted to True but 0 not converted to False.
        keys = [
            "EQUIPOTENTIAL_SURFACE_NUMBER",
            "EQUIPOTENTIAL_SURFACE_COORDINATE",
            "INDUCTANCE_MODE",
            "ELECTRIC_SOLVER",
        ]
        self.operations.update({key: 1 for key in keys if self.operations[key]})
        self.operations.update(
            {key: 0 for key in keys if self.operations[key] == False}
        )

        _ = {
            True: self.__manage_equipotential_surfaces_coordinate,
            False: self.__delete_equipotential_inputs,
        }

        # Calls method self.__manage_equipotential_surfaces_coordinate if
        # EQUIPOTENTIAL_SURFACE_FLAG is True.
        conductorlogger.debug(
            f"{self.operations['EQUIPOTENTIAL_SURFACE_FLAG']=};call method {_[self.operations['EQUIPOTENTIAL_SURFACE_FLAG']].__name__}\n"
        )
        _[self.operations["EQUIPOTENTIAL_SURFACE_FLAG"]]()

        conductorlogger.debug(f"After equipotetial surface(s) definition\n")

        # CREATE grid for the i-th conductor
        self.grid_features = dict()
        self.grid_input = pd.read_excel(
            os.path.join(self.BASE_PATH, self.file_input["GRID_DEFINITION"]),
            sheet_name="GRID",
            skiprows=2,
            header=0,
            index_col=0,
            usecols=["Variable name", self.identifier],
            dtype="object",
        )[self.identifier].to_dict()

        # Load all the sheets in file conductor_coupling.xlsx as a dictionary of dataframes.
        self.dict_df_coupling = pd.read_excel(
            os.path.join(self.BASE_PATH, self.file_input["STRUCTURE_COUPLING"]),
            sheet_name=None,
            skiprows=1,
            header=0,
            index_col=0,
        )

        # Alias for self.dict_df_coupling["contact_perimeter_flag"]
        interf_flag = self.dict_df_coupling["contact_perimeter_flag"].to_numpy()
        # Remove all nonzero values in interf_flag and convert matrix to array.
        interf_flag = interf_flag[interf_flag.nonzero()]

        # Check if user declared variable contact perimeter for some of the 
        # conductor components.
        if (
            self.file_input["VARIABLE_CONTACT_PERIMETER"] != "none"
            and any(interf_flag == -1)
        ):
            # Correctly declared a variable contact perimeter: load all sheets 
            # in file variable_contact_perimeter.xlsx in a dictionary of 
            # dataframes.
            self.dict_df_variable_contact_perimeter = pd.read_excel(
                os.path.join(
                    self.BASE_PATH,
                    self.file_input["VARIABLE_CONTACT_PERIMETER"]
                ),
                sheet_name=None,
                header=0,
            )

            # Loop to restore original column name (the engine of method 
            # pd.read_excel would append a number to columns with the same 
            # name, e.g. JK_1, JK_1.1,...; this loop will remove the .1,..., 
            # part of the column name).
            for df in self.dict_df_variable_contact_perimeter.values():
                df.columns = [col_name.split(".")[0] for col_name in df.columns]

        elif (
            self.file_input["VARIABLE_CONTACT_PERIMETER"] != "none"
            and all(interf_flag == 1)
            ):
            # User provided a file for variable contact perimeter but in sheet 
            # contact_perimeters_flag none of the interfaces has the flag -1.
            raise ValueError("User provides a file for variable contact perimeter but in sheet contact_perimeter_flag none of the interfaces has the valid flag for the variable contact perimeter (-1). \nIf user wants to assign variable contact perimeters, please check sheet contact_perimeter_flag in file conductor_coupling.xlsx.\nIf user does not want to assign a variable contact perimeter flag, please check sheet CONDUCTOR_files in file conductor_definition.xlsx and replace the file name in row VARIABLE_CONTACT_PERIMETER with 'none'.")
        elif (
            self.file_input["VARIABLE_CONTACT_PERIMETER"] == "none"
            and any(interf_flag == -1)
            ):
            # User prescribed the flag for variable contact perimeters without 
            # providing an auxiliary file to read the variable contact 
            # perimeters.
            raise FileNotFoundError("User prescribed the flag for variable contact perimeters without providing an auxiliary file to read the variable contact perimeters.\nIf user wants to assign variable contact perimeters, please check sheet CONDUCTOR_files in file conductor_definition.xlsx and provide a valid file name in row VARIABLE_CONTACT_PERIMETER with the values of the variable contact perimeters.\nIf user does not want to assign a variable contact perimeter flag, please check sheet contact_perimeter_flag in file conductor_coupling.xlsx and replace all -1 flag with value 1.")

        # Checks on auxiliary input file variable_contact_perimeter.xlsx
        self.dict_df_variable_contact_perimeter = (
            self.__check_variable_contact_perimeter()
        )

        # Dictionary declaration (cdp, 09/2020)
        self.inventory = dict()
        # call method Conductor_components_instance to make instance of conductor components (cdp, 11/2020)
        # conductorlogger.debug(
        #     f"Before call method {self.conductor_components_instance.__name__}"
        # )
        self.conductor_components_instance(simulation)
        # conductorlogger.debug(
        #     f"After call method {self.conductor_components_instance.__name__}"
        # )

        self.__get_total_cross_section()

        # Call private method __coordinates to build grid coordinates.
        # conductorlogger.debug(f"Before call method {self.__coordinates.__name__}")
        self.__coordinates(simulation)
        # conductorlogger.debug(f"After call method {self.__coordinates.__name__}")

        self.__update_grid_features()

        # Call private method __initialize_attributes to initialize all the other useful and necessary attributes of class Conductor.
        # conductorlogger.debug(
        #     f"Before call method {self.__initialize_attributes.__name__}"
        # )
        self.__initialize_attributes(simulation)
        # conductorlogger.debug(
        #     f"After call method {self.__initialize_attributes.__name__}"
        # )


    # end method __init__ (cdp, 11/2020)

    def __str__(self):
        pass

    def __repr__(self):
        return f"{self.__class__.__name__}(Type: {self.KIND}, identifier: {self.identifier})"
    
    def __check_thermal_contact_resistance_values(self:Self):

        """Private method that checks that values in sheet thermal_contact_resistance of input file conductor_coupling.xlsx are larger or equal than zero.

        Args:
            self (Self): conductor object.
        
        Raises:
            ValueError: if negative values for the thermal contact resistance are provided by the user.
        """

        # Alias.
        therm_cont_res = self.dict_df_coupling["thermal_contact_resistance"].to_numpy()
        # Get index with negative values for thermal contact resistances.
        negative_idx = np.nonzero(therm_cont_res < 0.0)
        # Check if negative_idx is not empty.
        if negative_idx:
            # negative_idx is not empty: raise ValueError.
            raise ValueError(f"Thermal contact resistance should be >= 0.0. Please check index {negative_idx} in sheet thermal_contact_resistance of input file {self.file_input['STRUCTURE_COUPLING']}")

    def __check_variable_contact_perimeter(self:Self)->dict:
        """Private method that performs checks on user defined auxiliary input file variable_contact_perimeter.xlsx.

        Args:
            self (Self): conductor object.

        Returns:
            dict: cleaned collection of data to interpolate the variable contact perimeter.
        """
        
        # Checks the column heading in self.dict_df_variable_contact_perimeter.
        self.__check_heading_variable_contact_perimeter()
        # Checks consistencty between self.dict_df_variable_contact_perimeter and self.dict_df_coupling["contact_perimeter_flag"] (the reference one).
        self.__check_variable_contact_perimeter_consistency()
        # Check if user provided valid coordinates to perform variable contact 
        # perimeter interpolation alogn conductor spatial discretization.
        self.__check_variable_contact_perimeter_coordinate()
        # Checks if there are sheets in excess in file 
        # variable_contact_perimeter.xlsx of if in valid sheets there are 
        # columns in exces and removes them.
        v_c_p = self.__check_variable_contact_perimeter_surplus_info()
        
        return v_c_p


    def __check_heading_variable_contact_perimeter(self:Self):
        """Private method that checks if user defined repeated headings in any of the sheets of auxiliary file variable_contact_perimeter.xlsx

        Args:
            self (Self): conductor object.

        Raises:
            ValueError: if there are repeated headings in any of the sheets of auxiliary file variable_contact_perimeter.xlsx
        """
        
        sheets = dict()
        # Loop to check if there are repeated headings in any of the sheets in 
        # file variable_contact_perimeters.xlsx
        for comp_id,df in self.dict_df_variable_contact_perimeter.items():
            headers = df.columns.to_list()
            # Remove duplicate items exploiting set primitive.
            unique_headers = set(headers)
            # Check if there are repeated headings.
            if len(unique_headers) < len(headers):
                # Found repeated headings: update the dictionary of sheets with 
                # repeated headings in file variable_contact_perimeter.xlsx
                sheets[comp_id] = [
                    col for col in unique_headers if headers.count(col) > 1
                ]
        
        # Raise error if sheets is not empty.
        if sheets:
            raise ValueError(f"Found repeated headings. Please check remove repeated headings in file {self.file_input['VARIABLE_CONTACT_PERIMETER']} as described below:\n{sheets}")

    def __check_variable_contact_perimeter_consistency(self:Self):

        """Private method that checks the consistecy between sheet contact_perimeter_flags in input file conductor_coupling.xlsx and the user defined auxiliary input file variable_contact_perimeter.xlsx. The assumption is that sheet in file conductor_coupling.xlsx is correct.

        Args:
            self (Self): conductor object.

        Raises:
            KeyError: if a sheet is totally missing in auxiliary input file variable_contact_perimeter.xlsx, i.e. user forget to define a full set of interfaces with a variable contact perimeter.
            ValueError: if in an existing sheet of auxiliary input file variable_contact_perimeter.xlsx, some interfaces which are defined with a variable contact perimeter are missing.
        """

        # Aliases.
        cont_peri_flag = self.dict_df_coupling["contact_perimeter_flag"]
        identifiers = cont_peri_flag.index.to_list()
        
        missing_var_cont_peri = dict()

        # Loop on rows of sheet contact_perimeter_flags.
        for row_idx, row_name in enumerate(identifiers):
            # Loop on colums of sheet contact_perimeter_flags, scan only the 
            # upper triangular matrix, excluding main diagonal.
            for col_idx, col_name in enumerate(
                identifiers[row_idx+1:],row_idx+1
            ):
                # Check if there is iterface with flag for variable contact 
                # perimeter.
                if cont_peri_flag.iat[row_idx,col_idx] == VARIABLE_CONTACT_PERIMETER:
                    # Found an interface with flag for variable contact 
                    # perimeter: make checks on auxiliary file 
                    # variable_contact_perimeters.xlsx
                    # Check if sheet exists in variable_contact_perimeters.xlsx.
                    if row_name in self.dict_df_variable_contact_perimeter:
                        # Sheet exist.
                        cont_peri_def = self.dict_df_variable_contact_perimeter[row_name]
                        if row_name not in missing_var_cont_peri:
                            missing_var_cont_peri[row_name] = list()
                        # Check if component identifier (col_name) is included 
                        # in sheet headers.
                        if col_name not in cont_peri_def.columns.to_list():
                            # Component identifier (col_name) is not included 
                            # in sheet headers: update a dictionary to build 
                            # sutable error message.
                            missing_var_cont_peri[row_name].append(col_name)
                    else:
                        # Sheet does not exist: raise KeyError.
                        raise KeyError(f"User forgets to define a full set of interfaces with a variable contact perimeter. Please check auxiliary input file {self.file_input['VARIABLE_CONTACT_PERIMETER']}, missing sheet {row_name}.")
        
        # Filter missing_var_cont_peri on the only not empty list exploiting 
        # dictionary comprehension.
        missing_var_cont_peri = {key: value for key,value in missing_var_cont_peri.items() if value}
        # Check if missing_var_cont_peri is not empty.
        if missing_var_cont_peri:
            # missing_var_cont_peri is not empty: there are missing interfaces 
            # in some sheets of auxiliary file variable_contact_perimeter.xlsx.
            raise ValueError(f"Found missing interfaces with a variable contact perimeter. Please, in auxiliary input file {self.file_input['VARIABLE_CONTACT_PERIMETER']}, add the columns reported below:\n{missing_var_cont_peri}.")

    def __check_variable_contact_perimeter_coordinate(self:Self):

        """Private method that checks the consistency of the spatial cooridinates user provides to make interpolation of the variable contact perimeter in file variable_contact_perimeter.xlsx.

        Args:
            self (Self): conductor object.

        Raises:
            ValueError: if less than two coordinates are provided.
            ValueError: if any of the coordinates provided is negative
            ValueError: if first coordinate is not equal to 0.0
            ValueError: if last coordinate is larger that conductor length.
        """
    
        # Aliases.
        var_cont_peri = self.dict_df_variable_contact_perimeter

        for sheet_name, df in var_cont_peri.items():
            zcoord = df.iloc[:,0].to_numpy()
            # Check number of items in array zcoord.
            if zcoord.size < 2:
                # Wrong number of items in array zcoord.
                raise ValueError(f"User must provide at least two coordinates to define the variable contact perimeter. Please, check in sheet {sheet_name} of file {self.file_input['VARIABLE_CONTACT_PERIMETER']}.\n")
            
            # Check if coordinates are positive.
            if any(zcoord < 0.0):
                row_idx = np.nonzero(zcoord < 0.0)[0] + 1
                raise ValueError(f"Spatial coordinates for variable contact perimeter interpolation must be positive. Please, check rows {row_idx} in sheet {sheet_name} of file {self.file_input['VARIABLE_CONTACT_PERIMETER']}.\n ")

            # Check first item value in zcoord.
            if zcoord[0] != 0.0:
                raise ValueError(f"First z coordinate value should be 0.0. Please, check row 2 in sheet {sheet_name} of file {self.file_input['VARIABLE_CONTACT_PERIMETER']}.\n ")
            
            # Check last item value in zcoord.
            if zcoord[-1] > self.inputs["ZLENGTH"]:
                raise ValueError(f"Last z coordinate value should be lower or equal than the conductor length ({self.inputs['ZLENGTH']} m). Please, check row {zcoord.size + 1} in sheet {sheet_name} of file {self.file_input['VARIABLE_CONTACT_PERIMETER']}.\n ")

    def __check_variable_contact_perimeter_surplus_info(self:Self)->dict:
        """Private method that checks if there are any surplus sheets in user defined auxiliary input file variable_contact_perimeter.xlsx and removes them. Moreover the method checks if there are surplus columns from valid sheets in the same file and removes them. Reference sheet names and colums names cames from attribute self.dict_df_coupling["contact_perimeter_flag"].

        Args:
            self (Self): conductor object.

        Returns:
            dict: cleaned collection of data to interpolate the variable contact perimeter.
        """

        # Aliases.
        cont_peri_flag = self.dict_df_coupling["contact_perimeter_flag"]
        identifiers = cont_peri_flag.index.to_list()
        var_cont_peri = self.dict_df_variable_contact_perimeter

        # Build the test set of sheets from var_cont_peri
        sheets = set(var_cont_peri.keys())
        # Initialize empty reference set for sheets
        reference_sheets = set()

        # Loop on rows of sheet contact_perimeter_flags.
        for row_idx, row_name in enumerate(identifiers):
            # Initialize reference set of columns
            reference_columns = set()

            # Loop on colums of sheet contact_perimeter_flags, scan only the 
            # upper triangular matrix, excluding main diagonal.
            for col_idx, col_name in enumerate(
                identifiers[row_idx+1:],row_idx+1
            ):
                # Check if there is iterface with flag for variable contact 
                # perimeter.
                if cont_peri_flag.iat[row_idx,col_idx] == VARIABLE_CONTACT_PERIMETER:
                    # Found an interface with flag for variable contact 
                    # perimeter: update set reference_sheets
                    reference_sheets.add(row_name)
                    # Update set reference_columns
                    reference_columns.add(col_name)
            
            if row_name in var_cont_peri:
                # Build the test set of columns from a dataframe in 
                # var_cont_peri removing the first header (z_coord).
                columns = set(var_cont_peri[row_name].columns.to_list()[1:])
                # Compare set columns agaist set referece_columns: get the 
                # elements in columns that are not in reference_columns, i.e. 
                # the surplus columns information.
                column_diff = columns.difference(reference_columns)
                # Check if columns_diff is not empyt
                if column_diff:
                    # Remove extra columns from var_cont_peri[row_name].
                    var_cont_peri[row_name].drop(columns=column_diff,inplace=True)
                    warnings.warn(f"Removed surplus columns {column_diff} from sheet {row_name} in {self.file_input['VARIABLE_CONTACT_PERIMETER']}.\n")
            
        # Compare set sheets agaist set referece_sheets: get the elements in 
        # sheet that are not in reference:sheets, i.e. the surplus information.
        sheet_diff = sheets.difference(reference_sheets)
        # Check if sheet_diff is not empyt
        if sheet_diff:
            # Remove extra sheets from var_cont_peri.
            for sheet in sheet_diff:
                var_cont_peri.pop(sheet)
                warnings.warn(f"Removed surplus sheet {sheet} from in {self.file_input['VARIABLE_CONTACT_PERIMETER']}.\n")

        return var_cont_peri

    def __delete_equipotential_inputs(self: Self):
        """Private method that deletes input values EQUIPOTENTIAL_SURFACE_NUMBER and EQUIPOTENTIAL_SURFACE_COORDINATE if they are not needed.

        Args:
            self (Self): conductor object.
        """
        pass
        # del self.operations["EQUIPOTENTIAL_SURFACE_NUMBER"]
        # del self.operations["EQUIPOTENTIAL_SURFACE_COORDINATE"]

    def __convert_equipotential_surface_coordinate_to_array(self: Self):
        """Private method used as switch to convert values corresponding to key EQUIPOTENTIAL_SURFACE_COORDINATE to numpy array according to the original type (integer for single value or string for multiple values).

        Args:
            self (Self): conductor object.
        """
        _ = {int: self.__int_to_array, str: self.__str_to_array}

        _[type(self.operations["EQUIPOTENTIAL_SURFACE_COORDINATE"])]()

    def __int_to_array(self: Self):
        """Private method that convert integer value of key EQUIPOTENTIAL_SURFACE_COORDINATE to numpy array; used when one coordinate is assigned.

        Args:
            self (Self): conductor object.
        """
        self.operations["EQUIPOTENTIAL_SURFACE_COORDINATE"] = np.array(
            [self.operations["EQUIPOTENTIAL_SURFACE_COORDINATE"]], dtype=float
        )

    def __str_to_array(self: Self):
        """Private method that convert sting value of key EQUIPOTENTIAL_SURFACE_COORDINATE to numpy array, used when more than one coordinate is assigned.

        Args:
            self (Self): conductor object.
        """
        self.operations["EQUIPOTENTIAL_SURFACE_COORDINATE"] = np.array(
            self.operations["EQUIPOTENTIAL_SURFACE_COORDINATE"].split(","), dtype=float
        )

    def __checks_equipotential_surface_coordinate(self: Self):
        """Private method that make consistency checks on input values EQUIPOTENTIAL_SURFACE_COORDINATE and EQUIPOTENTIAL_SURFACE_NUMBER.

        Args:
            self (Self): conductor object.

        Raises:
            ValueError: raise error if the number of assinged coordinates is different from the number of declared equipotential surfaces.
            ValueError: raises value error if equipotential coordinate is exceeded conductor length.
            ValueError: raises value error if equipotential coordinate is negative.
        """

        if (
            len(self.operations["EQUIPOTENTIAL_SURFACE_COORDINATE"])
            != self.operations["EQUIPOTENTIAL_SURFACE_NUMBER"]
        ):
            raise ValueError(
                f"The number of the equipotential surfaces coordinates must be equal to the number of declared equipotential surfaces:\n{self.operations['EQUIPOTENTIAL_SURFACE_COORDINATE']=};\n{self.operations['EQUIPOTENTIAL_SURFACE_NUMBER']=}\n"
            )
        if (
            np.max(self.operations["EQUIPOTENTIAL_SURFACE_COORDINATE"])
            > self.inputs["ZLENGTH"]
        ):
            raise ValueError(
                f"Equipotential surface coordinate cannot exceed conductor length:\n{np.max(self.operations['EQUIPOTENTIAL_SURFACE_COORDINATE'])=}m;\n{self.inputs['ZLENGTH']=}m"
            )
        if np.min(self.operations["EQUIPOTENTIAL_SURFACE_COORDINATE"]) < 0.0:
            raise ValueError(
                f"Equipotential surface coordinate must be positive:\n{np.min(self.operations['EQUIPOTENTIAL_SURFACE_COORDINATE'])=}m"
            )

    def __manage_equipotential_surfaces_coordinate(self: Self):
        """Private method that manages the input values for the definition of the equipotential surfaces.

        Args:
            self (Self): conductor object.
        """

        self.__convert_equipotential_surface_coordinate_to_array()
        self.__checks_equipotential_surface_coordinate()

    def conductor_components_instance(self, simulation):
        """
        Method that makes instances of the conductor components defined with workbook conductorc_input.xlsx and conductcor_operation.xlsx (cdp, 11/2020)
        """
        dict_file_path = dict(
            input=os.path.join(self.BASE_PATH, self.file_input["STRUCTURE_ELEMENTS"]),
            operation=os.path.join(self.BASE_PATH, self.file_input["OPERATION"]),
        )
        # Load workbook conductor_input.xlsx.
        wb_input = load_workbook(dict_file_path["input"], data_only=True)
        # Load workbook conductor_operation.xlsx.
        wb_operations = load_workbook(dict_file_path["operation"], data_only=True)

        listOfComponents = wb_input.get_sheet_names()
        self.inventory["FluidComponent"] = ComponentCollection("CHAN")
        self.inventory["StrandMixedComponent"] = ComponentCollection("STR_MIX")
        self.inventory["StrandStabilizerComponent"] = ComponentCollection("STR_STAB")
        self.inventory["StackComponent"] = ComponentCollection("STACK")
        self.inventory["JacketComponent"] = ComponentCollection("Z_JACKET")
        self.inventory["StrandComponent"] = ComponentCollection()
        self.inventory["SolidComponent"] = ComponentCollection()
        self.inventory["all_component"] = ComponentCollection()
        for sheetID in listOfComponents:
            sheet = wb_input[sheetID]
            sheetOpar = wb_operations[sheetID]
            # Call function check_object_number to check that the number of objects defined in sheets of file conductor_input.xlsx and conductor_operation.xlsx are the same.
            check_object_number(
                self,
                dict_file_path["input"],
                dict_file_path["operation"],
                sheet,
                sheetOpar,
            )
            # Call function check_repeated_headings to check if there are repeaded headings in sheet of file conductor_input.xlsx.
            check_repeated_headings(dict_file_path["input"], sheet)
            # Call function check_repeated_headings to check if there are repeaded headings in sheet of file conductor_operation.xlsx.
            check_repeated_headings(dict_file_path["operation"], sheetOpar)
            # Call function check_headers to check if sheets of file conductor_input.xlsx and conductor_operation.xlsx have exactly the same headers.
            check_headers(
                self,
                dict_file_path["input"],
                dict_file_path["operation"],
                sheet,
                sheetOpar,
            )
            kindObj = sheet.cell(row=1, column=1).value  # sheet["A1"].value
            numObj = int(sheet.cell(row=1, column=2).value)  # sheet["B1"].value
            if kindObj == "CHAN":
                # Assign the total number of defined FluidComponent object to
                # attribute number of object ComponentCollection.
                self.inventory["FluidComponent"].number = numObj
                for ii in range(1, 1 + numObj):
                    # ["FluidComponent"].collection: list of FluidComponent
                    # objects;
                    # ["all_component"].collection list of all objects
                    self.inventory["FluidComponent"].collection.append(
                        FluidComponent(sheet, sheetOpar, ii, dict_file_path)
                    )
                    self.inventory["all_component"].collection.append(
                        self.inventory["FluidComponent"].collection[ii - 1]
                    )
                # end for ii (cdp, 09/2020)
            elif kindObj == "STACK":
                # Assign the total number of defined StackComponent object to
                # attribute number of object ComponentCollection.
                self.inventory["StackComponent"].number = numObj
                for ii in range(1, 1 + numObj):
                    # ["StackComponent"].collection: list of StackComponent objects;
                    # ["StrandComponent"].collection: list of StrandComponent objects;
                    # ["SolidComponent"].collection: list of SolidComponent objects;
                    # ["all_component"].collection: list of all objects
                    # (cdp, 09/2020)
                    self.inventory["StackComponent"].collection.append(
                        StackComponent(
                            simulation, sheet, ii, kindObj, dict_file_path, self
                        )
                    )
                    self.inventory["StrandComponent"].collection.append(
                        self.inventory["StackComponent"].collection[ii - 1]
                    )
                    self.inventory["SolidComponent"].collection.append(
                        self.inventory["StackComponent"].collection[ii - 1]
                    )
                    self.inventory["all_component"].collection.append(
                        self.inventory["StackComponent"].collection[ii - 1]
                    )
                # end for ii (cdp, 09/2020)
            elif kindObj == "STR_MIX":
                # Assign the total number of defined StrandMixedComponent object to
                # attribute number of object ComponentCollection.
                self.inventory["StrandMixedComponent"].number = numObj
                for ii in range(1, 1 + numObj):
                    # ["StrandMixedComponent"].collection: list of StrandMixedComponent objects;
                    # ["StrandComponent"].collection: list of StrandComponent objects;
                    # ["SolidComponent"].collection: list of SolidComponent objects;
                    # ["all_component"].collection: list of all objects
                    # (cdp, 09/2020)
                    self.inventory["StrandMixedComponent"].collection.append(
                        StrandMixedComponent(
                            simulation, sheet, ii, kindObj, dict_file_path, self
                        )
                    )
                    self.inventory["StrandComponent"].collection.append(
                        self.inventory["StrandMixedComponent"].collection[ii - 1]
                    )
                    self.inventory["SolidComponent"].collection.append(
                        self.inventory["StrandMixedComponent"].collection[ii - 1]
                    )
                    self.inventory["all_component"].collection.append(
                        self.inventory["StrandMixedComponent"].collection[ii - 1]
                    )
                # end for ii (cdp, 09/2020)
            elif kindObj == "STR_STAB":
                # Assign the total number of defined StrandStabilizerComponent object to
                # attribute number of object ComponentCollection.
                self.inventory["StrandStabilizerComponent"].number = numObj
                for ii in range(1, 1 + numObj):
                    # ["StrandStabilizerComponent"].collection: list of StrandStabilizerComponent objects;
                    # ["StrandComponent"].collection: list of StrandComponent objects;
                    # ["SolidComponent"].collection: list of SolidComponent objects;
                    # ["all_component"].collection: list of all objects
                    # (cdp, 09/2020)
                    self.inventory["StrandStabilizerComponent"].collection.append(
                        StrandStabilizerComponent(
                            simulation, sheet, ii, kindObj, dict_file_path, self
                        )
                    )
                    self.inventory["StrandComponent"].collection.append(
                        self.inventory["StrandStabilizerComponent"].collection[ii - 1]
                    )
                    self.inventory["SolidComponent"].collection.append(
                        self.inventory["StrandStabilizerComponent"].collection[ii - 1]
                    )
                    self.inventory["all_component"].collection.append(
                        self.inventory["StrandStabilizerComponent"].collection[ii - 1]
                    )
                # end for ii (cdp, 09/2020)
            elif kindObj == "Z_JACKET":
                # Assign the total number of defined JacketComponent object to
                # attribute number of object ComponentCollection.
                self.inventory["JacketComponent"].number = numObj
                # Z_JACKET since it must be the last object in the list (cdp, 06/2020)
                for ii in range(1, 1 + numObj):
                    # ["JacketComponent"].collection: list of JacketComponent objects;
                    # ["SolidComponent"].collection: list of SolidComponent objects;
                    # ["all_component"].collection: list of all objects
                    # (cdp, 09/2020)
                    self.inventory["JacketComponent"].collection.append(
                        JacketComponent(
                            simulation, sheet, ii, kindObj, dict_file_path, self
                        )
                    )
                    self.inventory["SolidComponent"].collection.append(
                        self.inventory["JacketComponent"].collection[ii - 1]
                    )
                    self.inventory["all_component"].collection.append(
                        self.inventory["JacketComponent"].collection[ii - 1]
                    )
                # end for ii (cdp, 09/2020)
            else:
                raise NameError(
                    f"ERROR in method {self.__init__.__name__} of class {self.__class__.__name__}: kind object {kindObj} does not exist.\nPlease check cell A1 in sheet {sheet} of file {self.file_input['STRUCTURE_ELEMENTS']}.\n"
                )
            # end if kindObj (cdp, 11/2020)
        # end for sheetID (cdp, 11/2020)

        # Total number of StrandComponent objects (cdp, 09/2020)
        self.inventory["StrandComponent"].number = (
            self.inventory["StrandMixedComponent"].number
            + self.inventory["StrandStabilizerComponent"].number
            + self.inventory["StackComponent"].number
        )
        # Total number of SolidComponent objects (cdp, 09/2020)
        self.inventory["SolidComponent"].number = (
            self.inventory["StrandComponent"].number
            + self.inventory["JacketComponent"].number
        )
        # Total number of Conductor component objects (cdp, 09/2020)
        self.inventory["all_component"].number = (
            self.inventory["FluidComponent"].number
            + self.inventory["SolidComponent"].number
        )

    # end method Conductor_components_instance (cdp, 11/2020)

    def __get_total_cross_section(self):
        """Private method that evaluates: 1) the total cross section of strands and stacks object of the conductor; 2) the total cross section of superconducting materials of the conductor."""

        self.total_so_cross_section = np.array(
            [
                obj.inputs["CROSSECTION"]
                for obj in self.inventory["StrandComponent"].collection
            ]
        ).sum()
        self.total_sc_cross_section = np.array(
            [
                obj.cross_section["sc"]
                for obj in self.inventory["StrandComponent"].collection
                if isinstance(obj, (StackComponent, StrandMixedComponent))
            ]
        ).sum()

    def __coordinates(self,simulation):
        """Private method that allows to evaluate the grid coordinates and assign them to the conductor objects and its comonents according to the value of flag grid_input["ITYMSH"]."""
        if self.grid_input["ITYMSH"] >= 0:
            n_nod = self.grid_input["NELEMS"] + 1
            file_path = os.path.join(self.BASE_PATH, self.file_input["GRID_DEFINITION"])
            self.grid_features["N_nod"] = check_max_node_number(n_nod, self, file_path)
            for comp in self.inventory["all_component"].collection:
                build_coordinates_of_barycenter(simulation,self,comp)
            # Evaluate conductor spatial discretization along z axis. This is used for the thermal fluid-dynamic solution.
            conductor_spatial_discretization(simulation,self)
        else:
            # Call function user_defined_grid: makes ckecks on the user defined
            # grid and then assinge the coordinates to the conductor components.
            user_defined_grid(self)

    def __build_equation_idx(self):
        """Private method that evaluates the index of the velocity, pressure and temperature equation of the FluidComponent objects, collecting them in a dictionary of NamedTuple, together with the index of the temperature equation of the SolidComponent objects stored as integer in the same dictionary.
        """
        
        # Constructor of the namedtuple to store the index of the equations for 
        # FluidComponent objects.
        Fluid_eq_idx = namedtuple(
            "Fluid_eq_idx",
            ("velocity","pressure","temperature")
        )

        # self.equation_index -> dict: collection of NamedTuple with the index
        # of velocity, pressure and temperaure equation for FluidComponent
        # objects and of integer for the index of the temperature equation of
        # SolidComponent.
        
        # Build dictionary of NamedTuple with the index of the equations for 
        # FluidComponent objects exploiting dictionary comprehension.
        self.equation_index = {
            fcomp.identifier:Fluid_eq_idx(
                # velocity equation index
                velocity=fcomp_idx,
                # pressure equation index
                pressure=fcomp_idx + self.inventory["FluidComponent"].number,
                # temperature equation index
                # Exploit left binary shift, equivalent to:
                # fcomp_idx + 2 * conductor.inventory["FluidComponent"].number
                temperature=(
                    fcomp_idx
                    + (self.inventory["FluidComponent"].number << 1)
                )
            )
            for fcomp_idx,fcomp in enumerate(
                self.inventory["FluidComponent"].collection
            )
        }
        
        # Update dictionary equation_index with integer corresponding to the 
        # index of the equations for SolidComponent objects exploiting 
        # dictionary comprehension and dictionary method update.
        self.equation_index.update(
            {
                scomp.identifier: scomp_idx + self.dict_N_equation[
                    "FluidComponent"
                ]
                for scomp_idx,scomp in enumerate(
                    self.inventory["SolidComponent"].collection
                )
            }
        )

    def __build_multi_index(self) -> pd.MultiIndex:
        """Private method that builds multindex used in pandas dataframes used to store the nodal coordinates and the connectivity (matrix) of each conductor component.

        Returns:
            pd.MultiIndex: pandas multindex with 'Kind' (parent class) and 'Identifier' (the component identifier).
        """
        identifiers = [
            obj.identifier for obj in self.inventory["all_component"].collection
        ]
        kinds = list()
        for obj in self.inventory["all_component"].collection:
            if isinstance(obj, StrandComponent):
                kinds.append(StrandComponent.__name__)
            else:
                kinds.append(obj.__class__.__name__)

        cat_kind = pd.CategoricalIndex(
            np.tile(kinds, self.grid_input["NELEMS"] + 1),
            dtype="category",
            ordered=True,
            categories=["FluidComponent", "StrandComponent", "JacketComponent"],
        )
        cat_ids = pd.CategoricalIndex(
            np.tile(identifiers, self.grid_input["NELEMS"] + 1),
            dtype="category",
            ordered=True,
            categories=identifiers,
        )

        return pd.MultiIndex.from_arrays(
            [cat_kind, cat_ids], names=["Kind", "Identifier"]
        )

    def __build_multi_index_current_carriers(self) -> pd.MultiIndex:
        """Private method that builds multindex used in pandas dataframes used to store the nodal coordinates and the connectivity (matrix) only for conductor components of kind strand (StrandMixedComponent, StrandStabilizerComponent and StrandSuperconductroComponent).

        Returns:
            pd.MultiIndex:  pandas multindex with 'Kind' (parent class) and 'Identifier' (the component identifier).
        """
        identifiers = [
            obj.identifier for obj in self.inventory["StrandComponent"].collection
        ]
        kinds = [
            obj.__class__.__name__
            for obj in self.inventory["StrandComponent"].collection
        ]

        cat_kind = pd.CategoricalIndex(
            np.tile(kinds, self.grid_input["NELEMS"] + 1),
            dtype="category",
            ordered=True,
            categories=[
                "StrandMixedComponent",
                "StrandStabilizerComponent",
                "StackComponent",
            ],
        )
        cat_ids = pd.CategoricalIndex(
            np.tile(identifiers, self.grid_input["NELEMS"] + 1),
            dtype="category",
            ordered=True,
            categories=identifiers,
        )

        return pd.MultiIndex.from_arrays(
            [cat_kind, cat_ids], names=["Kind", "Identifier"]
        )

    def __initialize_attributes(self: Self, simulation: object):
        """Private method that initializes usefull attributes of conductor object.

        Args:
            self (Self):

        Args:
            self (Self): conductor object
            simulation (object): simulation object.

        Raises:
            ValueError: raise error if spatial coordinates in sheet Time_evolutions of file conductor_diagnostic are larger than the conductor length.
            ValueError: raise error if time values in sheet Spatial_distribution of file conductor diagnostic are larger than the end time of the simulation.
        """

        self.dict_topology = dict()  # dictionary declaration (cdp, 09/2020)
        self.dict_interf_peri = dict()  # dictionary declaration (cdp, 07/2020)
        # Call method Get_conductor_topology to evaluate conductor topology: \
        # interfaces between channels, channels and solid components and between \
        # solid components(cdp, 09/2020)
        self.get_conductor_topology(simulation.environment)
        
        # Call private method __get_conductor_interfaces to get the interfaces 
        # between conductor components. This method could replace 
        # get_conductor_topology but this change must carefully discusse with 
        # both prof Savoldi and prof Savino.
        self.__get_conductor_interfaces(simulation.environment)

        self.dict_node_pt = dict()
        self.dict_Gauss_pt = dict()

        self.heat_rad_jk = dict()
        self.heat_exchange_jk_env = dict()

        # **NUMERICS**
        # evaluate value of theta_method according to flag METHOD (cdo, 08/2020)
        # Adams Moulton value is temporary and maybe non correct
        _ = dict(BE=1.0, CE=0.5, AM4=1.0 / 24.0)
        self.theta_method = _[self.inputs["ELECTRIC_METHOD"]]
        self.electric_theta = _[self.inputs["ELECTRIC_METHOD"]]
        conductorlogger.debug(f"Defined electric_theta\n")
        ## Evaluate parameters useful in function \
        # Transient_solution_functions.py\STEP (cdp, 07/2020)
        # dict_N_equation keys meaning:
        # ["FluidComponent"]: total number of equations for FluidComponent \
        # objects (cdp, 07/2020);
        # ["StrandComponent"]: total number of equations for StrandComponent objects (cdp, 07/2020);
        # ["JacketComponent"]: total number of equations for JacketComponent objects (cdp, 07/2020);
        # ["SolidComponent"]: total number of equations for SolidComponent \
        # objects (cdp, 07/2020);
        # ["NODOFS"]: total number of equations for for each node, i.e. Number Of \
        # Degrees Of Freedom, given by: \
        # 3*(number of channels) + (number of strands) + (number of jackets) \
        # (cdp, 07/2020);
        self.dict_N_equation = dict(
            FluidComponent=3 * self.inventory["FluidComponent"].number,
            StrandComponent=self.inventory["StrandComponent"].number,
            JacketComponent=self.inventory["JacketComponent"].number,
            SolidComponent=self.inventory["SolidComponent"].number,
        )
        # necessary since it is not allowed to use the value of a dictionary key \
        # before that the dictionary is fully defined (cdp, 09/2020)
        self.dict_N_equation.update(
            NODOFS=self.dict_N_equation["FluidComponent"]
            + self.dict_N_equation["SolidComponent"]
        )
        # Exploit left binary shift, equivalent to:
        # self.dict_N_equation["NODOFS2"] = 2 * self.dict_N_equation["NODOFS"]
        self.dict_N_equation["NODOFS2"] = self.dict_N_equation["NODOFS"] << 1
        # dict_band keys meaning:
        # ["Half"]: half band width, including main diagonal (IEDOFS) (cdp, 09/2020)
        # ["Main_diag"]: main diagonal index within the band (IHBAND) (cdp, 09/2020)
        # ["Full"]: full band width, including main diagonal (IBWIDT) (cdp, 09/2020)
        self.dict_band = dict(
            Half=2 * self.dict_N_equation["NODOFS"],
            Main_diag=2 * self.dict_N_equation["NODOFS"] - 1,
            Full=4 * self.dict_N_equation["NODOFS"] - 1,
        )
        # self.MAXDOF = self.dict_N_equation["NODOFS"]*MAXNOD
        self.EQTEIG = np.zeros(self.dict_N_equation["NODOFS"])
        # dict_norm keys meaning:
        # ["Solution"]: norm of the solution (cdp, 09/2020)
        # ["Change"]: norm of the solution variation wrt the previous time step \
        # (cdp, 09/2020)
        self.dict_norm = dict(
            Solution=np.zeros(self.dict_N_equation["NODOFS"]),
            Change=np.zeros(self.dict_N_equation["NODOFS"]),
        )
        
        # Call method __build_equation_idx to build attribute equation_index;
        # self.equation_index -> dict: collection of NamedTuple with the index
        # of velocity, pressure and temperaure equation for FluidComponent
        # objects and of integer for the index of the temperature equation of
        # SolidComponent. This is used in funcion step to solve the thermal 
        # hydraulic problem.
        self.__build_equation_idx()

        # evaluate attribute EIGTIM exploiting method Aprior (cdp, 08/2020)
        self.aprior()
        path_diagnostic = os.path.join(self.BASE_PATH, self.file_input["OUTPUT"])
        # Load the content of column self.ID of sheet Space in file conductors_disgnostic.xlsx as a series and convert to numpy array of float.
        self.Space_save = (
            pd.read_excel(
                path_diagnostic,
                sheet_name="Spatial_distribution",
                skiprows=2,
                header=0,
                usecols=[self.identifier],
                squeeze=True,
            )
            .dropna()
            .to_numpy()
            .astype(float)
        )
        # Adjust the user defined diagnostic.
        self.Space_save = set_diagnostic(
            self.Space_save, lb=0.0, ub=simulation.transient_input["TEND"]
        )
        # Check on spatial distribution diagnostic.
        if self.Space_save.max() > simulation.transient_input["TEND"]:
            raise ValueError(
                f"File {self.file_input['OUTPUT']}, sheet Space, conductor {self.identifier}: impossible to save spatial distributions at time {self.Space_save.max()} s since it is larger than the end time of the simulation {simulation.transient_input['TEND']} s.\n"
            )
        # End if self.Space_save.max() > simulation.transient_input["TEND"]
        # index pointer to save solution spatial distribution (cdp, 12/2020)
        self.i_save = 0
        # list of number of time steps at wich save the spatial discretization
        self.num_step_save = np.zeros(self.Space_save.shape, dtype=int)
        # Load the content of column self.identifier of sheet Time in file conductors_disgnostic.xlsx as a series and convert to numpy array of float.
        self.Time_save = (
            pd.read_excel(
                path_diagnostic,
                sheet_name="Time_evolution",
                skiprows=2,
                header=0,
                usecols=[self.identifier],
                squeeze=True,
            )
            .dropna()
            .to_numpy()
            .astype(float)
        )
        # Adjust the user defined diagnostic.
        self.Time_save = set_diagnostic(
            self.Time_save, lb=0.0, ub=self.inputs["ZLENGTH"]
        )
        # Check on time evolution diagnostic.
        if self.Time_save.max() > self.inputs["ZLENGTH"]:
            raise ValueError(
                f"File {self.file_input['OUTPUT']}, sheet Time, conductor {self.identifier}: impossible to save time evolutions at axial coordinate {self.Time_save.max()} s since it is ouside the computational domain of the simulation [0, {self.inputs['ZLENGTH']}] m.\n"
            )
        # End if self.Time_save.max() > self.inputs["ZLENGTH"]

        # declare dictionaries to store Figure and axes objects to constructi real \
        # time figures (cdp, 10/2020)
        self.dict_Figure_animation = dict(T_max=dict(), mfr=dict())
        self.dict_axes_animation = dict(T_max=dict(), mfr=dict())
        self.dict_canvas = dict(T_max=dict(), mfr=dict())
        self.color = ["r*", "bo"]

        # Introduced for the electric module.
        self.total_elements = (
            self.grid_input["NELEMS"] * self.inventory["all_component"].number
        )
        self.total_nodes = (self.grid_input["NELEMS"] + 1) * self.inventory[
            "all_component"
        ].number

        self.total_elements_current_carriers = (
            self.grid_input["NELEMS"] * self.inventory["StrandComponent"].number
        )
        self.total_nodes_current_carriers = (
            self.grid_input["NELEMS"] + 1
        ) * self.inventory["StrandComponent"].number

        # Convert to matrix (make them sparse matrices)
        # +1 keeps into account the Environment object in the
        # conductor_coupling workbook.
        self.electric_conductance = (
            self.dict_df_coupling["electric_conductance"]
            .iloc[
                self.inventory["FluidComponent"].number
                + 1 : self.inventory["FluidComponent"].number
                + self.inventory["StrandComponent"].number
                + 1,
                self.inventory["FluidComponent"].number
                + 1 : self.inventory["FluidComponent"].number
                + self.inventory["StrandComponent"].number
                + 1,
            ]
            .to_numpy()
        )
        del self.dict_df_coupling["electric_conductance"]
        self.electric_conductance_mode = (
            self.dict_df_coupling["electric_conductance_mode"]
            .iloc[
                self.inventory["FluidComponent"].number
                + 1 : self.inventory["FluidComponent"].number
                + self.inventory["StrandComponent"].number
                + 1,
                self.inventory["FluidComponent"].number
                + 1 : self.inventory["FluidComponent"].number
                + self.inventory["StrandComponent"].number
                + 1,
            ]
            .to_numpy()
        )
        del self.dict_df_coupling["electric_conductance_mode"]

        # Initialize resistance matrix to a dummy value (sparse matrix)
        self.electric_resistance_matrix = diags(
            10.0 * np.ones(self.total_elements_current_carriers),
            offsets=0,
            shape=(
                self.total_elements_current_carriers,
                self.total_elements_current_carriers,
            ),
            format="csr",
            dtype=float,
        )

        self.inductance_matrix = np.zeros(
            (
                self.total_elements_current_carriers,
                self.total_elements_current_carriers,
            )
        )

        self.electric_conductance_matrix = csr_matrix(
            (self.total_nodes_current_carriers, self.total_nodes_current_carriers),
            dtype=float,
        )

        self.build_electric_mass_matrix_flag = True
        self.electric_mass_matrix = lil_matrix(
            (
                self.total_elements_current_carriers
                + self.total_nodes_current_carriers,
                self.total_elements_current_carriers
                + self.total_nodes_current_carriers,
            ),
            dtype=float,
        )

        self.equipotential_node_index = np.zeros(
            (
                self.operations["EQUIPOTENTIAL_SURFACE_NUMBER"],
                self.inventory["StrandComponent"].number,
            ),
            dtype=int,
        )

        nn = 0
        for obj in self.inventory["StrandComponent"].collection:
            nn += obj.operations["FIX_POTENTIAL_NUMBER"]

        self.fixed_potential_index = np.zeros(nn, dtype=int)
        self.fixed_potential_value = np.zeros(nn)

        # Initialization moved in method build_electric_known_term_vector.
        # self.dict_node_pt["op_current"] = np.zeros(self.total_nodes_current_carriers)

        self.electric_known_term_vector = np.zeros(
            self.total_elements_current_carriers + self.total_nodes_current_carriers
        )

        self.electric_right_hand_side = np.zeros(
            self.total_elements_current_carriers + self.total_nodes_current_carriers
        )

        # Electric time initialization, to be understood where to actually do
        # this
        self.electric_time = 0.0  # s
        # Initialize the number of electric time steps to 0. This attribute 
        # will be updated at each electric time step; for each thermal time 
        # step it will start from 1. Value 0 is assumed only at initialization.
        self.cond_el_num_step = 0

        # conductorlogger.debug(
        #     f"Before call method {self.__initialize_mesh_dataframe.__name__}\n"
        # )
        self.__initialize_mesh_dataframe()
        # conductorlogger.debug(
        #     f"After call method {self.__initialize_mesh_dataframe.__name__}\n"
        # )

    def __initialize_mesh_dataframe(self):
        """Private method that initializes pandas dataframes used to store nodal coordinates and connectivity (matrix)."""

        # conductorlogger.debug(
        #     f"Before call method {self.__build_multi_index.__name__}\n"
        # )
        multi_index = self.__build_multi_index()
        # conductorlogger.debug(
        #     f"After call method {self.__build_multi_index.__name__}\n"
        # )
        # conductorlogger.debug(
        #     f"Before call method {self.__build_multi_index_current_carriers.__name__}\n"
        # )
        multi_index_current_carriers = self.__build_multi_index_current_carriers()
        # conductorlogger.debug(
        #     f"After call method {self.__build_multi_index_current_carriers.__name__}\n"
        # )

        self.nodal_coordinates = pd.DataFrame(
            dict(
                x=np.zeros(self.total_nodes),
                y=np.zeros(self.total_nodes),
                z=np.zeros(self.total_nodes),
            ),
            index=multi_index,
        )

        self.connectivity_matrix = pd.DataFrame(
            dict(
                start=np.zeros(self.total_elements, dtype=int),
                end=np.zeros(self.total_elements, dtype=int),
                identifiers=pd.Series(np.zeros(self.total_elements), dtype=str),
            ),
            index=multi_index[: -self.inventory["all_component"].number],
        )

        self.connectivity_matrix_current_carriers = pd.DataFrame(
            dict(
                start=np.zeros(self.total_elements_current_carriers, dtype=int),
                end=np.zeros(self.total_elements_current_carriers, dtype=int),
                identifiers=pd.Series(
                    np.zeros(self.total_elements_current_carriers), dtype=str
                ),
            ),
            index=multi_index_current_carriers[
                : -self.inventory["StrandComponent"].number
            ],
        )

    def conductors_coupling(self):
        pass

    def get_conductor_topology(self, environment):

        """
        Method that evaluate the detailed conductor topology (dict_topology) together with the contatc perimeter (dict_interf_peri). Possible configurations are:
        1) channels in hydraulic parallel;
        2) channels not in hydraulic parallel but in thermal contact;
        3) stand alone channels (no mass or energy exchange);
        4) channels in thermal contact with solid components;
        5) channels not in thermal contact with solid components;
        6) contact between solid components.

        dict_topology describes the full conductor topology, i.e. interface between channels, channels and solid components, solid components as well as the isolated channels; it is organized into three sub dictionaries accessed by keys "ch_ch"; "ch_sol"; "sol_sol".
        Each sub dictionary is characterized by a series of strings that uniquely determines which components are in contact, and a list of object constituted by all the components in contact. Keys of this dictionaries are the identifier of the first object in alphabetical order constituting the interface.

        dict_interf_peri holds the values of the contact perimenter. It is also subdivided into three sub dictionaries with the same name as above. It is important to notice that in case of contact between channels, the "Open" and "Close" keys are introduced.
        (cdp, 09/2020)
        """

        # Alias
        interf_flag = self.dict_df_coupling["contact_perimeter_flag"]

        # nested dictionaries declarations
        self.dict_topology["ch_ch"] = dict()
        self.dict_topology["ch_ch"]["Hydraulic_parallel"] = dict()
        self.dict_topology["ch_ch"]["Thermal_contact"] = dict()
        self.dict_topology["Standalone_channels"] = list()
        self.dict_interf_peri["ch_ch"] = dict()
        self.dict_interf_peri["ch_ch"]["Open"] = dict(
            nodal=dict(),
            Gauss=dict(),
        )
        self.dict_interf_peri["ch_ch"]["Close"] = dict(
            nodal=dict(),
            Gauss=dict(),
        )
        self.dict_topology["ch_sol"] = dict()
        self.dict_interf_peri["ch_sol"] = dict(
            nodal=dict(),
            Gauss=dict(),
        )
        self.dict_topology["sol_sol"] = dict()
        self.dict_interf_peri["sol_sol"] = dict(
            nodal=dict(),
            Gauss=dict(),
        )
        self.dict_interf_peri["env_sol"] = dict(
            nodal=dict(),
            Gauss=dict(),
        )

        # Call method Get_hydraulic_parallel to obtain the channels subdivision \
        # into groups of channels that are in hydraulic parallel.
        self.get_hydraulic_parallel()

        # Nested loop channel-channel (cdp, 09/2020)
        for rr, fluid_comp_r in enumerate(self.inventory["FluidComponent"].collection):
            for cc, fluid_comp_c in enumerate(
                self.inventory["FluidComponent"].collection[rr + 1 :], rr + 1
            ):
                if (
                    abs(interf_flag.at[
                        fluid_comp_r.identifier, fluid_comp_c.identifier
                    ]
                    ) == 1
                ):
                    # There is at least thermal contact between fluid_comp_r 
                    # and fluid_comp_c
                    # Assign the contact perimeter value
                    (
                        self.dict_interf_peri["ch_ch"]["Close"],
                        self.dict_interf_peri["ch_ch"]["Open"]
                    ) = self.__assign_contact_perimeter_fluid_comps(
                        fluid_comp_r.identifier,
                        fluid_comp_c.identifier,
                    )
                    if cc == rr + 1:
                        # declare dictionary flag_found (cdp, 09/2020)
                        flag_found = dict()
                    # Invoke method Get_thermal_contact_channels to search for channels \
                    # that are only in thermal contact (cdp, 09/2020)
                    flag_found = self.get_thermal_contact_channels(
                        rr, cc, fluid_comp_r, fluid_comp_c, flag_found
                    )
                # end abs(interf_flag.at[fluid_comp_r.identifier, fluid_comp_c.identifier]) == 1 (cdp, 09/2020)
            # end for cc (cdp, 09/2020)
            if (
                self.dict_topology["ch_ch"]["Thermal_contact"].get(
                    fluid_comp_r.identifier
                )
                != None
            ):
                # key fluid_comp_r.identifier exists (cdp, 09/2020)
                if (
                    len(
                        list(
                            self.dict_topology["ch_ch"]["Thermal_contact"][
                                fluid_comp_r.identifier
                            ].keys()
                        )
                    )
                    - 3
                    > 0
                ):
                    # There are channels that are in thermal contact (cdp, 09/2020)
                    # Update the number of channels in thermal contact with fluid_comp_r: it \
                    # is the length of list group. The actual number of thermal contacts \
                    # can be larger, since some channels that are in thermal contact may \
                    # belong to groups of channels in hydraulic parallel; it is keep \
                    # into account by the key Actual_number. (cdp, 09/2020)
                    self.dict_topology["ch_ch"]["Thermal_contact"][
                        fluid_comp_r.identifier
                    ].update(
                        Number=len(
                            self.dict_topology["ch_ch"]["Thermal_contact"][
                                fluid_comp_r.identifier
                            ]["Group"]
                        )
                    )
                    # Assign values to key Actual_number (cdp, 09/2020)
                    self.dict_topology["ch_ch"]["Thermal_contact"][
                        fluid_comp_r.identifier
                    ].update(
                        Actual_number=len(
                            list(
                                self.dict_topology["ch_ch"]["Thermal_contact"][
                                    fluid_comp_r.identifier
                                ].keys()
                            )
                        )
                        - 3
                        + 1
                    )
                else:
                    # There are not channels that are in thermal contact remove \
                    # key fluid_comp_r.identifier from dictionary \
                    # self.dict_topology["ch_ch"]["Thermal_contact"] (cdp, 09/2020)
                    self.dict_topology["ch_ch"]["Thermal_contact"].pop(
                        fluid_comp_r.identifier
                    )
            # end if self.dict_topology["ch_ch"]\
            # ["Thermal_contact"].get(fluid_comp_r.identifier) != None (cdp, 09/2020)
        # end for rr (cdp, 09/2020)
        # Call method Find_Standalone_channels to search for eventually stand \
        # alone channels (not in hydraulic parallel) (cdp, 09/2020)
        self.find_standalone_channels()
        # dummy dictionary to store channel-solid topology (cdp, 09/2020)
        dict_topology_dummy_ch_sol = dict()
        # dummy to optimize nested loop (cdp, 09/2020)
        dict_chan_s_comp_contact = dict()
        # Nested loop channel-solid (cdp, 09/2020)
        for _, fluid_comp_r in enumerate(self.inventory["FluidComponent"].collection):
            # List linked channels-solid initialization (cdp, 09/2020)
            list_linked_chan_sol = list()
            # Nested dictionary in dict_topology_dummy_ch_sol declaration \
            # dict_topology_dummy_ch_sol
            dict_topology_dummy_ch_sol[fluid_comp_r.identifier] = dict()
            for _, s_comp_c in enumerate(self.inventory["SolidComponent"].collection):
                if (
                    abs(interf_flag.at[
                            fluid_comp_r.identifier, s_comp_c.identifier
                        ]
                    ) == 1
                ):
                    # There is contact between fluid_comp_r and s_comp_c

                    self.dict_interf_peri["ch_sol"] = self.__assign_contact_perimeter_not_fluid_only(
                        fluid_comp_r.identifier,
                        s_comp_c.identifier,
                        "ch_sol",
                    )
                    
                    # Interface identification (cdp, 09/2020)
                    dict_topology_dummy_ch_sol[fluid_comp_r.identifier][
                        s_comp_c.identifier
                    ] = f"{fluid_comp_r.identifier}_{s_comp_c.identifier}"
                    # Call method Chan_sol_interfaces (cdp, 09/2020)
                    [
                        dict_chan_s_comp_contact,
                        list_linked_chan_sol,
                    ] = self.chan_sol_interfaces(
                        fluid_comp_r,
                        s_comp_c,
                        dict_chan_s_comp_contact,
                        list_linked_chan_sol,
                    )
                # end if abs(interf_flag.at[fluid_comp_r.identifier, s_comp_c.identifier]) == 1: (cdp, 09/2020)
            # end for cc (cdp, 09/2020)
            # Call method Update_interface_dictionary to update dictionaries \
            # (cdp, 09/2020)
            [
                dict_topology_dummy_ch_sol,
                dict_chan_s_comp_contact,
            ] = self.update_interface_dictionary(
                fluid_comp_r,
                dict_topology_dummy_ch_sol,
                dict_chan_s_comp_contact,
                list_linked_chan_sol,
            )
        # end for rr (cdp, 09/2020)
        self.dict_topology.update(ch_sol=dict_topology_dummy_ch_sol)
        # dummy dictionary to store solid-solid topology (cdp, 09/2020)
        dict_topology_dummy_sol = dict()
        # dummy to optimize nested loop (cdp, 09/2020)
        dict_s_comps_contact = dict()
        # Nested loop solid-solid (cdp, 09/2020)
        for rr, s_comp_r in enumerate(self.inventory["SolidComponent"].collection):
            # List linked solids initialization (cdp, 09/2020)
            list_linked_solids = list()
            # Nested dictionary in dict_topology_dummy_sol declaration \
            # dict_topology_dummy_sol
            dict_topology_dummy_sol[s_comp_r.identifier] = dict()
            for _, s_comp_c in enumerate(
                self.inventory["SolidComponent"].collection[rr + 1 :]
            ):
                if (
                    abs(interf_flag.at[
                            s_comp_r.identifier, s_comp_c.identifier
                        ]
                    ) == 1
                ):
                    # There is contact between s_comp_r and s_comp_c 

                    self.dict_interf_peri["sol_sol"] = self.__assign_contact_perimeter_not_fluid_only(
                        s_comp_r.identifier,
                        s_comp_c.identifier,
                        "sol_sol",
                    )

                    self.dict_interf_peri["sol_sol"][
                        f"{s_comp_r.identifier}_{s_comp_c.identifier}"
                    ] = self.dict_df_coupling["contact_perimeter"].at[
                        s_comp_r.identifier, s_comp_c.identifier
                    ]
                    # Interface identification (cdp, 09/2020)
                    dict_topology_dummy_sol[s_comp_r.identifier][
                        s_comp_c.identifier
                    ] = f"{s_comp_r.identifier}_{s_comp_c.identifier}"
                    # Call method Chan_sol_interfaces (cdp, 09/2020)
                    [
                        dict_s_comps_contact,
                        list_linked_solids,
                    ] = self.chan_sol_interfaces(
                        s_comp_r, s_comp_c, dict_s_comps_contact, list_linked_solids
                    )
                # end if abs(interf_flag.iat[rr, cc]) == 1: (cdp, 09/2020)
            # end for cc (cdp, 09/2020)
            # Call method Update_interface_dictionary to update dictionaries \
            # (cdp, 09/2020)
            [
                dict_topology_dummy_sol,
                dict_s_comps_contact,
            ] = self.update_interface_dictionary(
                s_comp_r,
                dict_topology_dummy_sol,
                dict_s_comps_contact,
                list_linked_solids,
            )
            if (
                abs(interf_flag.at[environment.KIND,s_comp_r.identifier]) == 1
            ):
                if (
                    s_comp_r.inputs["Jacket_kind"] == "outer_insulation"
                    or s_comp_r.inputs["Jacket_kind"] == "whole_enclosure"
                ):
                    # There is an interface between environment and s_comp_r.
                    self.dict_interf_peri["env_sol"] = self.__assign_contact_perimeter_not_fluid_only(
                        environment.KIND,
                        s_comp_r.identifier,
                        "env_sol",
                    )
                else:
                    # Raise error
                    raise os.error(
                        f"JacketComponent of kind {s_comp_r.inputs['Jacket_kind']} can not have and interface with the environment.\n"
                    )
                # End if s_comp_r.inputs["Jacket_kind"]
        # end for rr (cdp, 09/2020)
        self.dict_topology.update(sol_sol=dict_topology_dummy_sol)

    def __assign_contact_perimeter_fluid_comps(
        self:Self,
        comp1_id:str,
        comp2_id:str,
        )-> tuple:
        """Private method that evaluates and assigns contact perimeters for interfaces between fluid components, distinguiscing between closed and open contact perimeters according to the value of the open perimeter fraction. Contact perimeter can be costant or variable; in the latter case it is evaluated by interpolation in the conductor spatial discretization along z direction (the axis of the conductor) starting from values loaded from auxiliary input file variable_contact_perimeter.xlsx.

        Args:
            self (Self): conductor object
            comp1_id (str): identifier of the first fluid component (row)
            comp2_id (str): identifier of the first fluid component (row)

        Returns:
            tuple: collection dictionay of numpy arrays with the value of the close and open contact perimeter of the interface; key nodal has the contact perimeter on nodal points, key Gauss has the contact perimeter on Gauss points.
        """

        # Aliases
        interf_peri_open = self.dict_interf_peri["ch_ch"]["Open"]
        interf_peri_close = self.dict_interf_peri["ch_ch"]["Close"]
        interf_flag = self.dict_df_coupling["contact_perimeter_flag"].at[
            comp1_id,comp2_id
        ]
        open_fraction = self.dict_df_coupling["open_perimeter_fract"].at[
            comp1_id, comp2_id
        ]

        if interf_flag == CONSTANT_CONTACT_PERIMETER:
            
            # Alias for constant contact perimeter.
            contact_perimeter = self.dict_df_coupling["contact_perimeter"].at[
            comp1_id, comp2_id
        ]

            # Assign constant contact perimeter value.
            # To be refactored!

            # Open contact perimeter in nodal points.
            interf_peri_open["nodal"][f"{comp1_id}_{comp2_id}"] = (
                contact_perimeter * open_fraction
            ) * np.ones(self.grid_features["N_nod"])
            # Open contact perimeter in Gauss points.
            interf_peri_open["Gauss"][f"{comp1_id}_{comp2_id}"] = (
                contact_perimeter
                * open_fraction
            ) * np.ones(self.grid_input["NELEMS"])

            # Close contact perimeter in nodal points.
            interf_peri_close["nodal"][f"{comp1_id}_{comp2_id}"] = (
                contact_perimeter * (1.0 - open_fraction)
                ) * np.ones(self.grid_features["N_nod"])
            # Open contact perimeter in Gauss points.
            interf_peri_close["Gauss"][f"{comp1_id}_{comp2_id}"] = (
                contact_perimeter * (1.0 - open_fraction)
                ) * np.ones(self.grid_input["NELEMS"])
        
        elif interf_flag == VARIABLE_CONTACT_PERIMETER:
            
            # Alias for variable contact perimeter: convert padas series into a 
            # numpy array.
            contact_perimeter = self.dict_df_variable_contact_perimeter[comp1_id].loc[:,comp2_id].to_numpy(dtype=float)
            
            # Assign variable contact perimeter value.
            # To be refactored!

            # Interpolation points.
            space_points = self.dict_df_variable_contact_perimeter[comp1_id].iloc[:,0].to_numpy(dtype=float)

            # Open contact perimeter.
            open_contact_perimeter = contact_perimeter * open_fraction
            # Build interpolator.
            open_interpolator = interpolate.interp1d(
                space_points,
                open_contact_perimeter,
                bounds_error=False,
                fill_value=open_contact_perimeter[-1],
                kind='linear',
            )
            # Do interpolation: nodal discretization points.
            interf_peri_open["nodal"][f"{comp1_id}_{comp2_id}"] = open_interpolator(self.grid_features["zcoord"])
            # Do interpolation: Gauss discretization points.
            interf_peri_open["Gauss"][f"{comp1_id}_{comp2_id}"] = open_interpolator(self.grid_features["zcoord_gauss"])

            # Close contact perimeter.
            close_contact_perimeter = contact_perimeter * (1.0 - open_fraction)
            # Build interpolator.
            close_interpolator = interpolate.interp1d(
                space_points,
                close_contact_perimeter,
                bounds_error=False,
                fill_value=close_contact_perimeter[-1],
                kind='linear',
            )
            # Do interpolation: nodal discretization points.
            interf_peri_close["nodal"][f"{comp1_id}_{comp2_id}"] = close_interpolator(self.grid_features["zcoord"])
            # Do interpolation: Gauss discretization points.
            interf_peri_close["Gauss"][f"{comp1_id}_{comp2_id}"] = close_interpolator(self.grid_features["zcoord_gauss"])
        
        return interf_peri_close, interf_peri_open

    def __assign_contact_perimeter_not_fluid_only(
        self:Self,
        comp1_id:str,
        comp2_id:str,
        interf_kind:str,
        )->dict:
        """Private method that evaluates and assigns contact perimeters for interfaces between fluid components and solid components (ch_sol), between solid components (sol_sol) and between environment and solid components (env_sol). Contact perimeter can be costant or variable; in the latter case it is evaluated by interpolation in the conductor spatial discretization along z direction (the axis of the conductor) starting from values loaded from auxiliary input file variable_contact_perimeter.xlsx.

        Args:
            self (Self): conductor object
            comp1_id (str): identifier of the first fluid component (row)
            comp2_id (str): identifier of the first fluid component (row)
            interf_kind (str): key to access nested dictionaries in attribute self.dict_interf_peri. Possible values: ch_sol, sol_sol and env_sol.

        Returns:
            dict: collection of numpy arrays with the value of contact perimeter of the interface; key nodal has the contact perimeter on nodal points, key Gauss has the contact perimeter on Gauss points.
        """

        # Aliases
        interf_peri = self.dict_interf_peri[interf_kind]
        interf_flag = self.dict_df_coupling["contact_perimeter_flag"].at[
            comp1_id,comp2_id
        ]

        if interf_flag == CONSTANT_CONTACT_PERIMETER:
            
            # Alias for constant contact perimeter.
            contact_perimeter = self.dict_df_coupling["contact_perimeter"].at[
            comp1_id, comp2_id
        ]

            # Assign constant contact perimeter value.
            # To be refactored!

            # Open contact perimeter in nodal points.
            interf_peri["nodal"][f"{comp1_id}_{comp2_id}"] = (
                contact_perimeter * np.ones(self.grid_features["N_nod"])
            )
            # Open contact perimeter in Gauss points.
            interf_peri["Gauss"][f"{comp1_id}_{comp2_id}"] = (
                contact_perimeter * np.ones(self.grid_input["NELEMS"])
            )
        
        elif interf_flag == VARIABLE_CONTACT_PERIMETER:
            
            # Alias for variable contact perimeter: convert padas series into a 
            # numpy array.
            contact_perimeter = self.dict_df_variable_contact_perimeter[comp1_id].loc[:,comp2_id].to_numpy(dtype=float)
            
            # Assign variable contact perimeter value.
            # To be refactored!

            # Interpolation points.
            space_points = self.dict_df_variable_contact_perimeter[comp1_id].iloc[:,0].to_numpy(dtype=float)

            # Build interpolator.
            interpolator = interpolate.interp1d(
                space_points,
                contact_perimeter,
                bounds_error=False,
                fill_value=contact_perimeter[-1],
                kind='linear',
            )
            # Do interpolation: nodal discretization points.
            interf_peri["nodal"][f"{comp1_id}_{comp2_id}"] = interpolator(
                self.grid_features["zcoord"]
            )
            # Do interpolation: Gauss discretization points.
            interf_peri["Gauss"][f"{comp1_id}_{comp2_id}"] = interpolator(
                self.grid_features["zcoord_gauss"]
            )
        
        return interf_peri

    def chan_sol_interfaces(
        self, comp_r, comp_c, dict_comp_interface, list_linked_comp
    ):

        """
        Method that evaluates interfaces between channels and solid components or between solids, and list them in a list of objects to be assigned to dict_topology. (cdp, 09/2020)
        """

        if dict_comp_interface.get(comp_r.identifier) == None:
            # No key called comp_r.identifier in dictionary dict_comp_interface \
            # (cdp, 09/2020)
            dict_comp_interface[comp_r.identifier] = list()
            # In this case necessarily we store both comp_r and comp_c \
            # (cdp, 09/2020)
            list_linked_comp.append(comp_r)
            list_linked_comp.append(comp_c)
        else:  # key comp_r.identifier already exist in dict_comp_interface
            # In this case store necessarily only comp_c (cdp, 09/2020)
            list_linked_comp.append(comp_c)
        # end if dict_comp_interface.get(comp_r.identifier) (cdp, 09/2020)
        return [dict_comp_interface, list_linked_comp]

    # end method Chan_sol_interfaces (cdp, 09/2020)

    def find_standalone_channels(self):

        """
        Method that searchs for possible isolated (not in hydraulic parallel) channels: search is on each channel in order to not miss anything (cdp, 09/2020)
        """

        # crate dictionary used to understand if channel is or not a stand alone one
        check_found = dict()

        ii = -1
        while ii < self.inventory["FluidComponent"].number - 1:
            ii = ii + 1
            fluid_comp = self.inventory["FluidComponent"].collection[ii]
            # loop on reference channels (cdp, 09/2020)
            check_found[fluid_comp.identifier] = dict(
                Hydraulic_parallel=False, Thermal_contact=False
            )
            for fluid_comp_ref in list(
                self.dict_topology["ch_ch"]["Hydraulic_parallel"].keys()
            ):
                # Search in Hydraulic parallel groups (cdp, 09/2020)
                if check_found[fluid_comp.identifier]["Hydraulic_parallel"] == False:
                    if (
                        fluid_comp
                        in self.dict_topology["ch_ch"]["Hydraulic_parallel"][
                            fluid_comp_ref
                        ]["Group"]
                    ):
                        # channel fluid_comp constitutes a group of channels in hydraulic \
                        # parallel thus it can not be a stand alone channel (cdp, 09/2020)
                        # Update dictionart check_found (cdp, 09/2020)
                        check_found[fluid_comp.identifier].update(
                            Hydraulic_parallel=True
                        )
            if check_found[fluid_comp.identifier]["Hydraulic_parallel"] == False:
                # Channel fluid_comp is not inside Hydraulic parallel groups (cdp, 09/2020)
                for fluid_comp_ref in list(
                    self.dict_topology["ch_ch"]["Thermal_contact"].keys()
                ):
                    # Search in Hydraulic parallel groups (cdp, 09/2020)
                    if check_found[fluid_comp.identifier]["Thermal_contact"] == False:
                        if (
                            fluid_comp
                            in self.dict_topology["ch_ch"]["Thermal_contact"][
                                fluid_comp_ref
                            ]["Group"]
                        ):
                            # channel fluid_comp constitutes a thermal contact thus it can not be \
                            # a stand alone channel (cdp, 09/2020)
                            # Update dictionart check_found (cdp, 09/2020)
                            check_found[fluid_comp.identifier].update(
                                Thermal_contact=True
                            )
            if (
                check_found[fluid_comp.identifier]["Hydraulic_parallel"] == False
                and check_found[fluid_comp.identifier]["Thermal_contact"] == False
            ):
                # fluid_comp is a stand alone channel since it does not belong to a group of \
                # channels in hydraulic parallel and it does not constitute a thermal \
                # contact (cdp, 09/2020)
                self.dict_topology["Standalone_channels"].append(fluid_comp)

    # 	N_channel_no_par = len(self.dict_topology["Standalone_channels"])
    # 	if N_channel_no_par == 0:
    # 		print("There are no isolated channels\n")
    # 	elif N_channel_no_par > 0 and N_channel_no_par < \
    # 			 self.inventory["FluidComponent"].number:
    # 		if N_channel_no_par == 1:
    # 			print(f"""There is {N_channel_no_par} channel that is not in hydraulic parallel: {self.dict_topology["Standalone_channels"][0].identifier}\n""")
    # 		else:
    # 			print(f"""There are {N_channel_no_par} channels that are not in hydraulic parallel: {self.dict_topology["Standalone_channels"][:].identifier}\n""")
    # 	elif N_channel_no_par == \
    # 			 self.inventory["FluidComponent"].number:
    # 		print("All channels are isolated\n")
    # 	else:
    # 		print(f"Something does not work\n")
    # end method Find_Standalone_channels (cdp, 09/2020)

    def update_interface_dictionary(
        self, comp, dict_topology_dummy, dict_contacts, list_contacts
    ):

        dict_contacts[comp.identifier] = list_contacts
        dict_topology_dummy[comp.identifier].update(Group=list_contacts)
        dict_topology_dummy[comp.identifier].update(Number=len(list_contacts))
        if dict_topology_dummy[comp.identifier]["Number"] == 0:
            # Removed empty keys from dictionaries (cdp, 09/2020)
            dict_topology_dummy.pop(comp.identifier)
            dict_contacts.pop(comp.identifier)
        return [dict_topology_dummy, dict_contacts]

    # end method Update_interface_dictionary (cdp, 09/2020)

    def get_hydraulic_parallel(self):

        """
        Method that interprets the information in table self.dict_df_coupling["open_perimeter_fract"] understanding if there are channels in hydraulic parallel and how the are organized into groups. The method returns a dictionary with:
        1) the identifier of the reference channel of each group
        2) a list of all the channels that belongs to a group
        3) for each group the IDs of all the linked channels organized into lists
        (cdp, 09/2020)
        """

        full_ind = dict()
        # get row and column index of non zero matrix elements (cdp, 09/2020)
        [full_ind["row"], full_ind["col"]] = np.nonzero(
            self.dict_df_coupling["open_perimeter_fract"].iloc[1:, 1:].to_numpy()
        )
        # USEFUL QUANTITIES AND VARIABLES (cdp, 09/2020)
        # array of the not considered array (cdp, 09/2020)
        already = dict(no=np.unique(np.union1d(full_ind["row"], full_ind["col"])))
        # list of the already considered array (cdp, 09/2020)
        already["yes"] = -1 * np.ones(already["no"].shape, dtype=int)
        # index to used to update key "yes" (cdp, 09/2020)
        already["ii_yes"] = 0
        # Define dictionary dict_topology. This is different from \
        # self.dict_topology which is a class Conductor attribute. A the end of \
        # this method self.dict_topology will be updated with the values in \
        # dict_topology (cdp, 09/2020)
        dict_topology = dict()
        # Define dictionary check to be sure that for each channel both searchs \
        # are performed (cdp, 09/2020)
        check = dict()

        Total_connections = len(full_ind["row"])
        total_connections_counter = 0
        group_counter = 0
        # loop until all channel connections are realized (cdp, 09/2020)
        while total_connections_counter < Total_connections:
            # get the reference channel: it is the one characterized by the minimum \
            # index value in array already["no"] (cdp, 09/2020)
            fluid_comp_ref_row_ind = min(already["no"])
            fluid_comp_ref = self.inventory["FluidComponent"].collection[
                fluid_comp_ref_row_ind
            ]
            # update dictionary already (cdp, 09/2020)
            already["no"] = np.delete(already["no"], 0, 0)
            if fluid_comp_ref_row_ind not in already["yes"]:
                already["yes"][already["ii_yes"]] = fluid_comp_ref_row_ind
                already.update(ii_yes=already["ii_yes"] + 1)
            # Construct check dictionary (cdp, 09/2020)
            check[fluid_comp_ref.identifier] = dict()
            # Get minimum and maximum index of array full_ind["row"] that correspond \
            # to the reference channel (cdp, 09/2020)
            boundary = dict(
                fluid_comp_ref_lower=min(
                    np.nonzero(full_ind["row"] == fluid_comp_ref_row_ind)[0]
                ),
                fluid_comp_ref_upper=max(
                    np.nonzero(full_ind["row"] == fluid_comp_ref_row_ind)[0]
                ),
            )
            # get all the channels that are directly in contact with reference \
            # channel (cdp, 09/2020)
            ind_direct = full_ind["col"][
                boundary["fluid_comp_ref_lower"] : boundary["fluid_comp_ref_upper"] + 1
            ]
            # Update dictionary dict_topology
            dict_topology[fluid_comp_ref.identifier] = dict(
                Ref_channel=fluid_comp_ref.identifier, Group=list(), Number=0
            )
            dict_topology[fluid_comp_ref.identifier]["Group"].append(fluid_comp_ref)
            for ch_index in ind_direct:
                # get channel (cdp, 09/2020)
                fluid_comp = self.inventory["FluidComponent"].collection[ch_index]
                # Construct check dictionary (cdp, 09/2020)
                check[fluid_comp_ref.identifier][fluid_comp.identifier] = dict(
                    row=False, col=False
                )
                # find the index in array already["no"] of the element that must be \
                # deleted (cdp, 09/2020)
                i_del = np.nonzero(already["no"] == ch_index)[0]
                # update total_connections_counter (cdp, 09/2020)
                total_connections_counter = total_connections_counter + 1
                # update dictionary already (cdp, 09/2020)
                already["no"] = np.delete(already["no"], i_del, 0)
                if ch_index not in already["yes"]:
                    already["yes"][already["ii_yes"]] = ch_index
                    already.update(ii_yes=already["ii_yes"] + 1)
                dict_topology[fluid_comp_ref.identifier][fluid_comp.identifier] = [
                    f"{fluid_comp_ref.identifier}_{fluid_comp.identifier}"
                ]
                dict_topology[fluid_comp_ref.identifier]["Group"].append(fluid_comp)
            # end for ii (cdp, 09/2020)
            # for each channel that is in direct contact with the reference one, \
            # search if it is in contact with other channels, constituting an \
            # indirect contact with the reference channel. This is done in a \
            # different loop because total_connections_counter must be fully updated \
            # (cdp, 09/2020)
            for ch_index in ind_direct:
                # get channel (cdp, 09/2020)
                fluid_comp = self.inventory["FluidComponent"].collection[ch_index]
                # Initialize key value "variable_lower" of dictionary boundary. This \
                # parameter is used to look only in the region of not directly \
                # connected channels and is updated to consider only the data below \
                # this index value. Initialization must be done at each iteration in \
                # order to not miss some index during the search. (cdp, 09/2020)
                boundary.update(variable_lower=boundary["fluid_comp_ref_upper"] + 1)
                if (
                    check[fluid_comp_ref.identifier][fluid_comp.identifier]["col"]
                    == False
                ):
                    # The search in array full_ind["col"] is not performed yet \
                    # (cdp, 09/2020)
                    total_connections_counter = self.search_on_ind_col(
                        ch_index,
                        full_ind,
                        dict_topology,
                        fluid_comp_ref,
                        fluid_comp,
                        check,
                        already,
                        total_connections_counter,
                        boundary,
                    )
                if (
                    check[fluid_comp_ref.identifier][fluid_comp.identifier]["row"]
                    == False
                ):
                    # The search in array full_ind["row"] is not performed yet \
                    # (cdp, 09/2020)
                    total_connections_counter = self.search_on_ind_row(
                        ch_index,
                        full_ind,
                        dict_topology,
                        fluid_comp_ref,
                        fluid_comp,
                        check,
                        already,
                        total_connections_counter,
                        boundary,
                    )
            # end for (cdp, 09/2020)
            # Sort list Group by channel identifier (cdp, 09/2020)
            dict_topology[fluid_comp_ref.identifier].update(
                Group=sorted(
                    dict_topology[fluid_comp_ref.identifier]["Group"],
                    key=lambda ch: ch.identifier,
                )
            )
            # Get the number of channels that are in hydraulic parallel for each \
            # reference channel (cdp, 10/2020)
            dict_topology[fluid_comp_ref.identifier].update(
                Number=len(dict_topology[fluid_comp_ref.identifier]["Group"])
            )
            if total_connections_counter == Total_connections:
                if group_counter == 0:
                    # Only one group of channels in hydraulic parallel (cdp, 09/2020)
                    group_counter = group_counter + 1
                    print(
                        f"There is only {group_counter} group of channels in hydraulic parallel.\n"
                    )
                elif group_counter > 0:
                    # There are a least two groups of channels in hydraulic parallel \
                    # (cdp, 09/2020)
                    group_counter = group_counter + 1
                    print(
                        f"There are {group_counter} groups of channels in hydraulic parallel\n"
                    )
            elif total_connections_counter < Total_connections:
                # The number of groups of channels in hydraulic parallel is > 1.
                # Repeat the above procedure. Keep in mind that the reference channel \
                # is evaluated as the one with the minimum identifier between the ones that \
                # are not connected yet (cdp, 09/2020)
                group_counter = group_counter + 1
            elif total_connections_counter > Total_connections:
                # Something odd occurs
                raise ValueError(
                    "ERROR! The counter of the total connections can not be larger than the number of total connections! Something odd occurs!\n"
                )
            # end if total_connections_counter (cdp, 09/2020)
        # end while (cdp, 09/2020)
        # Update key Hydraulic_parallel of dictionary dict_topology (cdp, 09/2020)
        self.dict_topology["ch_ch"].update(Hydraulic_parallel=dict_topology)

    # end method Get_hydraulic_parallel (cdp, 09/2020)

    def search_on_ind_col(
        self,
        ch_index,
        full_ind,
        dict_topology,
        fluid_comp_ref,
        fluid_comp_c,
        check,
        already,
        total_connections_counter,
        boundary,
    ):

        """
        method that search in array full_ind["col"] if there are other recall to channel fluid_comp_c (cdp, 09/2020)
        """

        # N.B ch_index is the channel that is in contact with the fluid_comp_ref or \
        # another channel (cdp, 09/2020)

        # The search in array full_ind["col"] will be performed so flag check \
        # [fluid_comp_ref.identifier][fluid_comp.identifier]["col"] can be set to True. (cdp, 09/2020)
        check[fluid_comp_ref.identifier][fluid_comp_c.identifier].update(col=True)
        # search for all the values that are equal to ch_index in array \
        # full_ind["col"], excluding the index of the direct contact region, and
        # store the corresponding indices (cdp, 09/2020)
        ind_found = (
            np.nonzero(full_ind["col"][boundary["variable_lower"] :] == ch_index)[0]
            + boundary["variable_lower"]
        )
        if len(ind_found) == 0:
            if (
                check[fluid_comp_ref.identifier][fluid_comp_c.identifier]["row"]
                == False
            ):
                # Search if there is some value equal to ch_index in the array \
                # full_ind["row"] calling method Search_on_ind_row (cdp, 09/2020)
                total_connections_counter = self.search_on_ind_row(
                    ch_index,
                    full_ind,
                    dict_topology,
                    fluid_comp_ref,
                    fluid_comp_c,
                    check,
                    already,
                    total_connections_counter,
                    boundary,
                )
            elif (
                check[fluid_comp_ref.identifier][fluid_comp_c.identifier]["row"] == True
            ):
                # no channel with smaller channel index is connected with ch_index \
                # (cdp, 09/2020)
                print(f"No other channels are connected to {fluid_comp_c.identifier}\n")
        elif len(ind_found) > 0:
            # there is at least one channel with smaller channel index that is \
            # connected with ch_index (cdp, 09/2020)
            # update key "variable_lower": the next search is done from this index \
            # value up to the end of arrays full_ind["row"] and full_ind["col"] \
            # (cdp, 09/2020)
            boundary.update(variable_lower=ind_found[0] + 1)
            # get the index of the linked channel(s) (col -> row) (cdp, 09/2020)
            ind_link = full_ind["row"][ind_found[0 : len(ind_found)]]
            # loop linked channels (cdp, 09/2020)
            jj = -1
            while jj < len(ind_link) - 1:
                jj = jj + 1
                ch_index = ind_link[jj]
                # get channel
                fluid_comp_r = self.inventory["FluidComponent"].collection[ch_index]
                if (
                    dict_topology[fluid_comp_ref.identifier].get(
                        fluid_comp_r.identifier
                    )
                    != None
                    and f"{fluid_comp_r.identifier}_{fluid_comp_c.identifier}"
                    in dict_topology[fluid_comp_ref.identifier][fluid_comp_r.identifier]
                ):
                    # Skip operations since the connection is already preformed \
                    # (cdp, 09/2020)
                    jj = jj + 1
                else:
                    if (
                        check[fluid_comp_ref.identifier].get(fluid_comp_r.identifier)
                        == None
                    ):
                        # Update dictionary check: add key fluid_comp_r.identifier (cdp, 09/2020)
                        check[fluid_comp_ref.identifier][
                            fluid_comp_r.identifier
                        ] = dict(row=False, col=False)
                    # construct channels link (cdp, 09/2020)
                    if (
                        dict_topology[fluid_comp_ref.identifier].get(
                            fluid_comp_r.identifier
                        )
                        == None
                    ):
                        # key fluid_comp_r.identifier does not exist, so it is added to the dictionary \
                        # and a list is created (cdp, 09/2020)
                        dict_topology[fluid_comp_ref.identifier][
                            fluid_comp_r.identifier
                        ] = [f"{fluid_comp_r.identifier}_{fluid_comp_c.identifier}"]
                        print(
                            dict_topology[fluid_comp_ref.identifier][
                                fluid_comp_r.identifier
                            ][-1]
                            + "\n"
                        )
                        # update total_connections_counter (cdp, 09/2020)
                        total_connections_counter = total_connections_counter + 1
                    else:
                        # key fluid_comp_r.identifier exists (cdp, 09/2020)
                        if (
                            f"{fluid_comp_r.identifier}_{fluid_comp_c.identifier}"
                            not in dict_topology[fluid_comp_ref.identifier][
                                fluid_comp_r.identifier
                            ]
                        ):
                            # List is updated only if the contact identifier is not already in the \
                            # list (cdp, 09/2020)
                            dict_topology[fluid_comp_ref.identifier][
                                fluid_comp_r.identifier
                            ].append(
                                f"{fluid_comp_r.identifier}_{fluid_comp_c.identifier}"
                            )
                            print(
                                dict_topology[fluid_comp_ref.identifier][
                                    fluid_comp_r.identifier
                                ][-1]
                                + "\n"
                            )
                            # update total_connections_counter (cdp, 09/2020)
                            total_connections_counter = total_connections_counter + 1
                    # end if dict_topology[fluid_comp_ref.identifier].get(fluid_comp_r.identifier) == None \
                    # (cdp, 09/2020)
                    # find the index in array already["no"] of the element that must be \
                    # deleted (cdp, 09/2020)
                    i_del = np.nonzero(already["no"] == ch_index)[0]
                    # update dictionary already (cdp, 09/2020)
                    already["no"] = np.delete(already["no"], i_del, 0)
                    if ch_index not in already["yes"]:
                        already["yes"][already["ii_yes"]] = ch_index
                        already.update(ii_yes=already["ii_yes"] + 1)
                    if (
                        fluid_comp_r
                        not in dict_topology[fluid_comp_ref.identifier]["Group"]
                    ):
                        # Add channel fluid_comp_r to list Group (cdp, 09/2020)
                        dict_topology[fluid_comp_ref.identifier]["Group"].append(
                            fluid_comp_r
                        )
                    # call method Search_on_ind_row with to search if channel fluid_comp_r \
                    # is linked to other channels (cdp, 09/2020)
                    total_connections_counter = self.search_on_ind_row(
                        ch_index,
                        full_ind,
                        dict_topology,
                        fluid_comp_ref,
                        fluid_comp_r,
                        check,
                        already,
                        total_connections_counter,
                        boundary,
                    )
                # end if dict_topology[fluid_comp_ref.identifier].get(fluid_comp_r.identifier) != None and \
                # f"{fluid_comp_r.identifier}_{fluid_comp_c.identifier}" in dict_topology[fluid_comp_ref.identifier][fluid_comp_r.identifier] \
                # (cdp, 09/2020)
            # end while (cdp, 09/2020)
        # end if len(ind_found) (cdp, 09/2020)
        return total_connections_counter

    # end method Search_on_ind_col (cdp, 09/2020)

    def search_on_ind_row(
        self,
        ch_index,
        full_ind,
        dict_topology,
        fluid_comp_ref,
        fluid_comp_r,
        check,
        already,
        total_connections_counter,
        boundary,
    ):

        """
        Method that search in array full_ind["row"] if there are other recall to channel fluid_comp_c (cdp, 09/2020)
        """

        # The search in array full_ind["row"] will be performed so flag check \
        # [fluid_comp_ref.identifier][fluid_comp.identifier]["row"] can be set to True. (cdp, 09/2020)
        check[fluid_comp_ref.identifier][fluid_comp_r.identifier].update(row=True)

        # search for all the values that are equal to ch_index in array \
        # full_ind["row"], excluding the index of the direct contact region, and \
        # store the corresponding indices (cdp, 09/2020)
        ind_found = (
            np.nonzero(full_ind["row"][boundary["variable_lower"] :] == ch_index)[0]
            + boundary["variable_lower"]
        )
        if len(ind_found) == 0:
            if (
                check[fluid_comp_ref.identifier][fluid_comp_r.identifier]["col"]
                == False
            ):
                # Search if there is some value equal to ch_index in the array \
                # full_ind["col"] calling method Search_on_ind_col (cdp, 09/2020)
                total_connections_counter = self.search_on_ind_col(
                    ch_index,
                    full_ind,
                    dict_topology,
                    fluid_comp_ref,
                    fluid_comp_r,
                    check,
                    already,
                    total_connections_counter,
                    boundary,
                )
            elif (
                check[fluid_comp_ref.identifier][fluid_comp_r.identifier]["col"] == True
            ):
                # no channel with larger channel index is connected with ch_index \
                # (cdp, 09/2020)
                print(f"No other channels are connected to {fluid_comp_r.identifier}\n")
        elif len(ind_found) > 0:
            # there is at least one channel with larger channel index that is \
            # connected with ch_index (cdp, 09/2020)
            # update key "variable_lower": the next search is done from this index \
            # value up to the end of arrays full_ind["row"] and full_ind["col"] \
            # (cdp, 09/2020)
            boundary.update(variable_lower=ind_found[0] + 1)
            # get the index of the linked channel(s) (row -> col) (cdp, 09/2020)
            ind_link = full_ind["col"][ind_found[0 : len(ind_found)]]
            # loop linked channels (cdp, 09/2020)
            jj = -1
            while jj < len(ind_link) - 1:
                jj = jj + 1
                ch_index = ind_link[jj]
                # get channel (cdp, 09/2020)
                fluid_comp_c = self.inventory["FluidComponent"].collection[ch_index]
                if (
                    dict_topology[fluid_comp_ref.identifier].get(
                        fluid_comp_c.identifier
                    )
                    != None
                    and f"{fluid_comp_c.identifier}_{fluid_comp_r.identifier}"
                    in dict_topology[fluid_comp_ref.identifier][fluid_comp_c.identifier]
                ):
                    # Skip operations since the connection is already preformed \
                    # (cdp, 09/2020)
                    jj = jj + 1
                else:
                    if (
                        check[fluid_comp_ref.identifier].get(fluid_comp_c.identifier)
                        == None
                    ):
                        # Update dictionary check: add key fluid_comp_c.identifier (cdp, 09/2020)
                        check[fluid_comp_ref.identifier][
                            fluid_comp_c.identifier
                        ] = dict(row=False, col=False)
                    # construct channels link (cdp, 09/2020)
                    if (
                        dict_topology[fluid_comp_ref.identifier].get(
                            fluid_comp_r.identifier
                        )
                        == None
                    ):
                        # key fluid_comp_c.identifier does not exist, so it is added to the dictionary \
                        # and a list is created (cdp, 09/2020)
                        dict_topology[fluid_comp_ref.identifier][
                            fluid_comp_r.identifier
                        ] = [f"{fluid_comp_r.identifier}_{fluid_comp_c.identifier}"]
                        print(
                            dict_topology[fluid_comp_ref.identifier][
                                fluid_comp_r.identifier
                            ][-1]
                            + "\n"
                        )
                        # update total_connections_counter (cdp, 09/2020)
                        total_connections_counter = total_connections_counter + 1
                    else:
                        # key fluid_comp_c.identifier exists (cdp, 09/2020)
                        if (
                            f"{fluid_comp_r.identifier}_{fluid_comp_c.identifier}"
                            not in dict_topology[fluid_comp_ref.identifier][
                                fluid_comp_r.identifier
                            ]
                        ):
                            # List is updated only if the contact identifier is not already in the \
                            # list (cdp, 09/2020)
                            dict_topology[fluid_comp_ref.identifier][
                                fluid_comp_r.identifier
                            ].append(
                                f"{fluid_comp_r.identifier}_{fluid_comp_c.identifier}"
                            )
                            print(
                                dict_topology[fluid_comp_ref.identifier][
                                    fluid_comp_r.identifier
                                ][-1]
                                + "\n"
                            )
                            # update total_connections_counter (cdp, 09/2020)
                            total_connections_counter = total_connections_counter + 1
                    # end if dict_topology[fluid_comp_ref.identifier].get(fluid_comp_c.identifier) == None \
                    # (cdp, 09/2020)
                    # find the index in array already["no"] of the element that must be \
                    # deleted (cdp, 09/2020)
                    i_del = np.nonzero(already["no"] == ch_index)[0]
                    # update dictionary already (cdp, 09/2020)
                    already["no"] = np.delete(already["no"], i_del, 0)
                    if ch_index not in already["yes"]:
                        already["yes"][already["ii_yes"]] = ch_index
                        already.update(ii_yes=already["ii_yes"] + 1)
                    if (
                        fluid_comp_c
                        not in dict_topology[fluid_comp_ref.identifier]["Group"]
                    ):
                        # Add channel fluid_comp_c to list Group (cdp, 09/2020)
                        dict_topology[fluid_comp_ref.identifier]["Group"].append(
                            fluid_comp_c
                        )
                    # call method Search_on_ind_col to search if channel fluid_comp_c is \
                    # linked to other channels (cdp, 09/2020)
                    total_connections_counter = self.search_on_ind_col(
                        ch_index,
                        full_ind,
                        dict_topology,
                        fluid_comp_ref,
                        fluid_comp_c,
                        check,
                        already,
                        total_connections_counter,
                        boundary,
                    )
                # end if dict_topology[fluid_comp_ref.identifier].get(fluid_comp_c.identifier) != None and \
                # f"{fluid_comp_c.identifier}_{fluid_comp_r.identifier}" in dict_topology[fluid_comp_ref.identifier][fluid_comp_c.identifier] \
                # (cdp, 09/2020)
            # end while (cdp, 09/2020)
        # end if len(ind_found) (cdp, 09/2020)
        return total_connections_counter

    # end method Search_on_ind_row (cdp, 09/2020)

    def get_thermal_contact_channels(
        self, rr, cc, fluid_comp_r, fluid_comp_c, flag_found
    ):

        """
        Method that recognize if there are some channels that are only in thermal contact with other channels. If one or both of the two considered channels belongs also to two different groups of channels in hydraulic parallel, they are not included in the list called Group, however the thermal contact is indicated in a suitable key-value pair. This is because channels that have both the properties of being in thermal contact and in hydraulic parallel should be threated considering the latter, while flow initialization is performed. (cdp, 09/2020)
        """

        if (
            self.dict_topology["ch_ch"]["Thermal_contact"].get(fluid_comp_r.identifier)
            == None
        ):
            # Update dictionary self.dict_topology["ch_ch"]["Thermal_contact"] \
            # (cdp, 09/2020)
            # key fluid_comp_r.identifier does not already exist (cdp, 09/2020)
            self.dict_topology["ch_ch"]["Thermal_contact"][
                fluid_comp_r.identifier
            ] = dict(Group=list(), Number=0, Actual_number=0)
        if (
            self.dict_df_coupling["open_perimeter_fract"].at[
                fluid_comp_r.identifier, fluid_comp_c.identifier
            ]
            == 0.0
        ):
            # There is only thermal contact between fluid_comp_r and fluid_comp_c \
            # (cdp, 09/2020)
            # Update dictionary self.dict_topology["ch_ch"]["Thermal_contact"] \
            # (cdp, 09/2020)
            self.dict_topology["ch_ch"]["Thermal_contact"][fluid_comp_r.identifier][
                fluid_comp_c.identifier
            ] = f"{fluid_comp_r.identifier}_{fluid_comp_c.identifier}"
            if len(list(self.dict_topology["ch_ch"]["Hydraulic_parallel"].keys())) > 0:
                # There is at least one group of channels in hydraulic parallel \
                # (cdp, 09/2020)
                if flag_found.get(fluid_comp_r.identifier) == None:
                    # initialize key fluid_comp_r.identifier of dictionary flag_found to False \
                    # only once (cdp, 09/2020)
                    flag_found[fluid_comp_r.identifier] = False
                # initialize key fluid_comp_c.identifier of dictionary flag_found to False for \
                # each value of cc (cdp, 09/2020)
                flag_found[fluid_comp_c.identifier] = False
                # Following lines check if one of this two channels belongs to a
                # group of channels in hydraulic parallel. In this case only the
                # one that does not belong to the group is added to the list of
                # channels with only thermal contact. (cdp, 09/2020)
                for fluid_comp_ref in list(
                    self.dict_topology["ch_ch"]["Hydraulic_parallel"].keys()
                ):
                    # Search if fluid_comp_r and fluid_comp_c are already inserted into two
                    # different groups of channels in parallel. (cdp, 09/2020)
                    if (
                        flag_found[fluid_comp_r.identifier] == False
                        and fluid_comp_r
                        in self.dict_topology["ch_ch"]["Hydraulic_parallel"][
                            fluid_comp_ref
                        ]["Group"]
                    ):
                        # fluid_comp_r is in found in a group of channels in hydraulic \
                        # parallel (cdp, 09/2020)
                        flag_found[fluid_comp_r.identifier] = True
                    # end if cc == rr + 1 (cdp, 09/2020)
                    if (
                        flag_found[fluid_comp_c.identifier] == False
                        and fluid_comp_c
                        in self.dict_topology["ch_ch"]["Hydraulic_parallel"][
                            fluid_comp_ref
                        ]["Group"]
                    ):
                        # fluid_comp_c is in found in a group of channels in hydraulic \
                        # parallel (cdp, 09/2020)
                        flag_found[fluid_comp_c.identifier] = True
                # end for fluid_comp_ref (cdp, 09/2020)
                if flag_found[fluid_comp_r.identifier] == False:
                    # fluid_comp_r is not in hydraulic parallel with other channels \
                    # (cdp, 09/2020)
                    self.dict_topology["ch_ch"]["Thermal_contact"][
                        fluid_comp_r.identifier
                    ]["Group"].append(fluid_comp_r)
                if flag_found[fluid_comp_c.identifier] == False:
                    # fluid_comp_c is not in hydraulic parallel with other channels
                    # (cdp, 09/2020)
                    self.dict_topology["ch_ch"]["Thermal_contact"][
                        fluid_comp_r.identifier
                    ]["Group"].append(fluid_comp_c)
            else:
                # there are no groups of channels in hydraulic parallel \
                # (cdp, 09/2020)
                if cc == rr + 1:
                    # Add fluid_comp_r to the list Group only once (cdp, 09/2020)
                    self.dict_topology["ch_ch"]["Thermal_contact"][
                        fluid_comp_r.identifier
                    ]["Group"].append(fluid_comp_r)
                # Add channel fluid_comp_c to the list Group (cdp, 09/2020)
                self.dict_topology["ch_ch"]["Thermal_contact"][fluid_comp_r.identifier][
                    "Group"
                ].append(fluid_comp_c)
            # end if len(list(self.dict_topology["ch_ch"]\
            # ["Hydraulic_parallel"].keys())) > 0 (cdp, 09/2020)
        # end if self.dict_df_coupling["open_perimeter_fract"].iat[rr, cc] == 0.0 (cdp, 09/2020)
        return flag_found

    # end method Get_thermal_contact_channels (cdp, 09/2020)

    def __get_conductor_interfaces(self,environment:object):
        """Private method that identifies interfaces between conductor components, storing information in conductor attribute interface.

        Args:
            environment (object): object with all the info that characterize the environment.

        Raises:
            ValueError: if jacket component is not of kind outer_insulation or wall_enclosure.
        """

        # Aliases
        fluid_components = self.inventory["FluidComponent"].collection
        solid_components = self.inventory["SolidComponent"].collection
        interf_flag = self.dict_df_coupling["contact_perimeter_flag"]

        # Namedtuple constructor definition.
        Interface_collection = namedtuple(
            "Interface_collection",
            (
                "fluid_fluid",
                "fluid_solid",
                "solid_solid",
                "env_solid",
            )
        )

        # Namedtuple constructor definition.
        Interface = namedtuple(
            "Interface",
            (
                "interf_name",
                "comp_1",
                "comp_2",
            )
        )

        # Namedtuple initialization: each field is an empty list to be filled 
        # with interfaces.
        self.interface = Interface_collection(
            fluid_fluid=list(),
            fluid_solid=list(),
            solid_solid=list(),
            env_solid=list()
        )

        # Loop on FluidComponents.
        for f_comp_a_idx,f_comp_a in enumerate(fluid_components):
            # Loop on FluidComponents: identify fluid-fluid interfaces.
            for f_comp_b in fluid_components[f_comp_a_idx+1:]:
                # Check for interfaces.
                if (
                    abs(
                        interf_flag.at[f_comp_a.identifier,f_comp_b.identifier]
                    ) == 1
                ):
                    # Build fluid-fluid interface.
                    self.interface.fluid_fluid.append(
                        Interface(
                            interf_name=f"{f_comp_a.identifier}_{f_comp_b.identifier}",
                            comp_1=f_comp_a, # shallow copy: no waste of memory!
                            comp_2=f_comp_b, # shallow copy: no waste of memory!
                        )
                    )
            # Loop on SolidComponent: identify fluid-solid interfaces.
            for s_comp in solid_components:
                # Check for iterfaces.
                if (
                    abs(
                        interf_flag.at[f_comp_a.identifier,s_comp.identifier]
                    ) == 1
                ):
                    # Build fluid-solid interface.
                    self.interface.fluid_solid.append(
                        Interface(
                            interf_name=f"{f_comp_a.identifier}_{s_comp.identifier}",
                            comp_1=f_comp_a, # shallow copy: no waste of memory!
                            comp_2=s_comp, # shallow copy: no waste of memory!
                        )
                    )
        # Loop on SolidComponent.
        for s_comp_a_idx,s_comp_a in enumerate(solid_components):
            # Loop on SolidComponent: identify solid-solid interfaces.
            for s_comp_b in solid_components[s_comp_a_idx+1:]:
                # Check for interfaces.
                if (
                    abs(
                        interf_flag.at[s_comp_a.identifier,s_comp_b.identifier]
                    ) == 1
                ):
                    # Build solid-solid interface.
                    self.interface.solid_solid.append(
                        Interface(
                            interf_name=f"{s_comp_a.identifier}_{s_comp_b.identifier}",
                            comp_1=s_comp_a, # shallow copy: no waste of memory!
                            comp_2=s_comp_b, # shallow copy: no waste of memory!
                        )
                    )
            # Check for environment-solid interfaces.
            if (
                abs(
                    interf_flag.at[environment.KIND,s_comp_a.identifier]
                ) == 1
            ):  
                # Check on jacket kind.
                if (
                    s_comp_a.inputs["Jacket_kind"] == "outer_insulation"
                    or s_comp_a.inputs["Jacket_kind"] == "whole_enclosure"
                ):
                    # Build env-solid interface.
                    self.interface.env_solid.append(
                        Interface(
                            interf_name=f"{environment.KIND}_{s_comp_a.identifier}",
                            # shallow copy: no waste of memory!
                            comp_1=environment,
                            comp_2=s_comp_a, # shallow copy: no waste of memory!
                        )
                    )
                else:
                    # Raise error
                    raise ValueError(f"JacketComponent of kind {s_comp_a.inputs['Jacket_kind']} can not have and interface with the environment.\n"
                    )

    ############################################################################### this method initialize the Conductor on the base of the input parameters
    ##############################################################################

    def initialization(self, simulation, gui):

        time_simulation = simulation.simulation_time[-1]
        sim_name = simulation.transient_input["SIMULATION"]
        # Total number of equations for each conductor (cdp, 09/2020)
        self.dict_N_equation["Total"] = (
            self.dict_N_equation["NODOFS"] * self.grid_features["N_nod"]
        )
        # initialize conductor time values, it can be different for different \
        # conductors since the conductor time step can be different (cdp, 10/202)
        self.cond_time = [time_simulation]
        # Initialize conductor time step counter
        self.cond_num_step = 0

        """
    if (imsourcefun.eq.0) then
      premed(icond)=dble((preinl(icond)+preout(icond))/2.)
      temmed(icond)=dble((teminl(icond)+temout(icond))/2.)
    endif
      CALL GETsource  (NNODES,ncond) ### si trova in 
    ENTALINLET = hhe(premed(icond),temmed(icond))
    Questo per il momento non lo considero!!!!! (cdp)
    """

        gen_flow(self)
        # C*SET THE INITIAL VALUE OF THE FLOW VARIABLES (LINEAR P AND T)

        temp_ave = np.zeros(self.grid_features["N_nod"])
        self.enthalpy_balance = 0.0
        self.enthalpy_out = 0.0
        self.enthalpy_inl = 0.0
        # For each channel evaluate following fluid properties (array):velocity, \
        # pressure, temperature, density, Reynolds, Prandtl (cdp, 06/2020)
        # N.B. questo loop si potrebbe fare usando map.
        for fluid_comp in self.inventory["FluidComponent"].collection:
            # Compute pressure, temperature and velocity in nodal points according to the initial conditions
            fluid_comp.coolant._eval_nodal_pressure_temperature_velocity_initialization(
                self
            )
            # Build namedtuples fluid_comp.inl_idx and fluid_comp.inl_idx with 
            # the index used to assign the inlet and outlet BC.
            fluid_comp.build_th_bc_index(self)

        ## For each solid component evaluate temperature (cdp, 07/2020)
        ## If needed read only the sub matrix describing channel - solid objects \
        ## contact (cdp, 07/2020)
        ## nested loop on channel - solid objects (cpd 07/2020)
        # for cc in range(self.inventory["SolidComponent"].number):
        # 	s_comp = self.inventory["SolidComponent"].collection[cc]
        # 	weight = self.dict_df_coupling["contact_perimeter"][0:\
        # 					 self.inventory["FluidComponent"].number, \
        # 					 cc + self.inventory["FluidComponent"].number]
        # 	s_comp.dict_node_pt["temperature"] = \
        # 																np.zeros(self.grid_features["N_nod"])
        # 	if np.sum(weight) > 0:
        # 		# evaluate SolidComponent temperature as the weighted average on \
        # 		# conctat_perimeter with channels (cpd 07/2020)
        # 		for rr in range(self.inventory["FluidComponent"].number):
        # 			fluid_comp = self.inventory["FluidComponent"].collection[rr]
        # 			s_comp.dict_node_pt["temperature"] = \
        # 																	s_comp.dict_node_pt["temperature"] + \
        # 																	fluid_comp.coolant.dict_node_pt["temperature"]*\
        # 																	weight[rr]/np.sum(weight)
        # 	else:
        # 		# evaluate SolidComponent temperature as the algebraic average of \
        # 		# channels temperature (cdp, 07/2020)
        # 		s_comp.dict_node_pt["temperature"] = temp_ave
        # 		# scalar (cdp, 07/2020)
        # 		s_comp.TEMOUT = s_comp.dict_node_pt["temperature"][-1]
        # 	# call function f_plot to make property plots (cdp, 07/2020)
        # 	#f_plot(s_comp, self.grid_features["zcoord"])

        # Call function SolidComponents_T_initialization to initialize \
        # SolidComponent temperature spatial distribution from FluidComponent \
        # temperature or from input values according to flag INTIAL (cdp, 12/2020)
        solid_components_temperature_initialization(self)

        # Nested loop jacket - jacket.
        for rr, jacket_r in enumerate(self.inventory["JacketComponent"].collection):
            # np array of shape (Node, 1) to avoid broadcasting error.
            jacket_r.radiative_heat_env = np.zeros(
                (jacket_r.dict_node_pt["temperature"].size, 1)
            )
            for _, jacket_c in enumerate(
                self.inventory["JacketComponent"].collection[rr + 1 :]
            ):
                jacket_r.radiative_heat_inn[
                    f"{jacket_r.identifier}_{jacket_c.identifier}"
                ] = np.zeros((jacket_r.dict_node_pt["temperature"].size, 1))
                jacket_c.radiative_heat_inn[
                    f"{jacket_r.identifier}_{jacket_c.identifier}"
                ] = np.zeros((jacket_c.dict_node_pt["temperature"].size, 1))
            # End for cc.
        # End for rr.

        # IOP=IOP0(icond)
        for obj in self.inventory["SolidComponent"].collection:
            # Compute fractions of the total current that flows in
            # superconductor cross section of each strand or stack object if in
            # superconducting regime and fractions of the total current that
            # flows in the total cross section of each strand or stack object
            # if in current sharing regime.
            obj.get_current_fractions(
                self.total_sc_cross_section, self.total_so_cross_section, self.inventory
            )

        # Initialize thermal hydraulic quantities in both nodal and Gauss 
        # points.
        self.operating_conditions_th_initialization(simulation)
        # Initialize electromagnetic quantities in both nodal and Gauss 
        # points.
        self.operating_conditions_em()

        # Loop on SolidComponent (cdp, 01/2021)
        # N.B. questo loop si potrebbe fare usando map.
        for s_comp in self.inventory["SolidComponent"].collection:
            # compute, average density, thermal conductivity, specifi heat at \
            # constant pressure and electrical resistivity at initial \
            # SolidComponent temperature in nodal points (cdp, 01/2021)
            s_comp.eval_sol_comp_properties(self.inventory)
        # end for s_comp.

        # Loop to initialize electric related quantities for each
        # SolidComponent object.
        # N.B. remember that JacketComponent objects do not carry current for
        # the time being so these quantities will remain 0.
        for obj in self.inventory["SolidComponent"].collection:
            obj.initialize_electric_quantities(self)

        # Initialize to zeros all quantities related to heat source in nodal 
        # points.
        self.__build_heat_source_nodal_pt(simulation)

        # ENERGY BALANCE FLUID COMPONENTS
        for fluid_comp in self.inventory["FluidComponent"].collection:
            # Evaluate the density (if necessary) and the mass flow rate in 
            # points (nodal = True by default)
            fluid_comp.coolant._compute_density_and_mass_flow_rates_nodal_gauss(self)
            temp_ave = (
                temp_ave
                + fluid_comp.coolant.dict_node_pt["temperature"]
                / self.inventory["FluidComponent"].number
            )
            # Enthalpy balance: dt*sum((mdot*w)_out - (mdot*w)_inl), used to 
            # check the imposition of SolidComponent temperature initial 
            # spatial distribution.
            # N.B. queste istruzioni posso inserirle in un metodo della classe.
            self.enthalpy_balance = self.enthalpy_balance + simulation.transient_input[
                "STPMIN"
            ] * (
                fluid_comp.coolant.dict_node_pt["mass_flow_rate"][-1]
                * fluid_comp.coolant.dict_node_pt["total_enthalpy"][-1]
                - fluid_comp.coolant.dict_node_pt["mass_flow_rate"][0]
                * fluid_comp.coolant.dict_node_pt["total_enthalpy"][0]
            )
            self.enthalpy_out = (
                self.enthalpy_out
                + simulation.transient_input["STPMIN"]
                * fluid_comp.coolant.dict_node_pt["mass_flow_rate"][-1]
                * fluid_comp.coolant.dict_node_pt["total_enthalpy"][-1]
            )
            self.enthalpy_inl = (
                self.enthalpy_inl
                + simulation.transient_input["STPMIN"]
                * fluid_comp.coolant.dict_node_pt["mass_flow_rate"][0]
                * fluid_comp.coolant.dict_node_pt["total_enthalpy"][0]
            )

        # Initialize the Energy of the SolidComponent (cdp, 12/2020)
        self.E_sol_ini = 0.0
        self.E_sol_fin = 0.0
        self.E_str_ini = 0.0
        self.E_str_fin = 0.0
        self.E_jk_ini = 0.0
        self.E_jk_fin = 0.0
        # Loop on SolidComponent to evaluate the total initial energy of \
        # SolidComponent, used to check the imposition of SolidComponent \
        # temperature initial spatial distribution (cdp, 12/2020)
        # N.B. questo loop si potrebbe fare usando map.
        for s_comp in self.inventory["SolidComponent"].collection:
            # N.B. queste istruzioni posso inserirle in un metodo della classe.
            self.E_sol_ini = self.E_sol_ini + s_comp.inputs["CROSSECTION"] * np.sum(
                (
                    self.grid_features["zcoord"][1 : self.grid_features["N_nod"]]
                    - self.grid_features["zcoord"][0:-1]
                )
                * s_comp.dict_Gauss_pt["total_density"]
                * s_comp.dict_Gauss_pt["total_isobaric_specific_heat"]
                * s_comp.dict_Gauss_pt["temperature"]
            )
            if s_comp.name != "Z_JACKET":
                self.E_str_ini = self.E_str_ini + s_comp.inputs["CROSSECTION"] * np.sum(
                    (
                        self.grid_features["zcoord"][1 : self.grid_features["N_nod"]]
                        - self.grid_features["zcoord"][0:-1]
                    )
                    * s_comp.dict_Gauss_pt["total_density"]
                    * s_comp.dict_Gauss_pt["total_isobaric_specific_heat"]
                    * s_comp.dict_Gauss_pt["temperature"]
                )
            else:
                self.E_jk_ini = self.E_jk_ini + s_comp.inputs["CROSSECTION"] * np.sum(
                    (
                        self.grid_features["zcoord"][1 : self.grid_features["N_nod"]]
                        - self.grid_features["zcoord"][0:-1]
                    )
                    * s_comp.dict_Gauss_pt["total_density"]
                    * s_comp.dict_Gauss_pt["total_isobaric_specific_heat"]
                    * s_comp.dict_Gauss_pt["temperature"]
                )
        # end for s_comp (cdp, 12/2020)

        # Construct and initialize dictionary dict_Step to correctly apply the \
        # method that solves the transient (cdp, 10/2020)
        if self.inputs["METHOD"] == "BE" or self.inputs["METHOD"] == "CN":
            # Backward Euler or Crank-Nicolson (cdp, 10/2020)
            self.dict_Step = dict(
                SYSLOD=np.zeros((self.dict_N_equation["Total"], 2)),
                SYSVAR=np.zeros((self.dict_N_equation["Total"], 1)),
            )
        elif self.inputs["METHOD"] == "AM4":
            # Adams-Moulton order 4 (cdp, 10/2020)
            self.dict_Step = dict(
                SYSLOD=np.zeros((self.dict_N_equation["Total"], 4)),
                SYSVAR=np.zeros((self.dict_N_equation["Total"], 3)),
                # AM4_AA: four matrices of size Full * Total
                AM4_AA=np.zeros(
                    (4,self.dict_band["Full"],self.dict_N_equation["Total"])
                ),
            )
        # end if self.inputs

        # Assign initial values to key SYSVAR (cdp, 10/2020)
        for jj, fluid_comp in enumerate(self.inventory["FluidComponent"].collection):
            # velocity (cdp, 10/2020)
            self.dict_Step["SYSVAR"][
                jj : self.dict_N_equation["Total"] : self.dict_N_equation["NODOFS"], 0
            ] = fluid_comp.coolant.dict_node_pt["velocity"]
            # pressure (cdp, 10/2020)
            self.dict_Step["SYSVAR"][
                jj
                + self.inventory["FluidComponent"].number : self.dict_N_equation[
                    "Total"
                ] : self.dict_N_equation["NODOFS"],
                0,
            ] = fluid_comp.coolant.dict_node_pt["pressure"]
            # temperature (cdp, 10/2020)
            self.dict_Step["SYSVAR"][
                jj
                + 2
                * self.inventory["FluidComponent"].number : self.dict_N_equation[
                    "Total"
                ] : self.dict_N_equation["NODOFS"],
                0,
            ] = fluid_comp.coolant.dict_node_pt["temperature"]
        # end for jj (cdp, 10/2020)
        for ll, comp in enumerate(self.inventory["SolidComponent"].collection):
            # solid components temperature (cdp, 10/2020)
            self.dict_Step["SYSVAR"][
                ll
                + self.dict_N_equation["FluidComponent"] : self.dict_N_equation[
                    "Total"
                ] : self.dict_N_equation["NODOFS"],
                0,
            ] = comp.dict_node_pt["temperature"]
        # end for ll (cdp, 10/2020)
        if self.dict_Step["SYSVAR"].shape[-1] > 1:
            # if this is true, it means that an higher order method than \
            # Crank-Nicolson is applied to solve the transient (cdp, 10/2020)
            for cc in range(1, self.dict_Step["SYSVAR"].shape[-1]):
                # Copy the values of the first colum in all the other columns, like \
                # they are the results of a dummy initial steady state (cdp, 10/2020)
                self.dict_Step["SYSVAR"][:, cc] = self.dict_Step["SYSVAR"][:, 0].copy()
            # end for cc (cdp, 10/2020)
        # end if self.dict_Step["SYSVAR"].shape[-1] (cdp, 10/2020)

        conductorlogger.debug(
            f"Before call function {save_geometry_discretization.__name__}.\n"
        )
        save_geometry_discretization(
            self.inventory["all_component"].collection,
            simulation.dict_path[f"Output_Initialization_{self.identifier}_dir"],
        )
        conductorlogger.debug(
            f"After call function {save_geometry_discretization.__name__}.\n"
        )

        conductorlogger.debug(f"Before call function {save_properties.__name__}.\n")
        # Call function Save_properties to save conductor inizialization
        save_properties(
            self, simulation.dict_path[f"Output_Initialization_{self.identifier}_dir"]
        )
        conductorlogger.debug(f"After call function {save_properties.__name__}.\n")

        # Call function update_real_time_plot
        update_real_time_plots(self)
        create_legend_rtp(self)

    # end method initialization

    ############################################################################

    ##### ELECTRIC PREPROCESSING ############

    def __build_nodal_coordinates(self, nn: int, key: str):
        """Private method that builds the dataframe with the nodal coordinates of all conductor components.

        Args:
            nn (int): starting value of the index
            key (str): key of the dictionary self.inventory; can be FluidComponent, StrandComponent, JacketComponent.
        """

        for ii, obj in enumerate(self.inventory[key].collection, nn):
            for coord in ["x", "y", "z"]:
                self.nodal_coordinates.iloc[
                    ii :: self.inventory["all_component"].number,
                    self.nodal_coordinates.columns.get_loc(coord),
                ] = obj.coordinate[coord]

    def __build_connectivity(self, nn: int, key: str):
        """Private method that builds the dataframe with the connections (start and end node of each elements) of all conductor components.

        Args:
            nn (int): starting value of the index
            key (str): key of the dictionary self.inventory; can be FluidComponent, StrandComponent, JacketComponent.
        """
        for ii, obj in enumerate(self.inventory[key].collection, nn):
            nodes = np.linspace(
                ii, ii + self.total_elements, self.grid_input["NELEMS"] + 1, dtype=int
            )
            self.connectivity_matrix.iloc[
                ii :: self.inventory["all_component"].number,
                self.connectivity_matrix.columns.get_loc("start"),
            ] = nodes[:-1]
            self.connectivity_matrix.iloc[
                ii :: self.inventory["all_component"].number,
                self.connectivity_matrix.columns.get_loc("end"),
            ] = nodes[1:]
            self.connectivity_matrix.iloc[
                ii :: self.inventory["all_component"].number,
                self.connectivity_matrix.columns.get_loc("identifiers"),
            ] = obj.identifier
        # End for

    def __build_connectivity_current_carriers(self):
        """Private method that builds the dataframe with the connections (start and end node of each elements) of StrandMixedComponent, StrandStabilizerComonent and StackComponent components."""
        for ii, obj in enumerate(self.inventory["StrandComponent"].collection):
            nodes = np.linspace(
                ii,
                ii + self.total_elements_current_carriers,
                self.grid_input["NELEMS"] + 1,
                dtype=int,
            )
            self.connectivity_matrix_current_carriers.iloc[
                ii :: self.inventory["StrandComponent"].number,
                self.connectivity_matrix_current_carriers.columns.get_loc("start"),
            ] = nodes[:-1]
            self.connectivity_matrix_current_carriers.iloc[
                ii :: self.inventory["StrandComponent"].number,
                self.connectivity_matrix_current_carriers.columns.get_loc("end"),
            ] = nodes[1:]
            self.connectivity_matrix_current_carriers.iloc[
                ii :: self.inventory["StrandComponent"].number,
                self.connectivity_matrix_current_carriers.columns.get_loc(
                    "identifiers"
                ),
            ] = obj.identifier

    def __compute_node_distance(self):
        """Private method that computes the distance between nodes thaking into account all the coordinates (x,y,z). Values are stored in attribute node_distance."""
        self.node_distance = (
            (
                (
                    self.nodal_coordinates.iloc[self.connectivity_matrix["end"], :]
                    - self.nodal_coordinates.iloc[self.connectivity_matrix["start"], :]
                )
                ** 2
            )
            .sum(axis=1)
            .apply(np.sqrt)
        )

        # noda_distance correction.
        if self.inventory["StrandComponent"].number == 1:
            # In this case attribute node_distance does not account for the 
            # twist pich of the strand, i.e. the strand is straight. Apply the 
            # costheta correction to account for the real distance between 
            # nodal points. This correction has an impact only in the 
            # evaluation of the quantity used in the electromagnetic module, as 
            # for instance electric resistance, electric conductance and 
            # inductance.
            # This is not necessary if there are more than one 
            # strand since in this case the helicoidal coordinates are used.
            self.node_distance = self.node_distance / self.inventory["StrandComponent"].collection[0].inputs["COSTETA"]

    def __compute_gauss_node_distance(self):
        """Private method that evaluates the distance between consecutive gauss node (the mid point of the element), thaking into account all the coordinates (x,y,z). Values are stored in attribute gauss_node_distance."""
        self.gauss_node_distance = np.zeros(self.total_nodes)
        # On the first cross section there is only the contribution from the
        # firts element
        self.gauss_node_distance[: self.inventory["all_component"].number] = (
            self.node_distance[: self.inventory["all_component"].number] / 2
        )

        # All the 'inner' distances are evaluated as
        # (l_k + l_(k+1))/2, for k in [0,total_nodes]
        self.gauss_node_distance[
            self.inventory["all_component"]
            .number : -self.inventory["all_component"]
            .number
        ] = (
            self.node_distance[: -self.inventory["all_component"].number]
            + self.node_distance[self.inventory["all_component"].number :]
        ) / 2

        # On the last cross section there is only the contribution from the
        # last element
        self.gauss_node_distance[-self.inventory["all_component"].number :] = (
            self.node_distance[-self.inventory["all_component"].number :] / 2
        )

    def __build_incidence_matrix(self):
        """Private method that builds the incidence matrix limited to components of kind StrandMixedComponent, StrandStabilizerComonent and StackComponent. Value stored in attribute incidence_matrix; the transposed incidence matrix is also evaluated and stored in attribute incidence_matrix_transposed. Thake adantage of sparse matrices.

        From MatLab code given by professor F. Freschi.
        """
        # Row pointer: get_loc returns a boolean array, irow is build with
        # values in 0:Ne for which the boolean is True (corresponds to a
        # CurrenCarrier index).
        irow = np.tile(
            np.r_[0 : self.total_elements_current_carriers],
            (2, 1),
        ).flatten("F")
        # Column pointer.
        jcol = (
            self.connectivity_matrix_current_carriers.iloc[:, 0:2]
            .to_numpy()
            .copy()
            .transpose()
            .flatten("F")
        )
        # Nonzeros values
        s = np.tile([-1, 1], self.total_elements_current_carriers)
        # Assemble matrix
        self.incidence_matrix = coo_matrix(
            (s, (irow, jcol)),
            shape=(
                self.total_elements_current_carriers,
                self.total_nodes_current_carriers,
            ),
        ).tocsr()

        self.incidence_matrix_transposed = self.incidence_matrix.T

    def __build_electric_resistance_matrix(self):
        """Private method that builds the elecrtic resistance matrix limited to components of kind StrandMixedComponent, StrandStabilizerComonent and StackComponent. Value stored in attribute electric_resistance_matrix. Thake adantage of sparse matrices."""

        resistance = np.zeros(self.total_elements_current_carriers)
        for ii, obj in enumerate(self.inventory["StrandComponent"].collection):
            resistance[
                ii :: self.inventory["StrandComponent"].number
            ] = obj.get_electric_resistance(self)

        self.electric_resistance_matrix = diags(
            resistance,
            offsets=0,
            shape=(
                self.total_elements_current_carriers,
                self.total_elements_current_carriers,
            ),
            format="csr",
            dtype=float,
        )

    def __contact_current_carriers_first_cross_section(self):
        """Private method that evaluates the he contact nodes between StrandMixedComponent, StrandStabilizerComonent and StackComponent components on the first conductor cross section exploiting the contact perimeter flag value in sheet contact_perimeter_flag of input file conductor_coupling.xlsx.
        Values stored in private attribute _contact_nodes_first.
        """

        # Alias
        interf_flag = self.dict_df_coupling["contact_perimeter_flag"]

        self._contact_nodes_first = np.array([])
        # 1+self.inventory["FluidComponent"].number keeps into account the
        # Environment component.
        for row in range(
            1 + self.inventory["FluidComponent"].number,
            1
            + self.inventory["FluidComponent"].number
            + self.inventory["StrandComponent"].number,
        ):
            ind = np.nonzero(abs(interf_flag.iloc[
                    row,
                    1
                    + self.inventory["FluidComponent"].number : 1
                    + self.inventory["FluidComponent"].number
                    + self.inventory["StrandComponent"].number,
                ].to_numpy()
            ) == 1
            )[0]

            # Reduce the row index to convert from the whole system to the
            # reduced one (only the StrandComponent components).
            row -= 1 + self.inventory["FluidComponent"].number

            if row == 0:
                self._contact_nodes_first = np.array(
                    [row * np.ones(ind.shape, dtype=int), ind]
                ).T
            else:
                self._contact_nodes_first = np.concatenate(
                    (
                        self._contact_nodes_first,
                        np.array([row * np.ones(ind.shape, dtype=int), ind]).T,
                    )
                )

    def __contact_current_carriers(self):
        """Private method that detects the contacts between components of kind StrandMixedComponent, StrandStabilizerComonent and StackComponent, starting from the information on the first cross section. For the time being the component twist is not taken into account. Values stored in attribute contact_nodes_current_carriers.
        Exploits method __contact_current_carriers_first_cross_section.
        """
        self.__contact_current_carriers_first_cross_section()
        contact_nodes_current_carriers = np.zeros(
            (
                (self.grid_input["NELEMS"] + 1) * self._contact_nodes_first.shape[0],
                self._contact_nodes_first.shape[1],
            ),
            dtype=int,
        )
        contact_nodes_current_carriers[
            : self._contact_nodes_first.shape[0], :
        ] = self._contact_nodes_first
        for ii in range(1, self.grid_input["NELEMS"] + 1):
            contact_nodes_current_carriers[
                ii
                * self._contact_nodes_first.shape[0] : (ii + 1)
                * self._contact_nodes_first.shape[0],
                :,
            ] = (
                contact_nodes_current_carriers[
                    (ii - 1)
                    * self._contact_nodes_first.shape[0] : ii
                    * self._contact_nodes_first.shape[0],
                    :,
                ]
                + self.inventory["StrandComponent"].number
            )

        self.contact_nodes_current_carriers = pd.DataFrame(
            contact_nodes_current_carriers,
            dtype=int,
            columns=["start", "end"],
        )

    def __build_contact_incidence_matrix(self):
        """Private method that builds the edge to node incidence matrix limited to components of kind StrandMixedComponent, StrandStabilizerComonent and StackComponent. Values stored in attribute contact_incidence_matrix. Expoit sparse matrix."""

        # Edge-to-node incidence matrix (referred to En)
        row_ind = np.tile(
            np.r_[0 : self.contact_nodes_current_carriers.shape[0]], (2, 1)
        ).flatten("F")
        col_ind = (
            self.contact_nodes_current_carriers.to_numpy().transpose().flatten("F")
        )  # which column
        self.contact_incidence_matrix = coo_matrix(
            (
                np.tile([-1, 1], self.contact_nodes_current_carriers.shape[0]),
                (row_ind, col_ind),
            )
        ).tocsr()

    def __evaluate_transversal_distance(self) -> np.ndarray:
        """Private method that evaluates distance along the direction ortoghonal to the z direction, between nodes of components of kind StrandMixedComponent, StrandStabilizerComonent and StackComponent that are in contact.

        Returns:
            np.ndarray: array with the evaluated distance.
        """
        # reset_index is used to reset the index to numerical values instead of object identifier in order to make the correct operations and have the correct shape: distance.shape = (self.total_nodes_current_carriers,)
        distance = (
            (
                (
                    self.nodal_coordinates.loc["StrandComponent"]
                    .iloc[self.contact_nodes_current_carriers["end"], :]
                    .reset_index(level="Identifier", drop=True)
                    - self.nodal_coordinates.loc["StrandComponent"]
                    .iloc[self.contact_nodes_current_carriers["start"], :]
                    .reset_index(level="Identifier", drop=True)
                )
                ** 2
            )
            .sum(axis="columns")
            .apply(np.sqrt)
            .to_numpy()
        )  # distance
        return distance

    def __evaluate_electric_conductance(self, distance: np.ndarray) -> np.ndarray:
        """Private method that evaluates the electric conductance for components of kind StrandMixedComponent, StrandStabilizerComonent and StackComponent that are in contact in transverse direction. According to the value in sheet electric_conductance_mode of input file conductro_coupling.xlsx the electric conductance is evaluated in different modes:
        1) exploits function __evaluate_electric_conductance_unit_length (electric conductance is defined per unit length);
        2) exploits function __evaluate_electric_conductance_not_unit_length (electric conductance is not defined per unit length).

        Args:
            distance (np.ndarray): distance along the direction ortoghonal to the z direction, between nodes of components of kind StrandMixedComponent, StrandStabilizerComonent and StackComponent that are in contact

        Returns:
            np.ndarray: matrix with the electric conductance values.
        """

        def __evaluate_electric_conductance_unit_length() -> np.ndarray:
            """Fuction that evaluates the electric conductance per unit length.

            el_cond = sigma*gauss_node_distance

            with sigma the electric conductance per unit length.

            Returns:
                np.ndarray: matrix with the electric conductance values.
            """
            # el_cond = sigma*gauss_node_distance
            return (
                self.electric_conductance[indexes[0], indexes[1]]
                * self.gauss_node_distance[
                    arr_ind[0] :: self.inventory["all_component"].number
                ]
            )

        def __evaluate_electric_conductance_not_unit_length() -> np.ndarray:
            """Fuction that evaluates the electric conductance when the electric conductivity sigma is not given per unit length.

            el_cond = sigma*contact_perimeter*gauss_node_distance/contact_distance

            Returns:
                np.ndarray: matrix with the electric conductance values.
            """
            # el_cond = sigma*contact_perimeter*gauss_node_distance
            # /contact_distance
            return (
                self.electric_conductance[indexes[0], indexes[1]]
                * self.dict_df_coupling["contact_perimeter"].iat[mat_ind[0], mat_ind[1]]
                * (
                    self.gauss_node_distance[
                        arr_ind[0] :: self.inventory["all_component"].number
                    ]
                    + self.gauss_node_distance[
                        arr_ind[1] :: self.inventory["all_component"].number
                    ]
                )
                / 2
                / distance[ii :: self._contact_nodes_first.shape[0]]
            )

        electric_conductance = np.zeros((self.contact_nodes_current_carriers.shape[0],))

        # Switch to use the correct function to evaluate the electric
        # conductance according to the flag set in sheet
        # electric_conductance_mode in file conductor_coupling.xlsx
        _ = {
            ELECTRIC_CONDUCTANCE_UNIT_LENGTH: __evaluate_electric_conductance_unit_length,
            ELECTRIC_CONDUCTANCE_NOT_UNIT_LENGTH: __evaluate_electric_conductance_not_unit_length,
        }

        # Evaluate electric conductance
        for ii, indexes in enumerate(self._contact_nodes_first):
            mat_ind = indexes + self.inventory["FluidComponent"].number + 1
            arr_ind = indexes + self.inventory["FluidComponent"].number

            electric_conductance[ii :: self._contact_nodes_first.shape[0]] = _[
                self.electric_conductance_mode[indexes[0], indexes[1]]
            ]()

        return electric_conductance

    def __build_electric_conductance_matrix(self):
        """Private method that builds the electric conductance matrix for components of kind StrandMixedComponent, StrandStabilizerComonent and StackComponent that are in contact along transverse direction. Values are stored in attribute electric_conductance_matrix.
        Exploits sparse matrix.
        """

        # reset_index is used to reset the index to numerical values instead of object identifier in order to make the correct operations and have the correct shape: distance.shape = (self.total_nodes_current_carriers,)
        distance = self.__evaluate_transversal_distance()

        electric_conductance = self.__evaluate_electric_conductance(distance)
        np.savetxt(
            "electric_conductance_unit_length.tsv", electric_conductance, delimiter="\t"
        )

        self.electric_conductance_diag_matrix = diags(
            electric_conductance,
            offsets=0,
            shape=(
                self.contact_nodes_current_carriers.shape[0],
                self.contact_nodes_current_carriers.shape[0],
            ),
            format="csr",
            dtype=float,
        )
        # Conductance matrix
        self.electric_conductance_matrix = (
            self.contact_incidence_matrix.T
            @ self.electric_conductance_diag_matrix
            @ self.contact_incidence_matrix
        )

    def electric_preprocessing(self):
        """Method that allows to evaluate most of the quatities and data structures needed for the electric calculation.
        Builds nodal coordinates and connectiviy dataframes, the connectivity matrix only for StrandComponent, the inicidence matrices in both longitudinal and transversal directions, the resistance matrix (logitudinal) and the conductance matrix (transverse direction).
        """

        nn = 0

        for key in ["FluidComponent", "StrandComponent", "JacketComponent"]:
            # Build nodal coordinates
            # conductorlogger.debug(
            #     f"Before call method {self.__build_nodal_coordinates.__name__} for {key} objects.\n"
            # )
            self.__build_nodal_coordinates(nn, key)
            # conductorlogger.debug(
            #     f"After call method {self.__build_nodal_coordinates.__name__} for {key} objects.\n"
            # )

            # Build connectivity matrix
            # conductorlogger.debug(
            #     f"Before call method {self.__build_connectivity.__name__} for {key} objects.\n"
            # )
            self.__build_connectivity(nn, key)
            # conductorlogger.debug(
            #     f"After call method {self.__build_connectivity.__name__} for {key} objects.\n"
            # )
            nn += self.inventory[key].number
            # End if
        # End for key

        # Build the connectivity matrix for the reduced system of components:
        # keeps into account only the StrandComponent ones.
        # conductorlogger.debug(
        #     f"Before call method {self.__build_connectivity_current_carriers.__name__}, operates on StrandComponents only.\n"
        # )
        self.__build_connectivity_current_carriers()
        # conductorlogger.debug(
        #     f"After call method {self.__build_connectivity_current_carriers.__name__}, operates on StrandComponents only.\n"
        # )

        # Convert index to categorical
        # conductorlogger.debug(
        #     f"Before convert index of dataframe self.connectivity_matrix to categorical.\n"
        # )
        self.connectivity_matrix.loc[:, "identifiers"] = self.connectivity_matrix.loc[
            :, "identifiers"
        ].astype("category")
        self.connectivity_matrix_current_carriers.loc[
            :, "identifiers"
        ] = self.connectivity_matrix_current_carriers.loc[:, "identifiers"].astype(
            "category"
        )
        # conductorlogger.debug(
        #     f"After convert index of dataframe self.connectivity_matrix to categorical.\n"
        # )

        # Compute node distance
        # conductorlogger.debug(
        #     f"Before call method {self.__compute_node_distance.__name__}.\n"
        # )
        self.__compute_node_distance()
        # conductorlogger.debug(
        #     f"After call method {self.__compute_node_distance.__name__}.\n"
        # )

        # conductorlogger.debug(
        #     f"Before call method {self.__compute_gauss_node_distance.__name__}.\n"
        # )

        # Compute gauss node distance
        self.__compute_gauss_node_distance()
        # conductorlogger.debug(
        #     f"After call method {self.__compute_gauss_node_distance.__name__}.\n"
        # )

        # conductorlogger.debug(
        #     f"Before call method {self.__build_incidence_matrix.__name__}.\n"
        # )
        # Build incidence matrix only for StrandComponent
        self.__build_incidence_matrix()
        # conductorlogger.debug(
        #     f"After call method {self.__build_incidence_matrix.__name__}.\n"
        # )

        # Build electric resistance matrix (for the first time)
        # conductorlogger.debug(
        #     f"Before call method {self.__build_electric_resistance_matrix.__name__}.\n"
        # )
        self.__build_electric_resistance_matrix()
        # conductorlogger.debug(
        #     f"After call method {self.__build_electric_resistance_matrix.__name__}.\n"
        # )

        if self.inventory["StrandComponent"].number > 1:
            # There are more than 1 StrandComponent objects, therefore there
            # are contacts between StrandComponent objects and matrices
            # contact_incidence_matrix and electric_conductance_matix can be
            # built. If there is only one StrandComponent object
            # contact_incidence_matrix can not be defined while
            # electric_conductance_matix is full of 0 from initialization.

            # Find contacts between StrandComponent objects.
            # conductorlogger.debug(
            #     f"Before call method {self.__contact_current_carriers.__name__}.\n"
            # )
            self.__contact_current_carriers()
            # conductorlogger.debug(
            #     f"After call method {self.__contact_current_carriers.__name__}.\n"
            # )

            # Build contact incidence matrix
            # conductorlogger.debug(
            #     f"Before call method {self.__build_contact_incidence_matrix.__name__}.\n"
            # )
            # this method builds the contact incidence matrix for current carriers
            # only
            self.__build_contact_incidence_matrix()
            # conductorlogger.debug(
            #     f"After call method {self.__build_contact_incidence_matrix.__name__}.\n"
            # )

            # Build electric conductance matrix
            # conductorlogger.debug(
            #     f"Before call method {self.__build_electric_conductance_matrix.__name__}.\n"
            # )
            self.__build_electric_conductance_matrix()
            # conductorlogger.debug(
            #     f"After call method {self.__build_electric_conductance_matrix.__name__}.\n"
            # )

        # Build electric stiffness matrix (for the first time)
        # conductorlogger.debug(
        #     f"Before call method {self.__build_electric_stiffness_matrix.__name__}.\n"
        # )
        self.__build_electric_stiffness_matrix()
        # conductorlogger.debug(
        #     f"After call method {self.__build_electric_stiffness_matrix.__name__}.\n"
        # )

        if self.build_electric_mass_matrix_flag == True:
            # Build electric mass matrix (for the first time)
            # conductorlogger.debug(
            #     f"Before call method {self.__build_electric_mass_matrix.__name__}.\n"
            # )
            self.__build_electric_mass_matrix()
            # conductorlogger.debug(
            #     f"After call method {self.__build_electric_mass_matrix.__name__}.\n"
            # )

        if (
            self.grid_input["ITYMSH"]
            != 3 | self.grid_input["ITYMSH"]
            != -1 & self.build_electric_mass_matrix_flag
            == True
        ):
            # Discretization grid does not change at each time step so there is
            # no need to build electric mass matrix at each thermal time step
            # because inductances will not change since they are evaluating
            # starting from the coordinates which are constant in this case: flag build_electric_mass_matrix_flag is therefore set to False.
            self.build_electric_mass_matrix_flag = False

        # Assign equivalue surfaces
        # conductorlogger.debug(
        #     f"Before call method {self.__assign_equivalue_surfaces.__name__}.\n"
        # )
        self.__assign_equivalue_surfaces()
        # conductorlogger.debug(
        #     f"After call method {self.__assign_equivalue_surfaces.__name__}.\n"
        # )

        # Assign fixed potential
        # conductorlogger.debug(
        #     f"Before call method {self.__assign_fix_potential.__name__}.\n"
        # )
        self.__assign_fix_potential()
        # conductorlogger.debug(
        #     f"After call method {self.__assign_fix_potential.__name__}.\n"
        # )

    def __build_electric_stiffness_matrix(self):
        """Private method that builds the electric stiffness matrix as a combination of the electric_resistance_matrix, incidence_matrix and electric_conductance_matrix. Exploit sparse matrix."""

        # Electric stiffness matrix initialization. Moved here from private
        # method __initialize_attributes because of matrix reduction carried
        # out in function fixed_value of module electric_auxiliary_functions.py.
        # Need to think about this part since, at the state of the art of the
        # code, the only matrix that will change at each thermal time step is
        # the elctric resistance matrix.
        self.electric_stiffness_matrix = lil_matrix(
            (
                self.total_elements_current_carriers
                + self.total_nodes_current_carriers,
                self.total_elements_current_carriers
                + self.total_nodes_current_carriers,
            ),
            dtype=float,
        )

        self.electric_stiffness_matrix[
            : self.total_elements_current_carriers,
            : self.total_elements_current_carriers,
        ] = self.electric_resistance_matrix
        self.electric_stiffness_matrix[
            : self.total_elements_current_carriers,
            self.total_elements_current_carriers :,
        ] = self.incidence_matrix
        self.electric_stiffness_matrix[
            self.total_elements_current_carriers :,
            : self.total_elements_current_carriers,
        ] = -self.incidence_matrix_transposed
        self.electric_stiffness_matrix[
            self.total_elements_current_carriers :,
            self.total_elements_current_carriers :,
        ] = self.electric_conductance_matrix

        self.electric_stiffness_matrix = self.electric_stiffness_matrix.tocsr(copy=True)

    def __assign_equivalue_surfaces(self):
        """Private method that assigns the prescribed equipotential surface of the conductor."""
        for ii, coord in enumerate(self.operations["EQUIPOTENTIAL_SURFACE_COORDINATE"]):
            # Find the index of the spatial discretization along z such that
            # z <= round(coord,n_digit_z); assing the last StrandComponent
            # number values to the i-th row of equipotential_node_index and add
            # self.total_elements_current_carriers to shift the indexes to the
            # correct position (in the portion of the array dedicated to the
            # current).

            self.equipotential_node_index[ii, :] = (
                np.nonzero(
                    self.nodal_coordinates.loc["StrandComponent", "z"].to_numpy()
                    <= round(coord, self.n_digit_z)
                )[0][-self.inventory["StrandComponent"].number :]
                + self.total_elements_current_carriers
            )

    def __assign_fix_potential(self):
        """Private method that assigns the value of the fixed potential on prescribed fixed potential surfaces."""
        jj = 0
        tol = 1e-10
        for kk, obj in enumerate(self.inventory["StrandComponent"].collection):
            if obj.operations["FIX_POTENTIAL_FLAG"]:
                # Assign potential values.
                self.fixed_potential_value[
                    jj : jj + obj.operations["FIX_POTENTIAL_NUMBER"]
                ] = obj.operations["FIX_POTENTIAL_VALUE"]
                # Find and assign the index corresponding to fix potential
                # coordinates.
                for ii, coord in enumerate(obj.operations["FIX_POTENTIAL_COORDINATE"], jj):
                    self.fixed_potential_index[ii] = (
                        (self.nodal_coordinates.loc["StrandComponent", "z"] - coord).abs()
                        <= tol
                    ).to_numpy().nonzero()[0][kk] + self.total_elements_current_carriers

                jj += obj.operations["FIX_POTENTIAL_NUMBER"]

    def eval_total_operating_current(self):
        """Method that evaluates the total electric current flowing in the conductor according to the value of flag I0_OP_MODE:
        * 0: constant value
        * -1: from auxiliary file (to be implemented)
        * -2: from user defined external function.
        """

        # Initialize condutctor operating current vector.
        self.dict_node_pt["op_current"] = np.zeros(self.total_nodes_current_carriers)
        # Rimuovere gli if

        if self.inputs["I0_OP_MODE"] == IOP_CONSTANT:
            # positive means entering the node.
            # This is already a boundary condition.
            # All the current enters the first node of the first current
            # carrier.
            self.dict_node_pt["op_current"][0] = self.inputs["I0_OP_TOT"]
            # All the current exits from the lats node of the last current
            # carrier.
            self.dict_node_pt["op_current"][-1] = -self.inputs["I0_OP_TOT"]
        elif self.inputs["I0_OP_MODE"] == IOP_FROM_FILE:
            # Loop on SolidComponent objects to sum inlet and outlet current in
            # order to get the total conductor inlet and outlet current.
            for obj in self.inventory["SolidComponent"].collection:
                self.dict_node_pt["op_current"][0] = (
                    self.dict_node_pt["op_current"][0]
                    + obj.dict_node_pt["op_current"][0]
                )
                self.dict_node_pt["op_current"][-1] = (
                    self.dict_node_pt["op_current"][-1]
                    + obj.dict_node_pt["op_current"][-1]
                )
            # Change sign to the outlet curren since it is exiting the
            # conductor.
            self.dict_node_pt["op_current"][-1] = -self.dict_node_pt["op_current"][-1]
        elif self.inputs["I0_OP_MODE"] == IOP_FROM_EXT_FUNCTION:

            # All the current enters the first node of the first current
            # carrier.
            self.dict_node_pt["op_current"][0] = custom_current_function(
                self.electric_time_step
            )
            # All the current exits from the lats node of the last current
            # carrier.
            self.dict_node_pt["op_current"][-1] = -custom_current_function(
                self.electric_time_step
            )
            # to be implemented.
        # End if self.dict_input["I0_OP_MODE"].

    def build_electric_known_term_vector(self):
        """Method that builds the known therm vector for the electric module."""
        self.electric_known_term_vector[
            self.total_elements_current_carriers :
        ] = self.dict_node_pt["op_current"]

    # START: INDUCTANCE ANALYTICAL EVALUATION

    def __inductance_analytical_calculation(self, mode: int = 2):
        """Private method that evaluates the magnetic inductance analytically. Self inductances can be evaluated with to different formulations exploiting input argument mode (see section Args for details).

        Args:
            mode (int,optional): flag to select the equation for the analytical evaluation of self inductance. 1: mode 1; 2: mode 2. Defaults to 2.
        """

        if mode != 1 and mode != 2:
            raise ValueError(
                f"{self.identifier}\nArgument 'mode' must be equal to {SELF_INDUCTANCE_MODE_1 = } or to {SELF_INDUCTANCE_MODE_2 = }. Current value {mode = } is not allowed. Please check sheet {self.workbook_sheet_name[2]} in file {self.workbook_name}.\n"
            )
        ABSTOL = 1e-6
        lmod = (
            (
                (
                    self.nodal_coordinates.iloc[
                        self.connectivity_matrix.loc[
                            "StrandComponent",
                            "end",
                        ],
                        :,
                    ]
                    - self.nodal_coordinates.iloc[
                        self.connectivity_matrix.loc[
                            "StrandComponent",
                            "start",
                        ],
                        :,
                    ]
                )
                ** 2
            )
            .sum(axis=1)
            .apply(np.sqrt)
        )

        mutual_inductance = np.zeros(self.inductance_matrix.shape)
        if self.inventory["StrandComponent"].number > 1:
            # There are more than 1 StrandComponent objects, therefore the
            # mutual inductances between StrandComponent objects can be
            # evaluated.
            # If there is only 1 StrandComponent object, the mutual inductance
            # matrixis set to 0 as from initialization.

            # Evaluate mutual inductances
            for ii in range(self.total_elements_current_carriers - 1):
                mutual_inductance = self.__mutual_inductance(
                    lmod, ii, mutual_inductance, ABSTOL
                )
            # end for

        # Switch to evalutae self inductance.
        self_inductance_switch = {
            SELF_INDUCTANCE_MODE_1: self.__self_inductance_mode1,
            SELF_INDUCTANCE_MODE_2: self.__self_inductance_mode2,
        }
        self_inductance = self_inductance_switch[mode](lmod)

        internal_inductance = lmod / 2.0

        self.inductance_matrix = (
            constants.mu_0
            / (4.0 * constants.pi)
            * (
                np.diag(self_inductance + internal_inductance)
                + self.inductance_matrix
                + self.inductance_matrix.T
            )
        )

    def __mutual_inductance(
        self, lmod: np.ndarray, ii: int, matrix: np.ndarray, abstol: float = 1e-6
    ) -> np.ndarray:
        """Private method that evaluates the mutual inductance analytically.

        Args:
            lmod (np.ndarray): array with the distance between strand component nodal nodes.
            ii (int): index of the i-th edge on which the mutual inductance is evaluated.
            matrix (np.ndarray): initialized matrix to store analytically evaluated mutual inductance values.
            abstol (float, optional): absolute tollerance to avoid rounding for segments in a plane. Defaults to 1e-6.

            Returns:
                np.ndarray: analytically evaluated mutual inductance.
        """

        jj = np.r_[ii + 1 : self.total_elements_current_carriers]
        ll = lmod[jj]
        mm = lmod[ii]
        len_jj = self.total_elements_current_carriers - (ii + 1)
        rr = dict(
            end_end=np.zeros(len_jj),
            end_start=np.zeros(len_jj),
            start_start=np.zeros(len_jj),
            start_end=np.zeros(len_jj),
        )

        for key in rr.keys():
            rr[key] = self.__vertex_to_vertex_distance(key, ii, jj)
        # End for key

        # Additional parameters
        alpha2 = (
            rr["start_end"] ** 2
            - rr["start_start"] ** 2
            + rr["end_start"] ** 2
            - rr["end_end"] ** 2
        )

        cos_eps = np.minimum(np.maximum(alpha2 / (2 * ll * mm), -1.0), 1.0)
        sin_eps = np.sin(np.arccos(cos_eps))

        dd = 4 * ll ** 2 * mm ** 2 - alpha2 ** 2
        mu = (
            ll
            * (
                2 * mm ** 2 * (rr["end_start"] ** 2 - rr["start_start"] ** 2 - ll ** 2)
                + alpha2 * (rr["start_end"] ** 2 - rr["start_start"] ** 2 - mm ** 2)
            )
            / dd
        )
        nu = (
            mm
            * (
                2 * ll ** 2 * (rr["start_end"] ** 2 - rr["start_start"] ** 2 - mm ** 2)
                + alpha2 * (rr["end_start"] ** 2 - rr["start_start"] ** 2 - ll ** 2)
            )
            / dd
        )
        d2 = rr["start_start"] ** 2 - mu ** 2 - nu ** 2 + 2 * mu * nu * cos_eps

        # avoid rounding for segments in a plane
        d2[d2 < abstol ** 2] = 0
        d0 = np.sqrt(d2)

        # solid angles
        omega = (
            np.arctan(
                (d2 * cos_eps + (mu + ll) * (nu + mm) * sin_eps ** 2)
                / (d0 * rr["end_end"] * sin_eps)
            )
            - np.arctan(
                (d2 * cos_eps + (mu + ll) * nu * sin_eps ** 2)
                / (d0 * rr["end_start"] * sin_eps)
            )
            + np.arctan(
                (d2 * cos_eps + mu * nu * sin_eps ** 2)
                / (d0 * rr["start_start"] * sin_eps)
            )
            - np.arctan(
                (d2 * cos_eps + mu * (nu + mm) * sin_eps ** 2)
                / (d0 * rr["start_end"] * sin_eps)
            )
        )
        omega[d0 == 0.0] = 0.0

        # contribution
        pp = np.zeros((len_jj, 5), dtype=float)
        pp[:, 0] = (ll + mu) * np.arctanh(mm / (rr["end_end"] + rr["end_start"]))
        pp[:, 1] = -nu * np.arctanh(ll / (rr["end_start"] + rr["start_start"]))
        pp[:, 2] = (mm + nu) * np.arctanh(ll / (rr["end_end"] + rr["start_end"]))
        pp[:, 3] = -mu * np.arctanh(mm / (rr["start_start"] + rr["start_end"]))
        pp[:, 4] = d0 * omega / sin_eps

        # filter odd cases (e.g. consecutive segments)
        pp[np.isnan(pp)] = 0.0
        pp[np.isinf(pp)] = 0.0

        # Mutual inductances
        matrix[ii, jj] = (
            2 * cos_eps * (pp[:, 0] + pp[:, 1] + pp[:, 2] + pp[:, 3])
            - cos_eps * pp[:, 4]
        )
        return matrix

    def __vertex_to_vertex_distance(
        self, key: str, ii: int, jj: np.ndarray
    ) -> np.ndarray:
        """Private method that evaluates the vertex to vertex distances. Possible definitions are stored in input argument key.

        Args:
            key (str): string with the definition for distance evaluation. Possible definitions are:
                1) start_end: distance bewteen start and end nodes;
                2) start_start: distance bewteen two start nodes;
                3) end_end: distance bewteen two end nodes;
                4) end_start: distance bewteen end and start nodes.

            ii (int): index of the i-th edge on which the mutual inductance is evaluated.
            jj (np.ndarray): array of index of the nodes for which distance must be evaluated.

        Returns:
            np.ndarray: vertex to vertex distances.
        """
        cols = key.split("_")
        return (
            (
                (
                    self.nodal_coordinates.iloc[
                        self.connectivity_matrix.loc["StrandComponent"].iloc[
                            jj,
                            self.connectivity_matrix.columns.get_loc(cols[0]),
                        ],
                        :,
                    ]
                    - self.nodal_coordinates.iloc[
                        self.connectivity_matrix.loc["StrandComponent"].iat[
                            ii,
                            self.connectivity_matrix.columns.get_loc(cols[1]),
                        ],
                        :,
                    ]
                )
                ** 2
            )
            .sum(axis="columns")
            .apply(np.sqrt)
        )

    def __self_inductance_mode1(self, lmod: np.ndarray) -> np.ndarray:
        """Private method that analytically evaluates self inductances according to mode 1.

        Args:
            lmod (np.ndarray): array with the distance between strand component nodal nodes.

        Returns:
            np.ndarray: self inductances.
        """
        self_inductance = np.zeros(lmod.shape)

        for ii, obj in enumerate(self.inventory["StrandComponent"].collection):
            self_inductance[ii :: self.inventory["StrandComponent"].number] = (
                2
                * lmod[ii :: self.inventory["StrandComponent"].number]
                * (
                    np.arcsinh(
                        lmod[ii :: self.inventory["StrandComponent"].number]
                        / obj.radius
                    )
                    - np.sqrt(
                        1.0
                        + (
                            obj.radius
                            / lmod[ii :: self.inventory["StrandComponent"].number]
                        )
                        ** 2
                    )
                    + obj.radius / lmod[ii :: self.inventory["StrandComponent"].number]
                )
            )
        return self_inductance

    def __self_inductance_mode2(self, lmod: np.ndarray) -> np.ndarray:
        """Private method that analytically evaluates self inductances according to mode 2.
        Args:
            lmod (np.ndarray): array with the distance between strand component nodal nodes.

        Returns:
            np.ndarray: self inductances.
        """
        self_inductance = np.zeros(lmod.shape)

        for ii, obj in enumerate(self.inventory["StrandComponent"].collection):

            self_inductance[ii :: self.inventory["StrandComponent"].number] = 2 * (
                lmod[ii :: self.inventory["StrandComponent"].number]
                * np.log(
                    (
                        lmod[ii :: self.inventory["StrandComponent"].number]
                        + np.sqrt(
                            lmod[ii :: self.inventory["StrandComponent"].number] ** 2
                            + obj.radius ** 2
                        )
                    )
                    / obj.radius
                )
                - np.sqrt(
                    lmod[ii :: self.inventory["StrandComponent"].number] ** 2
                    + obj.radius ** 2
                )
                + lmod[ii :: self.inventory["StrandComponent"].number] / 4
                + obj.radius
            )

        return self_inductance

    # END: INDUCTANCE ANALYTICAL EVALUATION

    # START: INDUCTANCE APPROXIMATE EVALUATION

    def __inductance_approximate_calculation(self):
        """Private method that approximate the inductance of the system. For an analytical evaluation of the inductance use private method __inductance_analytical_calculation."""

        ll = (
            self.nodal_coordinates.iloc[
                self.connectivity_matrix.loc[
                    "StrandComponent",
                    "end",
                ],
                :,
            ]
            - self.nodal_coordinates.iloc[
                self.connectivity_matrix.loc[
                    "StrandComponent",
                    "start",
                ],
                :,
            ]
        )
        lmod = (ll ** 2).sum(axis=1).apply(np.sqrt)

        internal_inductance = np.zeros(lmod.shape)
        mutual_inductance = np.zeros(self.inductance_matrix.shape)
        if self.inventory["StrandComponent"].number > 1:
            # There are more than 1 StrandComponent objects, therefore the
            # mutual inductances between StrandComponent objects can be
            # evaluated.
            # If there is only 1 StrandComponent object, the mutual inductance
            # matrixis set to 0 as from initialization.

            # Evaluate mutual inductance
            mutual_inductance = self.__mutual_inductance_approximate(
                ll, mutual_inductance
            )

        # Evaluate self inductance
        self_inductance = self.__self_inductance_approximate(lmod)
        # Evaluate internal inductance
        internal_inductance = lmod.to_numpy() / 2.0

        self.inductance_matrix = (
            constants.mu_0
            / (4.0 * constants.pi)
            * (
                np.diag(self_inductance + internal_inductance)
                + mutual_inductance
                + mutual_inductance.T
            )
        )

    def __mutual_inductance_approximate(
        self, ll: pd.DataFrame, matrix: np.ndarray
    ) -> np.ndarray:
        """Private method that evaluates an approximation of mutual inductance.

        Args:
            ll (pd.DataFrame): variable with useful values for the evaluation.
            matrix (np.ndarray): initialized matrix to store approximated mutual inductance values.

        Returns:
            np.ndarray: approximate value of mutual inductance.
        """

        ABSTOL = 1e-4
        RELTOL = 1e-4

        def __reverse_distance(xi, xj, qi, vi, qj, vj):
            return 1.0 / np.sqrt(
                (qi[0] + xi * vi[0] - qj[0] - xj * vj[0]) ** 2
                + (qi[1] + xi * vi[1] - qj[1] - xj * vj[1]) ** 2
                + (qi[2] + xi * vi[2] - qj[2] - xj * vj[2]) ** 2
            )

        for ii in range(self.total_elements_current_carriers):
            qi = (
                self.nodal_coordinates.iloc[
                    self.connectivity_matrix.loc[
                        "StrandComponent",
                        "start",
                    ],
                    :,
                ]
                .iloc[ii, :]
                .to_numpy()
            )
            vi = ll.iloc[ii, :].to_numpy()
            for jj in range(ii + 1, self.total_elements_current_carriers):
                qj = (
                    self.nodal_coordinates.iloc[
                        self.connectivity_matrix.loc[
                            "StrandComponent",
                            "start",
                        ],
                        :,
                    ]
                    .iloc[jj, :]
                    .to_numpy()
                )
                vj = ll.iloc[jj, :].to_numpy()
                matrix[ii, jj] = np.sum(vi * vj) * integrate.dblquad(
                    __reverse_distance,
                    0.0,
                    1.0,
                    0.0,
                    1.0,
                    (qi, vi, qj, vj),
                    ABSTOL,
                    RELTOL,
                )[0]

            # End for
        # End for
        return matrix

    def __self_inductance_approximate(self, lmod: np.ndarray) -> np.ndarray:
        """Private method that evaluates an approximation of self inductance.

        Args:
            lmod (np.ndarray): array with the distance between strand component nodal nodes.

        Returns:
            np.ndarray: approximate value of self inductance.
        """
        self_inductance = np.zeros(lmod.shape)

        for ii, obj in enumerate(self.inventory["StrandComponent"].collection):
            self_inductance[ii :: self.inventory["StrandComponent"].number] = 2 * (
                lmod[ii :: self.inventory["StrandComponent"].number]
                * np.log(
                    (
                        lmod[ii :: self.inventory["StrandComponent"].number]
                        + np.sqrt(
                            lmod[ii :: self.inventory["StrandComponent"].number] ** 2
                            + obj.radius ** 2
                        )
                    )
                    / obj.radius
                )
                - np.sqrt(
                    lmod[ii :: self.inventory["StrandComponent"].number] ** 2
                    + obj.radius ** 2
                )
                + lmod[ii :: self.inventory["StrandComponent"].number] / 4
                + obj.radius
            )
        # End for

        return self_inductance

    # END: INDUCTANCE APPROXIMATE EVALUATION

    def __build_electric_mass_matrix(self):
        """Private method that builds the electric mass matrix from the inductance matrix. Inductance matrix can be evaluated analytically or approximatey.

        Note: the other three blocks of the electric mass matrix are already set to zeros in the initialization.

        Raises:
            ValueError: raise error if mode is a not valid value.
        """

        if (
            self.operations["INDUCTANCE_MODE"] != 0
            and self.operations["INDUCTANCE_MODE"] != 1
        ):
            raise ValueError(
                f"{self.identifier = }\nArgument self.operations['INDUCTANCE_MODE'] should be equal to {APPROXIMATE_INDUCTANCE = } or {ANALYTICAL_INDUCTANCE = }. Current value ({self.operations['INDUCTANCE_MODE'] = }) is not allowed. Please check {self.workbook_sheet_name[2]} in file {self.workbook_name}.\n"
            )

        inductance_switch = {
            ANALYTICAL_INDUCTANCE: self.__inductance_analytical_calculation,
            APPROXIMATE_INDUCTANCE: self.__inductance_approximate_calculation,
        }

        inductance_switch[self.operations["INDUCTANCE_MODE"]]()

        self.electric_mass_matrix[
            : self.total_elements_current_carriers,
            : self.total_elements_current_carriers,
        ] = self.inductance_matrix

        self.electric_mass_matrix = self.electric_mass_matrix.tocsr(copy=True)

    def __get_electric_time_step(self):
        """Private method that evaluates the electric time step according to use definition.

        Raises:
            ValueError: if electric time step is negative.
            ValueError: if electric time step is larger and or equal than thermal time step.
        """

        # Rimuovere gli if.
        if self.inputs["ELECTRIC_TIME_STEP"] == None:
            self.electric_time_step = self.time_step / ELECTRIC_TIME_STEP_NUMBER
        else:
            if self.inputs["ELECTRIC_TIME_STEP"] < 0.0:
                raise ValueError[
                    f"Electric time step must be > 0.0 s; current value is: {self.inputs['ELECTRIC_TIME_STEP']=}s\n"
                ]
            if self.inputs["ELECTRIC_TIME_STEP"] > self.time_step:
                raise ValueError[
                    f"Electric time step must be < than thermal time step; current values are: {self.time_step=}s\n; {self.inputs['ELECTRIC_TIME_STEP']=}s\n"
                ]
            self.electric_time_step = self.inputs["ELECTRIC_TIME_STEP"]

        self.electric_time_end = self.time_step

    def build_right_hand_side(self, foo: np.ndarray, bar: np.ndarray, idx: int):
        """Method that builds the right hand side of the transient electric equation.

        Args:
            foo (np.ndarray): matrix.
            bar (np.ndarray): reduced electric known therm vector.
            idx (int): index array of the fixed value spatial coordinates to assign boundary conditions.
        """

        # Build right hand side
        self.electric_right_hand_side = (
            self.electric_theta * self.electric_known_term_vector
            + (1.0 - self.electric_theta) * self.electric_known_term_vector_old
            + foo @ self.electric_solution
        )

        # Reduce rhs
        if self.operations["EQUIPOTENTIAL_SURFACE_FLAG"]:
            for _, row in enumerate(self.equipotential_node_index):
                self.electric_right_hand_side[row[0]] = np.sum(
                    self.electric_right_hand_side[row]
                )

        self.electric_right_hand_side = self.electric_right_hand_side[idx] - bar

    def __electric_solution_reorganization(self):
        """Private method that reorganizes the electric solution. Specifically it:
        * extracts edge current (current along StrandComponent) and nodal potentials from electric solution array;
        * computes voltage difference along StrandComponent;
        * assignes current along StrandComponent and voltage difference along StrandComponent.
        """

        # Extract edge current (current along StrandComponent) from
        # electric solution array.
        current_along = self.electric_solution[: self.total_elements_current_carriers]

        # Extract nodal potentials from electric solution array
        self.nodal_potential = self.electric_solution[
            self.total_elements_current_carriers :
        ]

        # Compute voltage difference along StrandComponent.
        delta_voltage_along = -self.incidence_matrix @ self.nodal_potential

        # Loop to assign values to each StrandComponent.
        for ii, obj in enumerate(self.inventory["StrandComponent"].collection):
            obj.dict_Gauss_pt["current_along"] = current_along[
                ii :: self.inventory["StrandComponent"].number
            ]

            obj.dict_Gauss_pt["delta_voltage_along"] = delta_voltage_along[
                ii :: self.inventory["StrandComponent"].number
            ]

            # Compute voltage difference due to electric resistance.
            obj.dict_Gauss_pt["delta_voltag_along_R"] = (
                obj.dict_Gauss_pt["current_along"]
                * obj.dict_Gauss_pt["electric_resistance"]
                    )

    def __get_total_joule_power_electric_conductance(self):
        """Private method that evaluates total Joule power in each node of the spatial discretization associated to the electric conductance between StrandComponent objects. The method re-distribues computed values to each defined StrandComponent object."""

        # Loop to initialize total_powe_el_cond to 0 for each StrandComponent.
        for ii, obj in enumerate(self.inventory["StrandComponent"].collection):
            obj.dict_node_pt["total_power_el_cond"] = 0.0

        if self.inventory["StrandComponent"].number > 1:

            # There are more than 1 StrandComponent objects, therefore there
            # are contacts between StrandComponent objects power due to this
            # electric contacts can be evaluated as follows.
            # If there is only one StrandComponent object, total_power_el_cond
            # is equal to 0 as from initialization.

            # Compute voltage difference due to electric condutances across
            # StrandComponent objects.
            self.delta_voltage_across = np.real(
                self.contact_incidence_matrix @ self.nodal_potential
            )

            # Compute electric currrent that flows in electric conductances across
            # StrandComponent objects.
            self.current_across = np.real(
                np.conj(
                    self.electric_conductance_diag_matrix
                    @ self.contact_incidence_matrix
                    @ self.nodal_potential
                )
            )

            # Converison from instantaneous to effective values (used in sinusoidal
            # regime).
            if self.operations["ELECTRIC_SOLVER"] == STATIC_ELECTRIC_SOLVER:
                # Conversion of voltage difference across electric conductances:
                # Delta_V_across_eff = Delta_V_across / sqrt(2)
                self.delta_voltage_across = self.delta_voltage_across / np.sqrt(2.0)
                # Conversion of voltage difference across electric conductances:
                # I_across_eff = I_across / sqrt(2)
                self.current_across = self.current_across / np.sqrt(2.0)

            # Compute array with the Joule power in W due to each electric
            # conductance:
            # P_across = I_across * Delta_V_across
            # or (in sinusoidal regime):
            # P_across = I_across_eff * Delta_V_across_eff
            # Shape: (N_ct,1) with N_ct = N_c*N_n.
            # N_c = number of electric contact on a single cross section;
            # N_n = number of nodes (or number of cross sections);
            # N_ct = total number of electric contacts.
            joule_power_across = self.delta_voltage_across * self.current_across

            # Evaluate the contribution of the joule power due to conductances
            # between SolidComponent in each node of the spatial discretization.
            # The coefficient 1/2 halves the sum of the powers due to the
            # conductances referring to each node.
            # Shape : (N_nt,1) with N_nt = N_s*N_n.
            # N_s = number of strand objects (in future maybe also jacket);
            # N_n = number of nodes (or number of cross sections);
            # N_nt = total number of nodes.
            joule_power_across_node = (
                np.abs(self.contact_incidence_matrix.T) @ joule_power_across
            ) / 2

            # Loop to assign values to each StrandComponent.
            for ii, obj in enumerate(self.inventory["StrandComponent"].collection):
                obj.dict_node_pt["total_power_el_cond"] = joule_power_across_node[
                    ii :: self.inventory["StrandComponent"].number
                ]

    def __compute_voltage_sum(self):
        """Private method that evaluates the total voltage between to consecutive diagnostic spatial coordinates, assigning the computed value to the rightmost one.
        """
        # Get diagonostic spatial coordinates indexes exploiting list comprehension.
        foo = [np.max(
                    np.nonzero(
                        self.grid_features["zcoord_gauss"]
                        <= round(self.Time_save[ii], self.n_digit_z)
                    )
                )
                for ii in range(1, self.Time_save.size)]
        # Convert diagonostic spatial coordinates indexes to numpy array.
        ind_zcoord_gauss = np.array([0, *foo])

        # Loop to assign values to each StrandComponent.
        for ii, obj in enumerate(self.inventory["StrandComponent"].collection):
            # Loop to compute voltage sum for each pair of consecutive 
            # diagonostic spatial coordinates indexes. These values are saved 
            # as time evolutions at those spatial coordinates.
            for ii, idx in enumerate(ind_zcoord_gauss,1):
                if ii < ind_zcoord_gauss.shape[0]:
                    idxp = ind_zcoord_gauss[ii]
                    obj.dict_Gauss_pt["delta_voltage_along_sum"][idxp] = obj.dict_Gauss_pt["delta_voltage_along"][idx:idxp+1].sum()

    def electric_method(self):
        """Method that performs electric solution according to flag self.operations["ELECTRIC_SOLVER"]. Calls private method self.__electric_solution_reorganization to reorganize the electric solution.

        Raises:
            ValueError: if to flag self.operations["ELECTRIC_SOLVER"] user assigns not valid value.
        """

        if self.cond_num_step == 0:
            electric_steady_state_solution(self)
        else:
            self.__get_electric_time_step()
            # Always solve the electromagnetic problem as a transient problem. 
            # This is not general but edge cases are few and mostly "theoretic".
            electric_transient_solution(self)
        
        # Call method __electric_solution_reorganization: reorganize electric
        # solution and computes useful quantities used in the Joule power
        # evaluation.
        self.__electric_solution_reorganization()

        self.__compute_voltage_sum()
        # Call method __get_total_joule_power_electric_conductance to evaluate
        # the total Joule power in each node of the spatial discretization
        # associated to the electric conductance between StrandComponent
        # objects.
        self.__get_total_joule_power_electric_conductance()

    def __update_grid_features(self):
        """Private method that updates dictionary grid_features evaluating arrays delta_z, delta_z_tilde and zcoord_gauss as keys of dictionary self.grid_features. These arrays are used:
        * in method self.__assign_contact_perimeter_not_fluid_comps and self.__assign_contact_perimeter_not_fluid_only (zcoord_gauss);
        * in function step (delta_z);
        * in the joule power evaluation associated to electric resistance between StrandComponent objects (delta_z);
        * in the joule power evaluation associated to electric conductance between StrandComponent objects (delta_z_tilde);
        * in function step of module transient_solution_functions (zcoord_gauss);
        """

        # Define new key delta_z in dictionay self.grid_features (m).
        self.grid_features["delta_z"] = (
            self.grid_features["zcoord"][1:] - self.grid_features["zcoord"][:-1]
        )

        self.grid_features["dz_max"] = self.grid_features["delta_z"].max()
        self.grid_features["dz_min"] = self.grid_features["delta_z"].min()
        # Get the number of digits for rounding coordinates in order to find
        # indexes.
        self.__count_sigfigs(str(self.grid_features["dz_min"]))

        # Compute the coordintate of the Gauss point.
        self.grid_features["zcoord_gauss"] = (
            self.grid_features["zcoord"][:-1] + self.grid_features["zcoord"][1:]
        ) / 2

        # Define new key delta_z_tilde in dictionay self.grid_features (m).
        self.grid_features["delta_z_tilde"] = np.zeros(self.grid_features["N_nod"])
        self.grid_features["delta_z_tilde"][0] = (
            self.grid_features["zcoord"][1] - self.grid_features["zcoord"][0]
        ) / 2

        self.grid_features["delta_z_tilde"][1:-1] = (
            self.grid_features["zcoord"][2:] - self.grid_features["zcoord"][:-2]
        ) / 2

        self.grid_features["delta_z_tilde"][-1] = (
            self.grid_features["zcoord"][-1] - self.grid_features["zcoord"][-2]
        ) / 2

    def __count_sigfigs(self,numstr:str):
        """Private method that counts the number of significant digits.
        N.B: quite general but does not work for all cases; see 
        https://stackoverflow.com/questions/8101353/counting-significant-figures-in-python

        Args:
            numstr (str): number of which find the significant digits
        """
        decimal_norm = Decimal(numstr).normalize()
        if str(decimal_norm) == "0":
            # Deal with cases "0", "00", "00...0"
            self.n_digit_z = 0
        else:
            # Deal with cases specified in https://stackoverflow.com/questions/8101353/counting-significant-figures-in-python
            self.n_digit_z = len(Decimal(numstr).as_tuple().digits)
    
    def build_heat_source(self, simulation):
        """Method that builds heat source therms in nodal and Gauss points for 
        strand and jacket objects.

        Args:
            simulation (object): object with all information about the simulation.
        """
        # Refactor both private methods!
        self.__build_heat_source_nodal_pt(simulation)
        self.__build_heat_source_gauss_pt()

    def __build_heat_source_nodal_pt(self, simulation):
        """Private method that builds heat source therms in nodal points for 
        strand and jacket objects.

        Args:
            simulation (object): object with all information about the simulation.
        """
        
        # Alias
        interf_flag = self.dict_df_coupling["contact_perimeter_flag"]

        # Loop on StrandComponent objects.
        for strand in self.inventory["StrandComponent"].collection:
            if self.cond_num_step == 0 and strand.operations["IQFUN"] == 0:
                # Call method get_heat only once to initialize key EXTFLX of
                # dictionary dict_node_pt to zeros.
                strand.get_heat(self)
            elif strand.operations["IQFUN"] != 0:
                # Call method get_heat to evaluate external heating only if 
                # heating is on.
                strand.get_heat(self)
            
            # Call method jhtflx_new_0 to initialize JHTFLX to zeros for each 
            # conductor solid components.
            strand.jhtflx_new_0(self)
            # Evaluate joule power due to electric resistance along strand 
            # object.
            strand.get_joule_power_along(self)
            # Evaluate joule power due to electric conductance across strand 
            # object.
            strand.get_joule_power_across(self)
            # Call set_energy_counters to initialize EEXT and EJHT to zeros for 
            # each conductor solid components.
            strand.set_energy_counters(self)

        # Loop on JacketComponents objects.
        for rr, jacket in enumerate(self.inventory["JacketComponent"].collection):
            if self.cond_num_step == 0 and jacket.operations["IQFUN"] == 0:
                # Call method get_heat only once to initialize key EXTFLX of 
                # dictionary dict_node_pt to zeros.
                jacket.get_heat(self)
            elif jacket.operations["IQFUN"] != 0:
                # Call method get_heat to evaluate external heating only if 
                # heating is on.
                jacket.get_heat(self)

            # Call method jhtflx_new_0 to initialize JHTFLX to zeros for each 
            # conductor solid components.
            jacket.jhtflx_new_0(self)
            # Evaluate joule power due to electric resistance along jacket 
            # object.
            jacket.get_joule_power_along(self)
            # Evaluate joule power due to electric conductance across jacket 
            # object.
            jacket.get_joule_power_across(self)
            # Call set_energy_counters to initialize EEXT and EJHT to zeros for 
            # each conductor solid components.
            jacket.set_energy_counters(self)
            if (
                abs(interf_flag.at[
                    simulation.environment.KIND, jacket.identifier
                ]) == 1
            ):
                # Evaluate the external heat by radiation in nodal points.
                jacket._radiative_source_therm_env(self, simulation.environment)
            # End if abb(interf_flag)
            for _, jacket_c in enumerate(
                self.inventory["JacketComponent"].collection[rr + 1 :]
            ):
                if (
                    abs(
                        self.dict_df_coupling["HTC_choice"].at[
                            jacket.identifier, jacket_c.identifier
                        ]
                    )
                    == 3
                ):
                    # Evaluate the inner heat exchange by radiation in nodal 
                    # points.
                    jacket._radiative_heat_exc_inner(self, jacket_c)
                    jacket_c._radiative_heat_exc_inner(self, jacket)
                # End if abs.
            # End for jacket_c.
        # End for rr.

    def __build_heat_source_gauss_pt(self):
        """Private method that builds heat source therms in Gauss points for 
        strand and jacket objects."""

        # Loop on StrandComponent objects.
        for strand in self.inventory["StrandComponent"].collection:

            strand.dict_Gauss_pt["Q1"] = (
                strand.dict_node_pt["JHTFLX"][:-1]
                + strand.dict_node_pt["EXTFLX"][:-1]
                + strand.dict_node_pt["total_linear_power_el_cond"][:-1]
                + strand.dict_Gauss_pt["linear_power_el_resistance"]
            )

            strand.dict_Gauss_pt["Q2"] = (
                strand.dict_node_pt["JHTFLX"][1:]
                + strand.dict_node_pt["EXTFLX"][1:]
                + strand.dict_node_pt["total_linear_power_el_cond"][1:]
                + strand.dict_Gauss_pt["linear_power_el_resistance"]
            )
        
        # Loop on JacketComponents objects.
        for rr, jacket in enumerate(self.inventory["JacketComponent"].collection):

            jacket.dict_Gauss_pt["Q1"] = (
                jacket.dict_node_pt["JHTFLX"][:-1] + jacket.dict_node_pt["EXTFLX"][:-1]
            )
            jacket.dict_Gauss_pt["Q2"] = (
                jacket.dict_node_pt["JHTFLX"][1:] + jacket.dict_node_pt["EXTFLX"][1:]
            )
            # Add the radiative heat contribution with the environment.
            jacket.dict_Gauss_pt["Q1"] = (
                jacket.dict_Gauss_pt["Q1"] + jacket.radiative_heat_env[:-1]
            )
            jacket.dict_Gauss_pt["Q2"] = (
                jacket.dict_Gauss_pt["Q2"] + jacket.radiative_heat_env[1:]
            )

        # Separate nested loop is needed in order to define quantities 
        # dict_Gauss_pt["Q1"] and dict_Gauss_pt["Q2"] for component jacket_c in 
        # the previous loop, otherwise a key error will be raised.
        for rr, jacket in enumerate(self.inventory["JacketComponent"].collection):
            # Nested loop jacket - jacket.
            for _, jacket_c in enumerate(
                self.inventory["JacketComponent"].collection[rr + 1 :]
            ):
                key = f"{jacket.identifier}_{jacket_c.identifier}"
                # Add the radiative heat contribution between inner surface of the enclosure and inner jackets.
                jacket.dict_Gauss_pt["Q1"] = (
                    jacket.dict_Gauss_pt["Q1"] + jacket.radiative_heat_inn[key][:-1]
                )
                jacket.dict_Gauss_pt["Q2"] = (
                    jacket.dict_Gauss_pt["Q2"] + jacket.radiative_heat_inn[key][1:]
                )

                jacket_c.dict_Gauss_pt["Q1"] = (
                    jacket_c.dict_Gauss_pt["Q1"] + jacket_c.radiative_heat_inn[key][:-1]
                )
                jacket_c.dict_Gauss_pt["Q2"] = (
                    jacket_c.dict_Gauss_pt["Q2"] + jacket_c.radiative_heat_inn[key][1:]
                )
            # End for jacket_c.
        # end for jacket.
    
    def operating_conditions_th_initialization(self,simulation):
        """Method that evaluates thermal hydraulic (th) operating conditions in both nodal and in Gauss points.
        To be called at initialization only since it avoids a second call to method self.__update_grid_features, which is already called in method self.__init__.
        """

        self.get_transp_coeff(simulation)
        self.__eval_gauss_point_th(simulation)

    def operating_conditions_th(self,simulation):
        """Method that evaluates thermal hydraulic (th) operating conditions also in Gauss points."""

        self.__update_grid_features()
        self.get_transp_coeff(simulation)

        self.__eval_gauss_point_th(simulation)

    def operating_conditions_em(self):
        """Method that evaluates electromagnetic (em) operating conditions also in Gauss points."""

        for strand in self.inventory["StrandComponent"].collection:
            strand.get_current(self)
            strand.get_magnetic_field(self)
            strand.get_magnetic_field_gradient(self)
            
            # only for StrandMixedComponent and StackComponent objects
            if not isinstance(strand, StrandStabilizerComponent):
                if strand.inputs["superconducting_material"] == "Nb3Sn":
                    if self.cond_el_num_step <= 1:
                        # Evaluate strain only at initialization (0) and at 
                        # the first electric time step (1).
                        strand.get_eps(self)
                strand.get_superconductor_critical_prop(self)
                if (
                    strand.operations["TCS_EVALUATION"] == False
                    and self.cond_num_step == 0
                ):
                    strand.get_tcs()
                elif strand.operations["TCS_EVALUATION"] == True:
                    if self.cond_el_num_step <= 1:
                        # Evaluate current sharing temperature only at 
                        # initialization (0) and at the first electriC time 
                        # step (1).
                        strand.get_tcs()

        for jacket in self.inventory["JacketComponent"].collection:
            jacket.get_current(self)
            jacket.get_magnetic_field(self)

        self.__eval_gauss_point_em()

    def __eval_gauss_point_th(self, simulation):
        """
        Method that evaluates temperatures and transport coefficients at the Gauss point, i.e at the centre of the element.
        N.B. Fluid component properties are evaluated calling method get_transp_coeff.
        """

        # JacketComponent
        for jacket in self.inventory["JacketComponent"].collection:
            jacket.dict_Gauss_pt["temperature"] = (
                np.abs(
                    jacket.dict_node_pt["temperature"][:-1]
                    + jacket.dict_node_pt["temperature"][1:]
                )
                / 2.0
            )
        # end for rr.

        # StrandComponent
        for strand in self.inventory["StrandComponent"].collection:
            strand.dict_Gauss_pt["temperature"] = (
                np.abs(
                    strand.dict_node_pt["temperature"][:-1]
                    + strand.dict_node_pt["temperature"][1:]
                )
                / 2.0
            )

        # call method Get_transp_coeff to evaluate transport properties (heat
        # transfer coefficient and friction factor) in each Gauss point
        self.get_transp_coeff(simulation, flag_nodal=False)

    def __eval_gauss_point_em(self):
        """
        Private method that evaluates material properties and coefficients at the Gauss point, i.e at the centre of the element.
        """

        # JacketComponent
        for jacket in self.inventory["JacketComponent"].collection:
            
            jacket.get_magnetic_field(self, nodal=False)
            if self.cond_el_num_step <= 1:
                # Evaluate properties only at initialization (0) and at 
                # the first electric time step (1).
                jacket.eval_sol_comp_properties(self.inventory, nodal=False)
            else:
                # Update only electrical resistivity at each electri time step.
                jacket.dict_Gauss_pt["total_electrical_resistivity"] = jacket.jacket_electrical_resistivity(
                        jacket.dict_Gauss_pt
                    )

        # StrandComponent
        for strand in self.inventory["StrandComponent"].collection:
            strand.get_magnetic_field(self, nodal=False)

            # call method get_magnetic_field_gradient for each StrandComponent object
            strand.get_magnetic_field_gradient(self, nodal=False)
            # only for StrandMixedComponent and StackComponent objects
            if not isinstance(strand, StrandStabilizerComponent):
                if strand.inputs["superconducting_material"] == "Nb3Sn":
                    if self.cond_el_num_step <= 1:
                        # Evaluate strain only at initialization (0) and at 
                        # the first electric time step (1).
                        strand.get_eps(self,nodal=False)
                strand.get_superconductor_critical_prop(self,nodal=False)
                if (
                    strand.operations["TCS_EVALUATION"] == False
                    and self.cond_num_step == 0
                ):
                    strand.get_tcs(nodal=False)
                elif strand.operations["TCS_EVALUATION"] == True:
                    if self.cond_el_num_step <= 1:
                        # Evaluate current sharing temperature only at 
                        # initialization (0) and at the first electric time 
                        # step (1)
                        strand.get_tcs(nodal=False)
            # end if strand.name != self.inventory["StrandStabilizerComponent"].
            # name.
            # Evaluate SolidComponent properties
            if self.cond_el_num_step <= 1:
                # Evaluate properties only at initialization (0) and at 
                # the first electric time step (1).
                strand.eval_sol_comp_properties(self.inventory, nodal=False)
            else:
                # Update only electrical resistivity (stabilizer) at each 
                # electric time step.
                if isinstance(strand,StrandMixedComponent):
                    strand.dict_Gauss_pt["electrical_resistivity_stabilizer"] = strand.strand_electrical_resistivity_not_sc(
                            strand.dict_Gauss_pt
                        )
                elif isinstance(strand, StrandStabilizerComponent):
                    strand.dict_Gauss_pt["electrical_resistivity_stabilizer"] = strand.strand_electrical_resistivity(
                            strand.dict_Gauss_pt
                        )

    def post_processing(self, simulation):

        # bozza della funzione Post_processing

        # Loop on SolidComponent to evaluate the total final energy of \
        # SolidComponent, used to check the imposition of SolidComponent \
        # temperature initial spatial distribution (cdp, 12/2020)
        for s_comp in self.inventory["SolidComponent"].collection:
            self.E_sol_fin = self.E_sol_fin + s_comp.inputs["CROSSECTION"] * np.sum(
                (
                    self.grid_features["zcoord"][1 : self.grid_features["N_nod"]]
                    - self.grid_features["zcoord"][0:-1]
                )
                * s_comp.dict_Gauss_pt["total_density"]
                * s_comp.dict_Gauss_pt["total_isobaric_specific_heat"]
                * s_comp.dict_Gauss_pt["temperature"]
            )
            if s_comp.name != "Z_JACKET":
                self.E_str_fin = self.E_str_fin + s_comp.inputs["CROSSECTION"] * np.sum(
                    (
                        self.grid_features["zcoord"][1 : self.grid_features["N_nod"]]
                        - self.grid_features["zcoord"][0:-1]
                    )
                    * s_comp.dict_Gauss_pt["total_density"]
                    * s_comp.dict_Gauss_pt["total_isobaric_specific_heat"]
                    * s_comp.dict_Gauss_pt["temperature"]
                )
            else:
                self.E_jk_fin = self.E_jk_fin + s_comp.inputs["CROSSECTION"] * np.sum(
                    (
                        self.grid_features["zcoord"][1 : self.grid_features["N_nod"]]
                        - self.grid_features["zcoord"][0:-1]
                    )
                    * s_comp.dict_Gauss_pt["total_density"]
                    * s_comp.dict_Gauss_pt["total_isobaric_specific_heat"]
                    * s_comp.dict_Gauss_pt["temperature"]
                )
        # end for s_comp (cdp, 12/2020)

        # call method Mass_Energy_balance to get data for the space convergence \
        # (cdp, 09/2020)
        self.mass_energy_balance(simulation)
        # call function Save_convergence_data to save solution spatial \
        # distribution at TEND, together with the mass and energy balance results, \
        # to make the space convergence analisys (cdp, 12/2020)
        save_convergence_data(self, simulation.dict_path["Space_conv_output_dir"])
        # call function Save_convergence_data to save solution spatial \
        # distribution at TEND, together with the mass and energy balance results \
        # to make the time convergence analisys (cdp, 12/2020)
        save_convergence_data(
            self,
            simulation.dict_path["Time_conv_output_dir"],
            abs(simulation.n_digit_time),
            space_conv=False,
        )

    # end Post_processing

    def get_transp_coeff(self, simulation, flag_nodal=True):  # (cpd 06/2020)

        """
    Method that compute channels friction factor and heat transfer coefficient \
    between channels, channel and solid component, solid components, both in \
    nodal and Gauss points. (cpd 06/2020)
    """
        # Properties evaluation in each nodal point (cdp, 07/2020)
        if flag_nodal:
            self.dict_node_pt = self.eval_transp_coeff(simulation, self.dict_node_pt)
        # Properties evaluation in each Gauss point (cdp, 07/2020)
        elif flag_nodal == False:
            self.dict_Gauss_pt = self.eval_transp_coeff(
                simulation, self.dict_Gauss_pt, flag_nodal=False
            )

    # end method Get_transp_coeff (cdp, 09/2020)

    def eval_transp_coeff(self, simulation, dict_dummy, flag_nodal=True):

        """
        Method that actually computes channels friction factor and heat 
        transfer coefficient between channels, channel and solid component, 
        solid components, both in nodal and Gauss points.
        """

        # Alias
        interf_flag = self.dict_df_coupling["contact_perimeter_flag"]
        # loop to evaluate htc_steady for each channel according to its geometry (cpd 06/2020)
        for fluid_comp in self.inventory["FluidComponent"].collection:

            # Evaluate coolant properties in nodal points.
            fluid_comp.coolant._eval_properties_nodal_gauss(
                self, simulation.fluid_prop_aliases, flag_nodal
            )

            # Define dictionary to select nodal or gauss properties according to the value of flag_nodal.0
            dict_dummy_chan = {
                True: fluid_comp.coolant.dict_node_pt,
                False: fluid_comp.coolant.dict_Gauss_pt,
            }
            # Evaluate steady state heat transfer coefficient for each channel.
            fluid_comp.channel.eval_steady_state_htc(
                dict_dummy_chan[flag_nodal], nodal=flag_nodal
            )
            # Evaluate total friction factor for each channel.
            fluid_comp.channel.eval_friction_factor(
                dict_dummy_chan[flag_nodal]["Reynolds"], nodal=flag_nodal
            )
        # end for loop fluid_comp.

        # htc dummy dictionary and its sub-dictionary declaration (cpd 09/2020)
        dict_dummy["HTC"] = dict()
        dict_dummy["HTC"]["ch_ch"] = dict()
        dict_dummy["HTC"]["ch_ch"]["Open"] = dict()
        dict_dummy["HTC"]["ch_ch"]["Close"] = dict()
        dict_dummy["HTC"]["ch_sol"] = dict()
        dict_dummy["HTC"]["sol_sol"] = dict()
        dict_dummy["HTC"]["sol_sol"]["cond"] = dict()
        dict_dummy["HTC"]["sol_sol"]["rad"] = dict()
        dict_dummy["HTC"]["env_sol"] = dict()
        # Counters to check the number of the different possible kinds of interfaces (cdp, 09/2020)
        htc_len = 0
        for rr, fluid_comp_r in enumerate(self.inventory["FluidComponent"].collection):
            dict_dummy_chan_r = {
                True: fluid_comp_r.coolant.dict_node_pt,
                False: fluid_comp_r.coolant.dict_Gauss_pt,
            }
            # Read the submatrix containing information about channel - solid objects iterfaces (cdp, 06/2020)
            # nested loop on channel - solid objects (cpd 06/2020)
            for s_comp in self.inventory["SolidComponent"].collection:
                dict_dummy_comp = {
                    True: s_comp.dict_node_pt,
                    False: s_comp.dict_Gauss_pt,
                }
                # Multiplier used in both cases (positive and negative flag).
                mlt = self.dict_df_coupling["HTC_multiplier"].at[
                            fluid_comp_r.identifier, s_comp.identifier
                        ]
                # Rationale: compute dictionary vaules only if there is an interface \
                # (cdp, 09/2020)
                if (
                    abs(interf_flag.at[
                            fluid_comp_r.identifier, s_comp.identifier
                        ]
                    ) == 1
                ):
                    htc_len = htc_len + 1
                    # new channel-solid interface (cdp, 09/2020)
                    htc_Kapitza = np.zeros(
                        dict_dummy_chan_r[flag_nodal]["temperature"].shape
                    )
                    htc_transient = np.zeros(
                        dict_dummy_chan_r[flag_nodal]["temperature"].shape
                    )
                    htc_full_transient = np.zeros(
                        dict_dummy_chan_r[flag_nodal]["temperature"].shape
                    )
                    if (
                        self.dict_df_coupling["HTC_choice"].at[
                            fluid_comp_r.identifier, s_comp.identifier
                        ]
                        == 2
                    ):
                        htc_Kapitza = (
                            200.0
                            * (
                                dict_dummy_comp[flag_nodal]["temperature"]
                                + dict_dummy_chan_r[flag_nodal]["temperature"]
                            )
                            * (
                                dict_dummy_comp[flag_nodal]["temperature"] ** 2
                                + dict_dummy_chan_r[flag_nodal]["temperature"] ** 2
                            )
                        )
                        if self.cond_time[-1] > s_comp.operations["TQBEG"]:
                            # implementation fully correct only for fully implicit method \
                            # (cdp, 06/2020)
                            htc_transient = np.sqrt(
                                (
                                    dict_dummy_chan_r[flag_nodal][
                                        "total_thermal_conductivity"
                                    ]
                                    * dict_dummy_chan_r[flag_nodal]["total_density"]
                                    * dict_dummy_chan_r[flag_nodal][
                                        "total_isobaric_specific_heat"
                                    ]
                                )
                                / (
                                    np.pi
                                    * (self.cond_time[-1] - s_comp.operations["TQBEG"])
                                )
                            )
                            htc_full_transient = (htc_Kapitza * htc_transient) / (
                                htc_Kapitza + htc_transient
                            )
                        # Assign to the HTC key of dictionary dict_dummy the dictionary whit the information about heat trasfer coefficient betweent channel fluid_comp_r and solid s_comp. Interface identification is given by the key name itself: f"{fluid_comp_r.identifier}_{s_comp.identifier}". This inner dictionary consists of a single key-value pair. (cdp, 07/2020)
                        dict_dummy["HTC"]["ch_sol"][
                            self.dict_topology["ch_sol"][fluid_comp_r.identifier][
                                s_comp.identifier
                            ]
                        ] = np.maximum(
                            fluid_comp_r.channel.dict_htc_steady[flag_nodal] * mlt,
                            htc_full_transient,
                        )
                    elif (
                        self.dict_df_coupling["HTC_choice"].at[
                            fluid_comp_r.identifier, s_comp.identifier
                        ]
                        == -2
                    ):
                        dict_dummy["HTC"]["ch_sol"][
                            self.dict_topology["ch_sol"][fluid_comp_r.identifier][
                                s_comp.identifier
                            ]
                        ] = self.dict_df_coupling["contact_HTC"].at[
                            fluid_comp_r.identifier, s_comp.identifier
                        ] * np.ones(
                            dict_dummy_chan_r[flag_nodal]["temperature"].shape
                        ) * mlt
            # end loop on SolidComponent
            # nested loop on channel - channel objects (cdp, 06/2020)
            for _, fluid_comp_c in enumerate(
                self.inventory["FluidComponent"].collection[rr + 1 :]
            ):
                dict_dummy_chan_c = {
                    True: fluid_comp_c.coolant.dict_node_pt,
                    False: fluid_comp_c.coolant.dict_Gauss_pt,
                }
                # Multiplier used in both cases (positive and negative flag).
                mlt = self.dict_df_coupling["HTC_multiplier"].at[
                            fluid_comp_r.identifier, fluid_comp_c.identifier
                        ]
                if (
                    abs(interf_flag.at[
                            fluid_comp_r.identifier, fluid_comp_c.identifier
                        ]
                    ) == 1
                ):
                    # new channel-channel interface (cdp, 09/2020)
                    htc_len = htc_len + 1
                    # Construct interface name: it can be found also in dict_topology["ch_ch"] but a search in dictionaties "Hydraulic_parallel" and "Thermal_contact" should be performed, which makes thinks not easy to do; it is simpler to construct interface names combining channels identifier (cdp, 09/2020)
                    interface_name = (
                        f"{fluid_comp_r.identifier}_{fluid_comp_c.identifier}"
                    )
                    if (
                        self.dict_df_coupling["HTC_choice"].at[
                            fluid_comp_r.identifier, fluid_comp_c.identifier
                        ]
                        == 2
                    ):
                        # dummy
                        htc1 = fluid_comp_r.channel.dict_htc_steady[flag_nodal]
                        # dummy
                        htc2 = fluid_comp_c.channel.dict_htc_steady[flag_nodal]

                        dict_dummy["HTC"]["ch_ch"]["Open"][interface_name] = (
                            mlt * htc1 * htc2 / (htc1 + htc2)
                        )
                        cond_interface = thermal_conductivity_ss(
                            (
                                dict_dummy_chan_r[flag_nodal]["temperature"]
                                + dict_dummy_chan_c[flag_nodal]["temperature"]
                            )
                            / 2
                        )
                        R_wall = (
                            self.dict_df_coupling["interf_thickness"].at[
                                fluid_comp_r.identifier, fluid_comp_c.identifier
                            ]
                            / cond_interface
                        )
                        dict_dummy["HTC"]["ch_ch"]["Close"][interface_name] = mlt / (
                            1 / htc1 + 1 / htc2 + R_wall
                        )
                    elif (
                        self.dict_df_coupling["HTC_choice"].at[
                            fluid_comp_r.identifier, fluid_comp_c.identifier
                        ]
                        == -2
                    ):
                        # in this case it is assumed that both open and close hct have the same value (cdp, 07/2020)
                        dict_dummy["HTC"]["ch_ch"]["Open"][
                            interface_name
                        ] = self.dict_df_coupling["contact_HTC"].at[
                            fluid_comp_r.identifier, fluid_comp_c.identifier
                        ] * np.ones(
                            dict_dummy_chan_r[flag_nodal]["temperature"].shape
                        ) * mlt
                        dict_dummy["HTC"]["ch_ch"]["Close"][
                            interface_name
                        ] = self.dict_df_coupling["contact_HTC"].at[
                            fluid_comp_r.identifier, fluid_comp_c.identifier
                        ] * np.ones(
                            dict_dummy_chan_r[flag_nodal]["temperature"].shape
                        ) * mlt
            # end for loop cc
        # end for loop rr
        # nested loop on solid - solid objects (cdp, 06/2020)
        for rr, s_comp_r in enumerate(self.inventory["SolidComponent"].collection):
            dict_dummy_comp_r = {
                True: s_comp_r.dict_node_pt,
                False: s_comp_r.dict_Gauss_pt,
            }
            for _, s_comp_c in enumerate(
                self.inventory["SolidComponent"].collection[rr + 1 :]
            ):
                dict_dummy_comp_c = {
                    True: s_comp_c.dict_node_pt,
                    False: s_comp_c.dict_Gauss_pt,
                }
                # Multiplier used in both cases (positive and negative flag).
                mlt = self.dict_df_coupling["HTC_multiplier"].at[
                            s_comp_r.identifier, s_comp_c.identifier
                        ]
                if (
                    abs(interf_flag.at[
                            s_comp_r.identifier, s_comp_c.identifier
                        ]
                    ) == 1
                ):
                    dict_dummy["HTC"]["sol_sol"]["cond"][
                        self.dict_topology["sol_sol"][s_comp_r.identifier][
                            s_comp_c.identifier
                        ]
                    ] = np.zeros(dict_dummy_comp_r[flag_nodal]["temperature"].shape)
                    dict_dummy["HTC"]["sol_sol"]["rad"][
                        self.dict_topology["sol_sol"][s_comp_r.identifier][
                            s_comp_c.identifier
                        ]
                    ] = np.zeros(dict_dummy_comp_r[flag_nodal]["temperature"].shape)

                    # New solid-solid interface (cdp, 09/2020)
                    htc_len = htc_len + 1
                    if (
                        self.dict_df_coupling["HTC_choice"].at[
                            s_comp_r.identifier, s_comp_c.identifier
                        ]
                        == 1
                    ):
                        # Thermal contact.
                        htc_solid = 500.0
                        
                        dict_dummy["HTC"]["sol_sol"]["cond"][
                            self.dict_topology["sol_sol"][s_comp_r.identifier][
                                s_comp_c.identifier
                            ]
                        ] = (
                            mlt
                            * htc_solid
                            * np.ones(
                                dict_dummy_comp_r[flag_nodal]["temperature"].shape
                            )
                        )
                    elif (
                        self.dict_df_coupling["HTC_choice"].at[
                            s_comp_r.identifier, s_comp_c.identifier
                        ]
                        == -1
                    ):
                        # Thermal contact.
                        dict_dummy["HTC"]["sol_sol"]["cond"][
                            self.dict_topology["sol_sol"][s_comp_r.identifier][
                                s_comp_c.identifier
                            ]
                        ] = self.dict_df_coupling["contact_HTC"].at[
                            s_comp_r.identifier, s_comp_c.identifier
                        ] * np.ones(
                            dict_dummy_comp_r[flag_nodal]["temperature"].shape
                        ) * mlt
                    elif (
                        self.dict_df_coupling["HTC_choice"].at[
                            s_comp_r.identifier, s_comp_c.identifier
                        ]
                        == 3
                    ):
                        # Radiative heat transfer.
                        if (
                            s_comp_r.inputs["Emissivity"] > 0.0
                            and s_comp_c.inputs["Emissivity"] > 0.0
                            and self.dict_df_coupling["view_factors"].at[
                                s_comp_r.identifier, s_comp_c.identifier
                            ]
                            > 0.0
                        ):
                            # Evaluate the radiative heat transfer coefficient (assume that sr is the inner convex surface and sc the outer not convex surface, only in the comment below):
                            # A_sr*sigma*(T_sr^2 + T_sc^2)*(T_sr + T_sc)/((1 - emissivity_sr)/emissivity_sr + 1/F_sr_cs + (1 - emissivity_sc)/emissivity_sc*(A_sr/A_sc))
                            # Reciprocal of the view factor.
                            view_factor_rec = np.reciprocal(
                                self.dict_df_coupling["view_factors"].at[
                                    s_comp_r.identifier, s_comp_c.identifier
                                ]
                            )
                            if (
                                s_comp_r.inputs["Outer_perimeter"]
                                < s_comp_c.inputs["Inner_perimeter"]
                            ):
                                # Set the contact perimeter to the correct value (overwrite the value assigned in input file conductor_coupling.xlsx)
                                self.dict_df_coupling["contact_perimeter"].at[
                                    s_comp_r.identifier, s_comp_c.identifier
                                ] = s_comp_r.inputs["Outer_perimeter"]
                                dict_dummy["HTC"]["sol_sol"]["rad"][
                                    self.dict_topology["sol_sol"][s_comp_r.identifier][
                                        s_comp_c.identifier
                                    ]
                                ] = self._inner_radiative_htc(
                                    s_comp_r,
                                    s_comp_c,
                                    dict_dummy_comp_r[flag_nodal]["temperature"],
                                    dict_dummy_comp_c[flag_nodal]["temperature"],
                                    view_factor_rec,
                                )
                            elif (
                                s_comp_c.inputs["Outer_perimeter"]
                                < s_comp_r.inputs["Inner_perimeter"]
                            ):
                                # Set the contact perimeter to the correct value (overwrite the value assigned in input file conductor_coupling.xlsx)
                                self.dict_df_coupling["contact_perimeter"].at[
                                    s_comp_r.identifier, s_comp_c.identifier
                                ] = s_comp_c.inputs["Outer_perimeter"]
                                dict_dummy["HTC"]["sol_sol"]["rad"][
                                    self.dict_topology["sol_sol"][s_comp_r.identifier][
                                        s_comp_c.identifier
                                    ]
                                ] = self._inner_radiative_htc(
                                    s_comp_c,
                                    s_comp_r,
                                    dict_dummy_comp_c[flag_nodal]["temperature"],
                                    dict_dummy_comp_r[flag_nodal]["temperature"],
                                    view_factor_rec,
                                ) * mlt
                            # End if s_comp_r.inputs["Outer_perimeter"].
                        # End if emissivity.
                    elif (
                        self.dict_df_coupling["HTC_choice"].at[
                            s_comp_r.identifier, s_comp_c.identifier
                        ]
                        == -3
                    ):
                        # Radiative heat transfer from sheet contact_HTC of file conductor_coupling.xlsx.
                        dict_dummy["HTC"]["sol_sol"]["rad"][
                            self.dict_topology["sol_sol"][s_comp_r.identifier][
                                s_comp_c.identifier
                            ]
                        ] = self.dict_df_coupling["contact_HTC"].at[
                            s_comp_r.identifier, s_comp_c.identifier
                        ] * np.ones(
                            dict_dummy_comp_r[flag_nodal]["temperature"].shape
                        ) * mlt
                    # End if self.dict_df_coupling["HTC_choice"].at[s_comp_r.identifier, s_comp_c.identifier]
            # end for loop cc

            key = f"{simulation.environment.KIND}_{s_comp_r.identifier}"
            if (
                abs(interf_flag.at[
                        simulation.environment.KIND, s_comp_r.identifier
                    ]
                ) == 1
            ):
                # New environment-solid interface
                htc_len = htc_len + 1
                if (
                    s_comp_r.inputs["Jacket_kind"] == "outer_insulation"
                    or s_comp_r.inputs["Jacket_kind"] == "whole_enclosure"
                ):
                    # Heat transfer with the environment by radiation and/or by convection.
                    # Initialize dictionary.
                    dict_dummy["HTC"]["env_sol"][key] = dict(
                        conv=np.zeros(
                            dict_dummy_comp_r[flag_nodal]["temperature"].size
                        ),
                        rad=np.zeros(dict_dummy_comp_r[flag_nodal]["temperature"].size),
                    )
                    if (
                        self.dict_df_coupling["HTC_choice"].at[
                            simulation.environment.KIND, s_comp_r.identifier
                        ]
                        == 2
                    ):
                        # Heat transfer by convection: heat transfer coefficient evaluated from air properties.
                        if self.inputs["Is_rectangular"]:
                            # Rectangular conductor
                            dict_dummy["HTC"]["env_sol"][key]["conv"] = dict(
                                side=np.zeros(
                                    dict_dummy_comp_r[flag_nodal]["temperature"].size
                                ),
                                bottom=np.zeros(
                                    dict_dummy_comp_r[flag_nodal]["temperature"].size
                                ),
                                top=np.zeros(
                                    dict_dummy_comp_r[flag_nodal]["temperature"].size
                                ),
                            )
                            # Evaluate side bottom and top surfaces htc.
                            (
                                dict_dummy["HTC"]["env_sol"][key]["conv"]["side"],
                                dict_dummy["HTC"]["env_sol"][key]["conv"]["bottom"],
                                dict_dummy["HTC"]["env_sol"][key]["conv"]["top"],
                            ) = simulation.environment.eval_heat_transfer_coefficient(
                                self, dict_dummy_comp_r[flag_nodal]["temperature"]
                            ) * mlt
                        else:
                            # Circular conductor
                            if flag_nodal == False:
                                # Compute only in gauss node to avoid error
                                dict_dummy["HTC"]["env_sol"][key]["conv"] = (
                                    simulation.environment.eval_heat_transfer_coefficient(
                                        self,
                                        dict_dummy_comp_r[flag_nodal]["temperature"],
                                    )
                                    * self.inputs["Phi_conv"]
                                )
                    elif (
                        self.dict_df_coupling["HTC_choice"].at[
                            simulation.environment.KIND, s_comp_r.identifier
                        ]
                        == -2
                    ):
                        # Heat transfer by convection: from sheet contact_HTC of file conductor_coupling.xlsx.
                        dict_dummy["HTC"]["env_sol"][key]["conv"] = (
                            self.dict_df_coupling["contact_HTC"].at[
                                simulation.environment.KIND, s_comp_r.identifier
                            ]
                            * self.inputs["Phi_conv"]
                            * np.ones(
                                dict_dummy_comp_r[flag_nodal]["temperature"].shape
                            ) * mlt
                        )
                    elif (
                        self.dict_df_coupling["HTC_choice"].at[
                            simulation.environment.KIND, s_comp_r.identifier
                        ]
                        == 3
                    ):
                        # Heat transfer by radiation.
                        # Evaluate radiative heat transfer coefficient invoking method eval_weighted_radiative_htc.
                        dict_dummy["HTC"]["env_sol"][key][
                            "rad"
                        ] = self.eval_weighted_radiative_htc(
                            simulation,
                            s_comp_r,
                            dict_dummy_comp_r[flag_nodal]["temperature"],
                        ) * mlt
                    elif (
                        self.dict_df_coupling["HTC_choice"].at[
                            simulation.environment.KIND, s_comp_r.identifier
                        ]
                        == -3
                    ):
                        # Heat transfer by radiation: from sheet contact_HTC of file conductor_coupling.xlsx.
                        dict_dummy["HTC"]["env_sol"][key]["rad"] = (
                            self.dict_df_coupling["contact_HTC"].at[
                                simulation.environment.KIND, s_comp_r.identifier
                            ]
                            * self.inputs["Phi_rad"]
                            * np.ones(
                                dict_dummy_comp_r[flag_nodal]["temperature"].shape
                            ) * mlt
                        )
                    elif (
                        self.dict_df_coupling["HTC_choice"].at[
                            simulation.environment.KIND, s_comp_r.identifier
                        ]
                        == 4
                    ):
                        # Heat transfer by radiation and free convection: code evaluation
                        dict_dummy["HTC"]["env_sol"][key]["conv"] = (
                            simulation.environment.eval_heat_transfer_coefficient(
                                self, dict_dummy_comp_r[flag_nodal]["temperature"]
                            )
                            * self.inputs["Phi_conv"] * mlt
                        )

                        dict_dummy["HTC"]["env_sol"][key][
                            "rad"
                        ] = self.eval_weighted_radiative_htc(
                            simulation,
                            s_comp_r,
                            dict_dummy_comp_r[flag_nodal]["temperature"],
                        ) * mlt
                    elif (
                        self.dict_df_coupling["HTC_choice"].at[
                            simulation.environment.KIND, s_comp_r.identifier
                        ]
                        == -4
                    ):
                        # Heat transfer by radiation and free convection: from sheet contact_HTC of file conductor_coupling.xlsx.
                        # Questo va ragionato meglio: secondo me devo trovare il modo di distinguere i due contributi anche se in input sono dati come valore complessivo.
                        dict_dummy["HTC"]["env_sol"][key][
                            "conv"
                        ] = self.dict_df_coupling["contact_HTC"].at[
                            simulation.environment.KIND, s_comp_r.identifier
                        ] * np.ones(
                            dict_dummy_comp_r[flag_nodal]["temperature"].shape
                        ) * mlt
                    # End if self.dict_df_coupling["HTC_choice"].at[simulation.environment.KIND, s_comp_r.identifier]
                else:
                    # Raise error
                    raise os.error(
                        f"JacketComponent of kind {s_comp_r.inputs['Jacket_kind']} can not exchange heat by radiation and/or convection with the environment.\n"
                    )
                # End if s_comp_r.inputs["Jacket_kind"]
            # End if abs(intef_flag.at[simulation.environment.KIND, s_comp_r.identifier])
        # end for loop rr

        # Check number of evaluated interface htc (cdp, 06/2020)

        # Evaluate total number of user defined interfaces.
        total_interf_number = np.abs(interf_flag.to_numpy()).sum()
        if htc_len != total_interf_number:
            raise ValueError(
                f"ERROR!!! Number of interface and number of evaluated interface htc mismatch: {total_interf_number} != {htc_len}"
            )

        return dict_dummy

    # end method Eval_transp_coeff (cdp, 09/2020)

    def eval_weighted_radiative_htc(self, simulation, s_comp, temperature_jk):
        # Evaluate the weighted external radiative heat transfer coefficient: htc_rad*Phi_rad
        if s_comp.inputs["Emissivity"] <= 0.0:
            # Set the radiative heat transfer coefficient to zero: no radiative heat transfer.
            return np.zeros(temperature_jk.shape)  # W/m^2/K
        else:
            return (
                constants.Stefan_Boltzmann
                * s_comp.inputs["Emissivity"]
                * (
                    temperature_jk ** 2
                    + simulation.environment.inputs["Temperature"] ** 2
                )
                * (temperature_jk + simulation.environment.inputs["Temperature"])
                * self.inputs["Phi_rad"]
            )  # W/m^2/K
        # End if s_comp_r.inputs["Emissivity"] <= 0.

    # End method eval_weighted_radiative_htc.

    def _inner_radiative_htc(self, jk_i, jk_j, temp_i, temp_j, view_factor_rec):

        return (
            constants.Stefan_Boltzmann
            * (temp_i ** 2 + temp_j ** 2)
            * (temp_i + temp_j)
            / (
                (1.0 - jk_i.inputs["Emissivity"]) / jk_i.inputs["Emissivity"]
                + view_factor_rec
                + (1.0 - jk_j.inputs["Emissivity"])
                / jk_j.inputs["Emissivity"]
                * (jk_i.inputs["Outer_perimeter"] / jk_j.inputs["Inner_perimeter"])
            )
        )  # W/m^2/K

    # End method _inner_radiative_htc

    def aprior(self):

        # C * COMPUTE BY BISECTION THE PRODUCT LAMBDA * DELTA T TO ACHIEVE THE
        # C * DESIRED RELATIVE ACCURACY ON THE TIME STEP
        TIMACC = 1e-3
        X1 = 1.0
        X2 = 0.0
        RES = 1.0
        while abs(RES) > 1.0e-2 * TIMACC:
            self.EIGTIM = 0.5 * (X1 + X2)
            Y1 = (1.0 - (1.0 - self.theta_method) * self.EIGTIM) / (
                1.0 + self.theta_method * self.EIGTIM
            )
            Y2 = np.exp(-self.EIGTIM)
            RES = abs((Y1 / Y2) - 1.0) - TIMACC
            if RES >= 0.0:
                X1 = self.EIGTIM
            elif RES < 0.0:
                X2 = self.EIGTIM

    # end Aprior

    def mass_energy_balance(self, simulation):

        """
        Function that operform mass and energy balance on conductor (cdp, 09/2020)
        """

        # Alias
        interf_flag = self.dict_df_coupling["contact_perimeter_flag"]

        self.mass_balance = 0.0  # mass balance initialization (cdp, 09/2020)
        self.energy_balance = 0.0  # energy balance initialization (cdp, 09/2020)
        self.inner_pow = 0.0
        self.outer_pow = 0.0
        for fluid_comp in self.inventory["FluidComponent"].collection:
            # Mass balance (cdp, 09/2020)
            self.mass_balance = self.mass_balance + self.time_step * (
                fluid_comp.coolant.dict_node_pt["mass_flow_rate"][0]
                - fluid_comp.coolant.dict_node_pt["mass_flow_rate"][-1]
            )
            # Energy balance: sum(mdot_inl*(w_inl + v_inl^2/2) - \
            # mdot_out*(w_out + v_out^2/2)) (cdp, 09/2020)
            self.energy_balance = self.energy_balance + self.time_step * (
                fluid_comp.coolant.dict_node_pt["mass_flow_rate"][0]
                * (
                    fluid_comp.coolant.dict_node_pt["total_enthalpy"][0]
                    + fluid_comp.coolant.dict_node_pt["velocity"][0] ** 2 / 2.0
                )
                - fluid_comp.coolant.dict_node_pt["mass_flow_rate"][-1]
                * (
                    fluid_comp.coolant.dict_node_pt["total_enthalpy"][-1]
                    + fluid_comp.coolant.dict_node_pt["velocity"][-1] ** 2 / 2.0
                )
            )
            self.inner_pow = self.inner_pow + fluid_comp.coolant.dict_node_pt[
                "mass_flow_rate"
            ][0] * (
                fluid_comp.coolant.dict_node_pt["total_enthalpy"][0]
                + fluid_comp.coolant.dict_node_pt["velocity"][0] ** 2 / 2.0
            )
            self.outer_pow = self.outer_pow + fluid_comp.coolant.dict_node_pt[
                "mass_flow_rate"
            ][-1] * (
                fluid_comp.coolant.dict_node_pt["total_enthalpy"][-1]
                + fluid_comp.coolant.dict_node_pt["velocity"][-1] ** 2 / 2.0
            )
        # End for fluid_comp.
        for jacket in self.inventory["JacketComponent"].collection:
            if (
                abs(interf_flag.loc[
                        simulation.environment.KIND, jacket.identifier
                    ]
                ) == 1
            ):
                key = f"{simulation.environment.KIND}_{jacket.identifier}"
                self.energy_balance = (
                    self.energy_balance
                    + self.time_step
                    * jacket.inputs["Outer_perimeter"]
                    * np.sum(
                        (
                            self.dict_Gauss_pt["HTC"]["env_sol"][key]["conv"]
                            + self.dict_Gauss_pt["HTC"]["env_sol"][key]["rad"]
                        )
                        * self.grid_features["delta_z"]
                        * (
                            simulation.environment.inputs["Temperature"]
                            - jacket.dict_Gauss_pt["temperature"]
                        )
                    )
                )
                self.inner_pow = self.inner_pow + jacket.inputs[
                    "Outer_perimeter"
                ] * np.sum(
                    (
                        self.dict_Gauss_pt["HTC"]["env_sol"][key]["conv"]
                        + self.dict_Gauss_pt["HTC"]["env_sol"][key]["rad"]
                    )
                    * self.grid_features["delta_z"]
                    * (
                        simulation.environment.inputs["Temperature"]
                        - jacket.dict_Gauss_pt["temperature"]
                    )
                )
            # End if.
        # End jacket.
        # Energy balance to check the correct management of SolidComponent forced \
        # initial temperature distribution (cdp, 12/2020)
        # E_residual = (self.E_sol_ini - self.E_sol_fin) - \
        #              (self.enthalpy_inl - self.enthalpy_out)
        # print(f"E_sol_ini = {self.E_sol_ini} J\n")
        # print(f"E_sol_fin = {self.E_sol_fin} J\n")
        # print(f"E_str_ini = {self.E_str_ini} J\n")
        # print(f"E_str_fin = {self.E_str_fin} J\n")
        # print(f"E_jk_ini = {self.E_jk_ini} J\n")
        # print(f"E_jk_fin = {self.E_jk_fin} J\n")
        # print(f"enthalpy_inl = {self.enthalpy_inl} J\n")
        # print(f"enthalpy_out = {self.enthalpy_out} J\n")
        # print(f"enthalpy_balance = {self.enthalpy_balance} J\n")
        # print(f"E_residual = {E_residual} J\n")

        print(f"Energy balance = {self.energy_balance} J\n")
        # Print outer inner power ration only if inner_pow is not 0.0
        if self.inner_pow != 0.0:
            print(f"Outer inner power ratio % = {1e2*self.outer_pow/self.inner_pow} ~")

    # end method Mass_Energy_balance (cdp, 09/2020)

    def compute_radiative_heat_exhange_jk(self):
        """Method that evaluates the radiative heat exchanged by radiation between jackets."""
        # Nested loop on jackets.
        for rr, jk_r in enumerate(self.inventory["JacketComponent"].collection):
            for _, jk_c in enumerate(
                self.inventory["JacketComponent"].collection[rr + 1 :]
            ):
                if (
                    abs(
                        self.dict_df_coupling["HTC_choice"].at[
                            jk_r.identifier, jk_c.identifier
                        ]
                    )
                    == 3
                ):
                    self.heat_rad_jk[f"{jk_r.identifier}_{jk_c.identifier}"] = (
                        self.dict_df_coupling["contact_perimeter"].at[
                            jk_r.identifier, jk_c.identifier
                        ]
                        * self.grid_features["delta_z"]
                        * self.dict_Gauss_pt["HTC"]["sol_sol"]["rad"][
                            self.dict_topology["sol_sol"][jk_r.identifier][
                                jk_c.identifier
                            ]
                        ]
                        * (
                            jk_r.dict_Gauss_pt["temperature"]
                            - jk_c.dict_Gauss_pt["temperature"]
                        )
                    )  # W
                # End if abs.
            # End for cc.
        # End for rr.

    # End method _compute_radiative_heat_exhange_jk.

    def compute_heat_exchange_jk_env(self, environment):
        """Method that computes the heat exchange between the outer surface of the conductor and the environment by convection and or radiation.

        Args:
            environment ([type]): [description]
        """

        # Alias
        interf_flag = self.dict_df_coupling["contact_perimeter_flag"]

        for jk in self.inventory["JacketComponent"].collection:
            if (
                abs(interf_flag.at[
                        environment.KIND, jk.identifier
                    ]
                ) == 1
            ):
                key = f"{environment.KIND}_{jk.identifier}"
                self.heat_exchange_jk_env[key] = (
                    self.dict_df_coupling["contact_perimeter"].at[
                        environment.KIND, jk.identifier
                    ]
                    * self.grid_features["delta_z"]
                    * (
                        self.dict_Gauss_pt["HTC"]["env_sol"][key]["conv"]
                        + self.dict_Gauss_pt["HTC"]["env_sol"][key]["rad"]
                    )
                    * (
                        environment.inputs["Temperature"]
                        - jk.dict_Gauss_pt["temperature"]
                    )
                )  # W
            # End if interf_flag.
        # End for jk.

    # End method _compute_heat_exchange_environment.

    def load_user_defined_quantity(self, simulation, key, sheet_name):
        """[summary]

        Args:
            conductor ([type]): [description]

        Returns:
            [type]: [description]
        """
        file_extension = self.file_input[key].split(".")[1]
        # Build the path to the file
        fname = os.path.join(self.BASE_PATH, self.file_input[key])
        # Call function used to open the file
        return simulation.func_open_aux[file_extension](
            fname, sheet_name, simulation.default_vals[file_extension]
        )

        # End method load_user_defined_quantity.
